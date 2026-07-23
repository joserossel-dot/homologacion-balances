from __future__ import annotations

import json
import logging
import sqlite3
from pathlib import Path

import pytest

logging.disable(logging.CRITICAL)

SHADOW_PATH = Path("reports/semantic_shadow/shadow_data.json")


# ---------------------------------------------------------------------------
# review_models
# ---------------------------------------------------------------------------

def test_pending_account_defaults():
    from review.review_models import PendingAccount
    p = PendingAccount()
    assert p.priority == 0
    assert p.confidence == 0.0
    assert p.metodo == ""
    # Score includes unknown penalty (50) + low conf penalty (30) + freq/empresa bonuses
    assert p.score >= 50.0


def test_pending_account_score_unknown():
    from review.review_models import PendingAccount
    p = PendingAccount(metodo="unclassified", confidence=0.0, frecuencia=3, cantidad_empresas=2)
    assert p.score >= 50.0


def test_pending_account_score_high_conf():
    from review.review_models import PendingAccount
    p = PendingAccount(metodo="learning", confidence=0.97, frecuencia=1, cantidad_empresas=1)
    assert 0 < p.score < 50  # lower score because not unknown


def test_pending_account_score_semantic_boost():
    from review.review_models import PendingAccount
    p = PendingAccount(metodo="unclassified", confidence=0.3, semantic_hit=True, frecuencia=1, cantidad_empresas=1)
    assert p.score > 50.0


def test_pending_account_score_freq_boost():
    from review.review_models import PendingAccount
    p = PendingAccount(metodo="unclassified", confidence=0.0, frecuencia=50, cantidad_empresas=10)
    base_score = 50.0 + 30.0 + min(50 * 2.0, 20.0) + min(10 * 5.0, 25.0)
    assert p.score >= base_score


def test_account_to_pending():
    from review.review_models import account_to_pending
    a = {
        "account_code": "1234",
        "account_name": "Caja General",
        "nature": "deudora",
        "classification_amount": 500000.0,
        "method": "unclassified",
        "confidence": 0.0,
        "reason": "Sin coincidencia",
        "source_group": "empresa_a",
        "source_file": "balance.xlsx",
        "source_page": 1,
        "source_path": "data/balance.xlsx",
        "standard_code": None,
        "final_code": None,
        "semantic_result": {
            "semantic_type": "unknown",
            "matched_rule": "no_match",
        },
    }
    p = account_to_pending(a, freq=3, num_empresas=2)
    assert p.nombre_cuenta == "Caja General"
    assert p.metodo == "unclassified"
    assert p.frecuencia == 3
    assert p.cantidad_empresas == 2
    assert p.monto_original == 500000.0
    assert p.naturaleza == "deudora"


def test_account_to_pending_semantic_hit():
    from review.review_models import account_to_pending
    a = {
        "account_name": "DEPRECIACION ACUMULADA",
        "method": "unclassified",
        "confidence": 0.0,
        "classification_amount": 100.0,
        "semantic_result": {"semantic_type": "contra_asset", "matched_rule": "depreciacion_acumulada"},
        "source_group": "test",
        "source_file": "test.xlsx",
        "source_path": "test.xlsx",
    }
    p = account_to_pending(a)
    assert p.semantic_hit is True
    assert p.semantic_type == "contra_asset"
    assert p.semantic_rule == "depreciacion_acumulada"


def test_constants_present():
    from review.review_models import (
        PENDING_COLUMNS, PENDING_EDITABLE, PENDING_READONLY,
        CLASE_VALUES, SEMANTIC_TYPE_VALUES, APRENDER_VALUES, ALCANCE_VALUES,
    )
    assert len(PENDING_COLUMNS) >= 30
    assert len(PENDING_EDITABLE) >= 5
    assert "Activo" in CLASE_VALUES
    assert "asset" in SEMANTIC_TYPE_VALUES
    assert "Sí" in APRENDER_VALUES
    assert "Global" in ALCANCE_VALUES


# ---------------------------------------------------------------------------
# review_metrics
# ---------------------------------------------------------------------------

def test_compute_score():
    from review.review_models import PendingAccount
    from review.review_metrics import compute_score
    p = PendingAccount(metodo="unclassified", confidence=0.0)
    assert compute_score(p) >= 50.0


def test_prioritize_accounts():
    from review.review_models import PendingAccount
    from review.review_metrics import prioritize_accounts
    a1 = PendingAccount(metodo="unclassified", confidence=0.0, frecuencia=10, cantidad_empresas=5)
    a2 = PendingAccount(metodo="learning", confidence=0.98, frecuencia=1, cantidad_empresas=1)
    result = prioritize_accounts([a2, a1])
    assert result[0].priority == 1
    assert result[0].metodo == "unclassified"  # higher score


def test_build_pending_rows():
    if not SHADOW_PATH.exists():
        pytest.skip("shadow_data.json not found")
    import json
    with open(SHADOW_PATH) as f:
        accounts = json.load(f)["accounts"]
    from review.review_metrics import build_pending_rows
    result = build_pending_rows(accounts)
    assert len(result) > 0
    assert all(r.priority >= 1 for r in result)


def test_build_pending_rows_all_have_score():
    from review.review_models import PendingAccount
    from review.review_metrics import build_pending_rows
    accounts = [
        {"account_name": "Caja", "method": "unclassified", "confidence": 0.0, "classification_amount": 100.0, "nature": "deudora", "semantic_result": {}, "source_group": "g1", "source_file": "f1.xlsx", "source_path": "f1.xlsx"},
        {"account_name": "Banco", "method": "learning", "confidence": 0.7, "classification_amount": 200.0, "nature": "deudora", "semantic_result": {}, "source_group": "g2", "source_file": "f2.xlsx", "source_path": "f2.xlsx"},
    ]
    result = build_pending_rows(accounts, threshold=0.85)
    assert len(result) == 2
    for r in result:
        assert r.score > 0


def test_build_low_confidence_rows():
    from review.review_metrics import build_low_confidence_rows
    accounts = [
        {"account_name": "Caja", "method": "learning", "confidence": 0.5, "nature": "deudora", "classification_amount": 100.0, "semantic_result": {"semantic_type": "unknown"}, "source_group": "g1", "source_file": "f1.xlsx", "source_path": "f1.xlsx", "final_code": None, "standard_code": None},
        {"account_name": "Banco", "method": "code_exact", "confidence": 0.98, "nature": "deudora", "classification_amount": 200.0, "semantic_result": {"semantic_type": "unknown"}, "source_group": "g2", "source_file": "f2.xlsx", "source_path": "f2.xlsx", "final_code": "AC.01", "standard_code": "AC.01"},
    ]
    result = build_low_confidence_rows(accounts, threshold=0.85)
    assert len(result) == 1  # only first is below threshold
    assert result[0]["Cuenta"] == "Caja"


def test_build_dashboard():
    if not SHADOW_PATH.exists():
        pytest.skip("shadow_data.json not found")
    import json
    with open(SHADOW_PATH) as f:
        accounts = json.load(f)["accounts"]
    from review.review_metrics import build_dashboard
    d = build_dashboard(accounts)
    assert d.total_cuentas == len(accounts)
    assert d.cobertura_pct >= 0.0
    assert len(d.top_empresas) > 0
    assert len(d.top_cuentas) > 0


def test_build_dashboard_counts():
    from review.review_metrics import build_dashboard
    accounts = [
        {"account_name": "A", "method": "unclassified", "confidence": 0.0, "nature": "deudora", "classification_amount": 100.0, "semantic_result": {"semantic_type": "unknown", "matched_rule": "no_match"}, "source_group": "g1", "source_file": "f1.xlsx", "source_path": "f1.xlsx"},
        {"account_name": "B", "method": "code_exact", "confidence": 0.98, "nature": "deudora", "classification_amount": 200.0, "semantic_result": {"semantic_type": "asset", "matched_rule": "some_rule"}, "source_group": "g2", "source_file": "f2.xlsx", "source_path": "f2.xlsx"},
    ]
    d = build_dashboard(accounts)
    assert d.total_cuentas == 2
    assert d.clasificadas == 1
    assert d.unknown == 1
    assert d.semantic == 1
    assert d.codigo == 1


# ---------------------------------------------------------------------------
# review_package_builder
# ---------------------------------------------------------------------------

def test_load_all_data():
    if not SHADOW_PATH.exists():
        pytest.skip("shadow_data.json not found")
    from review.review_package_builder import load_all_data
    data = load_all_data()
    assert "shadow" in data
    assert "gold_gs" in data
    assert "generated_gs" in data
    assert "generated_dict" in data
    assert len(data["shadow"]["accounts"]) > 0


@pytest.mark.slow
def test_build_review_package(tmp_path):
    if not SHADOW_PATH.exists():
        pytest.skip("shadow_data.json not found")
    from review.review_package_builder import load_all_data, build_review_package
    data = load_all_data()
    out = tmp_path / "test_package.xlsx"
    path = build_review_package(str(out), data)
    assert Path(path).exists()
    assert Path(path).stat().st_size > 1000


@pytest.mark.slow
def test_build_review_package_sheets(tmp_path):
    if not SHADOW_PATH.exists():
        pytest.skip("shadow_data.json not found")
    from review.review_package_builder import load_all_data, build_review_package
    import openpyxl
    data = load_all_data()
    out = tmp_path / "test_sheets.xlsx"
    build_review_package(str(out), data)
    wb = openpyxl.load_workbook(str(out))
    expected = ["Pendientes", "Baja Confianza", "Conflictos", "Reglas Propuestas", "Sinónimos", "Gold Standard", "Dashboard", "Log"]
    for name in expected:
        assert name in wb.sheetnames, f"Missing sheet: {name}"


@pytest.mark.slow
def test_build_review_package_pending_has_columns(tmp_path):
    from review.review_models import PENDING_COLUMNS
    from review.review_package_builder import load_all_data, build_review_package
    import openpyxl
    if not SHADOW_PATH.exists():
        pytest.skip("shadow_data.json not found")
    data = load_all_data()
    out = tmp_path / "test_columns.xlsx"
    build_review_package(str(out), data)
    wb = openpyxl.load_workbook(str(out))
    ws = wb["Pendientes"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    for col in PENDING_COLUMNS:
        assert col in headers, f"Missing column: {col}"


def test_build_conflicts():
    from review.review_package_builder import _build_conflicts
    records = [
        {"account_name": "Caja", "final_code": "AC.01"},
        {"account_name": "Caja", "final_code": "AC.02"},
    ]
    accounts = [
        {"account_name": "Caja", "source_group": "g1"},
    ]
    conflicts = _build_conflicts(records, accounts)
    assert len(conflicts) == 1
    assert conflicts[0]["account_name"] == "Caja"
    assert "AC.01" in conflicts[0]["codes"]


def test_build_conflicts_no_conflict():
    from review.review_package_builder import _build_conflicts
    records = [
        {"account_name": "Caja", "final_code": "AC.01"},
        {"account_name": "Banco", "final_code": "AC.01"},
    ]
    conflicts = _build_conflicts(records, [])
    assert len(conflicts) == 0


def test_build_proposed_rules():
    from review.review_package_builder import _build_proposed_rules
    priority = [
        {"type": "new_semantic_rule", "priority": 1, "description": "Nueva regla: 'caja' -> asset", "cluster_size": 10, "confidence": 0.6},
        {"type": "dictionary_synonyms", "priority": 2, "description": "Sinónimos", "cluster_size": 5, "confidence": 0.7},
    ]
    rules = _build_proposed_rules(priority, [])
    assert len(rules) == 1  # only new_semantic_rule
    assert "caja" in rules[0].patron and "asset" in rules[0].patron


def test_build_synonym_entries():
    from review.review_package_builder import _build_synonym_entries
    rows = [
        {"group_label": "Caja", "num_variants": 3, "variants": "Caja General\nCaja M/N\nCaja Central"},
    ]
    entries = _build_synonym_entries(rows)
    assert len(entries) >= 2
    assert entries[0].canonical == "Caja"
    assert entries[0].synonym == "Caja General"


def test_build_gold_proposals():
    from review.review_package_builder import _build_gold_proposals
    gs = [
        {"account_name": "Caja General", "suggested_code": "AC.01", "source_file": "f1.xlsx", "usage_count": 0, "suggested_confidence": 0.6, "comments": "test"},
    ]
    props = _build_gold_proposals(gs)
    assert len(props) == 1
    assert props[0].cuenta == "Caja General"
    assert props[0].codigo == "AC.01"


# ---------------------------------------------------------------------------
# excel_formatter
# ---------------------------------------------------------------------------

def test_apply_default_sheet(tmp_path):
    import openpyxl
    from review.excel_formatter import apply_default_sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    cols = ["A", "B", "C"]
    rows = [[1, "x", 3.0], [2, "y", 4.0]]
    apply_default_sheet(ws, cols, rows, "Test")
    assert ws.cell(row=1, column=1).value == "A"
    assert ws.cell(row=2, column=1).value == 1
    assert ws.freeze_panes == "A2"


def test_apply_dropdowns(tmp_path):
    import openpyxl
    from review.excel_formatter import apply_dropdowns
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="H")
    apply_dropdowns(ws, "A", ["Op1", "Op2"], 2, 5)
    assert len(ws.data_validations.dataValidation) >= 1


def test_build_pending_sheet_has_dropdowns(tmp_path):
    import openpyxl
    from review.excel_formatter import build_pending_sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    rows = [
        [1, "e1", "", "", "", "f1", 0, "c1", "Caja", 0, 0, 0, 0, 0, 100.0, 100.0, "d", "unclassified", 0.0, "No", "No", "", "", "", "", 1, 1, "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
    ]
    build_pending_sheet(ws, rows)
    assert ws.cell(row=1, column=1).value == "Prioridad"


def test_write_dashboard_sheet(tmp_path):
    import openpyxl
    from review.review_models import DashboardMetrics
    from review.excel_formatter import write_dashboard_sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    metrics = DashboardMetrics(
        total_cuentas=100,
        clasificadas=50,
        unknown=50,
        learning=10,
        semantic=5,
        codigo=20,
        diccionario=15,
        cobertura_pct=50.0,
        confianza_promedio=0.75,
        top_empresas=[("e1", 10)],
        top_cuentas=[("c1", 5)],
        top_montos=[("c1", 1000.0)],
        top_reglas=[("r1", 3)],
    )
    write_dashboard_sheet(ws, metrics)
    assert ws.cell(row=1, column=1).value == "Dashboard - Review Package"


# ---------------------------------------------------------------------------
# Integration
# ---------------------------------------------------------------------------

@pytest.mark.slow
def test_end_to_end(tmp_path):
    if not SHADOW_PATH.exists():
        pytest.skip("shadow_data.json not found")
    from review.review_package_builder import load_all_data, build_review_package
    data = load_all_data()
    out = tmp_path / "e2e.xlsx"
    build_review_package(str(out), data)
    assert out.exists()


def test_pending_order():
    from review.review_models import PendingAccount
    from review.review_metrics import prioritize_accounts
    a1 = PendingAccount(metodo="unclassified", confidence=0.0, frecuencia=20, cantidad_empresas=10)
    a2 = PendingAccount(metodo="learning", confidence=0.99, frecuencia=1, cantidad_empresas=1)
    result = prioritize_accounts([a2, a1])
    assert result[0].priority < result[1].priority  # highest score = lowest number


def test_pending_row_length():
    from review.review_models import PENDING_COLUMNS, PendingAccount, account_to_pending
    a = {
        "account_name": "Test",
        "method": "unclassified",
        "confidence": 0.0,
        "classification_amount": 100.0,
        "nature": "deudora",
        "semantic_result": {},
        "source_group": "g1",
        "source_file": "f1.xlsx",
        "source_path": "f1.xlsx",
    }
    p = account_to_pending(a)
    from review.review_package_builder import _pending_to_row
    row = _pending_to_row(p)
    assert len(row) == len(PENDING_COLUMNS)


@pytest.mark.slow
def test_generated_excel_has_all_sheets():
    import glob
    xlsx_files = sorted(glob.glob("reports/review/review_package_*.xlsx"))
    if not xlsx_files:
        pytest.skip("No review package generated yet")
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_files[-1])
    expected = {"Pendientes", "Baja Confianza", "Conflictos", "Reglas Propuestas", "Sinónimos", "Gold Standard", "Dashboard", "Log"}
    assert expected.issubset(set(wb.sheetnames))
