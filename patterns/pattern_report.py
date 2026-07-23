from __future__ import annotations

import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Any

from patterns.pattern_matcher import PatternMatch

log = logging.getLogger(__name__)


class PatternReport:
    def __init__(self, output_dir: str | Path = "reports/patterns") -> None:
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def generate_all(
        self,
        results: list[dict[str, Any]],
        coverage: dict[str, Any],
    ) -> dict[str, Path]:
        paths: dict[str, Path] = {}
        paths["coverage_md"] = self.generate_coverage_md(results, coverage)
        try:
            paths["statistics"] = self.generate_statistics_xlsx(
                results, coverage
            )
        except Exception as e:
            log.warning("Could not generate statistics xlsx: %s", e)
        try:
            paths["unmatched"] = self.generate_unmatched_xlsx(results)
        except Exception as e:
            log.warning("Could not generate unmatched xlsx: %s", e)
        try:
            paths["examples"] = self.generate_examples_xlsx(results)
        except Exception as e:
            log.warning("Could not generate examples xlsx: %s", e)
        try:
            paths["json"] = self.generate_json(results, coverage)
        except Exception as e:
            log.warning("Could not generate json: %s", e)
        return paths

    def generate_coverage_md(
        self,
        results: list[dict[str, Any]],
        coverage: dict[str, Any],
    ) -> Path:
        path = self.output_dir / "pattern_coverage.md"
        lines: list[str] = [
            "# Pattern Engine — Informe de Cobertura\n",
            f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n",
            "---\n",
        ]
        lines.append("## Resumen General\n")
        lines.append(f"- Total cuentas analizadas: {coverage['total_accounts']}")
        lines.append(
            f"- Cuentas no clasificadas: {coverage['total_unclassified']}"
        )
        lines.append(
            f"- Coinciden con al menos un patrón: {coverage['matched_all']} "
            f"({coverage['coverage_all_pct']}%)"
        )
        lines.append(
            f"- No clasificadas con patrón: {coverage['matched_unclassified']} "
            f"({coverage['coverage_unclassified_pct']}%)"
        )
        lines.append("")

        lines.append("## Cobertura por Familia (todas las cuentas)\n")
        lines.append("| Familia | Coincidencias |")
        lines.append("|---------|---------------|")
        for family, count in coverage.get("family_counts", {}).items():
            lines.append(f"| {family} | {count} |")
        lines.append("")

        lines.append(
            "## Cobertura por Familia (solo no clasificadas)\n"
        )
        lines.append("| Familia | Coincidencias |")
        lines.append("|---------|---------------|")
        for family, count in coverage.get(
            "family_unclassified", {}
        ).items():
            lines.append(f"| {family} | {count} |")
        lines.append("")

        lines.append("## Familias de Patrones\n")
        families_seen: set[str] = set()
        for r in results:
            for m in r["matches"]:
                if m.family not in families_seen:
                    families_seen.add(m.family)
                    lines.append(
                        f"- **{m.family}** (`{m.pattern_id}`): "
                        f"confianza={m.confidence:.0%}, "
                        f"prioridad={m.priority}"
                    )

        lines.append("")
        lines.append(
            "## Recomendaciones\n"
        )
        if coverage["coverage_unclassified_pct"] < 50:
            lines.append(
                "- La cobertura en no clasificadas es baja "
                f"({coverage['coverage_unclassified_pct']}%). "
                "Se recomienda agregar más familias de patrones."
            )
        else:
            lines.append(
                "- La cobertura es adecuada. "
                "Se recomienda refinar patrones existentes."
            )
        if coverage["coverage_unclassified_pct"] > 80:
            lines.append(
                "- Alta cobertura: considerar integrar patrones "
                "al pipeline de clasificación."
            )

        path.write_text("\n".join(lines) + "\n", encoding="utf-8")
        log.info("Coverage report written to %s", path)
        return path

    def generate_statistics_xlsx(
        self,
        results: list[dict[str, Any]],
        coverage: dict[str, Any],
    ) -> Path | None:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            log.warning("openpyxl not available")
            return None

        path = self.output_dir / "pattern_statistics.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Resumen"
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )

        headers = ["Métrica", "Valor"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        metrics = [
            ("Total cuentas", coverage["total_accounts"]),
            ("No clasificadas", coverage["total_unclassified"]),
            ("Matched (todas)", coverage["matched_all"]),
            ("Matched (no clasificadas)", coverage["matched_unclassified"]),
            ("Cobertura total", f"{coverage['coverage_all_pct']}%"),
            (
                "Cobertura no clasificadas",
                f"{coverage['coverage_unclassified_pct']}%",
            ),
            ("Familias", len(coverage["family_counts"])),
        ]
        for row, (metric, value) in enumerate(metrics, 2):
            ws.cell(row=row, column=1, value=metric)
            ws.cell(row=row, column=2, value=value)

        ws2 = wb.create_sheet("Por Familia")
        ws2.cell(row=1, column=1, value="Familia")
        ws2.cell(row=1, column=2, value="Coincidencias (total)")
        ws2.cell(row=1, column=3, value="Coincidencias (no clasif)")
        for col in range(1, 4):
            cell = ws2.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill

        all_families = set(coverage.get("family_counts", {}).keys())
        all_families |= set(
            coverage.get("family_unclassified", {}).keys()
        )
        for row, family in enumerate(sorted(all_families), 2):
            ws2.cell(row=row, column=1, value=family)
            ws2.cell(
                row=row,
                column=2,
                value=coverage.get("family_counts", {}).get(family, 0),
            )
            ws2.cell(
                row=row,
                column=3,
                value=coverage.get("family_unclassified", {}).get(
                    family, 0
                ),
            )

        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 20
        ws2.column_dimensions["A"].width = 30
        ws2.column_dimensions["B"].width = 30
        ws2.column_dimensions["C"].width = 30

        wb.save(str(path))
        log.info("Statistics written to %s", path)
        return path

    def generate_unmatched_xlsx(
        self, results: list[dict[str, Any]]
    ) -> Path | None:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            return None

        path = self.output_dir / "pattern_unmatched.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sin Coincidencia"
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="C00000", end_color="C00000", fill_type="solid"
        )

        headers = [
            "account_name",
            "normalized_name",
            "method",
            "nature",
            "source_group",
        ]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        row = 2
        for r in results:
            if r["total_matches"] == 0:
                acct = r["account"]
                ws.cell(row=row, column=1, value=acct.get("account_name", ""))
                ws.cell(row=row, column=2, value=r["normalized_name"])
                ws.cell(row=row, column=3, value=acct.get("method", ""))
                ws.cell(row=row, column=4, value=acct.get("nature", ""))
                ws.cell(row=row, column=5, value=acct.get("source_group", ""))
                row += 1

        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col_idx)].width = 40

        wb.save(str(path))
        log.info("Unmatched written to %s", path)
        return path

    def generate_examples_xlsx(
        self, results: list[dict[str, Any]]
    ) -> Path | None:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            return None

        path = self.output_dir / "pattern_examples.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ejemplos"
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="548235", end_color="548235", fill_type="solid"
        )

        headers = [
            "account_name",
            "normalized_name",
            "family",
            "pattern_id",
            "confidence",
            "explanation",
            "method",
            "is_classified",
        ]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        row = 2
        for r in results:
            for m in r["matches"]:
                acct = r["account"]
                ws.cell(row=row, column=1, value=acct.get("account_name", ""))
                ws.cell(row=row, column=2, value=r["normalized_name"])
                ws.cell(row=row, column=3, value=m.family)
                ws.cell(row=row, column=4, value=m.pattern_id)
                ws.cell(row=row, column=5, value=m.confidence)
                ws.cell(row=row, column=6, value=m.explanation)
                ws.cell(row=row, column=7, value=acct.get("method", ""))
                ws.cell(row=row, column=8, value="Sí" if r["is_classified"] else "No")
                row += 1

        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col_idx)].width = 40

        wb.save(str(path))
        log.info("Examples written to %s", path)
        return path

    def generate_json(
        self,
        results: list[dict[str, Any]],
        coverage: dict[str, Any],
    ) -> Path | None:
        path = self.output_dir / "pattern_analysis.json"

        serializable = self._make_serializable(results, coverage)
        with open(path, "w", encoding="utf-8") as f:
            json.dump(serializable, f, indent=2, ensure_ascii=False)
        log.info("JSON written to %s", path)
        return path

    def _make_serializable(
        self,
        results: list[dict[str, Any]],
        coverage: dict[str, Any],
    ) -> dict:
        ser_results: list[dict] = []
        for r in results:
            ser_matches = [
                {
                    "pattern_id": m.pattern_id,
                    "family": m.family,
                    "confidence": m.confidence,
                    "priority": m.priority,
                    "matched_patterns": m.matched_patterns,
                    "tokens_found": m.tokens_found,
                    "explanation": m.explanation,
                }
                for m in r["matches"]
            ]
            acct = r["account"]
            entry: dict = {
                "account_name": acct.get("account_name", ""),
                "normalized_name": r["normalized_name"],
                "total_matches": r["total_matches"],
                "is_classified": r["is_classified"],
                "matches": ser_matches,
                "method": acct.get("method", ""),
                "nature": acct.get("nature", ""),
                "source_group": acct.get("source_group", ""),
            }
            if r["best_match"]:
                bm = r["best_match"]
                entry["best_match"] = {
                    "pattern_id": bm.pattern_id,
                    "family": bm.family,
                    "confidence": bm.confidence,
                }
            ser_results.append(entry)

        return {
            "generated_at": datetime.now().isoformat(),
            "total_results": len(results),
            "coverage": coverage,
            "results": ser_results,
        }
