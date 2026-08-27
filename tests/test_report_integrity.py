from io import BytesIO
from types import SimpleNamespace

import pandas as pd
import pytest
from openpyxl import load_workbook
from streamlit.testing.v1 import AppTest

from app_validacion import _codigo_compatible_con_origen, _control_emision
from pipeline.homologation_pipeline import HomologationPipeline
from reporting_integrity import catalogo_local, conciliar_resultados


def cuenta(nombre, monto, origen, codigo):
    return dict(nombre_original=nombre, codigo_original=nombre, monto=monto,
                origen_columna=origen, origen_columna_efectiva=origen,
                codigo_clasificado=codigo, es_total=False,
                requiere_revision=False, metodo='validacion_humana', confianza=1.0,
                nombre_revision_usuario='')


def afuminsal(mal=True):
    # Importes del reporte AFUMINSAL: la cuota social se había incluido en ER.04.
    return [
        cuenta('Activos', 29943255, 'activo', 'AC.01'),
        cuenta('Facturas por pagar', 3177262, 'pasivo', 'PC.01'),
        cuenta('Capital social', 15201792, 'pasivo', 'PAT.01'),
        cuenta('Resultados acumulados', 1653256, 'pasivo', 'PAT.03'),
        cuenta('Gastos administración', 16376428, 'perdida', 'ER.04'),
        cuenta('Gastos financieros', 19306, 'perdida', 'ER.09'),
        cuenta('Ingreso Cuotas Sociales', 21623055, 'ganancia', 'ER.04' if mal else 'ER.17'),
        cuenta('Ingresos financieros', 4089280, 'ganancia', 'ER.12'),
        cuenta('Otros ingresos', 594344, 'ganancia', 'ER.17'),
    ]


@pytest.mark.parametrize('codigo,columna,monto,permitido', [
    ('ER.04', 'ganancia', 100, False), ('ER.04', 'perdida', 100, True),
    ('ER.01', 'perdida', 100, False), ('ER.01', 'ganancia', 100, True),
    ('ER.04', 'ganancia', -100, True), ('ER.01', 'perdida', -100, True),
    ('ER.18', 'ganancia', 100, False), ('ER.18', 'perdida', 100, True),
    ('ER.14', 'perdida', 100, True), ('ER.14', 'ganancia', 100, True),
])
def test_naturaleza_individual_y_compartida(codigo, columna, monto, permitido):
    assert _codigo_compatible_con_origen(codigo, columna, monto) is permitido
    if monto > 0:
        assert HomologationPipeline._is_code_allowed_for_tipo(codigo, columna.upper()) is permitido


def test_afuminsal_no_confunde_cuadratura_con_resultado_correcto():
    rows = afuminsal()
    original = pd.DataFrame(rows)
    control = _control_emision(original, catalogo_local(), {'cuadra': True})
    assert not control['definitivo']
    assert control['resultado']['resultado_origen'] == 9910945
    assert control['resultado']['resultado_homologado'] == -33335165
    assert control['resultado']['diferencia'] == -43246110
    assert control['incidencias']['Cuenta'].tolist() == ['Ingreso Cuotas Sociales']
    pd.testing.assert_frame_equal(original, pd.DataFrame(rows))
    corrected = _control_emision(pd.DataFrame(afuminsal(False)), catalogo_local(), {'cuadra': True})
    assert corrected['definitivo']
    assert corrected['resultado']['diferencia'] == 0


def test_errores_compensados_siguen_bloqueados():
    df = pd.DataFrame([cuenta('Ingreso', 100, 'ganancia', 'ER.04'),
                       cuenta('Gasto', 100, 'perdida', 'ER.01')])
    control = _control_emision(df, catalogo_local(), {'cuadra': True})
    assert control['resultado']['diferencia'] == 0
    assert len(control['incidencias']) == 2
    assert not control['definitivo']


@pytest.mark.parametrize('cambio', [dict(requiere_revision=True),
    dict(codigo_clasificado=''), dict(codigo_clasificado='__EXCLUIR__'),
    dict(codigo_clasificado='ER.999'), dict(monto=float('nan'))])
def test_pendiente_o_excluida_no_se_certifica(cambio):
    rows = afuminsal(False)
    rows[-1].update(cambio)
    assert not _control_emision(pd.DataFrame(rows), catalogo_local(), {'cuadra': True})['definitivo']


def test_controles_impresos_no_suman_y_resultado_declarado_se_contrasta():
    rows = afuminsal(False)
    rows += [{**cuenta('SUMAS', 999999, 'ganancia', ''), 'es_total': True}]
    assert conciliar_resultados(rows, catalogo_local())['cuadra']
    rows += [cuenta('Utilidad del ejercicio', 123, 'pasivo', 'PAT.04')]
    result = conciliar_resultados(rows, catalogo_local())
    assert not result['cuadra']
    assert any('PAT.04' in msg for msg in result['problemas'])


def test_enforced_no_certifica_y_shadow_no_bloquea_por_si_solo():
    df = pd.DataFrame(afuminsal(False))
    assert not _control_emision(df, catalogo_local(), {'cuadra': True},
        SimpleNamespace(export_allowed=False, reasons=['Cobertura insuficiente']))['definitivo']
    assert _control_emision(df, catalogo_local(), {'cuadra': True},
        SimpleNamespace(export_allowed=True, reasons=['Shadow']))['definitivo']
    assert not _control_emision(df, catalogo_local(), {'cuadra': True},
        SimpleNamespace(export_allowed=False, reasons=[]))['definitivo']


@pytest.mark.parametrize('origen,monto', [('ganancia', 50), ('perdida', 50),
                                         ('ganancia', -50), ('perdida', -50)])
def test_resultado_mixto_conserva_signo_contable(origen, monto):
    result = conciliar_resultados([cuenta('Diferencia de cambio', monto, origen, 'ER.13')], catalogo_local())
    assert result['cuadra']
    assert result['resultado_homologado'] == (monto if origen == 'ganancia' else -monto)


def test_resultado_calculado_no_oculta_error_aunque_ambos_controles_coincidan():
    rows = afuminsal()
    rows += [cuenta('Utilidad calculada', 9910945, 'pasivo', 'PAT.04'),
             cuenta('Utilidad neta', 9910945, 'desconocido', 'ER.11')]
    result = conciliar_resultados(rows, catalogo_local())
    assert not result['cuadra']
    assert result['resultado_homologado'] == -33335165
    assert any('PAT.04' in p for p in result['problemas'])
    assert any('ER.11' in p for p in result['problemas'])


def test_atribuciones_controladora_y_no_controladores_no_duplican_resultado():
    rows = [
        cuenta('Ingresos', 110193, 'ganancia', 'ER.01'),
        cuenta('Costos y gastos', 113655, 'perdida', 'ER.02'),
        cuenta('Atribuible a propietarios de la controladora', 3488,
               'perdida', 'ER.20'),
        cuenta('Atribuible a participaciones no controladoras', 26,
               'ganancia', 'ER.21'),
    ]

    result = conciliar_resultados(rows, catalogo_local())

    assert result['cuadra']
    assert result['resultado_origen'] == -3462
    assert result['resultado_homologado'] == -3462


def review_app(rows):
    import streamlit as st
    import pandas as pd
    import app_validacion as app
    from reporting_integrity import catalogo_local
    st.session_state.setdefault('resultados', {'caso.pdf': pd.DataFrame(rows)})
    st.session_state.setdefault('diccionario', [])
    st.session_state.setdefault('correcciones', [])
    st.session_state.setdefault('persistidas', [])
    def persistir(**kwargs):
        st.session_state.persistidas.append(kwargs)
        return True
    def persistir_lote(items):
        st.session_state.persistidas.extend(items)
        return True
    prev, prev_lote = app._persistir_validacion, app._persistir_validaciones_lote
    app._persistir_validacion, app._persistir_validaciones_lote = persistir, persistir_lote
    try:
        app._tab_revision(st.session_state.resultados['caso.pdf'], catalogo_local(),
                          app.MotorHibridoLocal([]), 'caso.pdf')
    finally:
        app._persistir_validacion, app._persistir_validaciones_lote = prev, prev_lote


def all_view(rows):
    at = AppTest.from_function(review_app, args=(rows,)).run()
    assert not at.exception
    at.radio[0].set_value('Todas (incluye confirmadas y excluidas)').run()
    assert not at.exception
    return at


def test_ui_recupera_decision_manual_sin_pendientes():
    rows = [cuenta('Cuotas', 100, 'ganancia', 'ER.04'), cuenta('Cuotas', 100, 'ganancia', 'ER.12')]
    at = all_view(rows)
    selector = next(s for s in at.selectbox if s.label == 'Clasificación correcta')
    assert not any('ER.04' in option for option in selector.options)
    selector.set_value('ER.17').run()
    next(r for r in at.radio if r.label == '¿Aplicar esta clasificación?').set_value('Solo para este caso').run()
    next(b for b in at.button if b.label == '✅ Confirmar').click().run()
    assert not at.exception
    df = at.session_state.resultados['caso.pdf']
    assert df.at[0, 'codigo_clasificado'] == 'ER.17'
    assert df.at[1, 'codigo_clasificado'] == 'ER.12'
    assert at.session_state.historial_decisiones[0]['Clasificación anterior'] == 'ER.04'
    assert at.session_state.persistidas[0]['agregar_diccionario'] is False
    assert not at.session_state.diccionario


def test_ui_buscar_mas_no_permite_aprender_ingreso_como_gasto():
    at = all_view([cuenta('Cuotas', 100, 'ganancia', 'ER.17')])
    next(c for c in at.checkbox if 'Buscar más' in c.label).check().run()
    next(s for s in at.selectbox if s.label == 'Clasificación correcta').set_value('ER.04').run()
    next(b for b in at.button if b.label == '✅ Confirmar').click().run()
    assert not at.exception
    assert at.error
    assert at.session_state.resultados['caso.pdf'].at[0, 'codigo_clasificado'] == 'ER.17'
    assert not at.session_state.persistidas
    assert not at.session_state.diccionario


def test_ui_lote_incompatible_no_modifica_ninguna_cuenta():
    rows = [cuenta('Cuotas', 100, 'ganancia', 'ER.17'), cuenta('Gastos', 100, 'perdida', 'ER.04')]
    at = all_view(rows)
    next(b for b in at.button if 'Seleccionar todas' in b.label).click().run()
    next(s for s in at.selectbox if s.label.startswith('Clasificar todas')).set_value('ER.04').run()
    next(b for b in at.button if 'Confirmar lote' in b.label).click().run()
    assert not at.exception
    assert at.error
    pd.testing.assert_frame_equal(at.session_state.resultados['caso.pdf'], pd.DataFrame(rows))
    assert not at.session_state.persistidas


def test_ui_recupera_excluida():
    at = all_view([cuenta('Cuotas', 100, 'ganancia', '__EXCLUIR__')])
    next(s for s in at.selectbox if s.label == 'Clasificación correcta').set_value('ER.17').run()
    next(b for b in at.button if b.label == '✅ Confirmar').click().run()
    assert not at.exception
    assert at.session_state.resultados['caso.pdf'].at[0, 'codigo_clasificado'] == 'ER.17'


def test_ui_no_redefine_categoria_existente_para_evadir_control():
    at = all_view([cuenta('Cuotas', 100, 'ganancia', 'ER.17')])
    next(s for s in at.selectbox if s.label == 'Clasificación correcta').set_value('➕ NUEVA CATEGORÍA').run()
    next(t for t in at.text_input if t.label.startswith('Código (ej')).set_value('ER.04')
    next(t for t in at.text_input if t.label == 'Nombre de la categoría').set_value('Ingreso falso')
    next(s for s in at.selectbox if s.label == 'Categoría').set_value('resultado').run()
    next(b for b in at.button if b.label == '✅ Confirmar').click().run()
    assert not at.exception
    assert any('No se puede redefinir' in e.value for e in at.error)
    assert not at.session_state.persistidas


def test_ui_busqueda_vacia_no_oculta_acceso_a_confirmadas():
    at = AppTest.from_function(review_app, args=([cuenta('Cuotas', 100, 'ganancia', 'ER.17')],)).run()
    next(t for t in at.text_input if t.label == 'Buscar cuenta o código').set_value('Cuotas').run()
    assert not at.exception
    at.radio[0].set_value('Todas (incluye confirmadas y excluidas)').run()
    assert not at.exception
    assert any(s.label == 'Clasificación correcta' for s in at.selectbox)


def report_app(rows, depreciation="__resolved_without_adjustment__"):
    import streamlit as st
    import pandas as pd
    import app_validacion as app
    from extractor_metadata import MetadataEmpresa
    from reporting_integrity import catalogo_local
    st.session_state.setdefault('metadata_files', {'caso.pdf': MetadataEmpresa(moneda='$')})
    st.session_state.setdefault('company_periodos_seleccionados', ('2024',))
    # Estas pruebas históricas certifican el informe una vez resuelto el nuevo
    # control de depreciación desde notas.
    if depreciation == "__resolved_without_adjustment__":
        depreciation = {'mode': 'none'}
    depreciation_state = (
        {'caso.pdf': {'2024': depreciation}} if depreciation is not None else {}
    )
    st.session_state.setdefault('depreciation_reclassifications', depreciation_state)
    previous = app.st.download_button
    def capture(label, **kwargs):
        st.session_state['export_label'] = label
        st.session_state['export_kwargs'] = kwargs
    app.st.download_button = capture
    try:
        app._tab_balance(pd.DataFrame(rows), catalogo_local(), 'caso.pdf')
    finally:
        app.st.download_button = previous


def test_reporte_integra_depreciacion_de_notas_y_deja_trazabilidad():
    at = AppTest.from_function(
        report_app,
        args=(afuminsal(False), {
            'mode': 'notes', 'total': 500,
            'cost_of_sales': 0, 'administration': 500,
            'source': 'manual_notes',
        }),
    ).run(timeout=20)
    assert not at.exception
    assert 'BORRADOR' not in at.session_state.export_kwargs['file_name']
    wb = load_workbook(BytesIO(at.session_state.export_kwargs['data']), data_only=True)
    balance_rows = list(wb['Balance Normalizado'].iter_rows(values_only=True))
    administracion = next(row for row in balance_rows if row[0] == 'ER.04')
    depreciacion = next(row for row in balance_rows if row[0] == 'ER.07')
    assert administracion[2] == -16_375_928
    assert depreciacion[2] == -500
    summary_rows = list(wb['Resumen'].iter_rows(min_row=27, values_only=True))
    adjustment = next(
        row for row in summary_rows
        if row[3] == 'Depreciación del ejercicio informada desde notas'
    )
    assert adjustment[0] == 'ER.07'
    assert adjustment[4] is None
    assert adjustment[5] == -500
    controls = {row[0]: row[1] for row in wb['Control de emisión'].iter_rows(
        min_row=2, values_only=True,
    ) if row[0]}
    assert controls['Depreciación informada desde notas (2024)'] == 500
    assert controls['Rebaja de Gastos de Administración (2024)'] == 500


def test_reporte_pide_depreciacion_y_muestra_montos_sin_envio_intermedio():
    at = AppTest.from_function(
        report_app, args=(afuminsal(False), None),
    ).run(timeout=20)
    assert not at.exception
    tratamiento = next(s for s in at.selectbox if s.label == 'Tratamiento')
    tratamiento.select('notes').run(timeout=20)
    assert not at.exception
    assert {n.label for n in at.number_input} >= {
        'Depreciación total del período',
        'Incluida en Costo de Ventas',
        'Incluida en Gastos de Administración',
    }
    assert 'BORRADOR' in at.session_state.export_kwargs['file_name']


@pytest.mark.parametrize('mal', [True, False])
def test_reporte_real_ui_y_excel(mal, monkeypatch):
    monkeypatch.setenv('QUALITY_CONTROL_ENFORCE_EXPORT', 'false')
    at = AppTest.from_function(report_app, args=(afuminsal(mal),)).run(timeout=20)
    assert not at.exception
    download = at.session_state.export_kwargs
    assert ('BORRADOR' in download['file_name']) is mal
    wb = load_workbook(BytesIO(download['data']), data_only=True)
    assert ('BORRADOR' in wb['Balance Normalizado']['F1'].value) is mal
    assert wb['Balance Normalizado']['C7'].value == 'Monto Total ($)'
    assert 'Control de emisión' in wb.sheetnames
    rows = list(wb['Balance Normalizado'].iter_rows(min_row=8, max_row=7+len(catalogo_local()), values_only=True))
    gastos = next(row for row in rows if row[0] == 'ER.04')
    assert gastos[2] == (-37999483 if mal else -16376428)
    assert not any(m.label == 'Subtotal' and m.value == '52,613,358' for m in at.metric)
    next(b for b in at.button if b.label == 'Revisar o cambiar clasificaciones').click().run()
    assert not at.exception
    assert 'Cola de Revisión' in at.session_state.vista_trabajo_solicitada
