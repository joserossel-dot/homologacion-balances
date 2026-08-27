from io import BytesIO
import pandas as pd
from openpyxl import Workbook, load_workbook
from streamlit.testing.v1 import AppTest
from reporting_integrity import (
    catalogo_local, validar_reclasificacion_depreciacion,
)
from report_presentation import (
    add_report_sheets, apply_depreciation_reclassification, complete_catalog,
    ER_ORDER,
)
from test_report_integrity import report_app, afuminsal


def test_complete_catalog_includes_empty_categories_and_calculates_once():
    grouped = pd.DataFrame([
        dict(codigo_clasificado="ER.01", monto_total=100, num_cuentas=1),
        dict(codigo_clasificado="ER.02", monto_total=-20, num_cuentas=1),
        dict(codigo_clasificado="ER.04", monto_total=-10, num_cuentas=1),
        dict(codigo_clasificado="ER.10", monto_total=-7, num_cuentas=1),
        dict(codigo_clasificado="ER.17", monto_total=2, num_cuentas=1),
    ])
    original = grouped.copy(deep=True)
    table, formulas = complete_catalog(grouped, catalogo_local(), True)
    assert set(table.codigo_clasificado) == set(catalogo_local())
    values = table.set_index("codigo_clasificado").monto_total
    assert values["ER.03"] == 80
    assert values["ER.06"] == 70
    assert values["ER.19"] == 72
    assert values["ER.11"] == 65
    assert values["ANC.01.01"] == 0
    assert not set(formulas["ER.11"]).intersection({"ER.03", "ER.06", "ER.08", "ER.19", "ER.11"})
    assert table[table.categoria == "resultado"].codigo_clasificado.tolist() == list(ER_ORDER)
    pd.testing.assert_frame_equal(original, grouped)


def test_depreciation_from_notes_is_reclassified_without_changing_net_income():
    grouped = pd.DataFrame([
        dict(codigo_clasificado="ER.01", monto_total=12_000, num_cuentas=1),
        dict(codigo_clasificado="ER.02", monto_total=-8_000, num_cuentas=1),
        dict(codigo_clasificado="ER.04", monto_total=-2_000, num_cuentas=1),
    ])
    before = grouped.monto_total.sum()

    adjusted = apply_depreciation_reclassification(grouped, {
        "mode": "notes", "total": 500,
        "cost_of_sales": 350, "administration": 150,
    })
    values = adjusted.set_index("codigo_clasificado").monto_total

    assert values["ER.02"] == -7_650
    assert values["ER.04"] == -1_850
    assert values["ER.07"] == -500
    assert adjusted.monto_total.sum() == before
    assert grouped.set_index("codigo_clasificado").loc["ER.02", "monto_total"] == -8_000


def test_depreciation_notes_validation_requires_exact_and_available_allocation():
    assert validar_reclasificacion_depreciacion(
        500, 350, 150,
        costo_ventas_disponible=-8_000,
        gastos_administracion_disponible=-2_000,
    ) == []
    assert any("igual a la suma" in error for error in
               validar_reclasificacion_depreciacion(500, 350, 100))
    assert any("supera su saldo" in error for error in
               validar_reclasificacion_depreciacion(
                   500, 350, 150,
                   costo_ventas_disponible=-300,
                   gastos_administracion_disponible=-2_000,
               ))
    assert any("No existe un saldo" in error for error in
               validar_reclasificacion_depreciacion(
                   500, 0, 500,
                   costo_ventas_disponible=-8_000,
                   gastos_administracion_disponible=None,
               ))


def test_atribuciones_se_muestran_pero_no_entran_en_utilidad_calculada():
    grouped = pd.DataFrame([
        dict(codigo_clasificado="ER.01", monto_total=110193, num_cuentas=1),
        dict(codigo_clasificado="ER.02", monto_total=-113655, num_cuentas=1),
        dict(codigo_clasificado="ER.20", monto_total=-3488, num_cuentas=1),
        dict(codigo_clasificado="ER.21", monto_total=26, num_cuentas=1),
    ])

    table, formulas = complete_catalog(grouped, catalogo_local(), True)
    values = table.set_index("codigo_clasificado").monto_total

    assert values["ER.11"] == -3462
    assert values["ER.20"] == -3488
    assert values["ER.21"] == 26
    assert "ER.20" not in formulas["ER.11"]
    assert "ER.21" not in formulas["ER.11"]


def test_missing_income_is_not_reported_as_zero_profit():
    table, _ = complete_catalog(pd.DataFrame([
        dict(codigo_clasificado="AC.01", monto_total=100, num_cuentas=1),
    ]), catalogo_local(), False)
    net = table.set_index("codigo_clasificado").loc["ER.11"]
    assert pd.isna(net.monto_total)
    assert net.estado_presentacion == "No disponible"


def test_export_has_summary_metadata_units_formulas_and_full_income():
    at = AppTest.from_function(report_app, args=(afuminsal(False),)).run(timeout=20)
    assert not at.exception
    data = at.session_state.export_kwargs["data"]
    workbook = load_workbook(BytesIO(data), data_only=False)
    assert workbook.sheetnames[0] == "Resumen"
    summary = workbook["Resumen"]
    assert summary["B5"].value == "$"
    assert summary["B6"].value
    assert summary["B7"].value is None
    assert [summary.cell(26, column).value for column in range(1, 7)] == [
        "codigo_homologado", "nombre_homologado", "codigo_original",
        "nombre_original", "valor_extraido", "valor_homologado",
    ]
    detail = list(summary.iter_rows(min_row=27, max_row=35, max_col=6, values_only=True))
    assert len(detail) == len(afuminsal(False))
    gasto = next(row for row in detail if row[3] == "Gastos administración")
    assert gasto == ("ER.04", "Gastos de Administración", "Gastos administración",
                     "Gastos administración", 16376428, -16376428)
    assert summary.tables["CuentasClasificadas"].ref == "A26:F35"
    assert summary["E27"].number_format.startswith("#,##0")
    assert str(summary.print_title_rows) == "$1:$11"
    summary_rows = {r[0].value: r[1].value for r in summary}
    for label in ("TOTAL ACTIVOS", "TOTAL PASIVOS", "TOTAL PATRIMONIO", "Diferencia de cuadratura", "Cuadratura aritmética"):
        assert summary_rows[label].startswith("=")
    income = workbook["Estado de Resultados"]
    assert [r[0].value for r in income.iter_rows(min_row=3) if r[0].value] == list(ER_ORDER)
    assert income["C2"].value == "Importe ($)"
    assert income["C3"].number_format.startswith("#,##0")
    assert income["C3"].value.startswith("='Balance Normalizado'!")
    assert "Control de emisión" in workbook.sheetnames
    first_stamp = summary["B6"].value
    at.run(timeout=20)
    second = load_workbook(BytesIO(at.session_state.export_kwargs["data"]))
    assert second["Resumen"]["B6"].value == first_stamp


def test_comparative_report_keeps_two_periods_in_summary_detail_and_income():
    grouped = pd.DataFrame([
        dict(codigo_clasificado="AC.01", monto_total=100, monto_total_2018=80,
             num_cuentas=1),
        dict(codigo_clasificado="PC.01", monto_total=40, monto_total_2018=30,
             num_cuentas=1),
        dict(codigo_clasificado="PAT.01", monto_total=60, monto_total_2018=50,
             num_cuentas=1),
        dict(codigo_clasificado="ER.01", monto_total=25, monto_total_2018=20,
             num_cuentas=1),
        dict(codigo_clasificado="ER.04", monto_total=-10, monto_total_2018=-8,
             num_cuentas=1),
    ])
    complete, formulas = complete_catalog(
        grouped, catalogo_local(),
        {"monto_total": True, "monto_total_2018": True},
        amount_columns=["monto_total", "monto_total_2018"],
    )
    book = Workbook()
    balance = book.active
    balance.title = "Balance Normalizado"
    headers = ["Código", "Cuenta", "2019", "2018", "Cuentas", "Tipo"]
    for column, header in enumerate(headers, 1):
        balance.cell(7, column, header)
    for row_number, row in enumerate(complete.to_dict("records"), 8):
        balance.cell(row_number, 1, row["codigo_clasificado"])
        balance.cell(row_number, 2, row["nombre_estandar"])
        balance.cell(row_number, 3, row["monto_total"])
        balance.cell(row_number, 4, row["monto_total_2018"])
    detail = pd.DataFrame([{
        "Código Estándar": "AC.01", "Nombre Estándar": "Caja y Bancos",
        "Cód. Original": "1101", "Nombre": "Caja",
        "Monto Extraído 2019": 100, "Monto Normalizado 2019": 100,
        "Monto Extraído 2018": 80, "Monto Normalizado 2018": 80,
    }])
    meta = type("Meta", (), {
        "razon_social": "Comparativa", "rut": "", "periodo_desde": "",
        "periodo_hasta": "", "moneda": "$",
    })()

    add_report_sheets(
        book, complete, formulas, account_detail=detail, start_row=7, unit="$",
        meta=meta, processed_at="2026-08-26", source_name="comparativo.pdf",
        pages=[5, 6, 7], definitive=True, reasons=[],
        period_columns=[("2019", "monto_total"), ("2018", "monto_total_2018")],
    )

    summary = book["Resumen"]
    assert [summary.cell(11, column).value for column in range(1, 4)] == [
        "Control del balance", "2019 ($)", "2018 ($)",
    ]
    assert [summary.cell(26, column).value for column in range(1, 9)] == [
        "codigo_homologado", "nombre_homologado", "codigo_original",
        "nombre_original", "valor_extraido_2019", "valor_homologado_2019",
        "valor_extraido_2018", "valor_homologado_2018",
    ]
    assert summary.tables["CuentasClasificadas"].ref == "A26:H27"
    income = book["Estado de Resultados"]
    assert [income.cell(2, column).value for column in range(1, 6)] == [
        "Código", "Clasificación", "2019 ($)", "2018 ($)", "Tipo",
    ]
