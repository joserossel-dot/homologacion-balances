from copy import deepcopy
from io import BytesIO
from types import SimpleNamespace

import pandas as pd
import pytest
from openpyxl import load_workbook
from streamlit.testing.v1 import AppTest

import app_validacion as app
import parser_universal as parser


def sample():
    lines = [
        "110101 Banco Estado 150 50 0 0 100 0 0 0",
        "210101 Proveedores 0 40 0 40 0 40 0 0",
        "410101 Ventas 0 60 0 60 0 0 0 60",
        "Sumas 250 250 100 100 100 40 0 60",
        "Resultado positivo 0 0 0 0 0 60 60 0",
        "Sumas totales 250 250 100 100 100 100 60 60",
    ]
    return [parser.parsear_linea(line, i, parser.FormatoCodigo.COMPACTO, ".")
            for i, line in enumerate(lines)]


def test_finales_validados_no_certifican_ocho_columnas_ni_alteran_importes():
    cuentas = sample()
    originales = deepcopy(cuentas)
    cert = parser.certificar_extraccion_columnas(cuentas)
    assert cert.estado == "fallida"
    assert cert.columnas_finales_validadas
    assert app._permite_clasificar_extraccion(cert)
    assert cert.diferencias["debitos"] == -100
    assert cert.diferencias["saldo_deudor"] == -100
    assert cert.observaciones_auxiliares[0]["Revisar"] == "Saldo deudor / Saldo acreedor"
    assert cuentas == originales


def test_movimiento_erroneo_puede_ser_respaldado_por_saldo():
    cuentas = sample()
    cuentas[0].montos_columnas.update(debitos=50, saldo_deudor=100)
    cert = parser.certificar_extraccion_columnas(cuentas)
    assert cert.columnas_finales_validadas
    assert cert.observaciones_auxiliares[0]["Revisar"] == "Debe / Haber"


@pytest.mark.parametrize("column", parser.RAW_MONETARY_COLUMNS[4:])
def test_diferencia_en_cualquier_columna_final_bloquea(column):
    cuentas = sample()
    cuentas[3].montos_columnas[column] += 50
    cert = parser.certificar_extraccion_columnas(cuentas)
    assert not cert.columnas_finales_validadas
    assert not app._permite_clasificar_extraccion(cert)


@pytest.mark.parametrize("control", [4, 5])
def test_cierre_o_puente_incorrecto_no_habilita(control):
    cuentas = sample()
    cuentas[control].montos_columnas["pasivo"] += 50
    assert not parser.certificar_extraccion_columnas(cuentas).columnas_finales_validadas


@pytest.mark.parametrize("defecto", ["importe", "origen", "sin_respaldo", "doble_destino", "nan", "incompleta", "omitida"])
def test_no_habilita_filas_sin_respaldo(defecto):
    cuentas = sample()
    c = cuentas[0]
    if defecto == "importe":
        c.monto = 200
    elif defecto == "origen":
        c.origen_columna = parser.OrigenColumna.PASIVO
    elif defecto == "sin_respaldo":
        c.montos_columnas["debitos"] = 50
    elif defecto == "doble_destino":
        c.montos_columnas.update(activo=50, perdida=50)
        cuentas[3].montos_columnas.update(activo=50, perdida=50)
    elif defecto == "nan":
        c.montos_columnas["saldo_deudor"] = float("nan")
    elif defecto == "incompleta":
        del c.montos_columnas["saldo_deudor"]
    elif defecto == "omitida":
        cuentas.append(parser.parsear_linea("Sin codigo 10 0 10 0 10 0 0 0", 10, parser.FormatoCodigo.COMPACTO, "."))
    assert not parser.certificar_extraccion_columnas(cuentas).columnas_finales_validadas


def test_no_basta_que_los_errores_se_compensen_en_los_totales():
    cuentas = sample()
    extra = deepcopy(cuentas[0])
    extra.codigo, extra.linea = "110102", 10
    extra.montos_columnas.update(debitos=0, creditos=0, activo=50)
    extra.monto = 50
    cuentas[0].montos_columnas["activo"] = 50
    cuentas[0].monto = 50
    cuentas.append(extra)
    cert = parser.certificar_extraccion_columnas(cuentas)
    assert all(cert.diferencias[c] == 0 for c in parser.RAW_MONETARY_COLUMNS[4:])
    assert not cert.columnas_finales_validadas


def test_sin_control_independiente_no_habilita():
    cert = parser.certificar_extraccion_columnas(sample()[:3])
    assert not cert.columnas_finales_validadas
    assert not app._permite_clasificar_extraccion(cert)


@pytest.mark.parametrize("row,columns", [(0, ["debitos"]), (0, ["activo"]), (3, ["activo"])])
def test_no_usa_reconstruccion_como_respaldo_independiente(row, columns):
    cuentas = sample()
    cuentas[row].columnas_derivadas = columns
    assert not parser.certificar_extraccion_columnas(cuentas).columnas_finales_validadas


def test_perdida_y_puente_de_cierre_se_validan_con_signos_correctos():
    lines = [
        "110101 Caja 50 0 0 0 50 0 0 0",
        "210101 Capital 0 100 0 100 0 100 0 0",
        "310101 Gastos 50 0 50 0 0 0 50 0",
        "Sumas 200 200 100 100 50 100 50 0",
        "Resultado negativo 0 0 0 0 50 0 0 50",
        "Sumas totales 200 200 100 100 100 100 50 50",
    ]
    cuentas = [parser.parsear_linea(line, i, parser.FormatoCodigo.COMPACTO, ".")
               for i, line in enumerate(lines)]
    cert = parser.certificar_extraccion_columnas(cuentas)
    assert cert.columnas_finales_validadas
    assert cert.tipo_resultado == "perdida"
    assert cert.resultado_ejercicio == -50


def test_contracuenta_con_signo_negativo_conserva_importe():
    cuentas = sample()
    contra = deepcopy(cuentas[1])
    contra.codigo, contra.linea = "110102", 10
    contra.origen_columna = parser.OrigenColumna.ACTIVO
    contra.monto = -20
    contra.montos_columnas = dict(zip(parser.RAW_MONETARY_COLUMNS, [0, 20, 0, 20, -20, 0, 0, 0]))
    cuentas.append(contra)
    cuentas[3].montos_columnas["activo"] = 80
    cuentas[2].montos_columnas.update(creditos=40, saldo_acreedor=40, ganancia=40)
    cuentas[2].monto = 40
    cuentas[3].montos_columnas["ganancia"] = 40
    cuentas[4].montos_columnas.update(pasivo=40, perdida=40)
    cuentas[5].montos_columnas.update(activo=80, pasivo=80, perdida=40, ganancia=40)
    assert parser.certificar_extraccion_columnas(cuentas).columnas_finales_validadas
    assert contra.monto == -20


def test_diagnostico_banco_no_sugiere_borrar_activo_correcto():
    diag = app._diagnosticar_filas_extraccion(sample())[0]
    assert "Activo:" not in diag["valores_sugeridos"]
    assert "Saldo deudor: 0 → 100" in diag["valores_sugeridos"]


def test_exporta_advertencias_sin_declarar_ocho_columnas_certificadas():
    cert = parser.certificar_extraccion_columnas(sample())
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        pd.DataFrame({"Activo": [100]}).to_excel(writer, sheet_name="Balance Normalizado", index=False)
        app._exportar_advertencias_auxiliares(writer, cert)
    wb = load_workbook(buf)
    assert wb.sheetnames == ["Balance Normalizado", "Control de extracción"]
    values = [cell.value for row in wb["Control de extracción"] for cell in row]
    assert "No certificada: diferencias en movimientos o saldos" in values
    assert -100 in values
    assert "Banco Estado" in values
    assert wb["Balance Normalizado"]["A2"].value == 100


@pytest.mark.parametrize("suffix", ["pdf", "xlsx"])
@pytest.mark.parametrize("bad_final", [False, True])
def test_entrada_operativa_respeta_validacion_final(monkeypatch, suffix, bad_final):
    cuentas = sample()
    if bad_final:
        cuentas[3].montos_columnas["activo"] += 100
    cert = parser.certificar_extraccion_columnas(cuentas)
    result = SimpleNamespace(cuentas=cuentas, advertencias=[], certificacion_extraccion=cert)
    monkeypatch.setattr(app, "ParserPDF", lambda: SimpleNamespace(parsear=lambda path: result))
    monkeypatch.setattr(app, "parsear_excel", lambda file: cuentas)
    monkeypatch.setattr(app, "_safe_mode_enabled", lambda: False)
    state = SimpleNamespace(extraction_pending={})
    state.setdefault = lambda key, value: state.__dict__.setdefault(key, value)
    monkeypatch.setattr(app.st, "session_state", state)
    file = BytesIO(b"sample")
    file.name = "balance." + suffix
    extracted, _ = app._extraer_cuentas(file)
    assert bool(extracted) is not bad_final
    assert (file.name in state.extraction_pending) is bad_final
    assert state.extraction_certifications[file.name].estado == "fallida"


def test_advertencia_persiste_al_repetir_interacciones():
    at = AppTest.from_string('''
import streamlit as st
import app_validacion as app
from parser_universal import CertificacionExtraccion
st.session_state.setdefault("extraction_certifications", {"balance.pdf": CertificacionExtraccion(
    estado="fallida", columnas_finales_validadas=True,
    diferencias={"debitos": -100}, totales_calculados={"debitos": 150},
    totales_impresos={"debitos": 250},
)})
app._mostrar_advertencias_auxiliares(st.session_state.extraction_certifications["balance.pdf"])
st.button("Otra acción")
''').run()
    assert not at.exception
    assert "no se certifican las ocho columnas" in at.warning[0].value
    at.button[0].click().run()
    assert not at.exception
    assert "Clasificación habilitada" in at.warning[0].value


def test_correccion_manual_permite_seguir_con_advertencia_auxiliar():
    at = AppTest.from_string('''
import streamlit as st
import app_validacion as app
import parser_universal as p
if "extraction_pending" not in st.session_state:
    lines = ["110101 Caja 100 0 0 0 100 0 0 0", "210101 Capital 0 100 0 100 0 100 0 0", "Sumas 200 200 100 100 100 100 0 0"]
    cuentas = [p.parsear_linea(line, i, p.FormatoCodigo.COMPACTO, ".") for i, line in enumerate(lines)]
    st.session_state.extraction_pending = {"balance.pdf": p.ResultadoParseo(
        archivo="balance.pdf", formato_codigo=p.FormatoCodigo.COMPACTO, separador_miles=".",
        requirio_ocr=False, rotacion_aplicada=0, cuentas=cuentas,
        certificacion_extraccion=p.certificar_extraccion_columnas(cuentas))}
    st.session_state.extraction_resolved = {}
    st.session_state.resultados = {}
if "balance.pdf" in st.session_state.extraction_pending:
    app._mostrar_correccion_extraccion("balance.pdf")
else:
    app._mostrar_advertencias_auxiliares(st.session_state.extraction_certifications["balance.pdf"])
''').run()
    assert not at.exception
    next(b for b in at.button if "Verificar correcciones" in b.label).click().run()
    assert not at.exception
    assert "balance.pdf" in at.session_state.extraction_resolved
    assert not at.session_state.extraction_pending
    assert "Clasificación habilitada" in at.warning[0].value
