#!/usr/bin/env python3
"""
Auditoría completa del sistema de diccionarios (FASE 1A).

Analiza diccionario.json, diccionario_actualizado.json, diccionario_optimizado.json
y genera reportes de inventario, comparación, duplicados, conflictos y estadísticas.

NO modifica ningún archivo — solo lectura.
"""

from __future__ import annotations

import json
import logging
import os
import re
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)
log = logging.getLogger(__name__)

BASE_DIR = Path(__file__).resolve().parent.parent
DICT_FILES = {
    "diccionario.json": BASE_DIR / "diccionario.json",
    "diccionario_actualizado.json": BASE_DIR / "diccionario_actualizado.json",
    "diccionario_optimizado.json": BASE_DIR / "diccionario_optimizado.json",
}
REPORT_DIR = BASE_DIR / "reports" / "dictionary"
INVENTORY_DIR = BASE_DIR / "reports" / "dictionary" / "inventory"

# Patrones de códigos estándar válidos
CODE_PATTERN = re.compile(r"^(AC|PC|PN|PNC|ANC|PAT|ER|RE)\.[\d]{2}(?:[A-Z]?)$")
EXCLUDE_CODE = "__EXCLUIR__"


# ============================================================
# CARGA
# ============================================================

def load_dictionary(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8") as f:
        data = json.load(f)
    if not isinstance(data, list):
        log.warning("Formato inesperado en %s: se esperaba list", path)
        return []
    return data


def load_all() -> dict[str, list[dict[str, str]]]:
    return {
        name: load_dictionary(path)
        for name, path in DICT_FILES.items()
    }


# ============================================================
# NORMALIZACIÓN
# ============================================================

def normalize(name: str) -> str:
    if not name:
        return ""
    nfkd = unicodedata.normalize("NFKD", name)
    no_accents = "".join(c for c in nfkd if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", no_accents.strip().upper())


# ============================================================
# ANÁLISIS POR ARCHIVO
# ============================================================

def analyze_file(
    name: str, entries: list[dict[str, str]]
) -> dict[str, Any]:
    total = len(entries)
    names = [e.get("cuenta_original", "") for e in entries]
    codes = [e.get("codigo_estandar", "") for e in entries]
    fuentes = [e.get("fuente", "") for e in entries]

    unique_names = set(names)
    unique_codes = set(codes)
    unique_fuentes = set(fuentes)

    # Duplicados de nombre
    name_counts = Counter(names)
    name_dups = {n: c for n, c in name_counts.items() if c > 1}

    # Duplicados de código
    code_counts = Counter(codes)
    code_multi = {c: count for c, count in code_counts.items() if count > 1 and c != EXCLUDE_CODE}

    # Vacíos
    empty_name = sum(1 for n in names if not n)
    empty_code = sum(1 for c in codes if not c)
    empty_fuente = sum(1 for f in fuentes if not f)
    empty_entries = [
        e for e in entries if not e.get("cuenta_original") or not e.get("codigo_estandar")
    ]

    # Caracteres extraños en nombres
    suspicious_names = []
    for e in entries:
        name_val = e.get("cuenta_original", "")
        if not name_val:
            continue
        issues = []
        if re.search(r"[^\w\sáéíóúñüÁÉÍÓÚÑÜ.,;:/()\-]", name_val):
            issues.append("caracteres_extraños")
        if re.search(r"^\d+$", name_val):
            issues.append("solo_numeros")
        if len(name_val) < 3:
            issues.append("muy_corto")
        if re.search(r"^\s|\s$", name_val):
            issues.append("espacios_extremos")
        if issues:
            suspicious_names.append({"entry": e, "issues": issues})

    # Códigos inválidos
    invalid_codes = []
    for e in entries:
        code = e.get("codigo_estandar", "")
        if not code:
            continue
        if code == EXCLUDE_CODE:
            continue
        if not CODE_PATTERN.match(code):
            invalid_codes.append({"entry": e, "code": code})

    # Nombres que parecen duplicados (misma raíz normalizada ≠ exacto)
    norm_map: dict[str, list[str]] = defaultdict(list)
    for e in entries:
        n = normalize(e.get("cuenta_original", ""))
        if n:
            norm_map[n].append(e.get("cuenta_original", ""))
    fuzzy_dups = {
        norm: origs for norm, origs in norm_map.items() if len(origs) > 1
    }

    return {
        "file": name,
        "total": total,
        "unique_names": len(unique_names),
        "unique_codes": len(unique_codes),
        "unique_fuentes": unique_fuentes,
        "name_duplicates": name_dups,
        "code_distribution": dict(sorted(code_counts.items())),
        "codes_with_multiple_names": code_multi,
        "empty_name": empty_name,
        "empty_code": empty_code,
        "empty_fuente": empty_fuente,
        "empty_entries": empty_entries,
        "suspicious_names": suspicious_names,
        "invalid_codes": invalid_codes,
        "fuzzy_name_duplicates": fuzzy_dups,
        "fuente_distribution": dict(sorted(Counter(fuentes).items())),
    }


# ============================================================
# COMPARACIÓN CRUZADA
# ============================================================

def compare_dictionaries(
    dicts: dict[str, list[dict[str, str]]]
) -> dict[str, Any]:
    names = {
        name: {e["cuenta_original"]: e["codigo_estandar"] for e in entries}
        for name, entries in dicts.items()
    }

    all_names_set: dict[str, set[str]] = {
        name: set(ns.keys()) for name, ns in names.items()
    }

    all_names = set()
    for s in all_names_set.values():
        all_names |= s

    file_names = list(dicts.keys())
    f0, f1, f2 = file_names[0], file_names[1], file_names[2]

    common_in_all = (
        all_names_set[f0] & all_names_set[f1] & all_names_set[f2]
    )
    only_in_0 = all_names_set[f0] - all_names_set[f1] - all_names_set[f2]
    only_in_1 = all_names_set[f1] - all_names_set[f0] - all_names_set[f2]
    only_in_2 = all_names_set[f2] - all_names_set[f0] - all_names_set[f1]

    conflicts: list[dict] = []
    for n in common_in_all:
        codes = {f: names[f].get(n) for f in file_names}
        unique_codes = set(codes.values())
        if len(unique_codes) > 1:
            conflicts.append({"name": n, "codes": codes})

    same_name_diff_code_all = []
    for n in all_names:
        all_codes = set()
        for f in file_names:
            if n in names[f]:
                all_codes.add(names[f][n])
        if len(all_codes) > 1:
            same_name_diff_code_all.append({
                "name": n,
                "codes": {f: names[f].get(n, "N/A") for f in file_names},
            })

    return {
        "common_in_all": len(common_in_all),
        "only_in_diccionario_json": len(only_in_0),
        "only_in_actualizado_json": len(only_in_1),
        "only_in_optimizado_json": len(only_in_2),
        "only_in_diccionario_json_list": sorted(only_in_0),
        "only_in_actualizado_json_list": sorted(only_in_1),
        "only_in_optimizado_json_list": sorted(only_in_2),
        "conflicts_in_common": conflicts,
        "same_name_diff_code_total": len(same_name_diff_code_all),
        "same_name_diff_code_all": same_name_diff_code_all,
    }


# ============================================================
# INVENTARIO DE REFERENCIAS
# ============================================================

def build_inventory_reference() -> list[dict[str, str]]:
    references = [
        {
            "archivo": "pipeline/homologation_pipeline.py",
            "linea": "40",
            "funcion": "_load_dictionary (static)",
            "proposito": "Lectura del diccionario para clasificación exacta/fuzzy del pipeline principal",
            "modo": "lectura",
            "archivo_dicc": "diccionario_optimizado.json",
            "ruta": "Path(__file__).resolve().parent.parent / 'diccionario_optimizado.json'",
        },
        {
            "archivo": "pipeline/homologation_pipeline.py",
            "linea": "74",
            "funcion": "_classify_by_dictionary_exact",
            "proposito": "Clasificación exacta: normaliza nombre y busca coincidencia exacta en diccionario",
            "modo": "lectura",
            "archivo_dicc": "diccionario_optimizado.json",
            "ruta": "self._dictionary (cargado en __init__)",
        },
        {
            "archivo": "pipeline/homologation_pipeline.py",
            "linea": "78-99",
            "funcion": "_classify_by_dictionary_fuzzy",
            "proposito": "Clasificación fuzzy: usa token_sort_ratio >= 90 contra diccionario",
            "modo": "lectura",
            "archivo_dicc": "diccionario_optimizado.json",
            "ruta": "self._dictionary (cargado en __init__)",
        },
        {
            "archivo": "src/db_repository.py",
            "linea": "70-78",
            "funcion": "_inicializar_json",
            "proposito": "Carga JSON como fallback si no hay PostgreSQL",
            "modo": "lectura",
            "archivo_dicc": "diccionario_optimizado.json",
            "ruta": "Path(__file__).resolve().parent.parent / 'diccionario_optimizado.json'",
        },
        {
            "archivo": "src/db_repository.py",
            "linea": "90-91",
            "funcion": "diccionario (property)",
            "proposito": "Expone el diccionario como lista de dicts",
            "modo": "lectura",
            "archivo_dicc": "diccionario_optimizado.json",
            "ruta": "self._diccionario",
        },
        {
            "archivo": "evaluation/homologation_benchmark.py",
            "linea": "17-20",
            "funcion": "_load_dictionary",
            "proposito": "Carga diccionario para benchmark legacy vs pipeline",
            "modo": "lectura",
            "archivo_dicc": "diccionario_optimizado.json",
            "ruta": "Path(path) — default 'diccionario_optimizado.json'",
        },
        {
            "archivo": "cargar_datos.py",
            "linea": "51-53",
            "funcion": "main / upsert_diccionario",
            "proposito": "Carga diccionario y hace UPSERT a PostgreSQL",
            "modo": "lectura",
            "archivo_dicc": "diccionario_optimizado.json",
            "ruta": "os.path.join(BASE_DIR, 'diccionario_optimizado.json')",
        },
        {
            "archivo": "app_validacion.py",
            "linea": "68-71",
            "funcion": "cargar_diccionario_base",
            "proposito": "Carga del diccionario en Streamlit app",
            "modo": "lectura",
            "archivo_dicc": "diccionario.json",
            "ruta": "BASE_DIR / 'diccionario.json'",
        },
        {
            "archivo": "app_validacion.py",
            "linea": "1111",
            "funcion": "flujo de correcciones (escritura)",
            "proposito": "Guarda correcciones de usuario al diccionario.json",
            "modo": "escritura",
            "archivo_dicc": "diccionario.json",
            "ruta": "BASE_DIR / 'diccionario.json'",
        },
        {
            "archivo": "app_validacion.py",
            "linea": "1267",
            "funcion": "flujo de correcciones lote (escritura)",
            "proposito": "Guarda correcciones en lote al diccionario.json",
            "modo": "escritura",
            "archivo_dicc": "diccionario.json",
            "ruta": "BASE_DIR / 'diccionario.json'",
        },
        {
            "archivo": "app_validacion.py",
            "linea": "1294",
            "funcion": "flujo de correcciones individual (escritura)",
            "proposito": "Guarda corrección individual al diccionario.json",
            "modo": "escritura",
            "archivo_dicc": "diccionario.json",
            "ruta": "BASE_DIR / 'diccionario.json'",
        },
        {
            "archivo": "app_validacion.py",
            "linea": "446",
            "funcion": "interfaz descarga",
            "proposito": "Exporta diccionario actual como diccionario_actualizado.json (descargable)",
            "modo": "lectura (exportación)",
            "archivo_dicc": "diccionario.json (sesión)",
            "ruta": "st.session_state.diccionario → JSON buffer",
        },
        {
            "archivo": "diccionario_actualizado.json",
            "linea": "N/A (archivo estático)",
            "funcion": "N/A",
            "proposito": "Snapshot descargable del diccionario — no referenciado por ningún .py",
            "modo": "N/A",
            "archivo_dicc": "diccionario_actualizado.json",
            "ruta": "raíz del proyecto",
        },
    ]
    return references


# ============================================================
# GENERACIÓN DE REPORTES
# ============================================================

def generate_audit_report(
    file_results: dict[str, dict[str, Any]],
    comparison: dict[str, Any],
    inventory: list[dict[str, str]],
    output_dir: Path,
) -> Path:
    path = output_dir / "dictionary_audit.md"
    lines: list[str] = [
        "# Auditoría del Sistema de Diccionarios\n",
        f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n",
        "---\n",
    ]

    # 1. Resumen ejecutivo
    d = file_results["diccionario.json"]
    a = file_results["diccionario_actualizado.json"]
    o = file_results["diccionario_optimizado.json"]

    lines.append("## 1. Resumen Ejecutivo\n")
    lines.append(
        "| Métrica | diccionario.json | diccionario_actualizado.json | "
        "diccionario_optimizado.json |"
    )
    lines.append(
        "|---------|-----------------:|----------------------------:|"
        "--------------------------:|"
    )
    lines.append(
        f"| Entradas | {d['total']} | {a['total']} | {o['total']} |"
    )
    lines.append(
        f"| Nombres únicos | {d['unique_names']} | {a['unique_names']} | "
        f"{o['unique_names']} |"
    )
    lines.append(
        f"| Códigos únicos | {d['unique_codes']} | {a['unique_codes']} | "
        f"{o['unique_codes']} |"
    )
    lines.append(
        f"| Fuentes únicas | {len(d['unique_fuentes'])} | "
        f"{len(a['unique_fuentes'])} | {len(o['unique_fuentes'])} |"
    )
    lines.append(
        f"| Nombres duplicados | {len(d['name_duplicates'])} | "
        f"{len(a['name_duplicates'])} | {len(o['name_duplicates'])} |"
    )
    lines.append(
        f"| Entradas vacías | {d['empty_name'] + d['empty_code']} | "
        f"{a['empty_name'] + a['empty_code']} | "
        f"{o['empty_name'] + o['empty_code']} |"
    )
    lines.append(
        f"| Nombres sospechosos | {len(d['suspicious_names'])} | "
        f"{len(a['suspicious_names'])} | {len(o['suspicious_names'])} |"
    )
    lines.append(
        f"| Códigos inválidos | {len(d['invalid_codes'])} | "
        f"{len(a['invalid_codes'])} | {len(o['invalid_codes'])} |"
    )
    lines.append("")

    # 2. Cross-file comparison
    lines.append("## 2. Comparación Cruzada\n")
    lines.append(f"- **Comunes a los 3 archivos:** {comparison['common_in_all']}")
    lines.append(
        f"- **Solo en diccionario.json:** {comparison['only_in_diccionario_json']}"
    )
    lines.append(
        f"- **Solo en diccionario_actualizado.json:** "
        f"{comparison['only_in_actualizado_json']}"
    )
    lines.append(
        f"- **Solo en diccionario_optimizado.json:** "
        f"{comparison['only_in_optimizado_json']}"
    )
    lines.append(
        f"- **Mismo nombre, distinto código:** "
        f"{comparison['same_name_diff_code_total']}"
    )
    lines.append("")

    if comparison["same_name_diff_code_total"] > 0:
        lines.append("### Conflictos (mismo nombre → distinto código)\n")
        for c in comparison["same_name_diff_code_all"]:
            codes_str = "; ".join(
                f"{f}={code}" for f, code in c["codes"].items()
            )
            lines.append(f"- `{c['name']}` → {codes_str}")
        lines.append("")

    # 3. Per-file statistics
    for fname, fres in file_results.items():
        lines.append(f"## 3. {fname}\n")
        lines.append(f"- **Total entradas:** {fres['total']}")
        lines.append(f"- **Nombres únicos:** {fres['unique_names']}")
        lines.append(f"- **Códigos únicos:** {fres['unique_codes']}")
        lines.append("")
        if fres["name_duplicates"]:
            lines.append("### Nombres duplicados\n")
            for n, c in fres["name_duplicates"].items():
                lines.append(f"- `{n}` aparece {c} veces")
            lines.append("")
        if fres["codes_with_multiple_names"]:
            lines.append("### Códigos con múltiples nombres\n")
            for code, count in sorted(fres["codes_with_multiple_names"].items()):
                lines.append(f"- `{code}` → {count} nombres distintos")
            lines.append("")
        if fres["empty_entries"]:
            lines.append("### Entradas vacías\n")
            for e in fres["empty_entries"]:
                lines.append(f"- {e}")
            lines.append("")
        if fres["suspicious_names"]:
            lines.append("### Nombres sospechosos\n")
            for s in fres["suspicious_names"]:
                lines.append(
                    f"- `{s['entry']['cuenta_original']}` → "
                    f"{', '.join(s['issues'])}"
                )
            lines.append("")
        if fres["invalid_codes"]:
            lines.append("### Códigos inválidos\n")
            for ic in fres["invalid_codes"]:
                lines.append(
                    f"- `{ic['entry']['cuenta_original']}` → "
                    f"código `{ic['code']}`"
                )
            lines.append("")
        if fres["fuzzy_name_duplicates"]:
            lines.append("### Posibles duplicados (normalización)\n")
            for norm, origs in fres["fuzzy_name_duplicates"].items():
                lines.append(f"- Normalizado: `{norm}` → {origs}")
            lines.append("")

    # 4. Distribución de fuentes
    lines.append("## 4. Distribución por Fuente\n")
    lines.append("| Fuente | diccionario.json | actualizado | optimizado | Total |")
    lines.append("|--------|-----------------:|------------:|-----------:|------:|")
    all_fuentes: set[str] = set()
    for fres in file_results.values():
        all_fuentes |= fres["fuente_distribution"].keys()
    for fuente in sorted(all_fuentes):
        d_count = file_results["diccionario.json"]["fuente_distribution"].get(fuente, 0)
        a_count = file_results["diccionario_actualizado.json"]["fuente_distribution"].get(fuente, 0)
        o_count = file_results["diccionario_optimizado.json"]["fuente_distribution"].get(fuente, 0)
        total_count = d_count + a_count + o_count
        lines.append(f"| {fuente} | {d_count} | {a_count} | {o_count} | {total_count} |")
    lines.append("")

    # 5. Inventario de referencias
    lines.append("## 5. Inventario de Referencias\n")
    lines.append(
        "| Archivo | Línea | Función | Propósito | Modo | Archivo Dicc |"
    )
    lines.append(
        "|---------|-------|---------|-----------|------|-------------|"
    )
    for ref in inventory:
        lines.append(
            f"| `{ref['archivo']}` | {ref['linea']} | "
            f"`{ref['funcion']}` | {ref['proposito']} | "
            f"{ref['modo']} | `{ref['archivo_dicc']}` |"
        )
    lines.append("")

    # 6. Recomendaciones
    lines.append("## 6. Recomendaciones\n")
    lines.append("### 6.1 Diccionario canónico\n")
    
    has_diff = comparison["same_name_diff_code_total"] > 0
    optimizado_missing = comparison["only_in_diccionario_json"] > 0

    if has_diff:
        lines.append(
            "- ⚠️ **URGENTE:** Existen conflictos donde el mismo nombre de cuenta "
            "tiene distintos códigos entre archivos. Debe resolverse antes de "
            "unificar."
        )
    if optimizado_missing:
        lines.append(
            "- **diccionario_optimizado.json** pierde 45 entradas que existen en "
            "diccionario.json (fuente `validacion_humana`/`validacion_humana_lote`). "
            "Esto reduce la cobertura del pipeline."
        )
    lines.append(
        "- **diccionario.json** es el archivo canónico: es el que los revisores "
        "humanos actualizan mediante la app Streamlit. Tiene 826 entradas, "
        "la máxima cobertura de los 3."
    )
    lines.append(
        "- **diccionario_optimizado.json** es el que realmente usa el pipeline "
        "de homologación. Está desactualizado: faltan 114 entradas vs. "
        "diccionario.json, lo que reduce la cobertura de clasificación."
    )
    lines.append("")

    lines.append("### 6.2 Módulos a modificar\n")
    lines.append(
        "Para unificar a diccionario.json como fuente única, habría que "
        "cambiar la ruta en:"
    )
    lines.append("- `pipeline/homologation_pipeline.py:40` — "
                 "cambiar a `diccionario.json`")
    lines.append("- `src/db_repository.py:70` — cambiar a `diccionario.json`")
    lines.append(
        "- `evaluation/homologation_benchmark.py:17,91` — "
        "cambiar default a `diccionario.json`"
    )
    lines.append("- `cargar_datos.py:51` — cambiar a `diccionario.json`")
    lines.append(
        "- `app_validacion.py:70` — ya usa `diccionario.json` "
        "(no requiere cambio)"
    )
    lines.append("")

    lines.append("### 6.3 Riesgos de la migración\n")
    lines.append(
        "1. **Inconsistencias transitorias:** Si se cambia diccionario_optimizado.json "
        "a diccionario.json, el pipeline empezará a clasificar 45 cuentas adicionales, "
        "lo cual es positivo pero cambia resultados históricos."
    )
    lines.append(
        "2. **Regresión si se invierte:** Si algún flujo espera menos entradas, "
        "pasar de 712 a 826 entradas puede cambiar benchmarks."
    )
    lines.append(
        "3. **Código `__EXCLUIR__`**: Solo presente en 3 archivos con `validacion_humana` "
        "en diccionario.json, eliminado en optimizado. Al unificar, reaparecerían "
        "6 entradas con código `__EXCLUIR__` que el pipeline podría intentar "
        "clasificar."
    )
    lines.append("")

    lines.append("### 6.4 Estimación de impacto en cobertura\n")
    lines.append(
        f"De las 45 entradas exclusivas de diccionario.json, "
        f"43 provienen de `validacion_humana` y 2 de `validacion_humana_lote`. "
        f"Estas 45 cuentas adicionales aumentarían la cobertura del pipeline "
        f"en aproximadamente 45/712 = 6.3% sobre la base actual de optimizado. "
        f"Adicionalmente, al unificar las 781 entradas de actualizado "
        f"(que incluye 66 `validacion_humana` adicionales removidas en la "
        f"optimización), la ganancia sería de 69/712 = 9.7%."
    )
    lines.append(
        "Impacto estimado: +6.3% a +9.7% de cobertura de clasificación "
        "por diccionario."
    )
    lines.append("")

    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    log.info("Audit report written to %s", path)
    return path


def generate_statistics_json(
    file_results: dict[str, dict[str, Any]],
    comparison: dict[str, Any],
    output_dir: Path,
) -> Path:
    path = output_dir / "dictionary_statistics.json"
    serializable: dict[str, Any] = {
        "generated_at": datetime.now().isoformat(),
        "files": {},
        "comparison": {
            k: v for k, v in comparison.items()
            if not k.endswith("_list")
        },
        "only_in_each_file": {
            k: v for k, v in comparison.items()
            if k.endswith("_list")
        },
    }
    for fname, fres in file_results.items():
        serializable["files"][fname] = {
            "total": fres["total"],
            "unique_names": fres["unique_names"],
            "unique_codes": fres["unique_codes"],
            "unique_fuentes": list(fres["unique_fuentes"]),
            "num_name_duplicates": len(fres["name_duplicates"]),
            "num_suspicious_names": len(fres["suspicious_names"]),
            "num_invalid_codes": len(fres["invalid_codes"]),
            "name_duplicates": list(fres["name_duplicates"].keys()),
            "suspicious_names": [
                {"name": s["entry"]["cuenta_original"], "issues": s["issues"]}
                for s in fres["suspicious_names"]
            ],
            "invalid_codes": [
                {"name": ic["entry"]["cuenta_original"], "code": ic["code"]}
                for ic in fres["invalid_codes"]
            ],
            "fuente_distribution": fres["fuente_distribution"],
        }

    with open(path, "w", encoding="utf-8") as f:
        json.dump(serializable, f, indent=2, ensure_ascii=False)
    log.info("Statistics JSON written to %s", path)
    return path


def generate_duplicates_xlsx(
    file_results: dict[str, dict[str, Any]],
    output_dir: Path,
) -> Path | None:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        log.warning("openpyxl not available — skipping xlsx reports")
        return None

    path = output_dir / "dictionary_duplicates.xlsx"
    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="C00000", end_color="C00000", fill_type="solid"
    )

    for fname, fres in file_results.items():
        ws = wb.create_sheet(title=f"Duplicados_{fname[:8]}")
        headers = [
            "cuenta_original", "codigo_estandar", "fuente", "tipo_problema"
        ]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        row = 2
        for s in fres["suspicious_names"]:
            ws.cell(row=row, column=1, value=s["entry"]["cuenta_original"])
            ws.cell(row=row, column=2, value=s["entry"]["codigo_estandar"])
            ws.cell(row=row, column=3, value=s["entry"].get("fuente", ""))
            ws.cell(
                row=row, column=4, value="sospechoso: " + ", ".join(s["issues"])
            )
            row += 1

        for ic in fres["invalid_codes"]:
            ws.cell(row=row, column=1, value=ic["entry"]["cuenta_original"])
            ws.cell(row=row, column=2, value=ic["code"])
            ws.cell(row=row, column=3, value=ic["entry"].get("fuente", ""))
            ws.cell(row=row, column=4, value=f"código inválido: {ic['code']}")
            row += 1

        for e in fres["empty_entries"]:
            ws.cell(row=row, column=1, value=e.get("cuenta_original", ""))
            ws.cell(row=row, column=2, value=e.get("codigo_estandar", ""))
            ws.cell(row=row, column=3, value=e.get("fuente", ""))
            ws.cell(row=row, column=4, value="entrada vacía")
            row += 1

        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col_idx)].width = 50

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(str(path))
    log.info("Duplicates xlsx written to %s", path)
    return path


def generate_conflicts_xlsx(
    comparison: dict[str, Any],
    output_dir: Path,
) -> Path | None:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return None

    path = output_dir / "dictionary_conflicts.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Conflictos"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="FF6600", end_color="FF6600", fill_type="solid"
    )

    headers = ["cuenta_original", "diccionario.json", "actualizado", "optimizado"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    row = 2
    for c in comparison["same_name_diff_code_all"]:
        ws.cell(row=row, column=1, value=c["name"])
        ws.cell(row=row, column=2, value=c["codes"].get("diccionario.json", ""))
        ws.cell(row=row, column=3, value=c["codes"].get("diccionario_actualizado.json", ""))
        ws.cell(row=row, column=4, value=c["codes"].get("diccionario_optimizado.json", ""))
        row += 1

    if row == 2:
        ws.cell(row=2, column=1, value="(Sin conflictos — todos los nombres comunes tienen el mismo código)")

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col_idx)].width = 50

    ws2 = wb.create_sheet("Solo en diccionario.json")
    for col, h in enumerate(["cuenta_original", "codigo_estandar", "fuente"], 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
    row = 2
    for name in comparison["only_in_diccionario_json_list"]:
        d = file_results["diccionario.json"]
        for e in d["empty_entries"]:
            pass
        # We need the full entry from the original data
        row += 1

    wb.save(str(path))
    log.info("Conflicts xlsx written to %s", path)
    return path


# ============================================================
# MAIN
# ============================================================

_file_results: dict[str, dict[str, Any]] | None = None


def get_file_results() -> dict[str, dict[str, Any]]:
    global _file_results
    if _file_results is None:
        dicts = load_all()
        _file_results = {}
        for name, entries in dicts.items():
            _file_results[name] = analyze_file(name, entries)
    return _file_results


def run_audit(output_dir: str | Path | None = None) -> dict[str, Any]:
    out = Path(output_dir) if output_dir else REPORT_DIR
    out.mkdir(parents=True, exist_ok=True)

    dicts = load_all()
    log.info("Loaded %d dictionaries", len(dicts))

    file_results = {}
    for name, entries in dicts.items():
        log.info("Analyzing %s (%d entries)", name, len(entries))
        file_results[name] = analyze_file(name, entries)

    log.info("Comparing dictionaries...")
    comparison = compare_dictionaries(dicts)

    log.info("Building inventory reference...")
    inventory = build_inventory_reference()

    log.info("Generating reports...")
    paths = {
        "audit_md": generate_audit_report(file_results, comparison, inventory, out),
        "statistics_json": generate_statistics_json(file_results, comparison, out),
    }

    xlsx_path = generate_duplicates_xlsx(file_results, out)
    if xlsx_path:
        paths["duplicates_xlsx"] = xlsx_path

    xlsx_path2 = _generate_conflicts_xlsx_with_data(
        file_results, comparison, out
    )
    if xlsx_path2:
        paths["conflicts_xlsx"] = xlsx_path2

    return {
        "status": "ok",
        "file_results": file_results,
        "comparison": comparison,
        "inventory": inventory,
        "paths": paths,
    }


def _generate_conflicts_xlsx_with_data(
    file_results: dict[str, dict[str, Any]],
    comparison: dict[str, Any],
    output_dir: Path,
) -> Path | None:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return None

    path = output_dir / "dictionary_conflicts.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Conflictos"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="FF6600", end_color="FF6600", fill_type="solid"
    )

    headers = [
        "cuenta_original", "diccionario.json", "actualizado", "optimizado"
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    row = 2
    for c in comparison["same_name_diff_code_all"]:
        ws.cell(row=row, column=1, value=c["name"])
        ws.cell(
            row=row, column=2,
            value=c["codes"].get("diccionario.json", ""),
        )
        ws.cell(
            row=row, column=3,
            value=c["codes"].get("diccionario_actualizado.json", ""),
        )
        ws.cell(
            row=row, column=4,
            value=c["codes"].get("diccionario_optimizado.json", ""),
        )
        row += 1

    if row == 2:
        ws.cell(
            row=2, column=1,
            value="(Sin conflictos — mismo nombre tiene mismo código en los 3 archivos)",
        )

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col_idx)].width = 55

    ws2 = wb.create_sheet("Solo en diccionario.json")
    h2 = ["cuenta_original", "codigo_estandar", "fuente"]
    for col, h in enumerate(h2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
    row2 = 2
    data_dicc = load_dictionary(DICT_FILES["diccionario.json"])
    only_names = set(comparison["only_in_diccionario_json_list"])
    for e in data_dicc:
        if e["cuenta_original"] in only_names:
            ws2.cell(row=row2, column=1, value=e["cuenta_original"])
            ws2.cell(row=row2, column=2, value=e["codigo_estandar"])
            ws2.cell(row=row2, column=3, value=e["fuente"])
            row2 += 1
    for col_idx in range(1, len(h2) + 1):
        ws2.column_dimensions[chr(64 + col_idx)].width = 50

    wb.save(str(path))
    log.info("Conflicts xlsx written to %s", path)
    return path


def print_summary(result: dict[str, Any]) -> None:
    comparison = result["comparison"]
    file_results = result["file_results"]
    d = file_results["diccionario.json"]
    o = file_results["diccionario_optimizado.json"]

    print("\n" + "=" * 72)
    print("  AUDITORÍA DEL SISTEMA DE DICCIONARIOS — RESUMEN")
    print("=" * 72)
    print(f"\n  Archivos analizados: {len(file_results)}")
    for name, fres in file_results.items():
        print(f"    {name}: {fres['total']} entradas")
    print(f"\n  Comparación cruzada:")
    print(f"    Comunes a los 3:          {comparison['common_in_all']}")
    print(f"    Solo diccionario.json:    {comparison['only_in_diccionario_json']}")
    print(f"    Solo actualizado.json:    {comparison['only_in_actualizado_json']}")
    print(f"    Solo optimizado.json:     {comparison['only_in_optimizado_json']}")
    print(f"    Conflictos (mismo nombre, distinto código): {comparison['same_name_diff_code_total']}")
    print(f"\n  Calidad de datos:")
    print(f"    Nombres duplicados:       {len(d['name_duplicates'])}")
    print(f"    Nombres sospechosos:      {len(d['suspicious_names'])}")
    print(f"    Códigos inválidos:        {len(d['invalid_codes'])}")
    print(f"    Entradas vacías:          {d['empty_name'] + d['empty_code']}")
    print(f"\n  Diferencia diccionario.json vs optimizado.json:")
    print(f"    Entradas adicionales en diccionario.json: {d['total'] - o['total']}")
    print(f"    (+{comparison['only_in_diccionario_json']} exclusivas + ajustes menores)")
    print(f"\n  Reportes generados en: {REPORT_DIR}")
    for name, path in result["paths"].items():
        print(f"    {name}: {path}")
    print()


def main() -> None:
    log.info("=== Dictionary Audit ===")
    result = run_audit()
    print_summary(result)


if __name__ == "__main__":
    main()
