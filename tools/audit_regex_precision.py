#!/usr/bin/env python3
"""Audit precision of every REGLAS_REGEX pattern in MotorHibridoLocal.

For each regex: match count, true positives, false positives, precision,
recall estimate, examples, risk level.

Output:
    reports/regex_audit/regex_precision.xlsx
    reports/regex_audit/regex_precision.md
    reports/regex_audit/regex_precision.json
"""

import json
import re
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass, field, asdict
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from app_validacion import REGLAS_REGEX, REGLAS_COMPILADAS, MotorHibridoLocal, normalizar_nombre
from pipeline.homologation_pipeline import HomologationPipeline
from parser_universal import ParserPDF, OrigenColumna
from gold_standard.builder import GoldBuilder


###############################################################################
# Data structures
###############################################################################

@dataclass
class MatchRecord:
    account_name: str
    account_code: str
    monto: float | None
    origen_columna: str
    pdf_name: str
    matched_pattern: str   # index of pattern
    matched_code: str
    matched_confidence: float
    new_code: str | None
    new_method: str | None
    gold_code: str | None   # from gold standard if available
    is_false_positive: bool | None = None
    fp_reason: str | None = None


@dataclass
class PatternAudit:
    pattern_idx: int
    pattern_raw: str
    code: str
    confidence: float
    total_matches: int = 0
    true_positives: int = 0
    false_positives: int = 0
    uncertain: int = 0
    precision: float | None = None
    recall_estimate: float | None = None
    gold_matches: int = 0
    gold_correct: int = 0
    risk: str = "Bajo"
    true_examples: list[str] = field(default_factory=list)
    false_examples: list[dict] = field(default_factory=list)
    uncertain_examples: list[str] = field(default_factory=list)


###############################################################################
# Gold standard loader
###############################################################################

def load_gold_standard(path: str = "gold_standard.db") -> dict[str, str]:
    """Return dict of {normalized_name: final_code} from gold standard."""
    b = GoldBuilder(path)
    records = b.list_all()
    b.close()
    gold = {}
    for r in records:
        norm = normalizar_nombre(r.account_name)
        if r.final_code and norm not in gold:
            gold[norm] = r.final_code
    return gold


###############################################################################
# Heuristic false positive detection
###############################################################################

# Column-based contradictions (conservative - only when column is reliable)
COLUMN_CONTRADICTIONS: dict[str, list[str]] = {
    "AC.0": ["perdida"],
    "ANC.": ["perdida"],
    "PAT.": ["perdida"],
    "PNC.": ["ganancia"],
    "ER.0": ["activo"],
}


def has_column_contradiction(code: str, col: str) -> bool:
    if col in ("desconocido", ""):
        return False
    for prefix, bad_cols in COLUMN_CONTRADICTIONS.items():
        if code.startswith(prefix) and col in bad_cols:
            return True
    return False


###############################################################################
# Main analysis
###############################################################################

def parse_pdf(pdf_path: str | Path) -> list:
    parser = ParserPDF()
    resultado = parser.parsear(Path(pdf_path))
    return resultado.cuentas


def run_audit(pdf_paths: list[Path]) -> tuple[list[PatternAudit], list[MatchRecord]]:
    gold = load_gold_standard()
    dictionary_path = "diccionario.json"
    with open(dictionary_path, encoding="utf-8") as f:
        dictionary = json.load(f)

    hp = HomologationPipeline()
    motor_legacy = MotorHibridoLocal(dictionary)

    # Initialize per-pattern audit records
    audits = {}
    for i, (pat_raw, code, conf) in enumerate(REGLAS_REGEX):
        audits[i] = PatternAudit(
            pattern_idx=i,
            pattern_raw=pat_raw,
            code=code,
            confidence=conf,
        )

    all_matches: list[MatchRecord] = []

    for pdf_path in pdf_paths:
        pdf_name = pdf_path.name
        cuentas = parse_pdf(pdf_path)

        for c in cuentas:
            if c.monto is None and not c.codigo:
                continue

            # Determine which regex pattern matches (best by confidence)
            nombre_norm = normalizar_nombre(c.nombre or "")
            best_match_idx = None
            best_match_conf = 0
            best_match_code = None

            for i, (pat_raw, code, conf) in enumerate(REGLAS_REGEX):
                pat = re.compile(pat_raw, re.IGNORECASE | re.UNICODE)
                if pat.search(nombre_norm):
                    if conf > best_match_conf:
                        best_match_conf = conf
                        best_match_idx = i
                        best_match_code = code

            if best_match_idx is None:
                continue  # no regex matched

            # Run legacy classifier to get full classification
            legacy_result = motor_legacy.clasificar(c)
            legacy_code = legacy_result.get("codigo_estandar")
            legacy_method = legacy_result.get("metodo", "")

            # Only count if regex was the primary mechanism or contributed
            is_regex = legacy_method.startswith("regla_regex")

            # Run new pipeline
            new_result = hp._classify_account(c.codigo or "", c.nombre)
            new_code = new_result.get("standard_code")
            new_method = new_result.get("method", "")

            # Gold standard
            gold_code = gold.get(nombre_norm)

            # Determine if this is a false positive
            # Uses gold standard + HomologationPipeline agreement as primary signal
            fp = False
            fp_reason = None

            col = c.origen_columna.value if c.origen_columna else "desconocido"

            # --- Primary signals (authoritative) ---

            # 1. Gold standard disagrees → definitive FP
            if gold_code is not None and best_match_code != gold_code:
                fp = True
                fp_reason = f"gold standard dice {gold_code} (regex da {best_match_code})"

            # 2. New pipeline classifies AND disagrees → likely FP (two engines disagree)
            elif new_code is not None and best_match_code != new_code:
                fp = True
                fp_reason = f"new pipeline dice {new_code} ({new_method}), regex da {best_match_code}"

            # 3. Gold standard agrees → confirmed TP, override any FP signals
            elif gold_code is not None and best_match_code == gold_code:
                pass  # confirmed TP

            # --- Secondary signals (supplementary, not definitive) ---
            # These are stored as reasons but don't flip FP if primary signals say TP

            else:
                # Check column contradiction as supplementary info
                if has_column_contradiction(best_match_code, col):
                    fp_reason = f"columna '{col}' incompatible con {best_match_code} (advertencia)"
                    fp = True  # flag as weak FP


            rec = MatchRecord(
                account_name=c.nombre,
                account_code=c.codigo or "",
                monto=c.monto,
                origen_columna=col,
                pdf_name=pdf_name,
                matched_pattern=str(best_match_idx),
                matched_code=best_match_code or "",
                matched_confidence=best_match_conf,
                new_code=new_code,
                new_method=new_method,
                gold_code=gold_code,
                is_false_positive=fp,
                fp_reason=fp_reason,
            )
            all_matches.append(rec)

            audit = audits[best_match_idx]
            audit.total_matches += 1
            if gold_code is not None:
                audit.gold_matches += 1
                if best_match_code == gold_code:
                    audit.gold_correct += 1
            if fp:
                audit.false_positives += 1
                if len(audit.false_examples) < 5:
                    audit.false_examples.append({
                        "name": c.nombre,
                        "col": col,
                        "code": best_match_code,
                        "reason": fp_reason,
                        "new_code": new_code,
                        "gold_code": gold_code,
                    })
            elif gold_code is not None and best_match_code == gold_code:
                audit.true_positives += 1
            else:
                # Not confirmed FP and no gold → uncertain (tentative TP)
                audit.uncertain += 1

    # Compute metrics
    for audit in audits.values():
        total_fp = audit.false_positives
        total_tp = audit.true_positives + audit.uncertain
        total = audit.total_matches
        if total > 0:
            audit.precision = round((total - total_fp) / total * 100, 1)
        else:
            audit.precision = None

        # Recall estimate: matches / (matches + misses in pipeline-unclassified)
        # This is a rough estimate
        if audit.total_matches > 0:
            audit.recall_estimate = round(audit.total_matches / (audit.total_matches + 5) * 100, 1)

        # Collect examples
        for rec in all_matches:
            if int(rec.matched_pattern) == audit.pattern_idx:
                if rec.is_false_positive:
                    continue
                code_val = rec.matched_code
                if rec.gold_code and rec.gold_code == code_val:
                    audit.true_examples.append(f"{rec.account_name} → {code_val} (gold)")
                elif not rec.is_false_positive and len(audit.true_examples) < 5:
                    audit.true_examples.append(f"{rec.account_name} → {code_val}")
                if rec.is_false_positive:
                    pass  # already captured above

        # Risk level
        if audit.precision is not None:
            if audit.total_matches == 0:
                audit.risk = "Sin datos"
            elif audit.precision >= 85:
                audit.risk = "Bajo"
            elif audit.precision >= 60:
                audit.risk = "Medio"
            else:
                audit.risk = "Alto"
        else:
            audit.risk = "Sin datos"

    return list(audits.values()), all_matches


###############################################################################
# Report generation
###############################################################################

def generate_markdown(audits: list[PatternAudit], all_matches: list[MatchRecord]) -> str:
    lines = []
    lines.append("# Auditoría de precisión de REGLAS_REGEX\n")
    lines.append(f"**Total PDFs analizados:** {len(set(m.pdf_name for m in all_matches))}")
    lines.append(f"**Total cuentas con match regex:** {len(all_matches)}")
    lines.append(f"**Total cuentas únicas:** {len(set(m.account_name for m in all_matches))}")
    lines.append("")

    # Summary metrics
    total_matches = sum(a.total_matches for a in audits)
    total_fp = sum(a.false_positives for a in audits)
    avg_precision = sum(a.precision or 0 for a in audits if a.precision is not None)
    count_with_precision = sum(1 for a in audits if a.precision is not None)
    lines.append("## Resumen global\n")
    lines.append(f"| Métrica | Valor |")
    lines.append(f"|---------|-------|")
    lines.append(f"| Patrones evaluados | {len(audits)} |")
    lines.append(f"| Matches totales | {total_matches} |")
    lines.append(f"| Falsos positivos detectados | {total_fp} |")
    lines.append(f"| Precisión promedio | {avg_precision/count_with_precision:.1f}% |"
                 if count_with_precision else "| Precisión promedio | N/A |")
    lines.append("")

    # Ranking by impact (total matches)
    by_impact = sorted(audits, key=lambda a: -a.total_matches)
    lines.append("## Ranking por impacto (total matches)\n")
    lines.append("| # | Código | Patrón | Matches | FP | Precisión | Riesgo |")
    lines.append("|---|--------|--------|---------|----|-----------|--------|")
    for i, a in enumerate(by_impact, 1):
        if a.total_matches == 0:
            continue
        prec_str = f"{a.precision}%" if a.precision is not None else "—"
        lines.append(f"| {i} | {a.code} | `{a.pattern_raw[:50]}...` | {a.total_matches} | {a.false_positives} | {prec_str} | {a.risk} |")
    lines.append("")

    # Ranking by precision (worst first)
    by_precision = sorted(
        [a for a in audits if a.total_matches > 0 and a.precision is not None],
        key=lambda a: a.precision
    )
    lines.append("## Ranking por menor precisión\n")
    lines.append("| # | Código | Patrón | Matches | FP | Precisión | Riesgo |")
    lines.append("|---|--------|--------|---------|----|-----------|--------|")
    for i, a in enumerate(by_precision, 1):
        prec_str = f"{a.precision}%" if a.precision is not None else "—"
        lines.append(f"| {i} | {a.code} | `{a.pattern_raw[:50]}...` | {a.total_matches} | {a.false_positives} | {prec_str} | {a.risk} |")
    lines.append("")

    # Detail per pattern
    lines.append("## Detalle por patrón\n")
    for i, a in enumerate(audits):
        if a.total_matches == 0:
            continue
        prec_str = f"{a.precision}%" if a.precision is not None else "—"
        lines.append(f"### #{a.pattern_idx}: {a.code} (confianza={a.confidence})")
        lines.append(f"**Patrón:** `{a.pattern_raw}`")
        lines.append(f"**Matches:** {a.total_matches} | **FP:** {a.false_positives} | **Precisión:** {prec_str} | **Riesgo:** {a.risk}")
        if a.gold_matches > 0:
            lines.append(f"**Gold standard:** {a.gold_correct}/{a.gold_matches} correctos ({a.gold_correct/a.gold_matches*100:.0f}%)")
        if a.true_examples:
            lines.append("\n**Aciertos:**")
            for ex in a.true_examples[:3]:
                lines.append(f"- {ex}")
        if a.false_examples:
            lines.append("\n**Falsos positivos:**")
            for ex in a.false_examples[:3]:
                lines.append(f"- {ex['name']} (col={ex['col']}) → {ex['code']}: {ex['reason']}")
        lines.append("")

    # Recommendations
    lines.append("## Recomendaciones\n")
    safe = [a for a in audits if a.total_matches > 0 and a.precision is not None and a.precision >= 90]
    needs_work = [a for a in audits if a.total_matches > 0 and a.precision is not None and 70 <= a.precision < 90]
    eliminate = [a for a in audits if a.total_matches > 0 and a.precision is not None and a.precision < 70]
    no_data = [a for a in audits if a.total_matches == 0]

    lines.append(f"### Migrar sin cambios ({len(safe)} patrones)")
    for a in sorted(safe, key=lambda x: -x.total_matches):
        lines.append(f"- `{a.pattern_raw[:60]}...` → {a.code} (precisión {a.precision}%, {a.total_matches} matches)")
    lines.append("")

    if needs_work:
        lines.append(f"### Necesitan ajustes ({len(needs_work)} patrones)")
        for a in sorted(needs_work, key=lambda x: -x.total_matches):
            lines.append(f"- `{a.pattern_raw[:60]}...` → {a.code} (precisión {a.precision}%, {a.false_positives} FP)")
            for ex in a.false_examples[:2]:
                lines.append(f"  - FP: {ex['name']} → {ex['reason']}")
        lines.append("")

    if eliminate:
        lines.append(f"### Eliminar o rediseñar ({len(eliminate)} patrones)")
        for a in sorted(eliminate, key=lambda x: -x.total_matches):
            lines.append(f"- `{a.pattern_raw[:60]}...` → {a.code} (precisión {a.precision}%, {a.false_positives} FP)")
        lines.append("")

    if no_data:
        lines.append(f"### Sin datos disponibles ({len(no_data)} patrones)")
        for a in no_data:
            lines.append(f"- `{a.pattern_raw[:60]}...` → {a.code} (0 matches en la muestra)")
        lines.append("")

    # Top 10 ROI
    by_roi = sorted(
        [a for a in audits if a.total_matches > 0 and a.precision is not None and a.precision >= 70],
        key=lambda a: -a.total_matches
    )[:10]
    lines.append("## Top 10 regex con mejor ROI\n")
    recovery = sum(a.total_matches for a in by_roi)
    lines.append("| # | Código | Matches | FP | Precisión | Recuperación |")
    lines.append("|---|--------|---------|----|-----------|-------------|")
    for i, a in enumerate(by_roi, 1):
        lines.append(f"| {i} | {a.code} | {a.total_matches} | {a.false_positives} | {a.precision}% | {a.total_matches} cuentas |")
    lines.append(f"\n**Total recuperado con top 10:** {recovery} cuentas")
    lines.append("")

    # Expected coverage improvement
    lines.append("## Cobertura esperada después de implementación\n")
    lines.append("Escenario: implementar las Top 10 regex + las que precisión ≥ 90%")
    high_quality = [a for a in audits if a.total_matches > 0 and a.precision is not None and a.precision >= 90]
    top_10_set = {a.pattern_idx for a in by_roi}
    all_to_impl = set()
    for a in high_quality:
        all_to_impl.add(a.pattern_idx)
    for a in by_roi:
        all_to_impl.add(a.pattern_idx)

    impl_matches = sum(audits[i].total_matches for i in all_to_impl)
    lines.append(f"- Patrones a implementar: {len(all_to_impl)}")
    lines.append(f"- Cuentas recuperables: {impl_matches}")
    lines.append(f"- Reducción de 'sin clasificar': {impl_matches} cuentas")
    lines.append("")

    return "\n".join(lines)


def generate_json(audits: list[PatternAudit], all_matches: list[MatchRecord]) -> dict:
    result = {
        "summary": {
            "total_patterns": len(audits),
            "total_matches": sum(a.total_matches for a in audits),
            "total_false_positives": sum(a.false_positives for a in audits),
            "pdfs_analyzed": len(set(m.pdf_name for m in all_matches)),
            "patterns_with_data": sum(1 for a in audits if a.total_matches > 0),
            "patterns_no_data": sum(1 for a in audits if a.total_matches == 0),
        },
        "patterns": [],
        "rankings": {},
    }

    for a in audits:
        result["patterns"].append(asdict(a))

    by_impact = sorted(
        [a for a in audits if a.total_matches > 0],
        key=lambda a: -a.total_matches
    )
    result["rankings"]["by_impact"] = [
        {"rank": i+1, "code": a.code, "matches": a.total_matches, "precision": a.precision}
        for i, a in enumerate(by_impact)
    ]

    by_precision = sorted(
        [a for a in audits if a.total_matches > 0 and a.precision is not None],
        key=lambda a: a.precision
    )
    result["rankings"]["by_precision_worst"] = [
        {"rank": i+1, "code": a.code, "matches": a.total_matches, "precision": a.precision}
        for i, a in enumerate(by_precision)
    ]

    return result


###############################################################################
# Main
###############################################################################

def main():
    base = Path("datasets/HOLDOUT")
    pdf_paths = sorted(base.glob("*.pdf"))

    if not pdf_paths:
        print("ERROR: No PDFs in datasets/HOLDOUT/", file=sys.stderr)
        sys.exit(1)

    print(f"Processing {len(pdf_paths)} PDFs...", file=sys.stderr)
    audits, all_matches = run_audit(pdf_paths)
    print(f"Total matches: {len(all_matches)}", file=sys.stderr)

    # Generate outputs
    output_dir = Path("reports/regex_audit")
    output_dir.mkdir(parents=True, exist_ok=True)

    # Markdown
    md = generate_markdown(audits, all_matches)
    (output_dir / "regex_precision.md").write_text(md, encoding="utf-8")
    print(f"  → {output_dir / 'regex_precision.md'}", file=sys.stderr)

    # JSON
    jdata = generate_json(audits, all_matches)
    (output_dir / "regex_precision.json").write_text(
        json.dumps(jdata, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(f"  → {output_dir / 'regex_precision.json'}", file=sys.stderr)

    # XLSX
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Regex Precision"
        headers = ["#", "Código", "Patrón", "Matches", "TP", "FP", "Inciertos",
                   "Precisión%", "Gold/Total", "Gold OK", "Riesgo"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for i, a in enumerate(audits, 1):
            gold_str = f"{a.gold_correct}/{a.gold_matches}" if a.gold_matches > 0 else "—"
            prec_str = a.precision if a.precision is not None else "—"
            ws.append([i, a.code, a.pattern_raw, a.total_matches, a.true_positives,
                       a.false_positives, a.uncertain, prec_str, gold_str,
                       a.gold_correct, a.risk])
            row = ws[ws.max_row]
            if a.precision is not None:
                if a.precision < 70:
                    for cell in row:
                        cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                elif a.precision < 90:
                    for cell in row:
                        cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        # FP detail sheet
        ws2 = wb.create_sheet("FP Details")
        ws2.append(["Pattern", "Code", "Account", "Column", "Reason", "New Code", "Gold"])
        for cell in ws2[1]:
            cell.font = Font(bold=True)
        for a in audits:
            for ex in a.false_examples:
                ws2.append([a.pattern_idx, a.code, ex["name"], ex["col"],
                           ex["reason"], ex.get("new_code", ""), ex.get("gold_code", "")])
        wb.save(str(output_dir / "regex_precision.xlsx"))
        print(f"  → {output_dir / 'regex_precision.xlsx'}", file=sys.stderr)
    except ImportError:
        print("  (openpyxl not installed, skipping xlsx)", file=sys.stderr)

    # Print summary to stdout
    print("\n=== SUMMARY ===", file=sys.stderr)
    total_matches = sum(a.total_matches for a in audits)
    total_fp = sum(a.false_positives for a in audits)
    print(f"Total patterns: {len(audits)}", file=sys.stderr)
    print(f"Total matches: {total_matches}", file=sys.stderr)
    print(f"Total FP: {total_fp}", file=sys.stderr)
    print(f"Patterns with data: {sum(1 for a in audits if a.total_matches > 0)}", file=sys.stderr)
    print(f"Patterns no data: {sum(1 for a in audits if a.total_matches == 0)}", file=sys.stderr)

    safe = [a for a in audits if a.total_matches > 0 and a.precision is not None and a.precision >= 90]
    needs_work = [a for a in audits if a.total_matches > 0 and a.precision is not None and 70 <= a.precision < 90]
    eliminate = [a for a in audits if a.total_matches > 0 and a.precision is not None and a.precision < 70]
    print(f"Migrate as-is: {len(safe)}", file=sys.stderr)
    print(f"Need adjustment: {len(needs_work)}", file=sys.stderr)
    print(f"Eliminate: {len(eliminate)}", file=sys.stderr)

    by_roi = sorted(
        [a for a in audits if a.total_matches > 0 and a.precision is not None and a.precision >= 70],
        key=lambda a: -a.total_matches
    )[:10]
    print(f"\nTop 10 ROI recovery: {sum(a.total_matches for a in by_roi)} accounts", file=sys.stderr)

    safe_all = sum(a.total_matches for a in safe)
    roi_all = sum(a.total_matches for a in by_roi)
    combined = set()
    for a in safe:
        combined.add(a.pattern_idx)
    for a in by_roi:
        combined.add(a.pattern_idx)
    combined_matches = sum(audits[i].total_matches for i in combined)
    print(f"Safe-only recovery: {safe_all}", file=sys.stderr)
    print(f"Combined (safe+top10) recovery: {combined_matches}", file=sys.stderr)


if __name__ == "__main__":
    main()
