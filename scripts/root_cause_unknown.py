#!/usr/bin/env python3
"""Root-cause analysis of 5,392 DEv2 UNKNOWN accounts."""

from __future__ import annotations

import json
import re
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parent.parent

ROOT_CAUSES = [
    "ruido",
    "parser",
    "OCR",
    "concepto inexistente",
    "catálogo incompleto",
    "regex faltante",
    "concepto ambiguo",
    "nombre propio",
    "sigla",
    "abreviatura",
    "empresa",
    "error humano",
    "cuenta compuesta",
    "cuenta tributaria",
    "cuenta especial",
]

# ── Known patterns ──

RE_DATE = re.compile(r'\b\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\b')
RE_RUT = re.compile(r'\b\d{1,2}\.\d{3}\.\d{3}[-]\d\b')
RE_AMOUNT = re.compile(r'\b\d{1,3}(?:\.\d{3})*(?:,\d{2})?\b')
RE_PERCENT = re.compile(r'\d{1,3}(?:\.\d{3})*%')
RE_FOLIO = re.compile(r'[Ff][Oo][Ll][Ii][Oo]')
RE_NIVEL = re.compile(r'^[Nn][Ii][Vv][Ee][Ll]')
RE_HEADER_FOOTER = re.compile(
    r'^(?:Nivel|Desde|Del[ae]?|Folio|Total\s*P[aá]gina|Dela\s*P[aá]gina|'
    r'De\s*la\s*P[aá]gina|P[aá]gina\s*Anterior|Suma\s*anterior|Vienen|Van|'
    r'Son\s*\$|A[ñn]o\s*\d{4})',
    re.IGNORECASE
)

LEADING_GARBAGE = re.compile(r'^[lLoO0\|:Ee\[\]_=#\+~\^><]*\s*')
# Known all-caps acronyms (Chilean context)
KNOWN_ACRONYMS = {
    "AFP", "APV", "CCAF", "CMR", "CTE", "Cta", "Cta.",
    "FONASA", "FUT", "FUNT", "IPC", "ISAPRE", "ISAPRES",
    "IVA", "I.V.A.", "LIS", "Ley", "MIPYME", "MIPYMES",
    "PPM", "P.P.M.", "RUT", "SA", "S.A.", "SENCE",
    "SII", "SPA", "UF", "UF.", "USO", "US$",
    "UTF", "UTM", "VAN",
}
# Common abbreviations
KNOWN_ABBREVIATIONS = {
    "ACUM", "ACUM.", "ANT", "ANT.", "AP", "AP.",
    "CAP", "CAP.", "CC", "CC.", "Cta", "Cta.",
    "Ctas", "Ctas.", "CTE", "CTE.", "Dcta", "Dcta.",
    "DEP", "DEP.", "DIFF", "Doc", "Doc.",
    "Docs", "GAR", "GAR.", "GAST", "GAST.",
    "GTA", "GTA.", "IMPT", "IMPTO.", "IMP",
    "ING", "ING.", "INT", "INT.", "M/N", "M.N",
    "MO", "MO.", "MUEB", "MUEB.",
    "NAC", "NAC.", "NO", "NRO", "NRO.",
    "P", "P.", "PROV", "PROV.", "PROVISION",
    "Pje", "PP", "P.P.", "PROV", "PROV.",
    "PROVEED", "PROV.", "RET", "RET.",
    "SERV", "SERV.", "SOC", "SOC.",
    "SS", "SS.", "SUB", "SUB.",
    "T", "T.", "VARIOS", "VARIOS.",
}
# Tax-related keywords (Chile)
TAX_KEYWORDS = {
    "iva", "i.v.a.", "ppm", "p.p.m.", "renta", "impuesto",
    "impuestos", "sence", "credito sence", "cesantia", "cesantía",
    "tributaria", "tributario", "fiscal", "debito fiscal",
    "credito fiscal", "impuesto unico", "impuesto único",
    "impto unico", "impto único", "trabajadores",
    "segunda categoria", "segunda categoría", "honorarios retencion",
    "retencion", "retención", "iva fuera de plazo",
}

# Spanish stop words for cleaning
STOP = {"de", "del", "el", "la", "los", "las", "y", "e", "o", "a", "al",
        "por", "para", "con", "sin", "en", "un", "una", "su", "sus",
        "que", "es", "se", "no", "lo", "le", "les", "entre"}


def strip_accents(s: str) -> str:
    """Remove accents but keep other chars."""
    nfkd = unicodedata.normalize('NFKD', s)
    return ''.join(c for c in nfkd if not unicodedata.category(c).startswith('M'))


def count_words(text: str) -> list[str]:
    """Return cleaned words (lower, no punctuation)."""
    clean = re.sub(r'[^\w\s]', ' ', text)
    return [w for w in clean.lower().split() if w not in STOP]


def classify(name: str) -> tuple[str, str, str]:
    """
    Return (root_cause, sub_reason, explanation).
    Uses priority-based classification.
    """
    original = name
    stripped = name.strip()
    lowered = stripped.lower()

    # ── 1. RUIDO: Not even an account line ──
    if RE_HEADER_FOOTER.search(stripped):
        return ("ruido", "encabezado_pie",
                f"Texto de encabezado/pie de pagina: '{stripped[:50]}'")

    # All digits or nearly
    digit_ratio = sum(1 for c in stripped if c.isdigit()) / max(len(stripped), 1)
    if digit_ratio > 0.7 and len(stripped) > 3:
        return ("ruido", "solo_numeros",
                f"Principalmente numeros: '{stripped[:50]}'")

    # Month/year only
    month_pattern = r'^(?:enero|febrero|marzo|abril|mayo|junio|julio|agosto|septiembre|octubre|noviembre|diciembre)\s*\d{4}$'
    if re.match(month_pattern, lowered):
        return ("ruido", "mes_anno", f"Solo mes y año: '{stripped[:50]}'")

    # Just a year or range
    if re.match(r'^\d{4}\s*[-–—]\s*\d{4}$', stripped):
        return ("ruido", "rango_anno", f"Solo rango de años: '{stripped[:50]}'")

    # "Año 2020" style
    if re.match(r'^a[ñn]o\s+\d{4}', lowered):
        return ("ruido", "anno_solo", f"Solo indicación de año: '{stripped[:50]}'")

    # ── 2. RUIDO: PDF structural elements ──
    footer_triggers = [
        "de la pagina anterior", "dela página anterior", "dela pagina anterior",
        "pagina anterior", "página anterior", "total pagina", "total página",
        "total párina", "suma anterior", "vienen", "van", "son $",
        "hoja", "continuación", "continuacion", "anexo", "anexos",
        "nota", "notas", "cuadro", "detalle",
    ]
    for ft in footer_triggers:
        if ft in lowered:
            return ("ruido", "estructura_pdf",
                    f"Elemento estructural PDF: '{stripped[:50]}'")

    # Path-like or URL
    if re.search(r'\\\\|//|www\.|\.html?|\.pdf', lowered):
        return ("ruido", "ruta_archivo", f"Ruta o URL: '{stripped[:50]}'")

    # ── 3. RUIDO: "Detalle de" or "Movimiento" headers ──
    if re.match(r'^(?:detalle|movimiento|resumen|lista|informe)', lowered):
        return ("ruido", "encabezado_lista",
                f"Encabezado de lista: '{stripped[:50]}'")

    # Names shorter than 4 chars with mostly special chars
    clean_chars = sum(1 for c in stripped if c.isalpha() or c.isdigit())
    if len(stripped) < 6 and clean_chars < 3:
        return ("ruido", "insuficiente",
                f"Nombre demasiado corto: '{stripped}'")

    # ── 4. PARSER: Has parser debris but account-like content ──
    has_folio = bool(RE_FOLIO.search(stripped))
    has_date = bool(RE_DATE.search(stripped))
    has_rut = bool(RE_RUT.search(stripped))
    has_amount = bool(RE_AMOUNT.search(stripped))
    has_pct = bool(RE_PERCENT.search(stripped))

    if has_folio:
        return ("parser", "folio_embebido",
                f"El nombre contiene 'Folio': '{stripped[:50]}'")

    if has_rut:
        return ("parser", "rut_embebido",
                f"El nombre contiene RUT: '{stripped[:50]}'")

    if has_date:
        return ("parser", "fecha_embebida",
                f"El nombre contiene fecha: '{stripped[:50]}'")

    # Multiple amounts (like balance values embedded)
    amt_matches = RE_AMOUNT.findall(stripped)
    if len(amt_matches) >= 2:
        return ("parser", "montos_embebidos",
                f"El nombre contiene {len(amt_matches)} montos: '{stripped[:50]}'")

    if has_pct:
        return ("parser", "porcentaje_embebido",
                f"El nombre contiene %: '{stripped[:50]}'")

    # ── 5. PARSER: codes mixed in ──
    code_pattern = r'\b\d[-.]\d{2}[-.]\d{2}[-.]\d{2}\b'
    if re.search(code_pattern, stripped):
        return ("parser", "codigo_embebido",
                f"Contiene código contable: '{stripped[:50]}'")

    # "Desde ... hasta" patterns with dates
    if re.search(r'\bdesde\b.*\bhasta\b', lowered) and has_date:
        return ("parser", "rango_fechas",
                f"Rango de fechas embebido: '{stripped[:50]}'")

    # ── 6. PARSER: leading digits + text (e.g. "1 REMUNERACIONES", "0 ANIVEL") ──
    leading_digits_match = re.match(r'^(\d+)\s+(\S)', stripped)
    if leading_digits_match and len(leading_digits_match.group(2)) > 2:
        return ("parser", "digito_inicial",
                f"Digito inicial '{leading_digits_match.group(1)}' antes de texto: '{stripped[:50]}'")

    # Trailing digit ranges (section numbers or codes)
    if re.search(r'\s+\d+[-–—]\d+\s*$', stripped):
        return ("parser", "rango_numeros_final",
                f"Rango numerico al final: '{stripped[:50]}'")

    # ── 7. OCR: leading garbage chars ──
    # Match ONE leading lowercase letter that is NOT a word (o, l, i prefixed)
    # or any leading non-alpha pattern before a known account name
    cleaned = LEADING_GARBAGE.sub('', stripped).strip()
    leading_removed = len(stripped) - len(cleaned)
    if leading_removed > 0 and len(cleaned) > 4:
        garbage_start = stripped[:leading_removed]
        # Check: garbage chars should NOT form a valid Spanish word start
        garbage_alpha = ''.join(c for c in garbage_start if c.isalpha())
        cleaned_first = ''.join(c for c in cleaned[:8] if c.isalpha()).lower()
        if len(garbage_alpha) > 0 and len(garbage_alpha) <= 3:
            # Valid Spanish words starting with common garbage letters
            valid_starts = {'o', 'l', 'e', '0', 's', 'd', 'c'}
            if garbage_alpha.lower() in valid_starts:
                # Verify: if we keep the first garbage letter, does it form a real word?
                combined = (garbage_alpha + cleaned_first[:4]).lower()
                # Common Spanish words that might be valid
                common_words = {'otro', 'otros', 'obra', 'obras', 'lote', 'lotes',
                                'los', 'las', 'el', 'ella', 'ello', 'este', 'esta',
                                'ese', 'esa', 'eso', 'del', 'de', 'e', 'o', 'y',
                                'si', 'se', 'su', 'sus', 'c', 'cc', 's', 'd'}
                if combined not in common_words and len(stripped) > len(cleaned) + 1:
                    return ("OCR", "basura_inicial",
                            f"Basura inicial '{garbage_start}' en: '{stripped[:50]}' -> '{cleaned[:50]}'")

    # ── 8. OCR: merged words in ALL-CAPS ──
    # Detect: HORASEXTRAS, REMUNERACIONESPOR, PROVISIONINTERESES
    words = stripped.split()
    for w in words:
        if len(w) > 15 and w.isalpha() and w == w.upper():
            # Try to find word boundaries by checking known suffixes/prefixes
            lowers = w.lower()
            # Known concept words to detect merges
            known_parts = ['remuneracion', 'remuneracione', 'provision', 'horas',
                          'extra', 'extras', 'cesantia', 'feriado', 'telefono',
                          'insumos', 'artculos', 'gastos', 'gasto', 'total']
            matches = [kp for kp in known_parts if kp in lowers]
            if len(matches) >= 2:
                return ("OCR", "palabras_fusionadas",
                        f"Palabras fusionadas: '{w}' -> {' '.join(matches)}: '{stripped[:50]}'")
            # Generic: long all-caps might contain multiple words
            if len(lowers) > 20:
                return ("OCR", "palabras_fusionadas_posible",
                        f"Posible fusion: '{w}' en: '{stripped[:50]}'")

    # ── 8. OCR: digit substitution in known words ──
    # Check for 5→S, 8→B, 0→O, l→I
    ocr_substitutions = {
        '0': 'o', '5': 's', '8': 'b', '1': 'i', '4': 'a',
    }
    if len(cleaned) > 4:
        substitutions_found = []
        for orig_char, sub_char in ocr_substitutions.items():
            if orig_char in cleaned:
                # Check if replacing makes sense
                test = cleaned.replace(orig_char, sub_char)
                words_orig = set(count_words(cleaned))
                words_test = set(count_words(test))
                if len(words_test) > len(words_orig):
                    substitutions_found.append(f"{orig_char}->{sub_char}")
        if substitutions_found:
            return ("OCR", "sustitucion_digito",
                    f"Posible sustitucion OCR {' '.join(substitutions_found)}: '{stripped[:50]}'")

    # Leading lowercase letters that should be uppercase (OCR ate uppercase)
    if len(cleaned) > 5 and cleaned[0].islower() and cleaned[0].isalpha() and cleaned[1:2].isupper():
        return ("OCR", "minuscula_inicial",
                f"Primera letra en minuscula (OCR): '{stripped[:50]}' -> '{cleaned[:50]}'")

    # ── 9. EMPRESA: Company names ──
    company_suffixes = [r'\bS\.?A\.?\b', r'\bLTDA\b', r'\bE\.?I\.?R\.?L\.?\b',
                        r'\bSPA\b', r'\bLIMITADA\b', r'\bCOMPAÑIA\b', r'\bCOMPANIA\b',
                        r'\bCOMPANY\b', r'\bCORP\b', r'\bINC\b', r'\bS\.?A\.?S\.?\b',
                        r'\bS\.?R\.?L\.?\b', r'\bCIA\b', r'\bCÍA\b', r'\b&CIA\b']
    for suffix in company_suffixes:
        if re.search(suffix, stripped, re.IGNORECASE):
            return ("empresa", "razon_social",
                    f"Razon social: '{stripped[:50]}'")

    # Address-like
    addr_patterns = [r'\bAVDA\b', r'\bAVENIDA\b', r'\bCALLE\b', r'\bPASEO\b',
                     r'\bN°\b', r'\bNUMERO\b', r'\bPISO\b',
                     r'\bCIUDAD\b', r'\bCOMUNA\b', r'\bREGIÓN\b',
                     r'\bFOLIO\s+UNICO\b', r'\bUNICO\s+NACIONAL\b']
    for pat in addr_patterns:
        if re.search(pat, stripped, re.IGNORECASE):
            return ("empresa", "direccion",
                    f"Direccion/domicilio: '{stripped[:50]}'")

    # ── 10. NOMBRE PROPIO: Personal names ──
    honorifics = [r'\bSR\.?\b', r'\bSRA\.?\b', r'\bSRTA\.?\b', r'\bDON\b', r'\bDOÑA\b']
    for h in honorifics:
        if re.search(h, stripped, re.IGNORECASE):
            return ("nombre propio", "tratamiento",
                    f"Nombre personal con tratamiento: '{stripped[:50]}'")

    # ── 11. SIGLA ──
    words_in_name = [w.strip('.:;,!?') for w in stripped.split() if len(w.strip('.:;,!?')) >= 2]
    if len(words_in_name) <= 3 and len(words_in_name) >= 1:
        all_acronyms = all(
            w.upper() in KNOWN_ACRONYMS or
            (w == w.upper() and w.isalpha() and len(w) <= 6)
            for w in words_in_name
        )
        if all_acronyms:
            matched = [w for w in words_in_name if w.upper() in KNOWN_ACRONYMS]
            rest = [w for w in words_in_name if w.upper() not in KNOWN_ACRONYMS]
            return ("sigla", "sigla_reconocida",
                    f"Sigla(s): {', '.join(words_in_name)}")

    # ── 12. ABREVIATURA ──
    # Has words with periods (abbreviations)
    abbrev_words = [w for w in stripped.split() if '.' in w and len(w) <= 8]
    if abbrev_words:
        known_abbrev = [w for w in abbrev_words if w.upper() in KNOWN_ABBREVIATIONS]
        if known_abbrev:
            return ("abreviatura", "abreviatura_conocida",
                    f"Abreviatura(s): {', '.join(known_abbrev)} en: '{stripped[:50]}'")
        if abbrev_words:
            return ("abreviatura", "abreviatura_no_reconocida",
                    f"Abreviatura(s) no estandar: {', '.join(abbrev_words)} en: '{stripped[:50]}'")

    # ── 13. CUENTA TRIBUTARIA ──
    for kw in TAX_KEYWORDS:
        if kw in lowered:
            return ("cuenta tributaria", f"tributaria_{kw.split()[0]}",
                    f"Concepto tributario: '{stripped[:50]}'")

    # ── 14. CUENTA COMPUESTA ──
    if re.search(r'\bY\b|\bE\b|\/\b', stripped, re.IGNORECASE):
        parts = re.split(r'\s+(?:Y|E)\s+', stripped, flags=re.IGNORECASE)
        if len(parts) >= 2 and all(
            len(p.split()) <= 3 and len(p) > 3
            for p in parts
        ):
            return ("cuenta compuesta", "union_y",
                    f"Cuenta compuesta con 'Y': '{stripped[:50]}'")

    # ── 15. CATÁLOGO INCOMPLETO ──
    # This is a catch for accounts that LOOK valid but aren't in catalog
    near_miss_triggers = [
        (r'mantencion', 'MANTENCIÓN debería ser MANTENIMIENTO'),
        (r'\bgt[eo]\b', 'GTO/GTE como abreviatura de GASTO'),
        (r'\bremuneracione\b', 'REMUNERACIONE truncado'),
    ]
    for pat, reason in near_miss_triggers:
        if re.search(pat, lowered):
            return ("catálogo incompleto", "variante_ortografica",
                    f"{reason}: '{stripped[:50]}'")

    # ── 16. CONCEPTO INEXISTENTE (catch-all for clean Spanish) ──
    # Check if it looks like a legitimate account: Spanish words, no garbage
    alpha_ratio = sum(1 for c in cleaned if c.isalpha()) / max(len(cleaned), 1)
    if alpha_ratio > 0.60 and len(cleaned) >= 8:
        # Has Spanish-looking words
        spanish_indicators = ['ción', 'sión', 'dad', 'miento', 'mient', 'mien',
                              'nte', 'ble', 'ción', 'ción', 'ción',
                              'aje', 'ero', 'era', 'ista', 'dor', 'dera']
        has_spanish = any(s in lowered for s in spanish_indicators)
        if has_spanish or alpha_ratio > 0.75:
            return ("concepto inexistente", "concepto_limpio",
                    f"Concepto limpio no cubierto por catalogo: '{stripped[:50]}'")

    # ── 17. REGEX FALTANTE ──
    regex_patterns = [
        (r'^(?:GASTOS?\s+DE\s+|GASTOS?\s+POR\s+)', 'GASTOS DE/POR ...'),
        (r'^(?:INGRESOS?\s+POR\s+|INGRESOS?\s+DE\s+)', 'INGRESOS POR/DE ...'),
        (r'^(?:PROVISION\s+|PROVISIÓN\s+)', 'PROVISION ...'),
        (r'^(?:PROVEEDOR(?:ES)?\s+)', 'PROVEEDORES ...'),
        (r'^(?:CUENTA\s+(?:CORRIENTE|CTE)\s+)', 'CUENTA CORRIENTE ...'),
        (r'\b(?:COSTO\s+(?:DE\s+)?VENTA|COSTO\s+DE\s+EXPLOTACI[OÓ]N)\b', 'COSTO DE VENTA/EXPLOTACION'),
    ]
    for pat, desc in regex_patterns:
        if re.search(pat, lowered):
            return ("regex faltante", f"patron_{re.sub(r'[^a-z]', '_', desc.lower())[:30]}",
                    f"Patron detectable por regex: {desc}: '{stripped[:50]}'")

    # ── 18. OCR FALLO: only digits and separators ──
    if digit_ratio > 0.5:
        return ("OCR", "solo_digitos",
                f"Mayoritariamente digitos (posible scanner): '{stripped[:50]}'")

    # ── 19. ERROR HUMANO ──
    # Check for common typos that aren't OCR
    typos = [
        (r'\bemdicato\b', 'EMDICATO -> MEDICATO'),
        (r'\bfrovicion\b', 'FROVICION -> PROVISION'),
        (r'\bfrovision\b', 'FROVISION -> PROVISION'),
        (r'\bcessantia\b', 'CESSANTIA -> CESANTIA'),
        (r'\bhorasextra\b', 'HORASEXTRA -> HORAS EXTRA'),
        (r'\bhorasextras\b', 'HORASEXTRAS -> HORAS EXTRA'),
    ]
    for pat, desc in typos:
        if re.search(pat, lowered):
            return ("error humano", "tipo",
                    f"Posible typo: {desc}: '{stripped[:50]}'")

    # ── 20. FALLBACK: CATÁLOGO INCOMPLETO (by default for clean-ish names) ──
    if alpha_ratio > 0.40 and len(cleaned) > 5:
        return ("catálogo incompleto", "no_cubierto",
                f"Nombre con contenido pero no en catalogo: '{stripped[:50]}'")

    # ── 21. LAST RESORT: RUINDO ──
    return ("ruido", "no_clasificado",
            f"No se pudo clasificar en otra categoria: '{stripped[:50]}'")


def main() -> None:
    print("=== Root-Cause Analysis: 5,392 DEv2 UNKNOWN Accounts ===\n")

    # Load
    with open(ROOT / "reports" / "decision_engine_shadow.json") as f:
        de = json.load(f)
    unknown = [r for r in de["rows"] if not r.get("de_v2_code")]
    print(f"UNKNOWN accounts to classify: {len(unknown)}")

    # Classify each
    classifications: list[dict[str, Any]] = []
    category_counts: Counter[str] = Counter()
    sub_counts: Counter[str] = Counter()
    source_counts: Counter[str] = Counter()
    file_category: dict[str, Counter] = defaultdict(Counter)
    type_category: dict[str, Counter] = defaultdict(Counter)

    for r in unknown:
        name = str(r.get("account_name", "")).strip()
        a_type = str(r.get("account_type", ""))
        source_file = r.get("source_file", r.get("documento", ""))

        cause, sub, explanation = classify(name)
        category_counts[cause] += 1
        sub_counts[f"{cause}::{sub}"] += 1
        source_counts[source_file] += 1
        file_category[source_file][cause] += 1
        type_category[a_type][cause] += 1

        classifications.append({
            "nombre_original": name,
            "account_code": r.get("account_code", ""),
            "root_cause": cause,
            "sub_cause": sub,
            "explanation": explanation,
            "source_file": source_file,
            "account_type": a_type,
            "phase1_method": r.get("phase1_method", ""),
        })

    # ── Aggregate & Pareto ──
    total = len(unknown)

    # Pareto: what % would be resolved per category?
    # Calculate cumulative and remaining
    categories_ranked = sorted(category_counts.items(), key=lambda x: -x[1])
    pareto: list[dict] = []
    cumulative = 0
    for cause, count in categories_ranked:
        pct = count / total * 100
        cumulative += pct
        remaining = max(0, 100 - cumulative)
        # ROI estimate: effort (1-5) / impact
        roi_effort = {
            "ruido": 1,
            "parser": 2,
            "OCR": 3,
            "concepto inexistente": 4,
            "catálogo incompleto": 3,
            "regex faltante": 2,
            "concepto ambiguo": 2,
            "nombre propio": 1,
            "sigla": 1,
            "abreviatura": 2,
            "empresa": 1,
            "error humano": 1,
            "cuenta compuesta": 3,
            "cuenta tributaria": 2,
            "cuenta especial": 5,
        }.get(cause, 3)
        # Impact = % of UNKNOWN that would be resolved
        # ROI = impact / effort
        roi = round(pct / roi_effort, 1) if roi_effort > 0 else 0
        pareto.append({
            "categoria": cause,
            "cuentas": count,
            "porcentaje": round(pct, 1),
            "porcentaje_acumulado": round(cumulative, 1),
            "porcentaje_restante": round(remaining, 1),
            "esfuerzo": roi_effort,
            "ROI": roi,
            "descripcion": desc_for_cause(cause, count, pct),
        })

    # Sort by ROI descending for prioritization
    pareto_sorted = sorted(pareto, key=lambda x: -x["ROI"])

    print(f"\n{'=' * 70}")
    print(f"  Pareto — ordered by ROI")
    print(f"{'=' * 70}")
    print(f"  {'Categoria':30s} {'Cuentas':>8s} {'%':>6s} {'Acum%':>6s} {'Esf.':>5s} {'ROI':>6s}")
    print(f"  {'-'*30} {'-'*8} {'-'*6} {'-'*6} {'-'*5} {'-'*6}")
    for p in pareto_sorted:
        print(f"  {p['categoria']:30s} {p['cuentas']:8d} {p['porcentaje']:5.1f}% {p['porcentaje_acumulado']:5.1f}% {p['esfuerzo']:5d} {p['ROI']:5.1f}")
    print()

    # ── Output ──
    output = {
        "metadata": {
            "version": "1.0",
            "date": "2026-07-22",
            "total_unknown": total,
            "categories": ROOT_CAUSES,
        },
        "pareto": pareto_sorted,
        "classifications": classifications,
        "summary": {
            "by_category": dict(category_counts.most_common()),
            "by_sub_cause": dict(sub_counts.most_common(50)),
            "by_source_file_source": dict(source_counts.most_common(20)),
        },
    }

    # Write JSON
    write_json(output)
    # Write MD
    write_md(output, pareto_sorted, category_counts)
    # Write XLSX
    write_xlsx(output, pareto_sorted, classifications)

    print(f"\nDone. Files written to reports/root_cause_unknown.*")
    verify(output)


def desc_for_cause(cause: str, count: int, pct: float) -> str:
    descriptions = {
        "ruido": f"Lineas estructurales PDF, encabezados, pies de pagina, totales — {count} cuentas ({pct:.1f}%) no son cuentas reales",
        "parser": f"Datos de parser embebidos en el nombre (folios, RUTs, montos, fechas) — {count} cuentas ({pct:.1f}%)",
        "OCR": f"Artefactos de OCR (basura inicial, palabras fusionadas, sustituciones digito/letra) — {count} cuentas ({pct:.1f}%)",
        "concepto inexistente": f"Nombre limpio en espanol no cubierto por el catalogo — {count} cuentas ({pct:.1f}%)",
        "catálogo incompleto": f"Variantes ortograficas o nombres cercanos a conceptos existentes — {count} cuentas ({pct:.1f}%)",
        "regex faltante": f"Patrones detectables via regex que no estan implementados — {count} cuentas ({pct:.1f}%)",
        "concepto ambiguo": f"Nombre que puede mapear a multiples codigos — {count} cuentas ({pct:.1f}%)",
        "nombre propio": f"Nombres de personas usados como cuentas — {count} cuentas ({pct:.1f}%)",
        "sigla": f"Acronimos sin expansion en catalogo (AFP, IVA, PPM, etc.) — {count} cuentas ({pct:.1f}%)",
        "abreviatura": f"Abreviaturas no estandarizadas (PROV., SERV., GTA., etc.) — {count} cuentas ({pct:.1f}%)",
        "empresa": f"Razones sociales, direcciones, datos de empresa — {count} cuentas ({pct:.1f}%)",
        "error humano": f"Errores de tipeo en el PDF original — {count} cuentas ({pct:.1f}%)",
        "cuenta compuesta": f"Dos conceptos unidos por 'Y' que deberian separarse — {count} cuentas ({pct:.1f}%)",
        "cuenta tributaria": f"Conceptos tributarios sin cobertura en catalogo — {count} cuentas ({pct:.1f}%)",
        "cuenta especial": f"Casos especiales no encajan en otras categorias — {count} cuentas ({pct:.1f}%)",
    }
    return descriptions.get(cause, cause)


def write_json(output: dict) -> None:
    path = ROOT / "reports" / "root_cause_unknown.json"
    with open(path, "w") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)
    print(f"  Wrote {path}  ({len(output['classifications'])} rows, {len(str(json.dumps(output)))/1024:.0f} KB)")


def write_md(output: dict, pareto: list[dict], cat_counts: Counter) -> None:
    total = output["metadata"]["total_unknown"]

    lines = [
        "# Root-Cause Analysis — DEv2 UNKNOWN Accounts",
        "",
        f"**Date:** 2026-07-22  ",
        f"**Total UNKNOWN:** {total} accounts  ",
        f"**Source:** `decision_engine_shadow.json` — DEv2 shadow run on 11,690 accounts  ",
        "",
        "---",
        "",
        "## Executive Summary",
        "",
        f"Of {total} UNKNOWN accounts, the root causes distribute as follows. "
        "Each account is assigned to exactly ONE root cause.",
        "",
        "**Quick wins (ROI ≥ 20):** categories where fixing 1 concept resolves 20+ accounts.  ",
        "**Long tail (ROI < 5):** categories requiring per-account effort.",
        "",
        "---",
        "",
        "## Pareto Analysis (sorted by ROI)",
        "",
        "| # | Categoria | Cuentas | % | Acumulado % | Esfuerzo (1-5) | ROI | Descripción |",
        "|---|---|---:|---:|---:|---:|---:|---|",
    ]
    for i, p in enumerate(pareto, 1):
        lines.append(
            f"| {i} | {p['categoria']} | {p['cuentas']} | {p['porcentaje']:.1f}% | "
            f"{p['porcentaje_acumulado']:.1f}% | {p['esfuerzo']} | {p['ROI']} | {p['descripcion']} |"
        )

    lines += [
        "",
        "---",
        "",
        "## Category Details",
        "",
    ]

    # Detail section for each category
    for p in pareto:
        cat = p["categoria"]
        count = p["cuentas"]
        pct = p["porcentaje"]

        # Get samples
        samples = [c for c in output["classifications"] if c["root_cause"] == cat][:15]

        # Get sub-reasons
        from collections import Counter
        sub_counts = Counter(c["sub_cause"] for c in output["classifications"] if c["root_cause"] == cat)

        lines += [
            f"### {cat.title()} ({count} · {pct:.1f}%)",
            "",
            p["descripcion"],
            "",
            "**Sub-causas:**",
            "",
        ]
        for sub, n in sub_counts.most_common():
            lines.append(f"- `{sub}`: {n} cuentas")
        lines += [
            "",
            "**Muestras:**",
            "",
        ]
        for s in samples[:10]:
            expl = s["explanation"][:80]
            lines.append(f"- `{s['nombre_original'][:60]}` → {expl}")
        if len(samples) > 10:
            lines.append(f"- ... and {count - 10} more")
        lines += [
            "",
            "---",
            "",
        ]

    lines += [
        "",
        "## Recommendations",
        "",
        "### Phase 1 — No-code fixes (ROI > 10)",
        "",
        "| Category | Action | Impact |",
        "|---|---|---|",
    ]
    for p in pareto:
        if p["ROI"] < 10:
            continue
        action = action_for_cause(p["categoria"])
        lines.append(f"| {p['categoria']} | {action} | -{p['porcentaje']:.1f}% UNKNOWN |")

    lines += [
        "",
        "### Phase 2 — Catalog & regex (ROI 5-10)",
        "",
        "| Category | Action | Impact |",
        "|---|---|---|",
    ]
    for p in pareto:
        if not (5 <= p["ROI"] < 10):
            continue
        action = action_for_cause(p["categoria"])
        lines.append(f"| {p['categoria']} | {action} | -{p['porcentaje']:.1f}% UNKNOWN |")

    lines += [
        "",
        "### Phase 3 — Long tail (ROI < 5)",
        "",
        "| Category | Action | Impact |",
        "|---|---|---|",
    ]
    for p in pareto:
        if p["ROI"] >= 5:
            continue
        action = action_for_cause(p["categoria"])
        lines.append(f"| {p['categoria']} | {action} | -{p['porcentaje']:.1f}% UNKNOWN |")

    lines += [
        "",
        "---",
        "",
        "## Methodology",
        "",
        "Each of the 5,392 UNKNOWN accounts was classified into exactly one root cause using rule-based detection:",
        "",
        "1. **ruido** — regex detection of PDF structural elements (Nivel, Desde, Folio, Página, Total, fechas solas)",
        "2. **parser** — detection of embedded parser artifacts (RUTs, amounts, dates in account names)",
        "3. **OCR** — leading garbage characters, merged words, digit-letter substitutions",
        "4. **empresa** — company suffixes (SA, Ltda, EIRL, SPA), addresses",
        "5. **nombre propio** — honorifics, personal names",
        "6. **sigla** — known Chilean acronyms (AFP, IVA, PPM, FONASA, etc.)",
        "7. **abreviatura** — words with periods, known abbreviations",
        "8. **cuenta tributaria** — tax keywords (impuesto, ppm, iva, retención)",
        "9. **cuenta compuesta** — two concepts joined by 'Y'",
        "10. **catálogo incompleto** — near-miss variants of existing concepts",
        "11. **regex faltante** — patterns detectable by regex (GASTOS DE ..., PROVISION ...)",
        "12. **concepto inexistente** — clean Spanish names not covered",
        "13. **error humano** — typos, misspellings",
        "14. **ruido (fallback)** — unclassifiable garbage",
        "",
        "**Note:** Categories are mutually exclusive. Classification priority follows the order above. "
        "Some accounts in lower-priority categories might also have characteristics of higher-priority ones.",
        "",
    ]

    path = ROOT / "reports" / "root_cause_unknown.md"
    with open(path, "w") as f:
        f.write("\n".join(lines) + "\n")
    print(f"  Wrote {path}  ({len(lines)} lines)")


def action_for_cause(cause: str) -> str:
    actions = {
        "ruido": "Agregar filtro de lineas estructurales PDF en preprocesamiento (Nivel, Desde, Folio, Página, Total)",
        "parser": "Mejorar parser para extraer solo el nombre de cuenta, no datos adjuntos",
        "OCR": "Agregar pipeline de limpieza OCR: eliminar basura inicial, separar palabras fusionadas, corregir sustituciones",
        "concepto inexistente": "Ampliar catalogo de conceptos con ~200-300 nuevas entradas",
        "catálogo incompleto": "Agregar variantes ortograficas y sinonimos al catalogo existente (~150 entradas)",
        "regex faltante": "Implementar ~30 patrones regex adicionales (GASTOS DE, PROVISION, INGRESOS POR, etc.)",
        "concepto ambiguo": "Implementar desambiguacion por tipo de cuenta o contexto",
        "nombre propio": "Filtrar o normalizar nombres de personas (bajo ROI, mantener como UNKNOWN)",
        "sigla": "Agregar ~50 siglas al catalogo con su expansion",
        "abreviatura": "Crear diccionario de abreviaturas (~80 entradas) con expansion automatica",
        "empresa": "Filtrar razones sociales del pipeline de clasificacion (no son cuentas contables)",
        "error humano": "Implementar corrector ortografico basico para variantes conocidas",
        "cuenta compuesta": "Dividir cuentas compuestas en dos conceptos separados",
        "cuenta tributaria": "Agregar ~30 conceptos tributarios faltantes al catalogo",
        "cuenta especial": "Analisis caso a caso",
    }
    return actions.get(cause, cause)


def write_xlsx(output: dict, pareto: list[dict], classifications: list[dict]) -> None:
    try:
        import openpyxl
    except ImportError:
        print("  Skipping XLSX (openpyxl not available)")
        return

    wb = openpyxl.Workbook()

    # Sheet 1: Pareto
    ws = wb.active
    ws.title = "Pareto"
    pareto_keys = ["categoria", "cuentas", "porcentaje", "porcentaje_acumulado",
                    "porcentaje_restante", "esfuerzo", "ROI", "descripcion"]
    ws.append(pareto_keys)
    for p in pareto:
        ws.append([p.get(k) for k in pareto_keys])

    # Sheet 2: All classifications
    ws2 = wb.create_sheet("Clasificaciones")
    cls_keys = ["nombre_original", "account_code", "root_cause", "sub_cause",
                "explanation", "source_file"]
    ws2.append(cls_keys)
    for c in classifications:
        ws2.append([c.get(k) for k in cls_keys])

    # Sheet 3: Summary by category
    ws3 = wb.create_sheet("Resumen por Causa")
    ws3.append(["Categoria", "Cuentas", "%", "Sub-causas principales"])
    from collections import Counter
    for p in pareto:
        sub_causes = Counter(c["sub_cause"] for c in classifications if c["root_cause"] == p["categoria"])
        top_subs = "; ".join(f"{s}:{n}" for s, n in sub_causes.most_common(5))
        ws3.append([p["categoria"], p["cuentas"], f"{p['porcentaje']}%", top_subs])

    path = ROOT / "reports" / "root_cause_unknown.xlsx"
    wb.save(str(path))
    print(f"  Wrote {path}")


def verify(output: dict) -> None:
    """Cross-check that all accounts are classified and categories are valid."""
    total = len(output["classifications"])
    cats = set(c["root_cause"] for c in output["classifications"])
    invalid = cats - set(ROOT_CAUSES)
    if invalid:
        print(f"  WARNING: Invalid categories: {invalid}")
    assert total == output["metadata"]["total_unknown"]
    assert not invalid, f"Invalid categories found: {invalid}"
    print(f"  Verified: {total} accounts, {len(cats)} categories, all valid.")


if __name__ == "__main__":
    main()
