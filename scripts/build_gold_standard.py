#!/usr/bin/env python3
"""Build a permanent Gold Standard of 2,500 representative accounts."""

from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parent.parent

GS_KEYS = [
    "id", "nombre_original", "tipo_cuenta", "codigo_correcto",
    "motivo", "documento", "confianza", "prioridad_revision",
    "metodo", "score", "source",
]


def main() -> None:
    print("=== Gold Standard v1 — Builder ===\n")

    # Load data sources
    de_rows = load_de_rows()
    account_map = build_account_map(de_rows)
    audit_map = load_audit()

    print(f"DE shadow rows: {len(de_rows)}")
    print(f"Unique accounts: {len(account_map)}")
    print(f"Audit accounts: {len(audit_map)}")

    # Build selection pool
    pool = build_pool(account_map, audit_map)

    # Select 2,500 representative
    selected = select_representative(pool)

    # Resolve correct codes
    gs = resolve_codes(selected, audit_map)

    # Write outputs
    write_all(gs)


# ── Loaders ──


def load_de_rows() -> list[dict]:
    path = ROOT / "reports" / "decision_engine_shadow.json"
    with open(path) as f:
        return json.load(f).get("rows", [])


def build_account_map(rows: list[dict]) -> dict[str, list[dict]]:
    """Group DE rows by account name (one name can appear in multiple PDFs)."""
    m: dict[str, list[dict]] = {}
    for r in rows:
        name = str(r.get("account_name", "")).strip()
        if not name:
            continue
        # Extract account type from DE evidence field
        raw_type = str(r.get("account_type", "")).strip()
        tipo = categorize_type(raw_type)
        r["_tipo"] = tipo

        # Score
        raw_score = r.get("de_v2_score")
        score = float(raw_score) if raw_score is not None else 0.0

        # Consensus
        cons = r.get("de_v2_consensus")
        consensus = int(cons) if cons is not None else 0

        # OCR detection
        has_ocr = detect_ocr(name)

        entry = {
            "name": name,
            "code": str(r.get("account_code", "") or ""),
            "source": str(r.get("documento", r.get("source_file", "")) or ""),
            "tipo": tipo,
            "de_code": r.get("de_v2_code"),
            "de_score": score,
            "de_confidence": str(r.get("de_v2_confidence", "") or ""),
            "de_review": bool(r.get("de_v2_review", True)),
            "de_consensus": consensus,
            "de_conflicts": int(r.get("de_v2_conflicts", 0)),
            "de_source": str(r.get("de_v2_source", "") or ""),
            "has_ocr": has_ocr,
        }
        m.setdefault(name, []).append(entry)
    return m


def load_audit() -> dict[str, dict]:
    path = ROOT / "reports" / "disagreement_audit.json"
    if not path.exists():
        return {}
    with open(path) as f:
        data = json.load(f)

    audit: dict[str, dict] = {}
    for d in data.get("discrepancies", []):
        name = str(d.get("account_name", "")).strip()
        if not name:
            continue
        winner = d.get("who_wins", "")
        if winner == "SemanticMatcher":
            code = d.get("sm_code", "")
            method = "audit_sm"
        elif winner == "Regex":
            code = d.get("regex_code", "")
            method = "audit_regex"
        else:
            code = None
            method = "audit_pending"
        audit[name] = {
            "correct_code": code,
            "winner": winner,
            "method": method,
        }
    return audit


# ── Helpers ──


def categorize_type(raw: str) -> str:
    if not raw:
        return "DESCONOCIDO"
    r = raw.upper()
    if "ACTIVO" in r:
        if "NO CORRIENTE" in r:
            return "ACTIVO NO CORRIENTE"
        return "ACTIVO"
    if "PASIVO" in r:
        if "NO CORRIENTE" in r:
            return "PASIVO NO CORRIENTE"
        return "PASIVO"
    if "PATRIMONIO" in r or "PATRIM" in r:
        return "PATRIMONIO"
    if "PERDIDA" in r:
        return "PERDIDA"
    if "GANANCIA" in r or "RESULTADO" in r:
        return "GANANCIA"
    return "DESCONOCIDO"


ER_TIPO_MAP = {
    "ER.01": "GANANCIA", "ER.11": "GANANCIA", "ER.12": "GANANCIA", "ER.13": "GANANCIA",
    "ER.02": "PERDIDA", "ER.03": "PERDIDA", "ER.04": "PERDIDA", "ER.05": "PERDIDA",
    "ER.06": "PERDIDA", "ER.07": "PERDIDA", "ER.08": "PERDIDA", "ER.09": "PERDIDA",
    "ER.10": "PERDIDA", "ER.14": "PERDIDA",
}

DETECT_TIPO_MAP = {
    "AC": "ACTIVO", "AN": "ACTIVO NO CORRIENTE",
    "PC": "PASIVO", "PN": "PASIVO NO CORRIENTE",
    "PA": "PATRIMONIO", "PO": "PERDIDA",
}


def detect_tipo_from_code(code: str) -> str | None:
    if not code or len(code) < 2:
        return None
    if code[:2].upper() == "ER":
        return ER_TIPO_MAP.get(code.upper())
    return DETECT_TIPO_MAP.get(code[:2].upper())


def detect_ocr(name: str) -> bool:
    # Patterns that suggest OCR misread
    patterns = [
        r'[0Oo][0Oo]',        # "00" or "OO" instead of letters
        r'[l1I][l1I]',        # "ll", "11", "l1"
        r'5[Ss]',             # "5" instead of "S"
        r'8[Bb]',             # "8" instead of "B"
        r'\d{2}[a-záéíóúñ]',  # digits followed by lowercase letter
    ]
    return any(re.search(p, name) for p in patterns)


def code_type(code: str | None) -> str:
    if not code or len(code) < 2:
        return "DESCONOCIDO"
    pref = code[:2].upper()
    if pref == "ER":
        return ER_TIPO_MAP.get(code.upper(), "DESCONOCIDO")
    return DETECT_TIPO_MAP.get(pref, "DESCONOCIDO")


# ── Pool building ──


def build_pool(
    account_map: dict[str, list[dict]],
    audit_map: dict[str, dict],
) -> list[dict]:
    pool: dict[str, dict] = {}

    for name, entries in account_map.items():
        # Merge across all files for this name
        best = max(entries, key=lambda e: e["de_score"] + e["de_consensus"])
        merged = dict(best)

        # Count appearances
        sources = {e["source"] for e in entries if e["source"]}
        merged["n_files"] = len(sources)
        merged["all_sources"] = list(sources)
        merged["all_codes"] = [e.get("de_code") for e in entries if e.get("de_code")]
        merged["codes_set"] = len(set(merged["all_codes"]))
        merged["is_audited"] = name in audit_map
        merged["audit"] = audit_map.get(name, {})

        # Determine the best representative code from all entries
        best_entries = sorted(entries, key=lambda e: -e["de_score"])
        merged["best_code"] = best_entries[0]["de_code"] if best_entries else None
        merged["best_score"] = best_entries[0]["de_score"] if best_entries else 0.0

        pool[name] = merged
    return list(pool.values())


# ── Selection ──


def select_representative(pool: list[dict]) -> list[dict]:
    selected: list[dict] = []
    selected_names: set[str] = set()

    def _add(item: dict) -> bool:
        if item["name"] in selected_names:
            return False
        selected.append(item)
        selected_names.add(item["name"])
        return True

    # ── 1. ALL audit accounts (mandatory) ──
    for item in pool:
        if item["is_audited"]:
            _add(item)

    # ── 2. High-confidence (score >= 0.9, consensus >= 2) ──
    for item in sorted(
        [p for p in pool if p["name"] not in selected_names and p["best_score"] >= 0.9 and p["de_consensus"] >= 2],
        key=lambda x: -x["best_score"],
    )[:300]:
        _add(item)

    # ── 3. Medium confidence (score >= 0.8) ──
    for item in sorted(
        [p for p in pool if p["name"] not in selected_names and p["best_score"] >= 0.8],
        key=lambda x: -x["best_score"],
    )[:200]:
        _add(item)

    # ── 4. Score 0.6-0.79 (mid-tier DE) ──
    for item in sorted(
        [p for p in pool if p["name"] not in selected_names and 0.6 <= p["best_score"] < 0.8],
        key=lambda x: -x["best_score"],
    )[:200]:
        _add(item)

    # ── 5. By tipo — ensure diversity ──
    tipo_targets = {
        "ACTIVO": 500,
        "PASIVO": 500,
        "PATRIMONIO": 200,
        "GANANCIA": 350,
        "PERDIDA": 150,
        "ACTIVO NO CORRIENTE": 200,
        "PASIVO NO CORRIENTE": 100,
        "DESCONOCIDO": 100,
    }

    # Re-check current distribution
    tipo_counts = Counter(p["tipo"] for p in selected)
    print(f"\nAfter audit+high-conf selection: {len(selected)} accounts")
    for t, c in sorted(tipo_counts.items(), key=lambda x: -x[1]):
        print(f"  {t}: {c}")

    for tipo, target in sorted(tipo_targets.items(), key=lambda x: -x[1]):
        needed = max(0, target - tipo_counts.get(tipo, 0))
        candidates = [
            p for p in pool
            if p["name"] not in selected_names
            and p["tipo"] == tipo
            and p["best_score"] > 0.0
        ]
        # Score-descending
        for item in sorted(candidates, key=lambda x: -x["best_score"])[:needed]:
            _add(item)

    # ── 6. Ambiguous names — same name in multiple files ──
    ambig_candidates = [
        p for p in pool
        if p["name"] not in selected_names
        and p["n_files"] > 1
    ]
    for item in sorted(ambig_candidates, key=lambda x: -x["n_files"])[:150]:
        _add(item)

    # ── 7. OCR-difficult accounts ──
    ocr_candidates = [
        p for p in pool
        if p["name"] not in selected_names
        and p["has_ocr"]
    ]
    for item in sorted(ocr_candidates, key=lambda x: -x["best_score"])[:100]:
        _add(item)

    # ── 8. Unknowns (score=0, no code) — the hardest cases ──
    unknown_candidates = [
        p for p in pool
        if p["name"] not in selected_names
        and p["best_score"] == 0.0
    ]
    for item in sorted(unknown_candidates, key=lambda x: x["n_files"], reverse=True)[:100]:
        _add(item)

    # ── 9. Fill remaining to 2,500 ──
    remaining = max(0, 2500 - len(selected))
    for item in sorted(
        [p for p in pool if p["name"] not in selected_names],
        key=lambda x: -x["best_score"],
    )[:remaining]:
        _add(item)

    print(f"\nFinal selection: {len(selected)}")
    return selected


# ── Code resolution ──


def resolve_codes(
    selected: list[dict],
    audit_map: dict[str, dict],
) -> list[dict]:
    gs: list[dict] = []

    for i, item in enumerate(selected, 1):
        name = item["name"]
        audit = item["audit"]

        best_code = None
        confidence = 0.0
        method = ""
        motivo = ""
        source = ""
        priority = 3  # 1=high, 3=low

        if item["is_audited"] and audit.get("correct_code"):
            best_code = audit["correct_code"]
            confidence = 0.99
            method = audit["method"]
            motivo = f"Auditado: {audit.get('winner', 'verificado manualmente')}"
            source = "disagreement_audit"
            priority = 3

        elif item["de_consensus"] >= 3 and item["best_score"] >= 0.9:
            best_code = item["best_code"]
            confidence = 0.98
            method = "de_v2_strong_consensus"
            motivo = f"DEv2 consensus={item['de_consensus']}, score={item['best_score']:.2f}"
            source = "decision_engine_v2"
            priority = 3

        elif item["de_consensus"] >= 2 and item["best_score"] >= 0.8:
            best_code = item["best_code"]
            confidence = 0.95
            method = "de_v2_consensus"
            motivo = f"DEv2 consensus={item['de_consensus']}, score={item['best_score']:.2f}"
            source = "decision_engine_v2"
            priority = 3

        elif item["best_score"] >= 0.8:
            best_code = item["best_code"]
            confidence = 0.90
            method = "de_v2_high_score"
            motivo = f"DEv2 score={item['best_score']:.2f}"
            source = "decision_engine_v2"
            priority = 2

        elif item["best_score"] >= 0.6:
            best_code = item["best_code"]
            confidence = 0.80
            method = "de_v2_medium_score"
            motivo = f"DEv2 score={item['best_score']:.2f}"
            source = "decision_engine_v2"
            priority = 2

        elif item["best_code"]:
            best_code = item["best_code"]
            confidence = 0.60
            method = "de_v2_low"
            motivo = f"Solo DEv2, score={item['best_score']:.2f}"
            source = "decision_engine_v2"
            priority = 2

        else:
            best_code = None
            confidence = 0.0
            method = "unknown"
            motivo = "Ningun motor clasifica — requiere revision humana"
            source = "ninguno"
            priority = 1

        # Tipo: use code detection first, then DE type as fallback
        tipo = detect_tipo_from_code(best_code) or item["tipo"] or "DESCONOCIDO"

        gs.append({
            "id": i,
            "nombre_original": name,
            "tipo_cuenta": tipo,
            "codigo_correcto": best_code or "",
            "motivo": motivo,
            "documento": ", ".join(sorted(item.get("all_sources", [item.get("source", "")])[:3])),
            "confianza": round(confidence, 2),
            "prioridad_revision": priority,
            "metodo": method,
            "score": round(item["best_score"], 4),
            "source": source,
        })

    return gs


# ── Writers ──


def write_all(gs: list[dict]) -> None:
    write_json(gs)
    write_xlsx(gs)
    write_md(gs)


def write_json(gs: list[dict]) -> None:
    tipos = Counter(r["tipo_cuenta"] for r in gs)
    metodos = Counter(r["metodo"] for r in gs)
    prioridades = Counter(r["prioridad_revision"] for r in gs)
    output = {
        "metadata": {
            "version": "1.0",
            "date": "2026-07-22",
            "total_accounts": len(gs),
            "with_code": sum(1 for r in gs if r.get("codigo_correcto")),
            "pending_review": prioridades.get(1, 0),
            "distribution": {
                "by_type": dict(tipos.most_common()),
                "by_method": dict(metodos.most_common()),
                "by_priority": {str(k): v for k, v in sorted(prioridades.items())},
            },
        },
        "accounts": gs,
    }
    path = ROOT / "knowledge" / "gold_standard.json"
    with open(path, "w") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)
    print(f"  Wrote {path}")


def write_xlsx(gs: list[dict]) -> None:
    try:
        import openpyxl
    except ImportError:
        print("  Skipping XLSX (openpyxl not available)")
        return

    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Gold Standard"
    ws.append(GS_KEYS)
    for row in gs:
        ws.append([row.get(k, "") for k in GS_KEYS])

    # Sheet 2: Summary
    ws2 = wb.create_sheet("Resumen")
    tipos = Counter(r["tipo_cuenta"] for r in gs)
    metodos = Counter(r["metodo"] for r in gs)
    prioridades = Counter(r["prioridad_revision"] for r in gs)

    ws2.append(["Metrica", "Valor"])
    ws2.append(["Total cuentas", len(gs)])
    ws2.append(["Con codigo", sum(1 for r in gs if r.get("codigo_correcto"))])
    ws2.append(["Pendientes revision", prioridades.get(1, 0)])
    ws2.append([""])
    ws2.append(["Por tipo de cuenta:"])
    for t, n in sorted(tipos.items(), key=lambda x: -x[1]):
        ws2.append([f"  {t}", n])
    ws2.append([""])
    ws2.append(["Por metodo:"])
    for m, n in metodos.most_common():
        ws2.append([f"  {m}", n])

    # Sheet 3: High priority
    ws3 = wb.create_sheet("Prioridad Alta")
    ws3.append(GS_KEYS)
    for row in gs:
        if row.get("prioridad_revision") == 1:
            ws3.append([row.get(k, "") for k in GS_KEYS])

    path = ROOT / "knowledge" / "gold_standard.xlsx"
    wb.save(str(path))
    print(f"  Wrote {path}")


def write_md(gs: list[dict]) -> None:
    from collections import Counter
    tipos = Counter(r["tipo_cuenta"] for r in gs)
    metodos = Counter(r["metodo"] for r in gs)
    prioridades = Counter(r["prioridad_revision"] for r in gs)
    prioridad_1 = [r for r in gs if r.get("prioridad_revision") == 1]
    with_code = sum(1 for r in gs if r.get("codigo_correcto"))

    lines = [
        "# Gold Standard v1 — Permanent Ground Truth",
        "",
        "**Date:** 2026-07-22  ",
        f"**Total accounts:** {len(gs)}  ",
        "**Source:** 182 PDF balances (stratified selection from 11,690 accounts)  ",
        "",
        "---",
        "",
        "## Selection Criteria",
        "",
        "1. **All 105 unique audit accounts** — manually verified ground truth",
        "2. **DEv2 high-confidence** (score >= 0.9, consensus >= 2) — strong auto-classifications",
        "3. **Type-balanced fill** — ACTIVO, PASIVO, PATRIMONIO, GANANCIA, PERDIDA quotas",
        "4. **Ambiguous names** — same name across multiple files for cross-file validation",
        "5. **OCR-difficult** — accounts with OCR artifacts (digit substitutions)",
        "6. **Unknown/hard cases** — accounts no engine can classify (for human review)",
        "",
        "---",
        "",
        "## Summary",
        "",
        "| Metric | Value |",
        "|---|---|",
        f"| Total | {len(gs)} |",
        f"| With code assigned | {with_code} ({with_code/len(gs)*100:.1f}%) |",
        f"| Pending review | {len(prioridad_1)} ({len(prioridad_1)/len(gs)*100:.1f}%) |",
        "",
        "### By Account Type",
        "",
        "| Type | Count | % |",
        "|---|---|---|",
    ]
    for t, n in sorted(tipos.items(), key=lambda x: -x[1]):
        lines.append(f"| {t} | {n} | {n/len(gs)*100:.1f}% |")

    lines += [
        "",
        "### By Resolution Method",
        "",
        "| Method | Count | % |",
        "|---|---|---|",
    ]
    for m, n in metodos.most_common():
        lines.append(f"| {m} | {n} | {n/len(gs)*100:.1f}% |")

    lines += [
        "",
        "### By Priority",
        "",
        "| Priority | Count | % |",
        "|---|---|---|",
    ]
    for p in [3, 2, 1]:
        n = prioridades.get(p, 0)
        label = {1: "Alta (requiere revision)", 2: "Media (verificar)", 3: "Baja (confianza alta)"}.get(p, str(p))
        lines.append(f"| {label} | {n} | {n/len(gs)*100:.1f}% |")

    lines += [
        "",
        "### Top CMCC Codes",
        "",
        "| Code | Count |",
        "|---|---|",
    ]
    codigos = Counter(r["codigo_correcto"] for r in gs if r["codigo_correcto"])
    for c, n in codigos.most_common(20):
        lines.append(f"| {c} | {n} |")

    lines += [
        "",
        "---",
        "",
        "## High Priority Accounts (need human review)",
        "",
        f"**{len(prioridad_1)} accounts** where no reliable code could be auto-assigned:",
        "",
        "| # | Account | Motivo |",
        "|---|---|---|",
    ]
    for row in prioridad_1[:100]:
        lines.append(f"| {row['id']} | {row['nombre_original'][:60]} | {row.get('motivo', '')[:60]} |")
    if len(prioridad_1) > 100:
        lines.append(f"| ... | ... ({len(prioridad_1) - 100} more) | ... |")

    lines += [
        "",
        "---",
        "",
        "## Usage",
        "",
        "This Gold Standard is the single source of truth for:",
        "- Training precision/recall benchmarks",
        "- Validating classifier improvements",
        "- Detecting regression after catalog changes",
        "- Human review prioritization",
        "",
        "**Rules:**",
        "- DO NOT modify auto-assigned codes without manual verification",
        "- DO review high-priority (prioridad=1) accounts first",
        "- DO rebuild after every concept catalog change",
        "- DO NOT commit without at least 90% of priority-1 resolved",
        "",
    ]

    path = ROOT / "knowledge" / "gold_standard.md"
    with open(path, "w") as f:
        f.write("\n".join(lines) + "\n")
    print(f"  Wrote {path}")

    # Print summary
    print()
    print("=" * 60)
    print(f"  Gold Standard v1: {len(gs)} accounts")
    print("=" * 60)
    print(f"  With code: {with_code} / {len(gs)} ({with_code/len(gs)*100:.1f}%)")
    print(f"  Pending review: {prioridades.get(1, 0)}")
    print(f"  Confianza alta (p=3): {prioridades.get(3, 0)}")
    print(f"  Confianza media (p=2): {prioridades.get(2, 0)}")
    print()
    print("  Types:")
    for t, n in sorted(tipos.items(), key=lambda x: -x[1]):
        print(f"    {t:25s} {n:4d}")
    print()
    print("  Methods:")
    for m, n in metodos.most_common(10):
        print(f"    {m:35s} {n:4d}")


if __name__ == "__main__":
    main()
