#!/usr/bin/env python3
"""Benchmark: compare DecisionEngineV2 against Phase 1 pipeline on 11,696 accounts."""

from __future__ import annotations

import json
import logging
import time
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from decision_v2.engine import DecisionEngineV2

logging.basicConfig(level=logging.WARNING, format="%(levelname)s %(message)s")
logging.getLogger("learning.engine").setLevel(logging.ERROR)
logging.getLogger("decision_v2.engine").setLevel(logging.WARNING)
logger = logging.getLogger(__name__)

ROOT = Path(__file__).resolve().parent.parent
REPORTS = ROOT / "reports"

# ── Category labels for comparison ──

CATEGORY_ORDER = [
    "ambos_correctos",
    "pipeline_correcto",
    "de_v2_correcto",
    "ambos_incorrectos",
    "necesita_revision",
    "parser",
    "ocr",
    "catalogo",
    "regex",
    "semantic",
    "codigo",
]

CATEGORY_LABELS = {
    "ambos_correctos": "Ambos correctos",
    "pipeline_correcto": "Pipeline correcto",
    "de_v2_correcto": "Decision Engine V2 correcto",
    "ambos_incorrectos": "Ambos incorrectos",
    "necesita_revision": "Necesita revisión humana",
    "parser": "Error de Parser",
    "ocr": "Error de OCR",
    "catalogo": "Error de Catálogo",
    "regex": "Error de Regex",
    "semantic": "Error Semántico",
    "codigo": "Error de Código",
}

# ── Ground truth for known accounts ──

KNOWN_CODES: dict[str, str] = {}
GS_PATH = ROOT / "gold_standard.db"
if GS_PATH.exists():
    import sqlite3
    try:
        conn = sqlite3.connect(str(GS_PATH))
        for row in conn.execute("SELECT codigo_estandar, nombre_cuenta FROM gold_standard"):
            KNOWN_CODES[row[1].lower().strip()] = row[0]
        conn.close()
    except Exception:
        pass
logger.info("Loaded %d gold standard entries for ground truth", len(KNOWN_CODES))


def main() -> None:
    # ── Load accounts ──
    src = REPORTS / "semantic_matcher_v1.json"
    if not src.exists():
        logger.error("semantic_matcher_v1.json not found at %s", src)
        return
    with open(src) as f:
        data = json.load(f)
    accounts = data["results"]

    # ── Phase 1 source mapping ──
    ph1_method_map = {
        m: f"phase1_{m}" for m in {
            r.get("phase1_method", "unclassified") for r in accounts
        }
    }

    # ── Init DEv2 ──
    de_v2 = DecisionEngineV2(
        concept_catalog_path=str(ROOT / "knowledge" / "concept_catalog.json"),
        dictionary_path=str(ROOT / "diccionario.json"),
        gold_standard_db=str(GS_PATH) if GS_PATH.exists() else "gold_standard.db",
    )

    # ── Process all accounts ──
    rows: list[dict[str, Any]] = []
    errors = 0
    t0 = time.perf_counter()

    for i, acct in enumerate(accounts):
        name = acct.get("account_name", "")
        code = acct.get("account_code", "")
        ph1_code = acct.get("phase1_code")
        ph1_method = acct.get("phase1_method", "unclassified")
        sm_code = acct.get("expected_cmcc")
        tipo = acct.get("tipo")

        try:
            de_result = de_v2.classify(
                account_name=name,
                account_code=code,
                account_tipo=None,
                tipo=tipo,
            )
        except Exception:
            errors += 1
            continue

        rows.append({
            "account_name": name,
            "account_code": code,
            "account_type": tipo,
            "phase1_code": ph1_code,
            "phase1_method": ph1_method,
            "sm_code": sm_code,
            "de_v2_code": de_result.final_code,
            "de_v2_score": de_result.final_score,
            "de_v2_confidence": de_result.confidence_label,
            "de_v2_review": de_result.review_required,
            "de_v2_source": de_result.decision_source,
            "de_v2_consensus": de_result.consensus_count,
            "de_v2_conflicts": de_result.conflict_count,
            "evidence": [e.to_dict() for e in de_result.evidence],
            "explanation": de_result.explanation,
        })

    elapsed = time.perf_counter() - t0
    n = len(rows)
    logger.info("Processed %d accounts in %.1fs (%.1f/s, %d errors)", n, elapsed, n / elapsed if elapsed else 0, errors)

    # ── Analysis ──
    analysis = analyze(rows)
    analysis["metadata"] = {
        "total_accounts": n,
        "processing_time_seconds": round(elapsed, 2),
        "errors": errors,
        "pipeline_label": "Phase 1 (code -> dict -> regex) [default flags]",
        "de_v2_label": "DecisionEngineV2 (GS + code + dict + SM + regex, weighted ensemble)",
        "source_file": str(src),
    }
    analysis["rows"] = rows

    # ── Reports ──
    REPORTS.mkdir(parents=True, exist_ok=True)
    generate_json(analysis)
    generate_md(analysis)
    generate_xlsx(analysis)
    generate_conflicts_xlsx(analysis)
    generate_metrics_json(analysis)

    print_summary(analysis)


# ═══════════════════════════════════════════════════════════════════
# ANALYSIS
# ═══════════════════════════════════════════════════════════════════

def analyze(rows: list[dict[str, Any]]) -> dict[str, Any]:
    analysis: dict[str, Any] = {}

    # ── Phase 3: Comparison per account ──
    comparisons = []
    for r in rows:
        ph1 = r["phase1_code"]
        dv2 = r["de_v2_code"]
        same = (ph1 == dv2) or (not ph1 and not dv2)

        inferred_ground_truth = _infer_ground_truth(r)

        if same:
            if not ph1:
                cat = "ambos_unknown"
            elif ph1 == inferred_ground_truth:
                cat = "ambos_correctos"
            else:
                cat = "ambos_incorrectos"
        else:
            # Different
            ph1_is_gt = ph1 == inferred_ground_truth if ph1 else False
            dv2_is_gt = dv2 == inferred_ground_truth if dv2 else False

            if ph1_is_gt and dv2_is_gt:
                cat = "ambos_correctos"
            elif ph1_is_gt:
                cat = "pipeline_correcto"
            elif dv2_is_gt:
                cat = "de_v2_correcto"
            elif not ph1 and dv2:
                # Pipeline unknown, DEv2 has a proposal
                if dv2 == inferred_ground_truth:
                    cat = "de_v2_correcto"
                else:
                    cat = "necesita_revision"
            elif ph1 and not dv2:
                if ph1 == inferred_ground_truth:
                    cat = "pipeline_correcto"
                else:
                    cat = "necesita_revision"
            else:
                # Both different from GT
                cat = "necesita_revision"

        comparisons.append({
            "account_name": r["account_name"],
            "account_code": r["account_code"],
            "account_type": r["account_type"],
            "phase1_code": ph1,
            "phase1_method": r.get("phase1_method", ""),
            "de_v2_code": dv2,
            "de_v2_score": r.get("de_v2_score", 0),
            "same": same,
            "inferred_gt": inferred_ground_truth,
            "category": cat,
        })

    analysis["comparisons"] = comparisons

    # ── Phase 4: Global metrics ──
    total = len(rows)

    ph1_unknown = sum(1 for r in rows if not r["phase1_code"])
    dv2_unknown = sum(1 for r in rows if not r["de_v2_code"])
    unknown_reduction = ph1_unknown - dv2_unknown
    unknown_reduction_pct = (unknown_reduction / ph1_unknown * 100) if ph1_unknown else 0

    cat_counts = Counter(c["category"] for c in comparisons)
    correct_by_ph1 = cat_counts.get("pipeline_correcto", 0) + cat_counts.get("ambos_correctos", 0)
    correct_by_dv2 = cat_counts.get("de_v2_correcto", 0) + cat_counts.get("ambos_correctos", 0)
    estimated_precision_ph1 = correct_by_ph1 / max(total - ph1_unknown, 1) * 100
    estimated_precision_dv2 = correct_by_dv2 / max(total - dv2_unknown, 1) * 100

    recall_ph1 = correct_by_ph1 / max(total, 1) * 100
    recall_dv2 = correct_by_dv2 / max(total, 1) * 100

    scores = [r["de_v2_score"] for r in rows if r["de_v2_code"]]
    avg_consensus = sum(r["de_v2_consensus"] for r in rows) / max(total, 1)

    conflict_dist = Counter(
        "ninguno" if r["de_v2_conflicts"] == 0 else
        "bajo" if r["de_v2_conflicts"] <= 2 else
        "medio" if r["de_v2_conflicts"] <= 4 else
        "alto"
        for r in rows
    )
    score_dist = Counter(
        "0.00" if r["de_v2_score"] == 0 else
        f"{r['de_v2_score']:.1f}"[:4]
        for r in rows
    )
    confidence_dist = Counter(r["de_v2_confidence"] for r in rows)
    review_needed = sum(1 for r in rows if r["de_v2_review"])
    auto_decisions = total - review_needed

    analysis["metrics"] = {
        "total_accounts": total,
        "phase1_unknown": ph1_unknown,
        "phase1_unknown_pct": round(ph1_unknown / total * 100, 2),
        "de_v2_unknown": dv2_unknown,
        "de_v2_unknown_pct": round(dv2_unknown / total * 100, 2),
        "unknown_reduction": unknown_reduction,
        "unknown_reduction_pct": round(unknown_reduction_pct, 2),
        "estimated_precision_phase1": round(estimated_precision_ph1, 2),
        "estimated_precision_de_v2": round(estimated_precision_dv2, 2),
        "estimated_recall_phase1": round(recall_ph1, 2),
        "estimated_recall_de_v2": round(recall_dv2, 2),
        "avg_consensus": round(avg_consensus, 2),
        "auto_decision_rate": round(auto_decisions / total * 100, 2) if total else 0,
        "human_review_rate": round(review_needed / total * 100, 2) if total else 0,
        "scores": {
            "mean": round(sum(scores) / len(scores), 4) if scores else 0,
            "min": round(min(scores), 4) if scores else 0,
            "max": round(max(scores), 4) if scores else 0,
        },
        "score_distribution": dict(score_dist),
        "confidence_distribution": dict(confidence_dist),
        "conflict_distribution": dict(conflict_dist),
    }

    # ── Phase 5: Difference categories ──
    analysis["category_counts"] = {k: cat_counts.get(k, 0) for k in CATEGORY_ORDER}
    analysis["category_counts"]["ambos_unknown"] = cat_counts.get("ambos_unknown", 0)

    # ── Phase 6: Quantitative answers ──
    q = quantitative_analysis(rows, comparisons)
    analysis["quantitative"] = q

    # ── Phase 7: Classifier decision matrix ──
    analysis["classifier_matrix"] = build_classifier_matrix(rows)

    return analysis


def _infer_ground_truth(row: dict[str, Any]) -> str | None:
    name_norm = row["account_name"].lower().strip()
    if name_norm in KNOWN_CODES:
        return KNOWN_CODES[name_norm]
    # SemanticMatcher's expected_cmcc is the best proxy
    sm_code = row.get("sm_code") or row.get("expected_cmcc")
    if sm_code and sm_code != "UNKNOWN":
        return sm_code
    # Phase 1 GS match
    if row.get("phase1_method", "").startswith("learning_"):
        return row["phase1_code"]
    return None


# ═══════════════════════════════════════════════════════════════════
# QUANTITATIVE ANALYSIS (Phase 6)
# ═══════════════════════════════════════════════════════════════════

def quantitative_analysis(
    rows: list[dict[str, Any]],
    comparisons: list[dict[str, Any]],
) -> dict[str, Any]:
    total = len(rows)
    same = sum(1 for c in comparisons if c["same"])
    improved = sum(1 for c in comparisons if c["category"] == "de_v2_correcto")
    worsened = sum(1 for c in comparisons if c["category"] == "pipeline_correcto")
    both_wrong = sum(1 for c in comparisons if c["category"] == "ambos_incorrectos")
    need_review = sum(1 for c in comparisons if c["category"] == "necesita_revision")

    # Classifier contribution
    def _ev_code(ev):
        return ev.get("proposed_code") if isinstance(ev, dict) else ev.proposed_code
    def _ev_src(ev):
        return ev.get("source") if isinstance(ev, dict) else ev.source

    source_counts: Counter[str] = Counter()
    for r in rows:
        src = r["de_v2_source"]
        dv2_code = r["de_v2_code"]
        if src.startswith("CONSENSUS_"):
            for ev in r.get("evidence", []):
                if _ev_code(ev) == dv2_code:
                    source_counts[_ev_src(ev)] += 1
        elif src.startswith("SOLO_"):
            solo_class = src.replace("SOLO_", "").lower()
            source_counts[solo_class] += 1
        elif src.startswith("PRIORITY_"):
            pri_class = src.replace("PRIORITY_", "").lower()
            source_counts[pri_class] += 1
        elif src.startswith("TB") and src != "TB5_HUMAN_REVIEW":
            for ev in r.get("evidence", []):
                if _ev_code(ev) == dv2_code:
                    source_counts[_ev_src(ev)] += 1
        elif src == "SM_TIER_1_WINS" or src == "SM_TIER_2_WINS":
            source_counts["sm_tier_1_2"] += 1
        elif src == "TB5_HUMAN_REVIEW":
            pass
        else:
            for ev in r.get("evidence", []):
                if _ev_code(ev) == dv2_code:
                    source_counts[_ev_src(ev)] += 1

    # Rule usage
    rule_counts: Counter[str] = Counter()
    for r in rows:
        src = r["de_v2_source"]
        if src == "SM_TIER_1_WINS" or src == "SM_TIER_2_WINS":
            rule_counts["R1"] += 1
        elif src.startswith("CONSENSUS_"):
            rule_counts["R2"] += 1
        elif src.startswith("PRIORITY_"):
            rule_counts["R3-R8"] += 1
        elif src.startswith("SOLO_"):
            rule_counts["R9"] += 1
        elif src.startswith("TB"):
            rule_counts["R10"] += 1
        elif src == "ALL_UNKNOWN" or src == "ALL_FILTERED":
            rule_counts["R11-R12"] += 1
        elif src == "TB5_HUMAN_REVIEW":
            rule_counts["R10_TB5"] += 1

    # False positive analysis by score threshold
    fp_by_threshold = {}
    for th in [0.95, 0.90, 0.85, 0.80, 0.70, 0.60, 0.50]:
        above = [c for c in comparisons if c.get("de_v2_score", 0) >= th]
        wrong = [c for c in above if c["category"] in ("necesita_revision", "ambos_incorrectos")]
        fp_by_threshold[f"ge_{th}"] = {
            "above_threshold": len(above),
            "estimated_wrong": len(wrong),
            "error_rate": round(len(wrong) / max(len(above), 1) * 100, 2),
        }

    # Optimal threshold (max recall with min error)
    best_threshold = 0.50
    best_tradeoff = 0
    for th in [0.95, 0.90, 0.85, 0.80, 0.70, 0.60, 0.50]:
        above = [c for c in comparisons if c.get("de_v2_score", 0) >= th]
        correct = sum(
            1 for c in above
            if c["category"] in ("de_v2_correcto", "ambos_correctos")
        )
        wrong = len(above) - correct
        if wrong <= correct * 0.05:  # 5% error tolerance
            score = correct - wrong * 2
            if score > best_tradeoff:
                best_tradeoff = score
                best_threshold = th

    return {
        "total_accounts": total,
        "improved": improved,
        "worsened": worsened,
        "unchanged": same,
        "improved_pct": round(improved / max(total, 1) * 100, 2),
        "worsened_pct": round(worsened / max(total, 1) * 100, 2),
        "unchanged_pct": round(same / max(total, 1) * 100, 2),
        "both_wrong": both_wrong,
        "need_human_review": need_review,
        "need_human_review_pct": round(need_review / max(total, 1) * 100, 2),
        "classifier_contribution": dict(source_counts.most_common(15)),
        "least_contributing": source_counts.most_common()[-5:] if source_counts else [],
        "rule_usage": dict(rule_counts.most_common()),
        "fp_by_threshold": fp_by_threshold,
        "optimal_threshold": best_threshold,
    }


# ═══════════════════════════════════════════════════════════════════
# CLASSIFIER MATRIX (Phase 7)
# ═══════════════════════════════════════════════════════════════════

def build_classifier_matrix(rows: list[dict[str, Any]]) -> dict[str, Any]:
    classifiers = [
        "code", "dict_exact", "dict_fuzzy",
        "sm_tier_1", "sm_tier_2", "sm_tier_4", "sm_tier_5", "sm_tier_6",
        "regex", "gs_exact", "gs_fuzzy",
    ]
    matrix: dict[str, dict[str, Any]] = {}

    for clf in classifiers:
        consulted = 0
        winner = 0
        discarded = 0
        contradicted = 0
        scores: list[float] = []

        for r in rows:
            found = False
            for ev in r.get("evidence", []):
                if isinstance(ev, dict):
                    src = ev.get("source", "")
                    pcode = ev.get("proposed_code")
                    escore = ev.get("score", 0)
                else:
                    src = ev.source
                    pcode = ev.proposed_code
                    escore = ev.score
                if src == clf and pcode:
                    consulted += 1
                    scores.append(escore)
                    if pcode == r["de_v2_code"]:
                        winner += 1
                    else:
                        contradicted += 1
                    found = True
                    break
            if not found:
                discarded += 1

        matrix[clf] = {
            "consulted": consulted,
            "winner": winner,
            "discarded": discarded if discarded > 0 else 0,
            "contradicted": contradicted,
            "avg_score": round(sum(scores) / max(len(scores), 1), 4),
            "win_rate": round(winner / max(consulted, 1) * 100, 2),
        }

    return matrix


# ═══════════════════════════════════════════════════════════════════
# REPORTS (Phase 8)
# ═══════════════════════════════════════════════════════════════════

def generate_json(analysis: dict[str, Any]) -> None:
    path = REPORTS / "decision_engine_shadow.json"
    # Remove bulky evidence from per-row data for JSON compactness
    compact = {k: v for k, v in analysis.items() if k != "row"}
    compact["row_count"] = len(analysis.get("rows", []))
    compact["sample_rows"] = analysis.get("rows", [])[:5]
    with open(path, "w") as f:
        json.dump(compact, f, indent=2, ensure_ascii=False)
    logger.info("Wrote %s", path)


def generate_metrics_json(analysis: dict[str, Any]) -> None:
    path = REPORTS / "decision_engine_metrics.json"
    metrics = dict(analysis.get("metrics", {}))
    metrics["quantitative"] = {k: v for k, v in analysis.get("quantitative", {}).items() if not isinstance(v, dict) or ("pct" in k or "rate" in k)}
    metrics["category_counts"] = analysis.get("category_counts", {})
    metrics["classifier_matrix"] = analysis.get("classifier_matrix", {})
    metrics["metadata"] = analysis.get("metadata", {})
    with open(path, "w") as f:
        json.dump(metrics, f, indent=2, ensure_ascii=False)
    logger.info("Wrote %s", path)


def generate_md(analysis: dict[str, Any]) -> None:
    path = REPORTS / "decision_engine_shadow.md"
    m = analysis.get("metrics", {})
    q = analysis.get("quantitative", {})
    cat = analysis.get("category_counts", {})
    cm = analysis.get("classifier_matrix", {})
    meta = analysis.get("metadata", {})

    lines = [
        "# Benchmark: Decision Engine V2 vs Pipeline Actual\n",
        f"**Date:** 2026-07-22  \n",
        f"**Accounts:** {m.get('total_accounts', 0)}  \n",
        f"**Pipeline:** {meta.get('pipeline_label', 'Phase 1')}  \n",
        f"**DecisionEngineV2:** {meta.get('de_v2_label', 'DEv2')}  \n",
        "",
        "---",
        "",
        "## FASE 4 — Métricas Globales",
        "",
        "| Métrica | Pipeline Actual | DecisionEngineV2 |",
        "|---|---|---|",
        f"| UNKNOWN | {m.get('phase1_unknown', 0)} ({m.get('phase1_unknown_pct', 0)}%) | {m.get('de_v2_unknown', 0)} ({m.get('de_v2_unknown_pct', 0)}%) |",
        f"| Reducción UNKNOWN | — | {m.get('unknown_reduction', 0)} ({m.get('unknown_reduction_pct', 0)}%) |",
        f"| Precisión estimada | {m.get('estimated_precision_phase1', 0)}% | {m.get('estimated_precision_de_v2', 0)}% |",
        f"| Recall estimado | {m.get('estimated_recall_phase1', 0)}% | {m.get('estimated_recall_de_v2', 0)}% |",
        "",
        "| Métrica | Valor |",
        "|---|---|",
        f"| Consenso promedio | {m.get('avg_consensus', 0)} |",
        f"| Auto-decisiones | {m.get('auto_decision_rate', 0)}% |",
        f"| Revisiones humanas requeridas | {m.get('human_review_rate', 0)}% |",
        f"| Score promedio | {m.get('scores', {}).get('mean', 0)} |",
        f"| Score min/max | {m.get('scores', {}).get('min', 0)} / {m.get('scores', {}).get('max', 0)} |",
        f"| Tiempo total | {meta.get('processing_time_seconds', 0)}s |",
        "",
        "### Distribución de Confianza",
        "",
    ]
    cd = m.get("confidence_distribution", {})
    for label in ["VERY_HIGH", "HIGH", "MEDIUM", "LOW", "UNKNOWN"]:
        n = cd.get(label, 0)
        pct = n / m.get("total_accounts", 1) * 100
        lines.append(f"- **{label}**: {n} ({pct:.1f}%)")

    lines += [
        "",
        "### Distribución de Conflictos",
        "",
    ]
    for cl, n in sorted(m.get("conflict_distribution", {}).items()):
        pct = n / m.get("total_accounts", 1) * 100
        lines.append(f"- **{cl}**: {n} ({pct:.1f}%)")

    lines += [
        "",
        "---",
        "",
        "## FASE 5 — Clasificación de Diferencias",
        "",
        "| Categoría | Cuentas | % |",
        "|---|---|---|",
    ]
    total = m.get("total_accounts", 1)
    for ck in CATEGORY_ORDER:
        n = cat.get(ck, 0)
        lines.append(f"| {CATEGORY_LABELS.get(ck, ck)} | {n} | {n / total * 100:.2f}% |")
    n_unk = cat.get("ambos_unknown", 0)
    lines.append(f"| Ambos UNKNOWN | {n_unk} | {n_unk / total * 100:.2f}% |")

    lines += [
        "",
        "---",
        "",
        "## FASE 6 — Análisis Cuantitativo",
        "",
        f"**Mejora vs Pipeline:** +{q.get('improved', 0)} cuentas ({q.get('improved_pct', 0)}%)  ",
        f"**Empeora vs Pipeline:** {q.get('worsened', 0)} cuentas ({q.get('worsened_pct', 0)}%)  ",
        f"**Sin cambios:** {q.get('unchanged', 0)} cuentas ({q.get('unchanged_pct', 0)}%)  ",
        f"**Ambos incorrectos:** {q.get('both_wrong', 0)}  ",
        f"**Necesita revisión humana:** {q.get('need_human_review', 0)} ({q.get('need_human_review_pct', 0)}%)  ",
        "",
        "### Contribución por Clasificador",
        "",
        "| Clasificador | Decisiones finales | % |",
        "|---|---|---|",
    ]
    contrib = q.get("classifier_contribution", {})
    for clf, n in contrib.items():
        pct = n / total * 100
        lines.append(f"| {clf} | {n} | {pct:.1f}% |")

    lines += [
        "",
        "### Clasificadores con menor contribución",
        "",
    ]
    for clf, n in q.get("least_contributing", []):
        lines.append(f"- {clf}: {n} decisiones")

    lines += [
        "",
        "### Uso de Reglas",
        "",
        "| Regla | Veces | % |",
        "|---|---|---|",
    ]
    for rule, n in sorted(q.get("rule_usage", {}).items()):
        pct = n / total * 100
        lines.append(f"| {rule} | {n} | {pct:.1f}% |")

    lines += [
        "",
        "### Falsos Positivos por Threshold",
        "",
        "| Threshold | ≥ Threshold | Estimados incorrectos | Error rate |",
        "|---|---|---|---|",
    ]
    for th, info in sorted(q.get("fp_by_threshold", {}).items()):
        lines.append(f"| {th} | {info.get('above_threshold', 0)} | {info.get('estimated_wrong', 0)} | {info.get('error_rate', 0)}% |")

    lines += [
        "",
        f"**Threshold óptimo:** {q.get('optimal_threshold', 0.50)}  ",
        "",
        "---",
        "",
        "## FASE 7 — Matriz de Decisiones por Clasificador",
        "",
        "| Clasificador | Consultado | Ganador | Descartado | Contradicho | Score Promedio | Win Rate |",
        "|---|---|---|---|---|---|---|",
    ]
    for clf in ["code", "dict_exact", "dict_fuzzy", "sm_tier_1", "sm_tier_2", "sm_tier_4", "sm_tier_5", "sm_tier_6", "regex", "gs_exact", "gs_fuzzy"]:
        info = cm.get(clf, {})
        lines.append(
            f"| {clf} | {info.get('consulted', 0)} | {info.get('winner', 0)} | "
            f"{info.get('discarded', 0)} | {info.get('contradicted', 0)} | "
            f"{info.get('avg_score', 0)} | {info.get('win_rate', 0)}% |"
        )

    lines += [
        "",
        "---",
        "",
        "## Conclusión",
        "",
        f"**GO / NO GO:** Evaluar según métricas arriba.",
    ]

    with open(path, "w") as f:
        f.write("\n".join(lines) + "\n")
    logger.info("Wrote %s", path)


def generate_xlsx(analysis: dict[str, Any]) -> None:
    try:
        import openpyxl
    except ImportError:
        logger.warning("openpyxl not available, skipping xlsx report")
        return

    wb = openpyxl.Workbook()

    # Sheet 1: Per-account comparison
    ws1 = wb.active
    ws1.title = "Comparación"
    ws1.append(["Cuenta", "Código", "Tipo", "Pipeline Actual", "Método Pipeline",
                 "DE v2 Código", "DE v2 Score", "DE v2 Confianza",
                 "Origen", "Consenso", "Conflictos", "Revisión", "Categoría"])

    for c in analysis.get("comparisons", []):
        r = next((x for x in analysis.get("rows", []) if x["account_name"] == c["account_name"]), {})
        ws1.append([
            c["account_name"],
            c["account_code"],
            c.get("account_type", ""),
            c["phase1_code"] or "",
            c.get("phase1_method", ""),
            c["de_v2_code"] or "",
            r.get("de_v2_score", 0),
            r.get("de_v2_confidence", ""),
            r.get("de_v2_source", ""),
            r.get("de_v2_consensus", 0),
            r.get("de_v2_conflicts", 0),
            "Sí" if r.get("de_v2_review") else "No",
            CATEGORY_LABELS.get(c.get("category", ""), c.get("category", "")),
        ])

    # Sheet 2: Metrics
    ws2 = wb.create_sheet("Métricas")
    m = analysis.get("metrics", {})
    ws2.append(["Métrica", "Valor"])
    for k, v in m.items():
        if isinstance(v, dict):
            for sk, sv in v.items():
                ws2.append([f"{k}.{sk}", sv])
        elif isinstance(v, list):
            ws2.append([k, json.dumps(v)])
        else:
            ws2.append([k, v])

    # Sheet 3: Classifier matrix
    ws3 = wb.create_sheet("Matriz Clasificadores")
    cm = analysis.get("classifier_matrix", {})
    ws3.append(["Clasificador", "Consultado", "Ganador", "Descartado", "Contradicho", "Score Prom", "Win Rate"])
    for clf, info in sorted(cm.items()):
        ws3.append([
            clf, info.get("consulted", 0), info.get("winner", 0),
            info.get("discarded", 0), info.get("contradicted", 0),
            info.get("avg_score", 0), f"{info.get('win_rate', 0)}%",
        ])

    # Sheet 4: Category counts
    ws4 = wb.create_sheet("Categorías")
    ws4.append(["Categoría", "Cuentas", "%"])
    total = m.get("total_accounts", 1)
    for ck in CATEGORY_ORDER:
        n = analysis.get("category_counts", {}).get(ck, 0)
        ws4.append([CATEGORY_LABELS.get(ck, ck), n, round(n / total * 100, 2)])
    n_unk = analysis.get("category_counts", {}).get("ambos_unknown", 0)
    ws4.append(["Ambos UNKNOWN", n_unk, round(n_unk / total * 100, 2)])

    path = REPORTS / "decision_engine_shadow.xlsx"
    wb.save(str(path))
    logger.info("Wrote %s", path)


def generate_conflicts_xlsx(analysis: dict[str, Any]) -> None:
    try:
        import openpyxl
    except ImportError:
        return

    comparisons = analysis.get("comparisons", [])
    conflicts = [c for c in comparisons if c.get("category") == "necesita_revision"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Conflictos"
    ws.append(["Cuenta", "Código", "Tipo", "Phase1", "DEv2", "Inferred GT"])

    for c in conflicts:
        ws.append([
            c["account_name"],
            c["account_code"],
            c.get("account_type", ""),
            c["phase1_code"] or "",
            c["de_v2_code"] or "",
            c.get("inferred_gt", ""),
        ])

    path = REPORTS / "decision_engine_conflicts.xlsx"
    wb.save(str(path))
    logger.info("Wrote %s (n=%d)", path, len(conflicts))


# ═══════════════════════════════════════════════════════════════════
# PRINT SUMMARY
# ═══════════════════════════════════════════════════════════════════

def print_summary(analysis: dict[str, Any]) -> None:
    m = analysis["metrics"]
    q = analysis["quantitative"]
    cat = analysis["category_counts"]
    meta = analysis.get("metadata", {})

    print()
    print("=" * 70)
    print("  BENCHMARK: DecisionEngineV2 vs Pipeline Actual")
    print("=" * 70)
    print(f"  Total cuentas:          {m['total_accounts']}")
    print(f"  Pipeline UNKNOWN:       {m['phase1_unknown']} ({m['phase1_unknown_pct']}%)")
    print(f"  DEv2 UNKNOWN:           {m['de_v2_unknown']} ({m['de_v2_unknown_pct']}%)")
    print(f"  Reducción UNKNOWN:      {m['unknown_reduction']} ({m['unknown_reduction_pct']}%)")
    print(f"  Precisión Pipe/DEv2:    {m['estimated_precision_phase1']}% / {m['estimated_precision_de_v2']}%")
    print(f"  Recall Pipe/DEv2:       {m['estimated_recall_phase1']}% / {m['estimated_recall_de_v2']}%")
    print(f"  Consenso promedio:      {m['avg_consensus']}")
    print(f"  Auto-decisiones:        {m['auto_decision_rate']}%")
    print(f"  Score promedio:         {m['scores']['mean']}")
    print(f"  Tiempo total:           {meta.get('processing_time_seconds', 0)}s")
    print()
    print("  ── Diferencias ──")
    print(f"  Mejora:                 +{q['improved']} ({q['improved_pct']}%)")
    print(f"  Empeora:                {q['worsened']} ({q['worsened_pct']}%)")
    print(f"  Sin cambios:            {q['unchanged']} ({q['unchanged_pct']}%)")
    print(f"  Revisión humana:        {q['need_human_review']} ({q['need_human_review_pct']}%)")
    print()
    print("  ── Categorías ──")
    for ck in CATEGORY_ORDER:
        n = cat.get(ck, 0)
        print(f"  {CATEGORY_LABELS.get(ck, ck):30s} {n:5d} ({n/m['total_accounts']*100:.2f}%)")
    print()
    print("  ── Threshold óptimo ──")
    print(f"  Score mínimo:           {q['optimal_threshold']}")
    print()
    for th, info in sorted(q.get("fp_by_threshold", {}).items()):
        print(f"  {th:10s}: {info['above_threshold']:5d} cuentas, {info['estimated_wrong']:4d} errores ({info['error_rate']:.1f}%)")
    print()
    print("  Reportes en:", REPORTS)
    print("=" * 70)


if __name__ == "__main__":
    main()
