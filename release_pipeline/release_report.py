from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from .pipeline_runner import PipelineRunner, DeploymentDecision
from .release_stage import StageResult, StageStatus
from .gate_engine import GateResult, GateStatus


class ReleaseReport:
    def __init__(self, runner: PipelineRunner):
        self.runner = runner
        self.context = runner.context
        self.stages = runner.stages
        self.gates = runner.gates
        self.decision = runner.decision
        self.reasons = runner.decision_reasons
        self.out_dir = Path("reports/release_pipeline")

    def generate_all(self):
        self.out_dir.mkdir(parents=True, exist_ok=True)
        self._generate_summary_xlsx()
        self._generate_stages_xlsx()
        self._generate_gates_xlsx()
        self._generate_decision_xlsx()
        self._generate_timeline_xlsx()
        self._generate_statistics_json()
        self._generate_markdown()

    def _generate_summary_xlsx(self):
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Summary"
            ws.append(["Metric", "Value"])
            ws.append(["Release ID", self.context.release_id])

            stage_pass = len([s for s in self.stages if s.status == StageStatus.PASS])
            stage_total = len([s for s in self.stages if s.status in (StageStatus.PASS, StageStatus.FAIL, StageStatus.ERROR)])
            ws.append(["Stages Passed", f"{stage_pass}/{stage_total}"])
            ws.append(["Decision", self.decision])
            ws.append(["Tests Passed", self.context.tests_passed])
            ws.append(["Tests Failed", self.context.tests_failed])
            ws.append(["Execution Time (s)", self.context.execution_time])
            ws.append(["Git Commit", self.context.git_commit])
            ws.append(["Git Branch", self.context.git_branch])
            wb.save(str(self.out_dir / "release_summary.xlsx"))
        except Exception:
            pass

    def _generate_stages_xlsx(self):
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Stages"
            ws.append(["Stage ID", "Stage Name", "Status", "Duration (s)", "Errors", "Warnings"])
            for s in self.stages:
                ws.append([s.stage_id, s.stage_name, s.status.value, round(s.duration, 3), len(s.errors), len(s.warnings)])
            wb.save(str(self.out_dir / "release_stages.xlsx"))
        except Exception:
            pass

    def _generate_gates_xlsx(self):
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Gates"
            ws.append(["Gate ID", "Gate Name", "Status", "Message", "Report"])
            for g in self.gates:
                ws.append([g.gate_id, g.gate_name, g.status.value, g.message, g.report_path])
            wb.save(str(self.out_dir / "gate_results.xlsx"))
        except Exception:
            pass

    def _generate_decision_xlsx(self):
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Decision"
            ws.append(["Field", "Value"])
            ws.append(["Decision", self.decision])
            ws.append(["Total Gates", len(self.gates)])
            ws.append(["Passed", len([g for g in self.gates if g.status == GateStatus.PASS])])
            ws.append(["Warnings", len([g for g in self.gates if g.status == GateStatus.WARNING])])
            ws.append(["Failed", len([g for g in self.gates if g.status == GateStatus.FAIL])])
            for r in self.reasons:
                ws.append([f"Reason: {r['gate']}", f"{r['status']}: {r['reason']}"])
            wb.save(str(self.out_dir / "deployment_decision.xlsx"))
        except Exception:
            pass

    def _generate_timeline_xlsx(self):
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Timeline"
            ws.append(["Stage", "Started", "Finished", "Duration (s)"])
            for s in self.stages:
                ws.append([s.stage_name, s.started_at, s.finished_at, round(s.duration, 3)])
            wb.save(str(self.out_dir / "execution_timeline.xlsx"))
        except Exception:
            pass

    def _generate_statistics_json(self):
        data = {
            "release_id": self.context.release_id,
            "timestamp": self.context.timestamp,
            "git_commit": self.context.git_commit,
            "git_branch": self.context.git_branch,
            "decision": self.decision,
            "stages": [s.to_dict() for s in self.stages],
            "gates": [g.to_dict() for g in self.gates],
            "reasons": self.reasons,
            "tests_passed": self.context.tests_passed,
            "tests_failed": self.context.tests_failed,
            "execution_time": self.context.execution_time,
        }
        with open(self.out_dir / "release_statistics.json", "w") as f:
            json.dump(data, f, indent=2, default=str)

    def _format_metric_change(self, label: str, value: float, direction_good: str) -> str:
        if value > 0:
            arrow = "⬆" if direction_good == "up" else "⬇"
            return f"{arrow} {label}: +{value}"
        elif value < 0:
            arrow = "⬇" if direction_good == "up" else "⬆"
            return f"{arrow} {label}: {value}"
        return f"{label}: no change"

    def _generate_markdown(self):
        quality = self.runner.context_dict.get("quality_metrics", {})
        drift = self.runner.context_dict.get("drift_metrics", {})
        alerts = self.runner.context_dict.get("alerts", [])
        validation = self.runner.context_dict.get("validation_metrics", {})
        knowledge = self.runner.context_dict.get("knowledge_metrics", {})

        lines = []
        lines.append("# Release Governance Report")
        lines.append("")
        lines.append(f"**Release ID**: {self.context.release_id}")
        lines.append(f"**Timestamp**: {self.context.timestamp}")
        lines.append(f"**Git Commit**: {self.context.git_commit}")
        lines.append(f"**Git Branch**: {self.context.git_branch}")
        lines.append(f"**Decision**: `{self.decision}`")
        lines.append(f"**Execution Time**: {self.context.execution_time}s")
        lines.append("")

        lines.append("## 1. Test Results")
        lines.append("")
        passed = self.context.tests_passed
        failed = self.context.tests_failed
        total = passed + failed
        lines.append(f"- **{total}** tests executed")
        lines.append(f"- **{passed}** passed")
        lines.append(f"- **{failed}** failed")
        lines.append("")

        lines.append("## 2. Stage Execution")
        lines.append("")
        lines.append("| Stage | Status | Duration | Errors | Warnings |")
        lines.append("|-------|--------|----------|--------|----------|")
        for s in self.stages:
            status_icon = "✅" if s.status == StageStatus.PASS else "❌" if s.status in (StageStatus.FAIL, StageStatus.ERROR) else "⚠️"
            lines.append(f"| {status_icon} {s.stage_name} | {s.status.value} | {s.duration:.2f}s | {len(s.errors)} | {len(s.warnings)} |")
        lines.append("")

        slowest = sorted(self.stages, key=lambda s: s.duration, reverse=True)
        if slowest:
            lines.append(f"**Slowest stage**: {slowest[0].stage_name} ({slowest[0].duration:.2f}s)")
            lines.append("")

        lines.append("## 3. Gate Results")
        lines.append("")
        for g in self.gates:
            icon = "✅" if g.status == GateStatus.PASS else "❌" if g.status == GateStatus.FAIL else "⚠️"
            lines.append(f"- {icon} **{g.gate_name}**: {g.message}")
        lines.append("")

        lines.append("## 4. Deployment Decision")
        lines.append("")
        if self.decision == DeploymentDecision.APPROVED:
            lines.append("### ✅ APPROVED")
            lines.append("")
            lines.append("All gates passed. The release may be deployed.")
        elif self.decision == DeploymentDecision.APPROVED_WITH_WARNINGS:
            lines.append("### ⚠️ APPROVED WITH WARNINGS")
            lines.append("")
            lines.append("All critical gates passed, but warnings were issued:")
            for r in self.reasons:
                if r["status"] == "WARNING":
                    lines.append(f"- **{r['gate']}**: {r['reason']}")
                    lines.append(f"  - Report: {r['report']}")
        else:
            lines.append("### ❌ BLOCKED")
            lines.append("")
            lines.append("Deployment blocked due to gate failures:")
            for r in self.reasons:
                lines.append(f"- **{r['gate']}** ({r['status']}): {r['reason']}")
                lines.append(f"  - Report: {r['report']}")
            lines.append("")
            lines.append("**Action required**: Fix the issues above and re-run the release check.")

        lines.append("")
        lines.append("## 5. Quality Metrics")
        lines.append("")
        lines.append(f"- Coverage: {quality.get('coverage', 'N/A')}%")
        lines.append(f"- Classified: {quality.get('classified', 'N/A')}")
        lines.append(f"- UNKNOWN: {quality.get('unknown', 'N/A')}")
        lines.append(f"- Parser Confidence: {quality.get('parser_confidence', 'N/A')}")
        lines.append(f"- Accuracy: {quality.get('accuracy', 'N/A')}")
        lines.append("")

        lines.append("## 6. Drift Metrics")
        lines.append("")
        lines.append(f"- Coverage drop: {drift.get('coverage_drop_pct', 0)}pp")
        lines.append(f"- UNKNOWN growth: {drift.get('unknown_growth_abs', 0)} ({drift.get('unknown_growth_pct', 0)}%)")
        lines.append(f"- Accuracy drop: {drift.get('accuracy_drop_pct', 0)}")
        lines.append("")

        lines.append("## 7. Knowledge Evolution")
        lines.append("")
        lines.append(f"- Classifications before: {knowledge.get('classifications_before', 'N/A')}")
        lines.append(f"- Classifications after: {knowledge.get('classifications_after', 'N/A')}")
        lines.append(f"- New classifications: {knowledge.get('new_classifications', 'N/A')}")
        lines.append(f"- Coverage before: {knowledge.get('coverage_before', 'N/A')}%")
        lines.append(f"- Coverage after: {knowledge.get('coverage_after', 'N/A')}%")
        lines.append("")

        lines.append("## 8. Alerts")
        lines.append("")
        critical = [a for a in alerts if isinstance(a, dict) and a.get("severity") == "CRITICAL"]
        high = [a for a in alerts if isinstance(a, dict) and a.get("severity") == "HIGH"]
        lines.append(f"- Critical alerts: {len(critical)}")
        lines.append(f"- High alerts: {len(high)}")
        lines.append("")
        for a in alerts:
            if isinstance(a, dict):
                icon = "🔴" if a.get("severity") == "CRITICAL" else "🟡" if a.get("severity") == "HIGH" else "🟢"
                lines.append(f"- {icon} **{a.get('severity', 'INFO')}**: {a.get('message', '')}")

        lines.append("")
        lines.append("## 9. Recommendations")
        lines.append("")
        if self.decision == DeploymentDecision.BLOCKED:
            lines.append("1. **Fix gate failures**: Address each failed gate before re-running release check.")
            lines.append("2. **Review evidence**: Check the referenced reports for each failure.")
            lines.append("3. **Re-run**: Execute `python scripts/release_check.py` after fixes.")
        elif self.decision == DeploymentDecision.APPROVED_WITH_WARNINGS:
            lines.append("1. **Review warnings**: Address warnings before the next release.")
            lines.append("2. **Monitor drift**: Track coverage and UNKNOWN trends.")
            lines.append("3. **Deploy with caution**: Monitor post-deployment metrics.")
        else:
            lines.append("1. **Proceed with deployment**: All quality gates passed.")
            lines.append("2. **Continue monitoring**: Run quality monitoring after deployment.")
            lines.append("3. **Update baselines**: Consider taking a new quality baseline.")

        report_path = self.out_dir / "release_report.md"
        with open(report_path, "w") as f:
            f.write("\n".join(lines))
