from __future__ import annotations

import json
import logging
import sqlite3
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook

from review.review_models import (
    PENDING_COLUMNS,
    PendingAccount,
    ProposedRule,
    SynonymEntry,
    GoldProposal,
    account_to_pending,
)
from review.review_metrics import (
    build_pending_rows,
    build_low_confidence_rows,
    build_dashboard,
    LOW_CONFIDENCE_THRESHOLD,
)
from review.excel_formatter import (
    build_pending_sheet,
    apply_default_sheet,
    write_dashboard_sheet,
)

log = logging.getLogger(__name__)


def load_all_data(
    shadow_path: str = "reports/semantic_shadow/shadow_data.json",
    priority_path: str = "reports/knowledge/knowledge_priority.xlsx",
    synonyms_path: str = "reports/knowledge/knowledge_synonyms.xlsx",
    gold_db_path: str = "gold_standard.db",
    generated_gs_path: str = "reports/knowledge/generated_gold_standard.json",
    generated_dict_path: str = "reports/knowledge/generated_dictionary_entries.json",
) -> dict[str, Any]:
    data: dict[str, Any] = {}

    # Shadow data
    with open(shadow_path) as f:
        data["shadow"] = json.load(f)

    # Priority
    try:
        import pandas as pd
        data["priority"] = pd.read_excel(priority_path).to_dict("records")
    except Exception as e:
        log.warning("Could not load priority: %s", e)
        data["priority"] = []

    # Synonyms
    try:
        import pandas as pd
        data["synonyms"] = pd.read_excel(synonyms_path).to_dict("records")
    except Exception as e:
        log.warning("Could not load synonyms: %s", e)
        data["synonyms"] = []

    # Gold database
    data["gold_gs"] = []
    try:
        conn = sqlite3.connect(gold_db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM gold_records ORDER BY usage_count DESC")
        cols = [d[0] for d in cursor.description]
        for row in cursor.fetchall():
            data["gold_gs"].append(dict(zip(cols, row)))
        conn.close()
    except Exception as e:
        log.warning("Could not load gold_standard.db: %s", e)

    # Generated gold standard
    try:
        with open(generated_gs_path) as f:
            data["generated_gs"] = json.load(f)
    except Exception as e:
        log.warning("Could not load generated gold standard: %s", e)
        data["generated_gs"] = []

    # Generated dictionary
    try:
        with open(generated_dict_path) as f:
            data["generated_dict"] = json.load(f)
    except Exception as e:
        log.warning("Could not load generated dictionary: %s", e)
        data["generated_dict"] = []

    return data


def build_review_package(
    output_path: str | Path,
    data: dict[str, Any] | None = None,
    *,
    confidence_threshold: float = LOW_CONFIDENCE_THRESHOLD,
) -> str:
    if data is None:
        data = load_all_data()

    wb = Workbook()

    accounts: list[dict] = data.get("shadow", {}).get("accounts", [])
    files: list[dict] = data.get("shadow", {}).get("files", [])

    # ── Sheet 1: Pendientes ──
    log.info("  Sheet 1: Pendientes")
    ws1 = wb.active
    ws1.title = "Pendientes"
    pending = build_pending_rows(accounts, threshold=confidence_threshold)
    pending_rows = [_pending_to_row(p) for p in pending]
    build_pending_sheet(ws1, pending_rows)

    # ── Sheet 2: Baja Confianza ──
    log.info("  Sheet 2: Baja Confianza")
    ws2 = wb.create_sheet("Baja Confianza")
    lc_rows = build_low_confidence_rows(accounts, threshold=confidence_threshold)
    lc_cols = ["Empresa", "Cuenta", "Método", "Confidence", "Clasificación actual", "Semantic", "Learning", "Código sugerido"]
    lc_data = [[r[c] for c in lc_cols] for r in lc_rows]
    apply_default_sheet(ws2, lc_cols, lc_data, "Baja Confianza")

    # ── Sheet 3: Conflictos ──
    log.info("  Sheet 3: Conflictos")
    ws3 = wb.create_sheet("Conflictos")
    conflicts = _build_conflicts(data.get("gold_gs", []), accounts)
    conf_cols = ["Cuenta", "Códigos", "Cantidad", "Empresas"]
    conf_rows = [[c["account_name"], ", ".join(c["codes"]), c["cantidad"], ", ".join(c["empresas"])] for c in conflicts]
    apply_default_sheet(ws3, conf_cols, conf_rows, "Conflictos")

    # ── Sheet 4: Reglas Propuestas ──
    log.info("  Sheet 4: Reglas Propuestas")
    ws4 = wb.create_sheet("Reglas Propuestas")
    rules = _build_proposed_rules(data.get("priority", []), accounts)
    rule_cols = ["Prioridad", "Patrón", "Frecuencia", "Empresas", "Código sugerido", "Confianza", "Aprobar"]
    rule_rows = [[r.priority, r.patron, r.frecuencia, r.empresas, r.codigo_sugerido, r.confianza, ""] for r in rules]
    apply_default_sheet(ws4, rule_cols, rule_rows, "Reglas Propuestas")
    _add_approve_dropdown(ws4, len(rules))

    # ── Sheet 5: Sinónimos ──
    log.info("  Sheet 5: Sinónimos")
    ws5 = wb.create_sheet("Sinónimos")
    syns = _build_synonym_entries(data.get("synonyms", []))
    syn_cols = ["Canonical", "Sinónimo", "Frecuencia", "Aprobar"]
    syn_rows = [[s.canonical, s.synonym, s.frecuencia, ""] for s in syns]
    apply_default_sheet(ws5, syn_cols, syn_rows, "Sinónimos")
    _add_approve_dropdown(ws5, len(syns))

    # ── Sheet 6: Gold Standard ──
    log.info("  Sheet 6: Gold Standard")
    ws6 = wb.create_sheet("Gold Standard")
    gs_props = _build_gold_proposals(data.get("generated_gs", []))
    gs_cols = ["Cuenta", "Código", "Empresas", "Uso", "Confianza", "Aprobar"]
    gs_rows = [[g.cuenta, g.codigo, g.empresas, g.uso, g.confianza, ""] for g in gs_props]
    apply_default_sheet(ws6, gs_cols, gs_rows, "Gold Standard")
    _add_approve_dropdown(ws6, len(gs_props))

    # ── Sheet 7: Dashboard ──
    log.info("  Sheet 7: Dashboard")
    ws7 = wb.create_sheet("Dashboard")
    metrics = build_dashboard(accounts)
    write_dashboard_sheet(ws7, metrics)

    # ── Sheet 8: Log ──
    log.info("  Sheet 8: Log")
    ws8 = wb.create_sheet("Log")
    log_cols = ["Fecha", "Revisor", "Cuenta", "Acción", "Comentario"]
    _write_header_only(ws8, log_cols)

    wb.save(str(output_path))
    log.info(f"  Saved: {output_path}")
    return str(output_path)


def _pending_to_row(p: PendingAccount) -> list:
    return [
        p.priority,
        p.empresa,
        p.rut,
        p.giro,
        p.year,
        p.archivo,
        p.pagina,
        p.codigo_original,
        p.nombre_cuenta,
        p.activo,
        p.pasivo,
        p.patrimonio,
        p.perdida,
        p.ganancia,
        p.monto_original,
        p.monto_absoluto,
        p.naturaleza,
        p.metodo,
        p.confidence,
        "Sí" if p.learning_hit else "No",
        "Sí" if p.semantic_hit else "No",
        p.semantic_type,
        p.semantic_rule,
        p.codigo_sugerido,
        p.cuenta_sugerida,
        p.frecuencia,
        p.cantidad_empresas,
        p.ultima_aparicion,
        p.observaciones,
        # Editable columns (empty for review)
        "",  # Código Final
        "",  # Nombre Final
        "",  # Clase
        "",  # Subclase
        "",  # Rubro
        "",  # Subrubro
        "",  # Tipo Semántico
        "",  # Contra Cuenta
        "",  # Aprender
        "",  # Alcance
        "",  # Justificación
        "",  # Observaciones
    ]


def _build_conflicts(gold_records: list[dict], accounts: list[dict]) -> list[dict]:
    code_map: dict[str, list[str]] = defaultdict(list)
    empresa_map: dict[str, set[str]] = defaultdict(set)
    for r in gold_records:
        name = r.get("account_name", "")
        code = r.get("final_code", "") or r.get("suggested_code", "")
        if name and code:
            code_map[name].append(code)
    for a in accounts:
        name = a.get("account_name", "")
        group = a.get("source_group", "") or ""
        if name:
            empresa_map[name].add(group)
    conflicts = []
    for name, codes in code_map.items():
        unique_codes = list(dict.fromkeys(codes))
        if len(unique_codes) >= 2:
            conflicts.append({
                "account_name": name,
                "codes": unique_codes,
                "cantidad": len(unique_codes),
                "empresas": list(empresa_map.get(name, set())),
            })
    return conflicts


def _build_proposed_rules(priority_rows: list[dict], accounts: list[dict]) -> list[ProposedRule]:
    # Count how many accounts each pattern would match
    rules: list[ProposedRule] = []
    for row in priority_rows:
        if row.get("type") != "new_semantic_rule":
            continue
        desc = row.get("description", "")
        patron = desc.replace("Nueva regla: ", "").replace(" -> ", " → ") if desc else row.get("representative", "")
        rules.append(ProposedRule(
            priority=row.get("priority", 0),
            patron=patron,
            frecuencia=row.get("cluster_size", 0),
            empresas=1,
            codigo_sugerido="",
            confianza=row.get("confidence", 0.0),
        ))
    return rules


def _build_synonym_entries(synonym_rows: list[dict]) -> list[SynonymEntry]:
    entries: list[SynonymEntry] = []
    for row in synonym_rows:
        label = row.get("group_label", "")
        variants_raw = row.get("variants", "")
        if isinstance(variants_raw, str):
            variants = [v.strip() for v in variants_raw.split("\n") if v.strip()]
        else:
            variants = variants_raw if isinstance(variants_raw, list) else [str(variants_raw)]
        for v in variants[:10]:
            entries.append(SynonymEntry(
                canonical=label,
                synonym=v,
                frecuencia=row.get("num_variants", 0),
                source_label=label,
            ))
    return entries


def _build_gold_proposals(generated_gs: list[dict]) -> list[GoldProposal]:
    props: list[GoldProposal] = []
    for entry in generated_gs:
        props.append(GoldProposal(
            cuenta=entry.get("account_name", ""),
            codigo=entry.get("suggested_code", ""),
            empresas=entry.get("source_file", ""),
            uso=entry.get("usage_count", 0),
            confianza=entry.get("suggested_confidence", 0.0),
            comments=entry.get("comments", ""),
        ))
    return props


def _write_header_only(ws: Any, columns: list[str]) -> None:
    from review.excel_formatter import HEADER_FILL, HEADER_FONT, HEADER_ALIGNMENT, BORDER_THIN
    for j, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=j, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = BORDER_THIN


def _add_approve_dropdown(ws: Any, num_rows: int) -> None:
    if num_rows < 1:
        return
    from review.excel_formatter import apply_dropdowns
    last_col = get_column_letter(ws.max_column)
    apply_dropdowns(ws, last_col, ["", "Sí", "No", "Revisar"], 2, num_rows + 1)


def get_column_letter(n: int) -> str:
    from openpyxl.utils import get_column_letter as gcl
    return gcl(n)
