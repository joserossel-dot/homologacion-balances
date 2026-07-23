from __future__ import annotations

from typing import Any

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import (
    Alignment, Border, Font, NamedStyle, PatternFill, Side, numbers,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGNMENT = Alignment(vertical="top", wrap_text=False)
BORDER_THIN = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
RED_FONT = Font(color="9C0006")
ORANGE_FILL = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
ORANGE_FONT = Font(color="9C6500")
YELLOW_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
BLUE_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
GREEN_FONT = Font(color="006100")
APPROVED_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
PENDING_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

SECTION_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")


def apply_default_sheet(ws: Any, columns: list[str], rows: list[list], sheet_name: str) -> None:
    _write_header(ws, columns)
    for i, row in enumerate(rows, 2):
        for j, val in enumerate(row, 1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.alignment = CELL_ALIGNMENT
            cell.border = BORDER_THIN
    _apply_table(ws, columns, len(rows))
    _freeze_panes(ws)
    _auto_width(ws, columns)


def _write_header(ws: Any, columns: list[str]) -> None:
    for j, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=j, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = BORDER_THIN


def _apply_table(ws: Any, columns: list[str], num_rows: int) -> None:
    if num_rows == 0:
        ws._current_row_for_auto_filter = ws.max_row + 1
        return
    num_cols = len(columns)
    ref = f"A1:{get_column_letter(num_cols)}{num_rows + 1}"
    try:
        tab = Table(display_name=f"Table_{ws.title.replace(' ', '_')}", ref=ref)
        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(tab)
    except Exception:
        ws.auto_filter.ref = ref


def _freeze_panes(ws: Any) -> None:
    ws.freeze_panes = "A2"


def _auto_width(ws: Any, columns: list[str]) -> None:
    for j, col_name in enumerate(columns, 1):
        max_len = max(
            len(str(ws.cell(row=i, column=j).value or ""))
            for i in range(1, min(ws.max_row + 1, 50))
        )
        width = min(max(max_len + 2, len(col_name) + 2), 30)
        ws.column_dimensions[get_column_letter(j)].width = width


def apply_conditional_pending(ws: Any, num_rows: int, confidence_col: str = "T", method_col: str = "R") -> None:
    if num_rows < 1:
        return
    last_row = num_rows + 1
    data_range = f"A2:{get_column_letter(ws.max_column)}{last_row}"

    # Red: unknown / unclassified
    unknown_formula = f'OR({method_col}2="unknown",{method_col}2="unclassified",{method_col}2="")'
    ws.conditional_formatting.add(data_range, FormulaRule(
        formula=[unknown_formula], fill=RED_FILL, font=RED_FONT,
    ))

    # Orange: confidence < 0.85 and not red
    orange_formula = f'AND({confidence_col}2<0.85,{confidence_col}2>0,NOT(OR({method_col}2="unknown",{method_col}2="unclassified",{method_col}2="")))'
    ws.conditional_formatting.add(data_range, FormulaRule(
        formula=[orange_formula], fill=ORANGE_FILL, font=ORANGE_FONT,
    ))

    # Green: confidence > 0.95
    green_formula = f'AND({confidence_col}2>=0.95,{method_col}2<>"unknown",{method_col}2<>"unclassified")'
    ws.conditional_formatting.add(data_range, FormulaRule(
        formula=[green_formula], fill=GREEN_FILL, font=GREEN_FONT,
    ))


def apply_dropdowns(ws: Any, column_letter: str, options: list[str], start_row: int, end_row: int) -> None:
    if end_row < start_row:
        return
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(options)}"',
        allow_blank=True,
    )
    dv.error = "Seleccione un valor válido"
    dv.errorTitle = "Valor inválido"
    dv.prompt = "Seleccione de la lista"
    dv.promptTitle = "Lista desplegable"
    ws.add_data_validation(dv)
    for row in range(start_row, end_row + 1):
        dv.add(ws[f"{column_letter}{row}"])


def build_pending_sheet(ws: Any, rows: list[list]) -> None:
    from review.review_models import PENDING_COLUMNS, PENDING_EDITABLE, PENDING_READONLY, CLASE_VALUES, SEMANTIC_TYPE_VALUES, APRENDER_VALUES, CONTRA_CUENTA_VALUES, ALCANCE_VALUES

    columns = PENDING_COLUMNS
    _write_header(ws, columns)
    for i, row in enumerate(rows, 2):
        for j, val in enumerate(row, 1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.alignment = CELL_ALIGNMENT
            cell.border = BORDER_THIN

    num_rows = len(rows)
    last_data_row = num_rows + 1

    # Table with auto-filter
    if num_rows > 0:
        num_cols = len(columns)
        ref = f"A1:{get_column_letter(num_cols)}{last_data_row}"
        try:
            tab = Table(display_name="Table_Pendientes", ref=ref)
            tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            ws.add_table(tab)
        except Exception:
            ws.auto_filter.ref = ref

    ws.freeze_panes = "A2"

    # Conditional formatting on method (col R=18) and confidence (col T=20)
    if num_rows > 0:
        _apply_pending_conditional(ws, last_data_row)

    # Dropdowns for editable columns
    col_map = {name: letter for name, letter in zip(columns, [get_column_letter(c) for c in range(1, len(columns) + 1)])}

    # Column mapping (0-indexed from PENDING_COLUMNS):
    # Código Final=30(AD), Nombre Final=31(AE), Clase=32(AF), Subclase=33(AG), Rubro=34(AH), Subrubro=35(AI)
    # Tipo Semántico=36(AJ), Contra Cuenta=37(AK), Aprender=38(AL), Alcance=39(AM), Justificación=40(AN), Observaciones=41(AO)

    if num_rows > 0:
        apply_dropdowns(ws, "AF", CLASE_VALUES, 2, last_data_row)
        apply_dropdowns(ws, "AJ", SEMANTIC_TYPE_VALUES, 2, last_data_row)
        apply_dropdowns(ws, "AL", APRENDER_VALUES, 2, last_data_row)
        apply_dropdowns(ws, "AK", CONTRA_CUENTA_VALUES, 2, last_data_row)
        apply_dropdowns(ws, "AM", ALCANCE_VALUES, 2, last_data_row)

    _auto_width(ws, columns)


def _apply_pending_conditional(ws: Any, last_row: int) -> None:
    data_range = f"A2:AR{last_row}"
    method_col = "R"  # column 18
    conf_col = "T"    # column 20
    semantic_col = "V"  # column 22

    unknown_fmt = f'OR({method_col}2="unknown",{method_col}2="unclassified",{method_col}2="")'
    ws.conditional_formatting.add(data_range, FormulaRule(
        formula=[unknown_fmt], fill=RED_FILL, font=RED_FONT,
    ))

    orange_fmt = f'AND({conf_col}2<0.85,{conf_col}2>0,NOT(OR({method_col}2="unknown",{method_col}2="unclassified",{method_col}2="")))'
    ws.conditional_formatting.add(data_range, FormulaRule(
        formula=[orange_fmt], fill=ORANGE_FILL, font=ORANGE_FONT,
    ))

    yellow_fmt = f'AND({semantic_col}2<>"unknown",{semantic_col}2<>"",NOT(OR({method_col}2="unknown",{method_col}2="unclassified",{method_col}2="")))'
    ws.conditional_formatting.add(data_range, FormulaRule(
        formula=[yellow_fmt], fill=YELLOW_FILL,
    ))

    green_fmt = f'AND({conf_col}2>=0.95,{method_col}2<>"unknown",{method_col}2<>"unclassified")'
    ws.conditional_formatting.add(data_range, FormulaRule(
        formula=[green_fmt], fill=GREEN_FILL, font=GREEN_FONT,
    ))

    # Blue fill for learning hits (column S=19)
    blue_fmt = f'{get_column_letter(19)}2=TRUE()'  # Learning Hit
    ws.conditional_formatting.add(data_range, FormulaRule(
        formula=[blue_fmt], fill=BLUE_FILL,
    ))


def clear_sheet(ws: Any) -> None:
    for row in ws.iter_rows():
        for cell in row:
            cell.value = None
    ws.conditional_formatting = []
    ws.data_validations = []


def write_dashboard_sheet(ws: Any, metrics: Any) -> None:
    ws.cell(row=1, column=1, value="Dashboard - Review Package").font = Font(bold=True, size=14)
    ws.merge_cells("A1:D1")

    metrics_data: list[tuple[str, str]] = [
        ("Total cuentas", str(metrics.total_cuentas)),
        ("Clasificadas", str(metrics.clasificadas)),
        ("Unknown", str(metrics.unknown)),
        ("Learning", str(metrics.learning)),
        ("Semantic", str(metrics.semantic)),
        ("Código", str(metrics.codigo)),
        ("Diccionario", str(metrics.diccionario)),
        ("Cobertura", f"{metrics.cobertura_pct}%"),
        ("Confianza promedio", f"{metrics.confianza_promedio:.2%}"),
    ]

    for i, (label, value) in enumerate(metrics_data, 3):
        cell_lbl = ws.cell(row=i, column=1, value=label)
        cell_lbl.font = Font(bold=True, size=10)
        cell_lbl.fill = SECTION_FILL
        cell_val = ws.cell(row=i, column=2, value=value)
        cell_val.font = Font(size=10)

    row = len(metrics_data) + 5
    if metrics.top_empresas:
        ws.cell(row=row, column=1, value="Top Empresas").font = Font(bold=True, size=12)
        ws.cell(row=row + 1, column=1, value="Empresa")
        ws.cell(row=row + 1, column=2, value="Cuentas")
        for i, (emp, cnt) in enumerate(metrics.top_empresas, row + 2):
            ws.cell(row=i, column=1, value=emp)
            ws.cell(row=i, column=2, value=cnt)

    row += len(metrics.top_empresas) + 3
    if metrics.top_cuentas:
        ws.cell(row=row, column=1, value="Top Cuentas").font = Font(bold=True, size=12)
        ws.cell(row=row + 1, column=1, value="Cuenta")
        ws.cell(row=row + 1, column=2, value="Frecuencia")
        for i, (name, cnt) in enumerate(metrics.top_cuentas, row + 2):
            ws.cell(row=i, column=1, value=name)
            ws.cell(row=i, column=2, value=cnt)

    row += len(metrics.top_cuentas) + 3
    if metrics.top_montos:
        ws.cell(row=row, column=1, value="Top Montos").font = Font(bold=True, size=12)
        ws.cell(row=row + 1, column=1, value="Cuenta")
        ws.cell(row=row + 1, column=2, value="Monto")
        for i, (name, amt) in enumerate(metrics.top_montos, row + 2):
            ws.cell(row=i, column=1, value=name)
            ws.cell(row=i, column=2, value=amt)

    row += len(metrics.top_montos) + 3
    if metrics.top_reglas:
        ws.cell(row=row, column=1, value="Top Reglas Semánticas").font = Font(bold=True, size=12)
        ws.cell(row=row + 1, column=1, value="Regla")
        ws.cell(row=row + 1, column=2, value="Veces")
        for i, (rule, cnt) in enumerate(metrics.top_reglas, row + 2):
            ws.cell(row=i, column=1, value=rule)
            ws.cell(row=i, column=2, value=cnt)

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 25
