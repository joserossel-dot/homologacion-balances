"""
app_validacion.py — Plataforma de homologación de balances tributarios chilenos

Ejecutar con:
    streamlit run app_validacion.py

Requiere en el mismo directorio:
    - catalogo_maestro.json
    - diccionario.json
    - parser_universal.py
    - clasificador_codigo_cuenta.py
    - reglas_especiales.py

Funcionalidad:
    1. Carga de archivo (PDF o Excel)
    2. Clasificación híbrida: código de cuenta → diccionario (exacto/fuzzy) → reglas regex
    3. Aplicación de reglas especiales (D1-D5 del análisis del vaciador)
    4. Cola de revisión para cuentas con confianza < umbral
    5. Balance normalizado agrupado por catálogo maestro
    6. Feedback loop: las correcciones se agregan al diccionario y son descargables
"""

import hashlib
import calendar
import json
import math
import os
import re
import time
from dataclasses import replace
from pathlib import Path
from datetime import date, datetime
from io import BytesIO
from html import escape
from document_scope import page_count, parse_pages, select_pdf, render_page
from report_presentation import complete_catalog, add_report_sheets

import pandas as pd
import streamlit as st
from rapidfuzz import fuzz, process
from dotenv import load_dotenv

from clasificador_codigo_cuenta import ClasificadorCodigo
from catalog_selection import opciones_clasificacion
from gold_standard.builder import GoldBuilder
from gold_standard.promotion import promote as promover_revisiones
from gold_standard.runtime import RuntimeGoldStorage
from gold_standard.runtime_manager import RuntimeManager
from gold_standard.runtime_stats import RuntimeStatistics
from reglas_especiales import (
    ProcesadorReglasEspeciales,
    calcular_patrimonio_efectivo,
    es_cuenta_socios,
)
from config.regex_rules import REGLAS_REGEX, REGLAS_COMPILADAS
from parser_universal import (
    ParserPDF, CuentaRaw, OrigenColumna, RAW_MONETARY_COLUMNS,
    FormatoCodigo, ResultadoParseo,
    certificar_extraccion_columnas, detectar_años_y_monedas, parsear_excel,
)
from parsers.column_interpretation import es_ingreso as es_ingreso_col, es_gasto as es_gasto_col
from parsers.account_type_resolver import (
    is_accumulated_result_name,
    is_contra_asset_name,
    is_patrimonial_reserve_name,
)
from extractor_metadata import extraer_metadata, MetadataEmpresa
from account_qualification import qualify_cuentas as _safe_qualify_cuentas, \
    safe_mode_enabled as _safe_mode_enabled
from persistence.neon_store import NeonKnowledgeStore
from reporting_integrity import (
    resultado_compatible, importe_resultado_homologado, conciliar_resultados,
)


MESES_SELECCION = [
    "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
]


def _valores_periodo_metadata(meta: MetadataEmpresa) -> tuple[str, int, int]:
    """Deriva mes, año y duración desde el período detectado."""
    cierre = None
    inicio = None
    for raw, target in ((meta.periodo_hasta, "cierre"), (meta.periodo_desde, "inicio")):
        if not raw:
            continue
        try:
            parsed = datetime.strptime(str(raw), "%d/%m/%Y").date()
        except ValueError:
            continue
        if target == "cierre":
            cierre = parsed
        else:
            inicio = parsed
    cierre = cierre or date.today()
    meses = 12
    if inicio:
        meses = (cierre.year - inicio.year) * 12 + cierre.month - inicio.month + 1
        meses = min(12, max(1, meses))
    return MESES_SELECCION[cierre.month - 1], cierre.year, meses


def _fechas_periodo_seleccionado(
    mes: str, anio: int, numero_meses: int,
) -> tuple[str, str]:
    """Convierte mes de cierre y duración a fechas inclusivas del período."""
    mes_numero = MESES_SELECCION.index(mes) + 1
    total_inicio = anio * 12 + mes_numero - int(numero_meses)
    anio_inicio, mes_inicio_cero = divmod(total_inicio, 12)
    mes_inicio = mes_inicio_cero + 1
    ultimo_dia = calendar.monthrange(int(anio), mes_numero)[1]
    return (
        f"01/{mes_inicio:02d}/{anio_inicio}",
        f"{ultimo_dia:02d}/{mes_numero:02d}/{int(anio)}",
    )


def _aplicar_metadata_confirmada(meta: MetadataEmpresa) -> MetadataEmpresa:
    """Aplica a cada archivo los datos generales confirmados por el usuario."""
    meta.rut = st.session_state.company_rut
    meta.razon_social = st.session_state.company_razon
    meta.giro = st.session_state.company_giro
    meta.moneda = st.session_state.company_moneda
    meta.mes_cierre = st.session_state.company_mes
    meta.anio_cierre = int(st.session_state.company_anio)
    meta.numero_meses = int(st.session_state.company_numero_meses)
    meta.periodos_detectados = tuple(
        st.session_state.get("company_periodos_detectados", ())
    )
    meta.periodos_seleccionados = tuple(
        st.session_state.get(
            "company_periodos_seleccionados", (str(meta.anio_cierre),),
        )
    )
    meta.periodo_desde, meta.periodo_hasta = _fechas_periodo_seleccionado(
        meta.mes_cierre, meta.anio_cierre, meta.numero_meses,
    )
    return meta


def _detectar_periodos_comparativos(
    lineas: list[str], anio_fallback: int,
) -> tuple[str, ...]:
    """Devuelve hasta dos años explícitos respetando el orden del documento."""
    years, _ = detectar_años_y_monedas(lineas)
    numeric = []
    for year in years:
        if re.fullmatch(r"(?:19|20)\d{2}", str(year)) and year not in numeric:
            numeric.append(str(year))
    fallback = str(int(anio_fallback))
    if not numeric:
        numeric.append(fallback)
    return tuple(numeric[:2])


def _periodos_seleccionados() -> tuple[str, ...]:
    selected = tuple(st.session_state.get("company_periodos_seleccionados", ()))
    if selected:
        return selected
    return (str(int(st.session_state.get("company_anio", date.today().year))),)


def _valor_periodo(
    montos_periodos: dict, periodo: str, posicion: int,
    monto_fallback: float | None = None,
) -> float | None:
    """Obtiene el valor anual sin confundir alias actual/anterior o moneda."""
    if periodo in montos_periodos:
        return montos_periodos[periodo]
    alias = "actual" if posicion == 0 else "anterior"
    if alias in montos_periodos:
        return montos_periodos[alias]
    return monto_fallback if posicion == 0 else None


def _montos_periodos_cuenta(cuenta: CuentaRaw) -> dict[str, float | None]:
    return {
        periodo: _valor_periodo(
            cuenta.montos_periodos, periodo, posicion, cuenta.monto,
        )
        for posicion, periodo in enumerate(_periodos_seleccionados())
    }


def _campos_periodos_cuenta(cuenta: CuentaRaw) -> tuple[float | None, dict]:
    valores = _montos_periodos_cuenta(cuenta)
    periodos = _periodos_seleccionados()
    principal = valores.get(periodos[0], cuenta.monto)
    campos = {
        f"monto_periodo_{periodo}": valores.get(periodo)
        for periodo in periodos
    }
    campos["monto_periodo_actual"] = valores.get(periodos[0])
    campos["monto_periodo_anterior"] = (
        valores.get(periodos[1]) if len(periodos) > 1 else None
    )
    return principal, campos


def _cuenta_tiene_saldo_seleccionado(cuenta: CuentaRaw) -> bool:
    valores = _montos_periodos_cuenta(cuenta).values()
    presentes = [float(valor) for valor in valores if valor is not None]
    if presentes:
        return any(valor != 0 for valor in presentes)
    return cuenta.monto is None or float(cuenta.monto) != 0

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURACIÓN
# ─────────────────────────────────────────────────────────────────────────────

BASE_DIR = Path(__file__).parent
load_dotenv(BASE_DIR / '.env')
UMBRAL_REVISION = 0.85  # bajo este valor, la cuenta va a la cola de revisión
USE_LEGACY_ENGINE = False  # True → MotorHibridoLocal (antiguo); False → HomologationPipeline (nuevo, default)
SHADOW_MODE = True  # True → ejecuta nuevo pipeline en paralelo sin afectar UI, guarda logs en logs/shadow/


def _build_date() -> str:
    configured = os.environ.get("APP_BUILD_DATE")
    if configured:
        return configured
    try:
        return (BASE_DIR / ".build_date").read_text(encoding="utf-8").strip()
    except OSError:
        return "desarrollo-local"

st.set_page_config(
    page_title="Homologación de Balances Tributarios",
    page_icon="📊",
    layout="wide",
)


# ─────────────────────────────────────────────────────────────────────────────
# CARGA DE CATÁLOGO Y DICCIONARIO
# ─────────────────────────────────────────────────────────────────────────────

@st.cache_data
def cargar_catalogo() -> dict:
    with open(BASE_DIR / 'catalogo_maestro.json', encoding='utf-8') as f:
        catalogo_local = json.load(f)
    store = NeonKnowledgeStore()
    if store.enabled:
        try:
            catalogo_neon = store.load_catalog()
            if catalogo_neon:
                return {
                    codigo: {**entrada, **catalogo_neon.get(codigo, {})}
                    for codigo, entrada in catalogo_local.items()
                } | {
                    codigo: entrada for codigo, entrada in catalogo_neon.items()
                    if codigo not in catalogo_local
                }
        except Exception:
            pass
    return catalogo_local


@st.cache_data
def cargar_diccionario_base() -> list[dict]:
    store = NeonKnowledgeStore()
    if store.enabled:
        try:
            diccionario = store.load_dictionary()
            if diccionario:
                return diccionario
        except Exception:
            pass
    with open(BASE_DIR / 'diccionario.json', encoding='utf-8') as f:
        return json.load(f)


def _persistir_validacion(
    *, nombre: str, codigo: str, fuente: str, agregar_diccionario: bool,
    sugerido: str | None = None, metodo: str | None = None,
    confianza: float | None = None, archivo: str = '',
) -> bool:
    """Persiste en Neon; retorna False para que el caller use fallback JSON."""
    store = NeonKnowledgeStore()
    if not store.enabled:
        return False
    try:
        store.save_validation(
            account_name=nombre,
            validated_code=codigo,
            source=fuente,
            suggested_code=sugerido,
            suggested_method=metodo,
            suggested_confidence=confianza,
            source_file=archivo,
            add_to_dictionary=agregar_diccionario,
        )
        cargar_diccionario_base.clear()
        return True
    except Exception:
        st.error("Neon no pudo guardar la validación; se usará respaldo local.")
        return False


def _persistir_validaciones_lote(validaciones: list[dict]) -> bool:
    """Persiste un lote completo en una única transacción Neon."""
    store = NeonKnowledgeStore()
    if not store.enabled:
        return False
    try:
        store.save_validations(validaciones)
        cargar_diccionario_base.clear()
        return True
    except Exception:
        st.error(
            "Neon no pudo guardar el lote; las clasificaciones permanecen "
            "en esta sesión y se usará el respaldo local del diccionario."
        )
        return False


def _persistir_catalogo(entry: dict) -> bool:
    store = NeonKnowledgeStore()
    if not store.enabled:
        return False
    try:
        store.save_catalog_entry(entry)
        cargar_catalogo.clear()
        return True
    except Exception:
        st.error("Neon no pudo guardar la categoría; se usará respaldo local.")
        return False


@st.cache_data(ttl=60)
def _neon_disponible() -> bool:
    return NeonKnowledgeStore().healthcheck()


# ─────────────────────────────────────────────────────────────────────────────
# REGLAS REGEX (patrones de mayor cobertura)
# ─────────────────────────────────────────────────────────────────────────────

PATRON_NO_CUENTA = re.compile(
    r'^(comprendido|periodo|per[ií]odo|desde|hasta|rut|r\.u\.t|balance|'
    r'fecha|p[aá]gina|hora|moneda|firma|declaro|art[ií]culo|situaci[oó]n|'
    r'cifras\s+expresadas|direcci[oó]n|comuna|^giro|raz[oó]n\s+soc|'
    r'^a\s+nivel|contabilidad\s+en)',
    re.IGNORECASE
)


def normalizar_nombre(nombre: str) -> str:
    n = nombre.lower().strip()
    n = re.sub(r"[^\w\sñáéíóú]", " ", n)
    n = re.sub(r"\s+", " ", n)
    return n.strip()


def _seleccion_lote_actualizada(
    seleccion_actual: set, indices: list, accion: str,
) -> set:
    """Calcula la selección en lote sin depender del ciclo de Streamlit."""
    seleccion = set(seleccion_actual)
    objetivos = set(indices)
    if accion == 'reemplazar':
        return objetivos
    if accion == 'agregar':
        return seleccion | objetivos
    if accion == 'quitar':
        return seleccion - objetivos
    raise ValueError(f"Acción de selección desconocida: {accion}")


def _reemplazar_seleccion_lote(indices: list, checkbox_prefix: str) -> None:
    """Sincroniza selección y widgets antes del único rerun normal."""
    anterior = set(st.session_state.get('lote_seleccion', set()))
    nueva = _seleccion_lote_actualizada(anterior, indices, 'reemplazar')
    st.session_state.lote_seleccion = nueva
    for idx in anterior | nueva:
        st.session_state[f"{checkbox_prefix}_{idx}"] = idx in nueva


def _alternar_seleccion_lote(idx, checkbox_key: str) -> None:
    """Actualiza una casilla sin encadenar un segundo st.rerun()."""
    accion = 'agregar' if st.session_state.get(checkbox_key, False) else 'quitar'
    st.session_state.lote_seleccion = _seleccion_lote_actualizada(
        set(st.session_state.get('lote_seleccion', set())), [idx], accion,
    )


def _asignar_estado_widget(key: str, value) -> None:
    """Permite que un botón actualice otro widget antes del rerun."""
    st.session_state[key] = value


def propagar_clasificacion_resultados(nombre_original: str, codigo_final: str, metodo: str):
    nombre_norm = normalizar_nombre(nombre_original)
    if 'resultados' in st.session_state and isinstance(st.session_state.resultados, dict):
        propagaciones = 0
        for fn in list(st.session_state.resultados.keys()):
            df_res = st.session_state.resultados[fn].copy()
            names = df_res['nombre_original'].fillna('').apply(normalizar_nombre)
            mask = names == nombre_norm
            mask_target = mask & (
                df_res['requiere_revision'] | 
                (df_res['codigo_clasificado'] == '') | 
                (df_res['confianza'] < 1.0)
            )
            if codigo_final != '__EXCLUIR__':
                compatibles = df_res.apply(
                    lambda row: _codigo_compatible_con_origen(
                        codigo_final,
                        row.get('origen_columna'),
                        row.get('monto'),
                        row.get('nombre_original'),
                    ),
                    axis=1,
                )
                mask_target &= compatibles
            if mask_target.any():
                df_res.loc[mask_target, 'codigo_clasificado'] = codigo_final
                df_res.loc[mask_target, 'metodo'] = metodo
                df_res.loc[mask_target, 'confianza'] = 1.0
                df_res.loc[mask_target, 'requiere_revision'] = False
                if 'origen' in df_res.columns:
                    df_res.loc[mask_target, 'origen'] = (
                        'Manual' if any(k in metodo for k in ('humana', 'manual')) else 'Código'
                    )
                    df_res.loc[mask_target, 'regla'] = metodo
                    df_res.loc[mask_target, 'evidencia'] = 'Propagación automática entre balances'
                st.session_state.resultados[fn] = df_res
                propagaciones += 1
        
        if propagaciones > 1:
            st.toast(f"Homologación propagada a {propagaciones - 1} otro(s) balance(s) 🔄", icon="🔄")


def _nombre_mostrar(row: pd.Series) -> str:
    return row.get('nombre_revision_usuario', '') or row['nombre_original']


# ─────────────────────────────────────────────────────────────────────────────
# EXPLICABILIDAD (P6) — trazabilidad auditable de cada decisión.
# Solo expone información existente: nunca modifica la decisión del motor.
# ─────────────────────────────────────────────────────────────────────────────

ORIGENES_EXPLICABILIDAD = [
    'Runtime', 'Gold', 'Código', 'Regex', 'CMCC', 'Semantic', 'Manual', 'Regla', 'Sin clasificar',
]

_BADGE_ORIGEN_COLORS = {
    'Runtime': '#7B1FA2', 'Gold': '#C2185B', 'Código': '#1565C0', 'Regex': '#00695C',
    'CMCC': '#EF6C00', 'Semantic': '#2E7D32', 'Manual': '#37474F',
    'Regla': '#F9A825', 'Sin clasificar': '#757575',
}


def _badge_origen(origen: str) -> str:
    color = _BADGE_ORIGEN_COLORS.get(origen or '', '#757575')
    label = origen or '—'
    return (
        f"<span style='background:{color}; color:white; padding:2px 8px; border-radius:4px; "
        f"font-size:0.72em; font-weight:600;'>{label}</span>"
    )


def _origen_desde_metodo_display(metodo: str) -> str:
    """Mapea el método de clasificación a la categoría de origen (solo lectura)."""
    m = (metodo or '').lower()
    if not m:
        return 'Sin clasificar'
    if m in ('sin_clasificar', 'unclassified', 'movement_only'):
        return 'Sin clasificar'
    if 'learning' in m or m.startswith('gold_') or m.startswith('runtime_'):
        return 'Gold'
    if 'cmcc' in m:
        return 'CMCC'
    if 'semantic' in m:
        return 'Semantic'
    if 'regex' in m:
        return 'Regex'
    if m.startswith('code') or m == 'codigo' or 'diccionario' in m or 'dictionary' in m:
        return 'Código'
    if 'decision' in m:
        return 'Regla'
    if 'propagad' in m:
        return 'Código'
    if any(k in m for k in ('validacion_humana', 'manual', 'excluido', 'lote')):
        return 'Manual'
    if 'regla_especial' in m or m == 'regla' or 'regla' in m:
        return 'Regla'
    if 'columna_ambiguo' in m:
        return 'Código'
    return 'Regla'


def _fmt_confianza(val) -> str:
    if val is None:
        return '—'
    if isinstance(val, float) and pd.isna(val):
        return '—'
    try:
        return f"{float(val):.0%}"
    except (TypeError, ValueError):
        return str(val)


def _learning_runtime_hit(account_name: str, hp) -> bool:
    """True si el learning_* vino del runtime (consulta read-only)."""
    try:
        rt = hp._learning_engine._runtime_lookup(account_name)
    except Exception:
        rt = None
    return rt is not None


def _origen_desde_clasif(clasif: dict, account_name: str, hp=None) -> str:
    """Origen (Runtime/Gold/Código/Regex/CMCC/Semantic) a partir del dict del motor."""
    metodo = (clasif.get('method') or '').lower()
    if metodo.startswith('learning_'):
        if hp is not None and _learning_runtime_hit(account_name, hp):
            return 'Runtime'
        return 'Gold'
    if metodo.startswith('cmcc'):
        return 'CMCC'
    if metodo.startswith('semantic'):
        return 'Semantic'
    if metodo.startswith('regex'):
        return 'Regex'
    if metodo.startswith('code') or metodo.startswith('diccionario') or metodo.startswith('dictionary'):
        return 'Código'
    if metodo.startswith('decision'):
        de = clasif.get('decision_engine') or {}
        src = (de.get('decision_source') or '').upper()
        if 'SM' in src:
            return 'Semantic'
        if 'REGEX' in src:
            return 'Regex'
        if 'DICT' in src:
            return 'Código'
        return 'Regla'
    if metodo in ('', 'unclassified', 'movement_only'):
        return 'Sin clasificar'
    return 'Regla'


def _resolver_tipo_cuenta(origen_columna, codigo) -> str | None:
    """Account type (read-only) para etapas sensibles al tipo. Never modifica nada."""
    if str(getattr(origen_columna, 'value', origen_columna) or '').lower() == 'patrimonio':
        return 'PATRIMONIO'
    try:
        from parser_universal import OrigenColumna as _OC
        from parsers.account_type_resolver import AccountTypeResolver
        origen = _OC(origen_columna) if origen_columna else _OC.DESCONOCIDO
        return AccountTypeResolver().resolve(origen_columna=origen, codigo=codigo).account_type.value
    except Exception:
        return None


def _resolver_tipos_permitidos(origen_columna) -> set[str]:
    """Tipos compatibles con la columna, conservando sus ambiguedades reales."""
    if str(getattr(origen_columna, 'value', origen_columna) or '').lower() == 'patrimonio':
        return {'PATRIMONIO'}
    try:
        from parser_universal import OrigenColumna as _OC
        from parsers.account_type_resolver import AccountTypeResolver
        origen = _OC(origen_columna) if origen_columna else _OC.DESCONOCIDO
        resultado = AccountTypeResolver().resolve(origen_columna=origen)
        return {tipo.value for tipo in resultado.allowed_types}
    except Exception:
        return set()


_CONTRA_ORIGEN = {
    'activo': 'pasivo',
    'pasivo': 'activo',
    'perdida': 'ganancia',
    'ganancia': 'perdida',
}


def _origen_efectivo(origen_columna, monto, nombre: str | None = None) -> str:
    """Naturaleza contable efectiva, preservando aparte la columna física.

    En balances de 8 columnas un importe negativo representa una contra cuenta:
    ACTIVO <-> PASIVO y PERDIDA <-> GANANCIA.
    """
    origen = getattr(origen_columna, 'value', origen_columna)
    origen = str(origen or 'desconocido').strip().lower()
    if is_patrimonial_reserve_name(nombre) or is_accumulated_result_name(nombre):
        return 'patrimonio'
    try:
        es_negativo = monto is not None and pd.notna(monto) and float(monto) < 0
    except (TypeError, ValueError):
        es_negativo = False
    return _CONTRA_ORIGEN.get(origen, origen) if es_negativo else origen


def _etiqueta_origen(origen_columna, monto, nombre: str | None = None) -> str:
    """Etiqueta auditable para la cola: extracción y naturaleza efectiva."""
    extraido = getattr(origen_columna, 'value', origen_columna)
    extraido = str(extraido or 'desconocido').strip().upper()
    efectivo = _origen_efectivo(origen_columna, monto, nombre).upper()
    if is_patrimonial_reserve_name(nombre) or is_accumulated_result_name(nombre):
        return f"{extraido} → PATRIMONIO (naturaleza de la cuenta)"
    if efectivo != extraido:
        return f"{extraido} → {efectivo} (monto negativo)"
    return extraido


def _es_contra_activo(nombre: str | None) -> bool:
    """Reconoce cuentas acreedoras que corrigen el valor de un activo."""
    return is_contra_asset_name(nombre)


def _es_partida_patrimonial(nombre: str | None) -> bool:
    return is_patrimonial_reserve_name(nombre) or is_accumulated_result_name(nombre)


def _monto_presentacion(codigo: str | None, monto, nombre: str | None = None,
                        origen=None, catalogo=None) -> float:
    """Aplica el signo contable sin alterar el importe extraído auditable."""
    valor = 0.0 if pd.isna(monto) else float(monto)
    if str(codigo or '').startswith('ER.') and origen is not None:
        resultado = importe_resultado_homologado(codigo, valor, origen, catalogo)
        if resultado is not None:
            return resultado
    if str(codigo or '') == 'PAT.10':
        return -abs(valor)
    if str(codigo or '').startswith('ANC') and _es_contra_activo(nombre):
        return -abs(valor)
    return valor


def _resultado_periodo(clasificadas: pd.DataFrame) -> float | None:
    """Calcula ganancia menos pérdida desde las columnas efectivas extraídas."""
    ganancias = 0.0
    perdidas = 0.0
    encontro_resultado = False
    for _, row in clasificadas.iterrows():
        origen = row.get('origen_columna_efectiva') or _origen_efectivo(
            row.get('origen_columna'), row.get('monto'), row.get('nombre_original')
        )
        monto = abs(float(row.get('monto') or 0.0))
        if origen == 'ganancia':
            ganancias += monto
            encontro_resultado = True
        elif origen == 'perdida':
            perdidas += monto
            encontro_resultado = True
    return ganancias - perdidas if encontro_resultado else None


def _codigo_compatible_con_origen(
        codigo: str | None, origen_columna, monto, nombre: str | None = None,
        catalogo=None) -> bool:
    """Impide sugerencias que contradigan la columna efectiva del balance."""
    origen_efectivo = _origen_efectivo(origen_columna, monto, nombre)
    if not resultado_compatible(codigo, origen_efectivo, catalogo):
        return False
    if str(codigo or '').startswith('ER.') and origen_efectivo in {'ganancia', 'perdida'}:
        return True
    # Una depreciación/amortización acumulada puede venir físicamente en la
    # columna Pasivo por su saldo acreedor, pero contablemente es contra-activo
    # y debe poder homologarse dentro del activo fijo (ANC).
    if (origen_efectivo == 'pasivo' and _es_contra_activo(nombre)
            and str(codigo or '').startswith('ANC')):
        return True
    if _es_partida_patrimonial(nombre) and str(codigo or '').startswith('PAT'):
        return True
    if (origen_efectivo == 'activo' and str(codigo or '') == 'PAT.10'
            and es_cuenta_socios(nombre)):
        return True
    tipos = _resolver_tipos_permitidos(origen_efectivo)
    if not tipos:
        return True
    from pipeline.homologation_pipeline import HomologationPipeline
    return any(
        HomologationPipeline._is_code_allowed_for_tipo(codigo, tipo)
        for tipo in tipos
    )


def _pendientes_revision(df: pd.DataFrame) -> pd.DataFrame:
    """Devuelve solo cuentas revisables: no totales y con saldo distinto de cero."""
    pendientes = df[df['requiere_revision'] | (df['codigo_clasificado'] == '')]
    pendientes = pendientes[~pendientes['es_total']]
    montos = pd.to_numeric(pendientes['monto'], errors='coerce')
    return pendientes[montos.isna() | montos.ne(0)]


def _con_saldo_relevante(df: pd.DataFrame) -> pd.DataFrame:
    """Excluye filas con saldo cero de clasificación, resumen y exportación."""
    montos = pd.to_numeric(df['monto'], errors='coerce')
    return df[montos.isna() | montos.ne(0)]


def _alternativas_revision(
    *, nombre: str, sugerido: str, confianza: float,
    origen_columna, monto, catalogo: dict, motor,
    limite: int = 3,
) -> list[dict]:
    """Rankea candidatos compatibles sin alterar la decisión del motor."""
    candidatos: dict[str, dict] = {}

    def agregar(codigo: str | None, score: float, fuente: str, evidencia: str):
        if not codigo or codigo not in catalogo:
            return
        if not _codigo_compatible_con_origen(
            codigo, origen_columna, monto, nombre, catalogo,
        ):
            return
        item = {
            "codigo": codigo,
            "nombre": catalogo[codigo].get("nombre_estandar", codigo),
            "score": max(0.0, min(float(score), 1.0)),
            "fuente": fuente,
            "evidencia": evidencia,
        }
        previous = candidatos.get(codigo)
        if previous is None or item["score"] > previous["score"]:
            candidatos[codigo] = item

    agregar(
        sugerido or None,
        float(confianza or 0.0),
        "Sugerencia actual",
        "Clasificación producida por el pipeline operativo.",
    )

    normalized = normalizar_nombre(nombre)
    dictionary_names = getattr(motor, "dic_lista", []) or []
    dictionary = getattr(motor, "dic_exacto", {}) or {}
    if normalized and dictionary_names:
        for matched_name, similarity, _ in process.extract(
            normalized, dictionary_names,
            scorer=fuzz.token_set_ratio, limit=max(limite * 3, 8),
        ):
            if similarity < 55:
                continue
            entry = dictionary.get(matched_name) or {}
            source = str(entry.get("fuente") or "Diccionario")
            human = any(token in source.lower() for token in (
                "human", "manual", "validacion", "analista",
            ))
            agregar(
                entry.get("codigo_estandar"),
                similarity / 100.0 + (0.03 if human else 0.0),
                "Neon · validación humana" if human else "Diccionario",
                f"Cuenta similar: {entry.get('cuenta_original', matched_name)} "
                f"({similarity:.0f}% de similitud).",
            )

    catalog_names = {
        codigo: normalizar_nombre(str(entry.get("nombre_estandar") or codigo))
        for codigo, entry in catalogo.items()
    }
    for codigo, catalog_name in catalog_names.items():
        similarity = fuzz.token_set_ratio(normalized, catalog_name)
        if similarity >= 58:
            agregar(
                codigo, similarity / 100.0, "Catálogo maestro",
                f"Nombre estándar similar: {catalogo[codigo].get('nombre_estandar', codigo)} "
                f"({similarity:.0f}%).",
            )

    return sorted(
        candidatos.values(),
        key=lambda item: (-item["score"], item["codigo"]),
    )[:max(1, int(limite))]


def _explicar_clasificacion(hp, account_code: str, account_name: str, *,
                            account_tipo: str | None = None,
                            origen_columna=None,
                            metodo_actual: str = '') -> list[dict]:
    """Reconstrucción read-only de las reglas evaluadas para la UI.

    No ejecuta ni modifica la decisión del motor: re-evalúa cada etapa para
    auditar la cadena de decisión. Cada lectura va envuelta en try/except para
    que un fallo en una etapa nunca rompa la UI.
    """
    etapas: list[dict] = []

    def _agregar(nombre, result, metodo):
        if result is None:
            etapas.append({
                'regla': nombre, 'coincidio': False, 'resultado': '—',
                'confianza': '—', 'detalle': 'Sin coincidencia', 'ganadora': False,
                '_metodo': metodo,
            })
            return
        detalle = result.get('reason') if isinstance(result, dict) else None
        etapas.append({
            'regla': nombre, 'coincidio': True,
            'resultado': result.get('standard_code') or '—',
            'confianza': _fmt_confianza(result.get('confidence')),
            'detalle': detalle or 'Coincidencia',
            'ganadora': False,
            '_metodo': metodo,
        })

    try:
        _agregar('1 · Código de cuenta',
                 hp._classify_by_code(account_code) if account_code else None, 'code')
    except Exception:
        pass
    try:
        _agregar('2 · Diccionario (exacto)',
                 hp._classify_by_dictionary_exact(account_name), 'dictionary_exact')
    except Exception:
        pass
    try:
        _agregar('3 · Diccionario (fuzzy)',
                 hp._classify_by_dictionary_fuzzy(account_name), 'dictionary_fuzzy')
    except Exception:
        pass
    try:
        _regex = None
        if hp._features.ENABLE_REGEX_FALLBACK:
            _regex = hp._classify_by_regex(account_name, account_tipo)
        _agregar('4 · Regex fallback', _regex, 'regex_fallback')
    except Exception:
        pass
    try:
        _rt = None
        if hp._learning_engine._runtime_path.exists():
            _rt = hp._learning_engine._runtime_lookup(account_name)
        if _rt is not None:
            _agregar('5 · Runtime (gold runtime)', _rt, f"learning_{_rt.get('source')}")
    except Exception:
        pass
    try:
        _gold = hp._learning_engine.best_match(account_name, use_runtime=False)
        if _gold.get('source') != 'none':
            _agregar('6 · Gold Standard', _gold, f"learning_{_gold.get('source')}")
    except Exception:
        pass
    try:
        if hp._features.ENABLE_CMCC:
            _cmcc = hp._cmcc_classifier.classify(account_name)
            _agregar('7 · CMCC', _cmcc, (_cmcc or {}).get('method', 'cmcc'))
    except Exception:
        pass
    try:
        if hp._features.ENABLE_SEMANTIC_MATCHER and hp._semantic_matcher is not None:
            _sm = hp._semantic_matcher.match(account_name, account_tipo)
            if _sm is not None and not _sm.is_unknown:
                _agregar(
                    '8 · Semantic',
                    {'standard_code': _sm.expected_cmcc,
                     'confidence': min(_sm.score, 0.99),
                     'reason': f"tier={_sm.match_tier} concept={_sm.concept_name}"},
                    f"semantic_{_sm.match_tier}",
                )
    except Exception:
        pass

    ma = (metodo_actual or '').lower()
    if ma.startswith('learning_'):
        for e in etapas:
            if str(e.get('_metodo', '')).startswith('learning_'):
                e['ganadora'] = True
                break
    else:
        for e in etapas:
            if e.get('_metodo') == ma:
                e['ganadora'] = True
                break
    return etapas


def _mostrar_detalle_cuenta(row: pd.Series, catalogo: dict, hp=None, close_key: str | None = None):
    if close_key:
        if st.button('✖ Cerrar detalle', key=f"close_{close_key}"):
            st.session_state.pop(close_key, None)
            st.rerun()

    st.markdown('#### 🔍 Detalle de clasificación')
    nombre = _nombre_mostrar(row)
    origen = row.get('origen') if 'origen' in row.index else ''
    origen = origen or _origen_desde_metodo_display(row.get('metodo', ''))
    regla = row.get('regla') if 'regla' in row.index else ''
    regla = regla or row.get('metodo', '')
    evidencia = row.get('evidencia') if 'evidencia' in row.index else ''
    evidencia = evidencia or row.get('nota', '') or 'Sin evidencia adicional.'
    tiempo = row.get('tiempo_clasificacion') if 'tiempo_clasificacion' in row.index else None

    c1, c2 = st.columns(2)
    with c1:
        st.markdown(f"**Nombre original:** {nombre}")
        nombre_norm = row.get('nombre_normalizado') if 'nombre_normalizado' in row.index else ''
        st.markdown(f"**Nombre normalizado:** `{nombre_norm or normalizar_nombre(nombre)}`")
        st.markdown(f"**Código final:** `{row.get('codigo_clasificado', '') or '(sin clasificar)'}`")
    with c2:
        st.markdown(f"**Regla ganadora:** `{regla}`")
        st.markdown(f"**Score:** **{_fmt_confianza(row.get('confianza'))}**")
        st.markdown(f"**Confianza:** {_fmt_confianza(row.get('confianza'))}")
        st.markdown(f"**Origen:** {_badge_origen(origen)}", unsafe_allow_html=True)
        tiempo_txt = f"{tiempo:.2f} ms" if tiempo is not None and pd.notna(tiempo) else '—'
        st.markdown(f"**Tiempo de clasificación:** {tiempo_txt}")
        st.markdown(f"**Runtime usado:** {'Sí' if origen == 'Runtime' else 'No'}")
        st.markdown(f"**Gold usado:** {'Sí' if origen == 'Gold' else 'No'}")

    st.markdown('**Evidencia:**')
    st.info(evidencia)

    if hp is not None:
        etapas = _explicar_clasificacion(
            hp,
            row.get('codigo_original', '') or '',
            nombre,
            origen_columna=row.get('origen_columna'),
            metodo_actual=regla,
        )
        if etapas:
            st.markdown('**Reglas evaluadas (auditoría read-only):**')
            df_etapas = pd.DataFrame([
                {k: v for k, v in e.items() if not k.startswith('_')} for e in etapas
            ])
            st.dataframe(df_etapas, use_container_width=True, hide_index=True)


def _tab_explicabilidad(df: pd.DataFrame, catalogo: dict, hp=None, archivo_nombre: str = ''):
    st.subheader('🔍 Explicabilidad de clasificación')
    st.caption(
        'Trazabilidad auditable por cuenta: código final, origen, score, regla, evidencia y '
        'confianza. La auditoría es de solo lectura y no modifica decisiones del motor.'
    )

    clasificadas = df[
        (df['codigo_clasificado'] != '') &
        (df['codigo_clasificado'] != '__EXCLUIR__') &
        (~df['es_total'])
    ].copy()
    if clasificadas.empty:
        st.info('No hay cuentas clasificadas para auditar.')
        return

    if 'origen' not in clasificadas.columns:
        clasificadas['origen'] = clasificadas['metodo'].map(_origen_desde_metodo_display)
    if 'regla' not in clasificadas.columns:
        clasificadas['regla'] = clasificadas['metodo']
    if 'evidencia' not in clasificadas.columns:
        clasificadas['evidencia'] = ''

    cf1, cf2 = st.columns(2)
    with cf1:
        filtro = st.text_input('Buscar cuenta (nombre o código)', key='exp_buscar')
    with cf2:
        origenes = ['Todos'] + sorted(o for o in clasificadas['origen'].dropna().unique() if o)
        sel_origen = st.selectbox('Origen', origenes, key='exp_origen')

    v = clasificadas
    if filtro:
        f = filtro.lower()
        v = v[
            v['nombre_original'].astype(str).str.lower().str.contains(f, na=False) |
            v['codigo_clasificado'].astype(str).str.lower().str.contains(f, na=False)
        ]
    if sel_origen != 'Todos':
        v = v[v['origen'] == sel_origen]
    v = v.sort_values(
        ['codigo_clasificado', 'monto'],
        key=lambda x: x.abs() if x.dtype.kind == 'f' else x,
        ascending=[True, False],
    )

    st.caption(f"`{archivo_nombre}` · {len(v)} cuenta(s) · filtro origen = {sel_origen}")

    for idx, row in v.iterrows():
        with st.container(border=True):
            cc1, cc2, cc3, cc4, cc5 = st.columns([1.2, 2.5, 1, 1, 1.5])
            with cc1:
                st.markdown(f"`{row['codigo_clasificado']}`")
            with cc2:
                st.markdown(f"**{_nombre_mostrar(row)}**")
            with cc3:
                st.markdown(_badge_origen(row.get('origen', '')), unsafe_allow_html=True)
            with cc4:
                st.markdown(f"**{_fmt_confianza(row.get('confianza'))}**")
            with cc5:
                st.markdown(f"`{row.get('regla', '')}`")
            if st.button('🔍 Ver detalle', key=f"det_exp_{idx}", use_container_width=True):
                st.session_state['detalle_explicabilidad_idx'] = idx
                st.rerun()

    det_idx = st.session_state.get('detalle_explicabilidad_idx')
    if det_idx is not None and det_idx in v.index:
        st.divider()
        _mostrar_detalle_cuenta(v.loc[det_idx], catalogo, hp=hp, close_key='detalle_explicabilidad_idx')


# ─────────────────────────────────────────────────────────────────────────────
# MOTOR DE CLASIFICACIÓN HÍBRIDA
# ─────────────────────────────────────────────────────────────────────────────

class MotorHibridoLocal:
    UMBRAL_CODIGO = 0.85
    NOMBRES_AMBIGUOS = {'arriendos', 'arriendo', 'intereses', 'interes', 'honorarios', 'comisiones', 'servicios'}
    UMBRAL_DICCIONARIO_EXACTO = 0.98
    UMBRAL_DICCIONARIO_FUZZY = 0.85
    UMBRAL_REGLA = 0.80

    def __init__(self, diccionario: list[dict]):
        self.clasificador_codigo = ClasificadorCodigo()
        self.reglas_especiales = ProcesadorReglasEspeciales()
        self.dic_exacto = {normalizar_nombre(d['cuenta_original']): d for d in diccionario}
        self.dic_lista = list(self.dic_exacto.keys())

    def clasificar(self, cuenta: CuentaRaw, giro_empresa: str | None = None) -> dict:
        nombre_norm = normalizar_nombre(cuenta.nombre)

        if (cuenta.origen_columna != OrigenColumna.DESCONOCIDO and nombre_norm in self.NOMBRES_AMBIGUOS):
            col = cuenta.origen_columna
            es_ing = es_ingreso_col(col)
            es_gas = es_gasto_col(col)
            MAPA_AMBIGUO = {
                'arriendos': ('ER.01', 'ER.04'), 'arriendo':  ('ER.01', 'ER.04'),
                'intereses': ('ER.12', 'ER.09'), 'interes':   ('ER.12', 'ER.09'),
                'honorarios':('ER.01', 'ER.04'), 'honorario': ('ER.01', 'ER.04'),
                'comisiones':('ER.01', 'ER.05'), 'servicios': ('ER.01', 'ER.04'),
            }
            if nombre_norm in MAPA_AMBIGUO:
                cod_ing, cod_gas = MAPA_AMBIGUO[nombre_norm]
                if es_ing:
                    return {'codigo_estandar': cod_ing, 'metodo': 'columna_ambiguo', 'confianza': 0.88, 'requiere_revision': False}
                if es_gas:
                    return {'codigo_estandar': cod_gas, 'metodo': 'columna_ambiguo', 'confianza': 0.88, 'requiere_revision': False}

        r_codigo = self.clasificador_codigo.clasificar(cuenta.codigo)
        if r_codigo and r_codigo.confianza >= self.UMBRAL_CODIGO:
            resultado = {'codigo_estandar': r_codigo.codigo_estandar, 'metodo': 'codigo', 'confianza': r_codigo.confianza}
        else:
            if nombre_norm in self.dic_exacto:
                d = self.dic_exacto[nombre_norm]
                resultado = {'codigo_estandar': d['codigo_estandar'], 'metodo': 'diccionario_exacto', 'confianza': self.UMBRAL_DICCIONARIO_EXACTO}
            else:
                match = process.extractOne(nombre_norm, self.dic_lista, scorer=fuzz.token_sort_ratio)
                if match and match[1] >= 90:
                    d = self.dic_exacto[match[0]]
                    resultado = {'codigo_estandar': d['codigo_estandar'], 'metodo': 'diccionario_fuzzy', 'confianza': round(0.80 + (match[1] - 90) * 0.01, 3)}
                else:
                    mejor = None
                    for patron, cod, conf in REGLAS_COMPILADAS:
                        if patron.search(nombre_norm):
                            if mejor is None or conf > mejor[1]:
                                mejor = (cod, conf)
                    if mejor:
                        resultado = {'codigo_estandar': mejor[0], 'metodo': 'regla_regex', 'confianza': mejor[1]}
                    else:
                        resultado = {'codigo_estandar': None, 'metodo': 'sin_clasificar', 'confianza': 0.0}

        codigo_actual = resultado.get('codigo_estandar')
        origen = cuenta.origen_columna
        if origen != OrigenColumna.DESCONOCIDO and codigo_actual:
            codigo_corregido = self._corregir_por_columna(nombre_norm, codigo_actual, origen, cuenta.monto)
            if codigo_corregido and codigo_corregido != codigo_actual:
                resultado['codigo_estandar'] = codigo_corregido
                resultado['metodo'] += '+columna'
                resultado['confianza'] = min(resultado['confianza'] + 0.05, 0.99)

        codigo_pre = resultado['codigo_estandar'] or 'AC.08'
        ajuste = self.reglas_especiales.aplicar(
            cuenta.nombre, codigo_pre, cuenta.monto, giro_empresa,
            origen_columna=cuenta.origen_columna,
        )
        if ajuste.aplica:
            resultado['codigo_estandar'] = ajuste.codigo_final
            resultado['metodo'] += f'+regla_especial({ajuste.flag})'
            resultado['nota_regla_especial'] = ajuste.nota

        resultado['requiere_revision'] = (
            resultado['confianza'] < UMBRAL_REVISION
            or (ajuste.aplica and ajuste.requiere_revision)
        )
        return resultado

    def _corregir_por_columna(self, nombre_norm, codigo_actual, origen, monto):
        from parser_universal import OrigenColumna as OC
        es_ingreso = es_ingreso_col(origen)
        es_gasto   = es_gasto_col(origen)

        AMBIGUOS = [
            ('arriendo', 'ER.01', 'ER.04'), ('interes', 'ER.12', 'ER.09'),
            ('comision', 'ER.01', 'ER.05'), ('servicio', 'ER.01', 'ER.04'),
            ('honorario', 'ER.01', 'ER.04'), ('diferencia de cambio', 'ER.15', 'ER.15'),
            ('correccion monetaria', 'ER.14', 'ER.14'), ('otras ganancias', 'ER.13', 'ER.13'),
            ('reajuste', 'ER.14', 'ER.14'),
        ]

        if (monto is not None and monto < 0 and origen == OC.ACTIVO and codigo_actual in ('AC.01',)
                and any(k in nombre_norm for k in ('banco','cta cte','cuenta corriente'))):
            return 'PC.02'

        for keyword, cod_ing, cod_gas in AMBIGUOS:
            if keyword in nombre_norm:
                if es_ingreso and codigo_actual not in (cod_ing, cod_gas): return cod_ing
                if es_gasto and codigo_actual not in (cod_ing, cod_gas): return cod_gas

        if origen in (OC.ACTIVO, OC.PASIVO) and codigo_actual and codigo_actual.startswith('ER'):
            return None

        return None


# ─────────────────────────────────────────────────────────────────────────────
# INTERFAZ DE USUARIO PRINCIPAL (MAIN)
# ─────────────────────────────────────────────────────────────────────────────

def _contenido_para_extraer(archivo) -> bytes:
    content = archivo.getvalue()
    if Path(archivo.name).suffix.lower() == ".pdf":
        pages = st.session_state.setdefault("document_pages", {}).get(archivo.name)
        if pages is not None:
            return select_pdf(content, pages)
    return content


def _confirmar_alcance_documentos(archivos) -> bool:
    if len({a.name for a in archivos}) != len(archivos):
        st.error("Hay archivos con el mismo nombre. Renómbrelos para mantener separadas sus páginas y correcciones.")
        return False
    signature = tuple((a.name, hashlib.sha256(a.getvalue()).hexdigest()) for a in archivos)
    if (st.session_state.get("document_scope_confirmed") == signature
            and not st.session_state.get("document_scope_editing", False)):
        # Conserva el valor del selector aunque esta vista se omita durante el
        # procesamiento; al volver a editar el alcance, Streamlit puede
        # reconstruir el mismo widget sin perder su estado.
        st.session_state.setdefault("document_scope_preview", archivos[0].name)
        return True
    st.subheader("Seleccione las páginas que se analizarán")
    st.info(
        "Revise el documento completo en el visor inferior. Si contiene el mismo balance "
        "en varios formatos, elija sólo uno. En informes auditados seleccione las páginas "
        "del estado y sus continuaciones, sin volver a incluir sus anexos o duplicados."
    )
    if st.session_state.get("document_scope_editing"):
        st.warning("Cambiar las páginas vuelve a procesar ese documento y reemplaza sus correcciones en esta sesión. Los documentos sin cambios conservan sus resultados.")
    preview_name = st.selectbox(
        "Documento a visualizar", [a.name for a in archivos],
        key="document_scope_preview",
    )
    selections = {}
    errors = []
    with st.form("document_scope"):
        for archivo in archivos:
            if Path(archivo.name).suffix.lower() != ".pdf":
                continue
            try:
                count = page_count(archivo.getvalue())
            except Exception:
                errors.append(f"No se pudo abrir {archivo.name} como PDF.")
                continue
            st.write(f"{archivo.name}: {count} páginas")
            key = hashlib.sha256(archivo.name.encode() + archivo.getvalue()).hexdigest()[:16]
            saved_pages = st.session_state.get("document_pages", {}).get(archivo.name, [])
            saved_subset = bool(saved_pages) and saved_pages != list(range(1, count + 1))
            mode = st.radio("Páginas a analizar", ["Todas", "Sólo las seleccionadas"],
                            index=1 if saved_subset else 0,
                            key=f"scope_mode_{key}", horizontal=True)
            text = st.text_input("Páginas o rangos", placeholder="1, 3-5",
                                 value=", ".join(map(str, saved_pages)) if saved_subset else "",
                                 key=f"scope_pages_{key}")
            try:
                selections[archivo.name] = list(range(1, count + 1)) if mode == "Todas" else parse_pages(text, count)
            except ValueError as exc:
                errors.append(f"{archivo.name}: {exc}")
        submitted = st.form_submit_button("Confirmar páginas y continuar")
    if submitted:
        if errors:
            for error in errors:
                st.error(error)
        else:
            previous = dict(st.session_state.get("document_scope_confirmed") or ())
            previous_pages = st.session_state.get("document_pages", {})
            for name, digest in signature:
                if previous.get(name) != digest or previous_pages.get(name) != selections.get(name):
                    revisions = st.session_state.setdefault("extraction_revisions", {})
                    revisions[name] = revisions.get(name, 0) + 1
                    st.session_state.metadata_confirmada = False
                    st.session_state.company_periodos_detectados = ()
                    st.session_state.company_periodos_seleccionados = ()
                    for state_key in ("resultados", "metadata_files", "document_intel", "extraction_pending",
                                      "extraction_resolved", "extraction_certifications", "processed_at",
                                      "quality_controls"):
                        st.session_state.setdefault(state_key, {}).pop(name, None)
            st.session_state.document_pages = selections
            st.session_state.document_scope_confirmed = signature
            st.session_state.document_scope_editing = False
            return True
    st.divider()
    _visor_documento(next(a for a in archivos if a.name == preview_name),
                     altura="58vh", mostrar_titulo=True)
    return False


def main():
    st.title("📊 Homologación de Balances Tributarios Chilenos")
    st.caption(
        "Carga uno o más balances (PDF o Excel) → clasificación híbrida automática "
        "(código → diccionario → reglas) → cola de revisión → balance normalizado."
    )

    catalogo = cargar_catalogo()
    dic_base = cargar_diccionario_base()

    if 'diccionario' not in st.session_state:
        st.session_state.diccionario = list(dic_base)
    if 'resultados' not in st.session_state or not isinstance(st.session_state.resultados, dict):
        st.session_state.resultados = {}
    if 'metadata_files' not in st.session_state:
        st.session_state.metadata_files = {}
    if 'document_intel' not in st.session_state:
        st.session_state.document_intel = {}
    if 'extraction_pending' not in st.session_state:
        st.session_state.extraction_pending = {}
    if 'extraction_resolved' not in st.session_state:
        st.session_state.extraction_resolved = {}
    if 'extraction_certifications' not in st.session_state:
        st.session_state.extraction_certifications = {}
    if 'correcciones' not in st.session_state:
        st.session_state.correcciones = []

    with st.sidebar:
        st.header("⚙️ Configuración")
        commit = os.environ.get("RENDER_GIT_COMMIT", "local")[:8]
        release_branch = os.environ.get("APP_RELEASE_BRANCH", "desarrollo-local")
        build_date = _build_date()
        persistence_label = "Neon conectado" if _neon_disponible() else "JSON local"
        st.caption(
            f"Rama `{release_branch}` · Commit `{commit}` · Build `{build_date}` · "
            f"Persistencia: **{persistence_label}**"
        )
        archivos = st.file_uploader(
            "Balances tributarios", type=['pdf', 'xlsx', 'xls'], accept_multiple_files=True
        )
        if archivos and st.session_state.get("document_scope_confirmed"):
            if st.button("Cambiar páginas a analizar"):
                st.session_state.document_scope_editing = True
                st.rerun()
        giro = st.selectbox(
            "Giro de la empresa (afecta regla D2-Terrenos)",
            ['Otro', 'Inmobiliaria', 'Construcción', 'Promotora'],
            help="Si el giro es inmobiliario/construcción, los terrenos en activo corriente se reclasifican como inventario."
        )
        giro_norm = None if giro == 'Otro' else giro.lower()

        st.divider()
        st.metric("Cuentas en diccionario", len(st.session_state.diccionario))
        st.metric("Códigos en catálogo", len(catalogo))

        if st.session_state.correcciones:
            st.divider()
            st.success(f"{len(st.session_state.correcciones)} correcciones pendientes")
            buf = json.dumps(st.session_state.diccionario, ensure_ascii=False, indent=2)
            st.download_button(
                "⬇️ Descargar diccionario actualizado",
                data=buf, file_name="diccionario_actualizado.json",
                mime="application/json"
            )

    if not archivos:
        st.info("⬆️ Carga uno o más archivos en la barra lateral para comenzar.")
        st.session_state.resultados = {}
        st.session_state.metadata_files = {}
        st.session_state.extraction_pending = {}
        st.session_state.extraction_resolved = {}
        st.session_state.extraction_certifications = {}
        st.session_state.metadata_confirmada = False
        st.session_state.document_scope_confirmed = None
        st.session_state.document_scope_editing = False
        st.session_state.document_pages = {}
        st.session_state.company_periodos_detectados = ()
        st.session_state.company_periodos_seleccionados = ()
        _mostrar_resumen_catalogo(catalogo)
        return

    if not _confirmar_alcance_documentos(archivos):
        return

    if not st.session_state.get('metadata_confirmada', False):
        first_file = archivos[0]
        with st.spinner(f"Detectando metadata de la empresa en {first_file.name}..."):
            lineas_encabezado = _extraer_lineas_encabezado(first_file)
            meta = extraer_metadata(lineas_encabezado)
            st.session_state.company_rut = meta.rut or ""
            st.session_state.company_razon = meta.razon_social or ""
            st.session_state.company_giro = meta.giro or "Otro"
            mes_detectado, anio_detectado, meses_detectados = _valores_periodo_metadata(meta)
            st.session_state.company_moneda = meta.moneda or "$"
            st.session_state.company_mes = mes_detectado
            st.session_state.company_anio = anio_detectado
            st.session_state.company_numero_meses = meses_detectados
            periodos_detectados = _detectar_periodos_comparativos(
                lineas_encabezado, anio_detectado,
            )
            st.session_state.company_periodos_detectados = periodos_detectados
            seleccion_guardada = tuple(
                st.session_state.get("company_periodos_seleccionados", ())
            )
            if not seleccion_guardada or not set(seleccion_guardada).issubset(
                periodos_detectados
            ):
                st.session_state.company_periodos_seleccionados = periodos_detectados

        st.subheader("📋 Confirma los datos de la empresa")
        st.caption("El sistema detectó los siguientes datos generales. Corrígelos si es necesario antes de continuar.")

        with st.form("form_empresa"):
            col1, col2 = st.columns(2)
            with col1:
                rut = st.text_input("RUT", value=st.session_state.company_rut)
                razon = st.text_input("Razón Social", value=st.session_state.company_razon)
            with col2:
                giro_list = ['Otro', 'Inmobiliaria', 'Construcción', 'Promotora']
                default_giro_idx = 0
                if st.session_state.company_giro in giro_list:
                    default_giro_idx = giro_list.index(st.session_state.company_giro)
                giro_sel = st.selectbox(
                    "Giro de la empresa (afecta regla D2-Terrenos)",
                    giro_list, index=default_giro_idx,
                    help="Escriba dentro del desplegable para filtrar opciones.",
                )

            st.markdown("##### Moneda y período informado")
            moneda_col, otra_col = st.columns(2)
            moneda_opciones = ["$", "M", "MM", "USD", "Otra"]
            moneda_actual = st.session_state.company_moneda
            moneda_indice = (
                moneda_opciones.index(moneda_actual)
                if moneda_actual in moneda_opciones else moneda_opciones.index("Otra")
            )
            with moneda_col:
                moneda_sel = st.selectbox(
                    "Moneda o unidad",
                    moneda_opciones,
                    index=moneda_indice,
                    help=(
                        "$: pesos; M: miles; MM: millones; USD: dólares. "
                        "Escriba para buscar una opción."
                    ),
                    filter_mode="fuzzy",
                )
            with otra_col:
                otra_moneda = st.text_input(
                    "Otra moneda o unidad",
                    value=(moneda_actual if moneda_actual not in moneda_opciones else ""),
                    placeholder="Ejemplo: EUR, UF, UTM",
                    help="Complete este campo únicamente si seleccionó Otra.",
                )

            periodo1, periodo2, periodo3 = st.columns(3)
            with periodo1:
                mes_sel = st.selectbox(
                    "Mes de cierre",
                    MESES_SELECCION,
                    index=MESES_SELECCION.index(st.session_state.company_mes),
                    help="Escriba el nombre del mes para encontrarlo.",
                    filter_mode="fuzzy",
                )
            anio_actual = date.today().year
            anios = list(range(anio_actual + 2, 1979, -1))
            if int(st.session_state.company_anio) not in anios:
                anios.append(int(st.session_state.company_anio))
                anios.sort(reverse=True)
            with periodo2:
                anio_sel = st.selectbox(
                    "Año de cierre",
                    anios,
                    index=anios.index(int(st.session_state.company_anio)),
                    help="Escriba el año para encontrarlo.",
                    filter_mode="fuzzy",
                )
            opciones_meses = list(range(1, 13))
            with periodo3:
                numero_meses_sel = st.selectbox(
                    "Número de meses del período",
                    opciones_meses,
                    index=opciones_meses.index(
                        int(st.session_state.company_numero_meses)
                    ),
                    help="Escriba un número entre 1 y 12 para encontrarlo.",
                    filter_mode="fuzzy",
                )

            periodos_detectados = tuple(
                st.session_state.get(
                    "company_periodos_detectados", (str(anio_sel),),
                )
            )
            if len(periodos_detectados) >= 2:
                actual, anterior = periodos_detectados[:2]
                opciones_periodo = {
                    f"Ambos períodos: {actual} y {anterior}": (actual, anterior),
                    f"Sólo {actual}": (actual,),
                    f"Sólo {anterior}": (anterior,),
                }
            else:
                unico = periodos_detectados[0]
                opciones_periodo = {f"Sólo {unico}": (unico,)}
            seleccion_actual = tuple(
                st.session_state.get(
                    "company_periodos_seleccionados", periodos_detectados,
                )
            )
            etiquetas_periodo = list(opciones_periodo)
            indice_periodo = next(
                (
                    indice for indice, etiqueta in enumerate(etiquetas_periodo)
                    if opciones_periodo[etiqueta] == seleccion_actual
                ),
                0,
            )
            alcance_periodos = st.selectbox(
                "Períodos a extraer",
                etiquetas_periodo,
                index=indice_periodo,
                help=(
                    "Ambos conserva una sola clasificación por cuenta y un "
                    "importe independiente para cada período."
                ),
            )

            submitted = st.form_submit_button("Confirmar y procesar todos los balances")
            if submitted:
                moneda_final = otra_moneda.strip() if moneda_sel == "Otra" else moneda_sel
                if not moneda_final:
                    st.error("Especifique la moneda o unidad cuando seleccione Otra.")
                else:
                    st.session_state.company_rut = rut
                    st.session_state.company_razon = razon
                    st.session_state.company_giro = giro_sel
                    st.session_state.company_moneda = moneda_final
                    st.session_state.company_mes = mes_sel
                    st.session_state.company_anio = int(anio_sel)
                    st.session_state.company_numero_meses = int(numero_meses_sel)
                    st.session_state.company_periodos_seleccionados = (
                        opciones_periodo[alcance_periodos]
                    )
                    st.session_state.company_anio = int(
                        st.session_state.company_periodos_seleccionados[0]
                    )
                    st.session_state.metadata_confirmada = True
                    st.session_state.resultados = {}
                    st.session_state.metadata_files = {}
                    st.rerun()
        st.divider()
        st.subheader("Documento original")
        st.caption(
            "Revise el encabezado y el período antes de confirmar. El visor ocupa "
            "al menos la mitad inferior de la pantalla."
        )
        _visor_documento(first_file, altura="58vh", mostrar_titulo=False)
        return

    company_giro_norm = None if st.session_state.company_giro == 'Otro' else st.session_state.company_giro.lower()

    nombres_subidos = [a.name for a in archivos]
    for k in list(st.session_state.resultados.keys()):
        if k not in nombres_subidos:
            st.session_state.resultados.pop(k, None)
            st.session_state.metadata_files.pop(k, None)
    for state_key in ('extraction_pending', 'extraction_resolved', 'extraction_certifications'):
        state = st.session_state.get(state_key, {})
        for k in list(state.keys()):
            if k not in nombres_subidos:
                state.pop(k, None)

    # ── Procesar archivos nuevos ──────────────────────────────────────────────
    if USE_LEGACY_ENGINE:
        # LEGACY PIPELINE (MotorHibridoLocal)
        for archivo in archivos:
            if archivo.name not in st.session_state.resultados:
                with st.spinner(f"Clasificando cuentas de {archivo.name}..."):
                    lineas_encabezado = _extraer_lineas_encabezado(archivo)
                    meta_indiv = _aplicar_metadata_confirmada(
                        extraer_metadata(lineas_encabezado)
                    )
                    st.session_state.metadata_files[archivo.name] = meta_indiv

                    if archivo.name in st.session_state.extraction_resolved:
                        cuentas = st.session_state.extraction_resolved.pop(archivo.name)
                        doc_ctx = st.session_state.document_intel.get(archivo.name)
                    else:
                        cuentas, doc_ctx = _extraer_cuentas(archivo)
                    st.session_state.document_intel[archivo.name] = doc_ctx
                    motor = MotorHibridoLocal(st.session_state.diccionario)
                    filas = []
                    for c in cuentas:
                        if c.monto is None and not c.codigo:
                            continue
                        if not _cuenta_tiene_saldo_seleccionado(c):
                            continue
                        if not c.codigo and PATRON_NO_CUENTA.match(c.nombre.strip()):
                            continue
                        monto_seleccionado, campos_periodos = (
                            _campos_periodos_cuenta(c)
                        )
                        cuenta_clasificacion = replace(
                            c, monto=monto_seleccionado,
                        )
                        _t0_legacy = time.perf_counter()
                        r = motor.clasificar(
                            cuenta_clasificacion, company_giro_norm,
                        )
                        _t1_legacy = (time.perf_counter() - _t0_legacy) * 1000
                        origen_efectivo = _origen_efectivo(
                            c.origen_columna, monto_seleccionado, c.nombre,
                        )
                        if (r.get('codigo_estandar') and
                                not _codigo_compatible_con_origen(
                                    r['codigo_estandar'], c.origen_columna,
                                    monto_seleccionado,
                                    c.nombre)):
                            r = {
                                **r,
                                'codigo_estandar': None,
                                'metodo': 'sin_clasificar+filtro_columna',
                                'confianza': 0.0,
                                'requiere_revision': True,
                            }
                        filas.append({
                            'linea': c.linea,
                            'codigo_original': c.codigo or '',
                            'nombre_original': c.nombre,
                            'nombre_normalizado': normalizar_nombre(c.nombre),
                            'monto': monto_seleccionado,
                            **campos_periodos,
                            'columnas_derivadas': ', '.join(c.columnas_derivadas),
                            'origen_columna': c.origen_columna.value,
                            'origen_columna_efectiva': origen_efectivo,
                            'es_total': c.es_total,
                            'codigo_clasificado': r['codigo_estandar'] or '',
                            'metodo': r['metodo'],
                            'confianza': r['confianza'],
                            'requiere_revision': (
                                r['requiere_revision'] or bool(c.columnas_derivadas)
                            ),
                            'nota': r.get('nota_regla_especial', ''),
                            'confianza_extraccion': c.confianza_extraccion,
                            'origen_columna_display': _etiqueta_origen(
                                c.origen_columna, monto_seleccionado, c.nombre,
                            ),
                            'nombre_revision_usuario': '',
                            'tipo_revision': '',
                            'origen': _origen_desde_metodo_display(r['metodo']),
                            'regla': r['metodo'],
                            'evidencia': r.get('nota_regla_especial', '') or r['metodo'],
                            'tiempo_clasificacion': round(_t1_legacy, 3),
                        })
                    df_file = pd.DataFrame(filas)
                    st.session_state.resultados[archivo.name] = df_file

                    # SHADOW MODE — homologación comparativa contra motor legacy
                    if SHADOW_MODE and Path(archivo.name).suffix.lower() == '.pdf':
                        import tempfile
                        from pipeline.homologation_pipeline import HomologationPipeline
                        from shadow.shadow_logger import ShadowLogger
                        tmp_shadow = tempfile.NamedTemporaryFile(suffix='.pdf', delete=False)
                        try:
                            tmp_shadow.write(archivo.read())
                            tmp_shadow.close()
                            sh_path = Path(tmp_shadow.name)
                            hp_shadow = HomologationPipeline()
                            sh_summary = hp_shadow.process(sh_path)
                            logger_sh = ShadowLogger()
                            comparisons = []
                            matches = 0
                            for sh_entry in sh_summary.get("classified", []):
                                aname = sh_entry["account_name"]
                                amatch = df_file[df_file["nombre_original"] == aname]
                                if amatch.empty:
                                    continue
                                comparisons.append(logger_sh.build_comparison(
                                    account_name=aname,
                                    account_code=sh_entry.get("account_code", ""),
                                    legacy_code=amatch.iloc[0]["codigo_clasificado"] or None,
                                    legacy_confidence=amatch.iloc[0]["confianza"],
                                    new_code=sh_entry.get("final_code") or sh_entry.get("standard_code"),
                                    new_confidence=sh_entry.get("confidence", 0.0),
                                    new_method=sh_entry.get("method", ""),
                                    learning_hit="learning" in sh_entry.get("method", ""),
                                ))
                                if comparisons[-1]["match"]:
                                    matches += 1
                            total_comp = len(comparisons)
                            match_rate = matches / total_comp if total_comp else 1.0
                            logger_sh.log(archivo.name, comparisons, match_rate)
                        finally:
                            sh_path.unlink(missing_ok=True)
                            archivo.seek(0)
    else:
        # NEW PIPELINE (HomologationPipeline)
        import logging as _logging
        _shadow_logger = _logging.getLogger("homologation_pipeline")
        from pipeline.homologation_pipeline import HomologationPipeline
        from adapters.account_adapter import AccountAdapter
        from interpreters.balance_interpreter import BalanceInterpreter

        archivos_nuevos = [
            archivo for archivo in archivos
            if archivo.name not in st.session_state.resultados
        ]
        hp = HomologationPipeline() if archivos_nuevos else None
        for archivo in archivos_nuevos:
            if archivo.name not in st.session_state.resultados:
                with st.spinner(f"Clasificando cuentas de {archivo.name}..."):
                    _t0 = time.perf_counter()
                    lineas_encabezado = _extraer_lineas_encabezado(archivo)
                    meta_indiv = _aplicar_metadata_confirmada(
                        extraer_metadata(lineas_encabezado)
                    )
                    st.session_state.metadata_files[archivo.name] = meta_indiv

                    if archivo.name in st.session_state.extraction_resolved:
                        cuentas = st.session_state.extraction_resolved.pop(archivo.name)
                        doc_ctx = st.session_state.document_intel.get(archivo.name)
                    else:
                        cuentas, doc_ctx = _extraer_cuentas(archivo)
                    st.session_state.document_intel[archivo.name] = doc_ctx
                    total_cuentas = len(cuentas)
                    filas = []
                    clasificadas = 0
                    learning_hits = 0
                    fallback_count = 0

                    for c in cuentas:
                        if c.monto is None and not c.codigo:
                            continue
                        if not _cuenta_tiene_saldo_seleccionado(c):
                            continue
                        if not c.codigo and PATRON_NO_CUENTA.match(c.nombre.strip()):
                            continue
                        monto_seleccionado, campos_periodos = _campos_periodos_cuenta(c)
                        cuenta_clasificacion = replace(c, monto=monto_seleccionado)
                        ab = AccountAdapter.from_cuenta_raw(cuenta_clasificacion)
                        interp = BalanceInterpreter(ab)
                        classification_amount = interp.classification_amount
                        origen_efectivo = _origen_efectivo(
                            c.origen_columna, classification_amount, c.nombre,
                        )
                        account_tipo = _resolver_tipo_cuenta(origen_efectivo, c.codigo)
                        if (origen_efectivo == 'pasivo'
                                and _es_contra_activo(c.nombre)):
                            account_tipo = 'ACTIVO'
                        if _es_partida_patrimonial(c.nombre):
                            account_tipo = 'PATRIMONIO'
                        if classification_amount is None:
                            codigo_clasificado = ""
                            metodo = "movement_only"
                            confianza = 0.0
                            requiere_revision = True
                            nota = ""
                            clasif_origen = 'Sin clasificar'
                            clasif_evidencia = ''
                            tiempo_clasif_ms = 0.0
                        else:
                            clasificadas += 1
                            _t0_clasif = time.perf_counter()
                            classification = hp._classify_account(
                                ab.account_code,
                                ab.account_name,
                                account_tipo=account_tipo,
                            )
                            tiempo_clasif_ms = round((time.perf_counter() - _t0_clasif) * 1000, 3)
                            adjustment = hp._rule_processor.aplicar(
                                nombre_cuenta=ab.account_name,
                                codigo_clasificado=classification.get("standard_code") or "",
                                monto=classification_amount,
                                origen_columna=c.origen_columna,
                            )
                            final_code = (
                                adjustment.codigo_final if adjustment.aplica
                                else classification.get("standard_code")
                            )
                            if (final_code and
                                    not _codigo_compatible_con_origen(
                                        final_code, c.origen_columna,
                                        classification_amount, c.nombre)):
                                final_code = None
                                classification = {
                                    **classification,
                                    "standard_code": None,
                                    "confidence": 0.0,
                                    "method": "unclassified_type_filter",
                                    "reason": (
                                        "Sugerencia descartada por ser incompatible con "
                                        f"la columna efectiva {account_tipo}"
                                    ),
                                }
                            codigo_clasificado = final_code or ""
                            metodo = classification.get("method", "")
                            confianza = classification.get("confidence", 0.0)
                            requiere_revision = (
                                confianza < UMBRAL_REVISION
                                or (adjustment.aplica and adjustment.requiere_revision)
                                or bool(c.columnas_derivadas)
                            )
                            nota = adjustment.nota if adjustment.aplica else ""
                            clasif_evidencia = classification.get("reason", "")
                            clasif_origen = _origen_desde_clasif(classification, ab.account_name, hp)
                            if metodo.startswith("learning_"):
                                learning_hits += 1
                            else:
                                fallback_count += 1

                        filas.append({
                            'linea': c.linea,
                            'codigo_original': c.codigo or '',
                            'nombre_original': c.nombre,
                            'nombre_normalizado': normalizar_nombre(c.nombre),
                            'monto': monto_seleccionado,
                            **campos_periodos,
                            'columnas_derivadas': ', '.join(c.columnas_derivadas),
                            'origen_columna': c.origen_columna.value,
                            'origen_columna_efectiva': origen_efectivo,
                            'es_total': c.es_total,
                            'codigo_clasificado': codigo_clasificado,
                            'metodo': metodo,
                            'confianza': confianza,
                            'requiere_revision': requiere_revision,
                            'nota': nota,
                            'confianza_extraccion': c.confianza_extraccion,
                            'origen_columna_display': _etiqueta_origen(
                                c.origen_columna, classification_amount, c.nombre),
                            'nombre_revision_usuario': '',
                            'tipo_revision': '',
                            'origen': clasif_origen,
                            'regla': metodo,
                            'evidencia': clasif_evidencia,
                            'tiempo_clasificacion': tiempo_clasif_ms,
                        })

                    df_file = pd.DataFrame(filas)
                    st.session_state.resultados[archivo.name] = df_file

                    _t1 = time.perf_counter()
                    _shadow_logger.info(
                        "archivo=%s cuentas=%d clasificadas=%d learning_hits=%d fallback=%d time=%.3fs",
                        archivo.name, total_cuentas, clasificadas,
                        learning_hits, fallback_count, _t1 - _t0,
                    )
        # P5.5 Runtime Observability: persistir métricas de uso para Runtime Analytics.
        try:
            st.session_state["runtime_metrics_last"] = hp._learning_engine.get_metrics()
        except Exception:  # noqa: BLE001 — observabilidad no debe interrumpir el flujo
            pass
    # After processing all uploaded files, propagate classifications across all balances
    if 'propagation_done' not in st.session_state:
        # Define helper to propagate classifications across balances
        def propagar_entre_balances():
            """Propaga códigos clasificados entre balances cargados."""
            # Build mapping from normalized account name to list of (file, index)
            name_map = {}
            for fname, df in st.session_state.resultados.items():
                for idx, row in df.iterrows():
                    nombre = row['nombre_original']
                    if not nombre:
                        continue
                    norm = normalizar_nombre(nombre)
                    name_map.setdefault(norm, []).append((fname, idx, row['codigo_clasificado']))
            # Propagar si existe alguna clasificación
            for norm, entries in name_map.items():
                classified_code = None
                for fname, idx, cod in entries:
                    if cod and cod not in ('', '__EXCLUIR__'):
                        classified_code = cod
                        break
                if classified_code:
                    for fname, idx, cod in entries:
                        row = st.session_state.resultados[fname].loc[idx]
                        if ((not cod or cod == '')
                                and _codigo_compatible_con_origen(
                                    classified_code,
                                    row.get('origen_columna'),
                                    row.get('monto'),
                                    row.get('nombre_original'),
                                )):
                            st.session_state.resultados[fname].at[idx, 'codigo_clasificado'] = classified_code
                            st.session_state.resultados[fname].at[idx, 'metodo'] = 'propagado_automático'
                            st.session_state.resultados[fname].at[idx, 'confianza'] = 1.0
                            st.session_state.resultados[fname].at[idx, 'requiere_revision'] = False
                            if 'origen' in st.session_state.resultados[fname].columns:
                                st.session_state.resultados[fname].at[idx, 'origen'] = 'Código'
                                st.session_state.resultados[fname].at[idx, 'regla'] = 'propagado_automático'
                                st.session_state.resultados[fname].at[idx, 'evidencia'] = 'Propagación automática entre balances'
        propagar_entre_balances()
        st.session_state['propagation_done'] = True

    with st.sidebar:
        st.divider()
        st.subheader("📁 Balances Cargados")
        for name in nombres_subidos:
            df_f = st.session_state.resultados.get(name)
            if df_f is not None and not df_f.empty:
                pendientes_f = _pendientes_revision(df_f)
                n_pendientes = len(pendientes_f)
                if n_pendientes > 0:
                    st.caption(f"🔸 `{name}` ({n_pendientes} pendientes)")
                else:
                    st.caption(f"✅ `{name}` (completo)")
            else:
                st.caption(f"⏳ `{name}` (procesando)")

        st.write("")
        archivo_activo_name = st.selectbox(
            "Ver y validar balance:", options=nombres_subidos,
            index=0 if nombres_subidos else None, key="archivo_activo_select"
        )

    try:
        archivo_activo = next(a for a in archivos if a.name == archivo_activo_name)
        df = st.session_state.resultados[archivo_activo_name]
        meta_activo = st.session_state.metadata_files.get(archivo_activo_name)
    except (NameError, Exception):
        class DummyDF: empty = False
        df = DummyDF()
        archivo_activo_name = ""
        archivo_activo = None
        meta_activo = None

    if df.empty:
        if archivo_activo_name in st.session_state.extraction_pending:
            _mostrar_etapa_correccion_extraccion(archivo_activo, archivo_activo_name)
        else:
            st.warning(f"No se extrajeron cuentas del archivo {archivo_activo_name}.")
        st.stop()

    with st.container(border=True):
        c1, c2, c3, c4 = st.columns(4)
        razon_social = getattr(st.session_state, "company_razon", "—")
        rut_empresa = getattr(st.session_state, "company_rut", "—")
        c1.markdown(f"**{razon_social or '—'}** \n`{rut_empresa or '—'}`")
        periodo_str = "—"
        if meta_activo:
            periodo_str = (
                f"{meta_activo.periodo_desde or '—'} → "
                f"{meta_activo.periodo_hasta or '—'} · "
                f"{meta_activo.numero_meses or '—'} meses · "
                f"{meta_activo.moneda or '—'}"
            )
        c2.markdown(f"**Período y moneda** \n{periodo_str}")
        giro_empresa = getattr(st.session_state, "company_giro", "—")
        c3.markdown(f"**Giro** \n{giro_empresa or '—'}")
        c4.markdown(f"**Archivo Activo** \n`{archivo_activo_name}`")

    _mostrar_advertencias_auxiliares(
        st.session_state.extraction_certifications.get(archivo_activo_name)
    )

    # Sprint 31 — Información del documento (análisis documental, solo lectura)
    _doc_ctx = st.session_state.document_intel.get(archivo_activo_name)
    if _doc_ctx is not None:
        _mostrar_informacion_documento(_doc_ctx)

    col_visor, col_trabajo = st.columns([1, 1], gap="medium")

    with col_visor:
        try:
            _visor_documento(archivo_activo)
        except (NameError, Exception):
            pass

    with col_trabajo:
        tab_km = "🧠 Knowledge Manager"
        vistas = [
            "📈 Resumen", "🔍 Cola de Revisión", "📋 Balance Normalizado",
            "📚 Diccionario", "🧠 Aprendizaje", "📊 Analytics",
            "📖 Conocimiento Documental", "📈 Inteligencia del Dataset",
            tab_km,
        ]
        vista_solicitada = st.session_state.pop("vista_trabajo_solicitada", None)
        if vista_solicitada in vistas:
            st.session_state["vista_trabajo"] = vista_solicitada
        vista = st.selectbox(
            "Vista de trabajo", vistas, key="vista_trabajo",
            help="Solo se ejecuta la vista seleccionada para evitar recargas innecesarias.",
        )

        if vista == "📈 Resumen":
            _tab_resumen(df)
        elif vista == "🔍 Cola de Revisión":
            _tab_revision(
                df, catalogo,
                motor=MotorHibridoLocal(st.session_state.diccionario),
                archivo_nombre=archivo_activo_name,
            )
        elif vista == "📋 Balance Normalizado":
            _tab_balance(df, catalogo, archivo_activo_name)
        elif vista == "📚 Diccionario":
            _tab_diccionario()
        elif vista == "🧠 Aprendizaje":
            _tab_aprendizaje()
        elif vista == "📊 Analytics":
            st.markdown("Analytics Dashboard (Work in Progress)")
        elif vista == "📖 Conocimiento Documental":
            _tab_conocimiento(archivo_activo, _doc_ctx, meta_activo)
        elif vista == "📈 Inteligencia del Dataset":
            _tab_inteligencia()
        elif vista == tab_km:
            _tab_knowledge_manager()


@st.cache_data(max_entries=12, show_spinner=False)
def _render_pdf_cached(content: bytes, page: int) -> bytes:
    return render_page(content, page)


@st.fragment
def _visor_documento(
    archivo, *, altura: str = "72vh", mostrar_titulo: bool = True,
):
    import tempfile, base64, io, platform, shutil, subprocess, glob
    from PIL import Image
    from pathlib import Path

    suffix = Path(archivo.name).suffix.lower()
    archivo.seek(0)
    if mostrar_titulo:
        st.markdown("#### 📄 Documento original")

    if suffix == '.pdf':
        content = archivo.getvalue()
        contenido_id = hashlib.sha256(content).hexdigest()[:16]
        try:
            n_paginas = page_count(content)
        except Exception:
            st.error("No se pudo abrir el PDF para visualizarlo.")
            return
        ctrl1, ctrl2, ctrl3 = st.columns(3)
        with ctrl1:
            pagina = st.number_input(f"Página (1-{n_paginas})", min_value=1,
                                     max_value=n_paginas, value=1, step=1,
                                     key=f"visor_pagina_{contenido_id}")
        with ctrl2:
            zoom = st.slider("Zoom", 50, 200, 100, 10, format="%d%%",
                             key=f"visor_zoom_{contenido_id}")
        with ctrl3:
            rotacion = st.select_slider("Rotación", options=[0, 90, 180, 270],
                                        key=f"visor_rot_{contenido_id}")
        try:
            img = Image.open(BytesIO(_render_pdf_cached(content, int(pagina))))
            if rotacion:
                img = img.rotate(-rotacion, expand=True)
            if zoom != 100:
                img = img.resize((int(img.width * zoom / 100), int(img.height * zoom / 100)))
            buf = io.BytesIO()
            img.save(buf, format="PNG")
            b64 = base64.b64encode(buf.getvalue()).decode()
        except Exception:
            st.error("No se pudo mostrar esta página. El documento original no se modificó.")
            return
        st.html(f"""
        <div style="height:{altura};min-height:50vh;overflow:auto;border:1px solid #d0d0d0;background:#f5f5f5;text-align:center">
          <img src="data:image/png;base64,{b64}" style="max-width:none" />
        </div>
        <p>Página {pagina} de {n_paginas} · {escape(archivo.name)}</p>
        """)

    elif suffix in ('.xlsx', '.xls'):
        archivo.seek(0)
        try:
            df_raw = pd.read_excel(archivo, header=None, dtype=str).fillna('')
            html_tabla = df_raw.to_html(index=False, header=False, border=0, classes='excel-visor')
            html_visor = f"""
            <style>
                .excel-visor td {{ padding: 3px 8px; border-bottom: 1px solid #eee; font-size: 12px; white-space: nowrap; font-family: monospace; }}
                .excel-visor tr:nth-child(even) {{ background: #f9f9f9; }}
            </style>
            <div style="height: {altura}; min-height: 50vh; overflow-y: auto; overflow-x: auto; border: 1px solid #d0d0d0; border-radius: 8px; background: white; padding: 8px;">{html_tabla}</div>
            <div style="font-size:12px; color:#888; text-align:center; margin-top:4px;">{archivo.name}</div>
            """
            st.html(html_visor)
        except Exception as e:
            st.error(f"No se pudo mostrar el Excel: {e}")
        finally:
            archivo.seek(0)


def _extraer_lineas_encabezado(archivo) -> list[str]:
    import tempfile
    suffix = Path(archivo.name).suffix.lower()
    archivo.seek(0)
    if suffix == '.pdf':
        try:
            import pdfplumber
            with pdfplumber.open(BytesIO(_contenido_para_extraer(archivo))) as pdf:
                texto = pdf.pages[0].extract_text() or ""
                return texto.split('\n')[:40]
        except Exception:
            return []
        finally:
            archivo.seek(0)
    else:
        archivo.seek(0)
        df = pd.read_excel(archivo, header=None, nrows=15).fillna('')
        archivo.seek(0)
        lineas = []
        for _, row in df.iterrows():
            vals = [str(v) for v in row if str(v) not in ('nan', 'None', '')]
            if vals: lineas.append(' '.join(vals))
        return lineas


def _documento_no_es_balance(signature) -> bool:
    """Gate conservador para anexos/otros con señal documental suficiente."""
    document_type = getattr(getattr(signature, 'document_type', None), 'value', '')
    return bool(
        document_type == 'OTRO'
        and not bool(getattr(signature, 'has_headers', False))
        and float(getattr(signature, 'confidence', 0.0) or 0.0) >= 0.30
    )


def _aplicar_correcciones_extraccion(
    cuentas: list[CuentaRaw], edited: pd.DataFrame,
) -> tuple[list[CuentaRaw], object]:
    """Aplica importes editados y vuelve a certificar las ocho columnas."""
    rows = {int(row["linea"]): row for _, row in edited.iterrows()}
    corrected: list[CuentaRaw] = []
    origin_by_column = {
        "activo": OrigenColumna.ACTIVO,
        "pasivo": OrigenColumna.PASIVO,
        "perdida": OrigenColumna.PERDIDA,
        "ganancia": OrigenColumna.GANANCIA,
    }

    def numeric(value) -> float:
        parsed = pd.to_numeric(value, errors="coerce")
        return 0.0 if pd.isna(parsed) else float(parsed)

    for cuenta in cuentas:
        row = rows.get(int(cuenta.linea))
        if row is None or not cuenta.montos_columnas:
            corrected.append(cuenta)
            continue
        if bool(row.get("excluir", False)) and not bool(row.get("total", cuenta.es_total)):
            continue
        amounts = {
            column: numeric(row.get(column, cuenta.montos_columnas.get(column, 0)))
            for column in RAW_MONETARY_COLUMNS
        }
        amount = None
        origin = OrigenColumna.DESCONOCIDO
        for column in ("activo", "pasivo", "perdida", "ganancia"):
            if amounts[column] != 0:
                amount = amounts[column]
                origin = origin_by_column[column]
                break
        derived = list(cuenta.columnas_derivadas)
        if amounts != cuenta.montos_columnas and "correccion_humana" not in derived:
            derived.append("correccion_humana")
        corrected.append(replace(
            cuenta,
            montos_columnas=amounts,
            monto=amount,
            origen_columna=origin,
            es_total=bool(row.get("total", cuenta.es_total)),
            columnas_derivadas=derived,
        ))
    certification = certificar_extraccion_columnas(
        corrected, metodo="revision_humana_8_columnas",
    )
    return corrected, certification


def _crear_cuenta_manual_extraccion(
    cuentas: list[CuentaRaw], *, codigo: str, nombre: str,
    montos: dict[str, float], es_total: bool = False,
) -> CuentaRaw:
    """Construye una fila auditable ingresada por el analista."""
    if not str(nombre or '').strip():
        raise ValueError("El nombre de la cuenta es obligatorio")
    amounts = {
        column: float(montos.get(column, 0.0) or 0.0)
        for column in RAW_MONETARY_COLUMNS
    }
    origin_by_column = {
        "activo": OrigenColumna.ACTIVO,
        "pasivo": OrigenColumna.PASIVO,
        "perdida": OrigenColumna.PERDIDA,
        "ganancia": OrigenColumna.GANANCIA,
    }
    amount = None
    origin = OrigenColumna.DESCONOCIDO
    for column in ("activo", "pasivo", "perdida", "ganancia"):
        if amounts[column] != 0:
            amount = amounts[column]
            origin = origin_by_column[column]
            break
    return CuentaRaw(
        linea=max((int(c.linea) for c in cuentas), default=-1) + 1,
        codigo=str(codigo or '').strip() or None,
        nombre=str(nombre).strip(),
        monto=amount,
        origen_columna=origin,
        es_total=bool(es_total),
        confianza_extraccion=1.0,
        montos_columnas=amounts,
        columnas_derivadas=["ingreso_manual_analista"],
    )


def _diagnosticar_filas_extraccion(
        cuentas: list[CuentaRaw], certification=None,
        tolerancia: float = 10.0) -> dict[int, dict]:
    """Explica por fila por qué falló la certificación y qué revisar.

    La función es deliberadamente informativa: no modifica importes, no excluye
    filas y no marca subtotales automáticamente. La decisión permanece en manos
    del analista.
    """
    inconsistent = set(
        getattr(certification, "filas_inconsistentes", []) or []
    )
    diagnostics: dict[int, dict] = {}
    footer_terms = (
        "firma", "representante legal", "contador", "auditor", "rut",
        "dirección", "direccion", "página", "pagina",
    )
    total_terms = (
        "subtotal", "total", "totales", "sumas", "utilidad del ejercicio",
        "pérdida del ejercicio", "perdida del ejercicio",
    )

    for cuenta in cuentas:
        if not cuenta.montos_columnas:
            continue
        if cuenta.es_total:
            diagnostics[int(cuenta.linea)] = {
                "prioridad": 3,
                "diagnostico": "Control reconocido, no se suma como cuenta",
                "accion_sugerida": (
                    "No necesita excluirlo. Se conserva para comprobar la cuadratura; "
                    "corrija sus importes sólo si difieren del documento."
                ),
                "valores_sugeridos": "",
            }
            continue
        values = {
            column: float(cuenta.montos_columnas.get(column, 0.0) or 0.0)
            for column in RAW_MONETARY_COLUMNS
        }
        movement_error = (
            values["debitos"] - values["creditos"]
            - values["saldo_deudor"] + values["saldo_acreedor"]
        )
        classification_error = (
            values["saldo_deudor"] + values["saldo_acreedor"]
            - values["activo"] - values["pasivo"]
            - values["perdida"] - values["ganancia"]
        )
        name = normalizar_nombre(cuenta.nombre)
        is_inconsistent = int(cuenta.linea) in inconsistent
        suggested: list[str] = []
        movement_candidates = {
            "Debe": values["creditos"] + values["saldo_deudor"] - values["saldo_acreedor"],
            "Haber": values["debitos"] - values["saldo_deudor"] + values["saldo_acreedor"],
            "Saldo deudor": values["debitos"] - values["creditos"] + values["saldo_acreedor"],
            "Saldo acreedor": values["creditos"] - values["debitos"] + values["saldo_deudor"],
        }
        current_movement = {
            "Debe": values["debitos"], "Haber": values["creditos"],
            "Saldo deudor": values["saldo_deudor"],
            "Saldo acreedor": values["saldo_acreedor"],
        }
        if abs(movement_error) > tolerancia:
            zero_first = [
                label for label, current in current_movement.items()
                if current == 0 and movement_candidates[label] > 0
            ]
            labels = zero_first or list(movement_candidates)
            suggested.extend(
                f"{label}: {current_movement[label]:,.0f} → "
                f"{movement_candidates[label]:,.0f}"
                for label in labels
                if movement_candidates[label] >= 0
                and abs(current_movement[label] - movement_candidates[label]) > tolerancia
            )
        classification_values = {
            "Activo": values["activo"], "Pasivo": values["pasivo"],
            "Pérdidas": values["perdida"], "Ganancias": values["ganancia"],
        }
        populated = [label for label, value in classification_values.items() if value != 0]
        saldo_total = values["saldo_deudor"] + values["saldo_acreedor"]
        neto_final = values["activo"] + values["perdida"] - values["pasivo"] - values["ganancia"]
        movimiento_respalda_final = (
            len(populated) == 1
            and abs(values["debitos"] - values["creditos"] - neto_final) <= tolerancia
        )
        if abs(classification_error) > tolerancia and len(populated) == 1 and not movimiento_respalda_final:
            label = populated[0]
            suggested.append(
                f"{label}: {classification_values[label]:,.0f} → {saldo_total:,.0f}"
            )

        if is_inconsistent and movimiento_respalda_final and abs(movement_error) > tolerancia:
            reason = "El movimiento respalda el importe final, pero el saldo intermedio difiere"
            action = "Revise Saldo deudor y Saldo acreedor; no cambie el importe final respaldado por Debe menos Haber."
            priority = 1
        elif is_inconsistent and abs(movement_error) > tolerancia:
            if abs(classification_error) > tolerancia:
                reason = "No coinciden movimiento, saldo y columna de clasificación"
                action = (
                    "Compare Debe/Haber y ambos saldos con el documento; después "
                    "deje el saldo en una sola columna contable."
                )
            else:
                reason = "Debe menos Haber no coincide con el saldo"
                action = (
                    "Revise Débitos, Créditos, Saldo deudor y Saldo acreedor."
                )
            priority = 1
        elif is_inconsistent and abs(classification_error) > tolerancia:
            reason = "El saldo no coincide con Activo/Pasivo/Pérdidas/Ganancias"
            action = (
                "Compare las cuatro columnas finales y deje el importe en la "
                "columna impresa correspondiente."
            )
            priority = 1
        elif any(term in name for term in footer_terms):
            reason = "Posible firma, identificación o pie de página"
            action = "Si no es una cuenta contable, marque Excluir."
            priority = 2
        elif any(name.startswith(term) for term in total_terms):
            reason = "Posible fila de control o resultado del ejercicio"
            action = (
                "Si corresponde a un subtotal, utilidad/pérdida o total impreso, "
                "marque Subtotal/total."
            )
            priority = 2
        else:
            reason = "Sin inconsistencia individual detectada"
            action = (
                "No cambie esta fila salvo que el importe difiera del documento."
            )
            priority = 3

        diagnostics[int(cuenta.linea)] = {
            "prioridad": priority,
            "diagnostico": reason,
            "accion_sugerida": action,
            "valores_sugeridos": "; ".join(suggested),
            "error_movimiento": round(movement_error, 2),
            "error_clasificacion": round(classification_error, 2),
        }
    return diagnostics


def _permite_clasificar_extraccion(certification) -> bool:
    return bool(certification) and (
        certification.estado in {"certificada", "parcial"}
        or bool(getattr(certification, "columnas_finales_validadas", False))
    )


def _tabla_control_auxiliar(certification) -> pd.DataFrame:
    labels = ("Debe", "Haber", "Saldo deudor", "Saldo acreedor")
    return pd.DataFrame([
        {
            "Columna": label,
            "Suma de cuentas": certification.totales_calculados.get(col, 0),
            "Subtotal impreso": certification.totales_impresos.get(col, 0),
            "Diferencia": certification.diferencias.get(col, 0),
        }
        for col, label in zip(RAW_MONETARY_COLUMNS[:4], labels)
        if abs(certification.diferencias.get(col, 0)) > 10
    ])


def _mostrar_advertencias_auxiliares(certification) -> None:
    if not (
        getattr(certification, "columnas_finales_validadas", False)
        and certification.estado == "fallida"
    ):
        return
    st.warning(
        "Clasificación habilitada: Activo, Pasivo, Pérdidas y Ganancias coinciden "
        "con los controles impresos y cuadran. Quedan diferencias en movimientos "
        "o saldos intermedios; no se certifican las ocho columnas. No cambie los "
        "importes finales ni excluya los totales para resolver esas diferencias."
    )
    with st.expander("Detalle de advertencias de movimientos y saldos"):
        tabla = _tabla_control_auxiliar(certification)
        if not tabla.empty:
            st.dataframe(tabla, hide_index=True, use_container_width=True)
        if certification.observaciones_auxiliares:
            st.dataframe(pd.DataFrame(certification.observaciones_auxiliares), hide_index=True, use_container_width=True)
        st.caption("Las cifras originales se conservan. Estas observaciones también se incluyen en el Excel exportado.")


def _exportar_advertencias_auxiliares(writer, certification) -> None:
    if not (
        getattr(certification, "columnas_finales_validadas", False)
        and certification.estado == "fallida"
    ):
        return
    sheet = "Control de extracción"
    resumen = pd.DataFrame([
        {"Control": "Importes para homologación", "Estado": "Cuatro columnas finales validadas contra controles impresos"},
        {"Control": "Certificación de ocho columnas", "Estado": "No certificada: diferencias en movimientos o saldos"},
        {"Control": "Tratamiento", "Estado": "Cifras originales conservadas; no se aplicaron ajustes automáticos"},
    ])
    resumen.to_excel(writer, sheet_name=sheet, index=False)
    tabla = _tabla_control_auxiliar(certification)
    startrow = len(resumen) + 3
    if not tabla.empty:
        tabla.to_excel(writer, sheet_name=sheet, index=False, startrow=startrow)
        startrow += len(tabla) + 3
    if certification.observaciones_auxiliares:
        pd.DataFrame(certification.observaciones_auxiliares).to_excel(
            writer, sheet_name=sheet, index=False, startrow=startrow,
        )


def _mostrar_etapa_correccion_extraccion(archivo, filename: str) -> None:
    """Corrección a ancho completo y documento en la parte inferior."""
    _mostrar_correccion_extraccion(filename)
    st.divider()
    st.subheader("Documento original")
    _visor_documento(archivo, altura="58vh", mostrar_titulo=False)


def _explicar_diferencia_controles(certification) -> str:
    """Distingue diferencias de movimientos de errores en saldos/clasificación."""
    diferencias = getattr(certification, "diferencias", {}) or {}
    if not set(RAW_MONETARY_COLUMNS).issubset(diferencias):
        return ""
    debe = diferencias["debitos"]
    haber = diferencias["creditos"]
    if (
        abs(debe) > 10 and abs(debe - haber) <= 10
        and all(abs(diferencias[column]) <= 10 for column in RAW_MONETARY_COLUMNS[2:])
        and not getattr(certification, "filas_inconsistentes", [])
    ):
        return (
            f"La diferencia de {abs(debe):,.0f} se repite en Debe y Haber; "
            "los saldos y las cuatro columnas de clasificación coinciden con "
            "el subtotal, y las cuentas cumplen sus identidades individuales. "
            "Revise los movimientos y el subtotal impreso con el emisor: esto "
            "no se corrige reclasificando cuentas ni excluyendo los controles. "
            "Si el subtotal del original es incorrecto, solicite un balance corregido."
        )
    return ""


def _config_importes_extraccion():
    labels = ["Debe", "Haber", "Saldo deudor", "Saldo acreedor",
              "Activo", "Pasivo", "Pérdidas", "Ganancias"]
    return {key: st.column_config.NumberColumn(label, format="localized")
            for key, label in zip(RAW_MONETARY_COLUMNS, labels)}


def _mostrar_correccion_extraccion(filename: str) -> None:
    """Editor seguro previo a homologación para extracciones OCR fallidas."""
    resultado = st.session_state.extraction_pending.get(filename)
    if resultado is None:
        return
    revisions = st.session_state.setdefault("extraction_revisions", {})
    revision = revisions.get(filename, 0)
    if revision:
        st.caption("Sus correcciones están guardadas; el diagnóstico muestra la validación actualizada.")
    st.error(
        "La extracción aún no puede certificarse. La clasificación está pausada "
        "para evitar que una lectura incorrecta llegue al balance homologado."
    )
    rows = []
    certification = getattr(resultado, "certificacion_extraccion", None)
    if getattr(certification, "estado", "") == "no_evaluable":
        st.warning(
            "Falta un subtotal impreso utilizable para validar. Eliminar los controles "
            "no resuelve el descuadre. Conserve o ingrese el subtotal del documento."
        )
    explicacion_control = _explicar_diferencia_controles(certification)
    if explicacion_control:
        st.warning(explicacion_control)
    inconsistent = set(getattr(certification, "filas_inconsistentes", []) or [])
    diagnostics = _diagnosticar_filas_extraccion(
        resultado.cuentas, certification,
    )
    for cuenta in resultado.cuentas:
        if not cuenta.montos_columnas:
            continue
        diagnosis = diagnostics.get(int(cuenta.linea), {})
        rows.append({
            "linea": cuenta.linea,
            "codigo": cuenta.codigo or "",
            "cuenta": cuenta.nombre,
            "inconsistente": int(cuenta.linea) in inconsistent,
            "prioridad": int(diagnosis.get("prioridad", 3)),
            "diagnostico": diagnosis.get("diagnostico", ""),
            "accion_sugerida": diagnosis.get("accion_sugerida", ""),
            "valores_sugeridos": diagnosis.get("valores_sugeridos", ""),
            "excluir": False,
            **{
                column: float(cuenta.montos_columnas.get(column, 0.0) or 0.0)
                for column in RAW_MONETARY_COLUMNS
            },
            "total": bool(cuenta.es_total),
        })
    if not rows:
        st.warning("OCR no produjo filas tabulares que puedan corregirse.")
        return
    source = pd.DataFrame(rows).sort_values(
        ["prioridad", "linea"], kind="stable",
    ).reset_index(drop=True)
    c1, c2, c3 = st.columns(3)
    c1.metric("Filas extraídas", len(source))
    c2.metric("Filas inconsistentes", len(inconsistent))
    c3.metric("Estado", "Bloqueada")
    if certification is not None and certification.razones:
        st.caption(" ".join(certification.razones))
        diferencias = getattr(certification, "diferencias", {}) or {}
        impresos = getattr(certification, "totales_impresos", {}) or {}
        calculados = getattr(certification, "totales_calculados", {}) or {}
        columnas_con_diferencia = [
            column for column, value in diferencias.items() if abs(value) > 10
        ]
        if columnas_con_diferencia:
            st.markdown("#### Qué no coincide")
            labels = {
                "debitos": "Débitos", "creditos": "Créditos",
                "saldo_deudor": "Saldo deudor", "saldo_acreedor": "Saldo acreedor",
                "activo": "Activo", "pasivo": "Pasivo",
                "perdida": "Pérdidas", "ganancia": "Ganancias",
            }
            st.dataframe(pd.DataFrame([
                {
                    "Columna": labels.get(column, column),
                    "Suma de cuentas": float(calculados.get(column, 0) or 0),
                    "Control extraído": float(impresos.get(column, 0) or 0),
                    "Diferencia": float(diferencias[column]),
                }
                for column in columnas_con_diferencia
            ]), use_container_width=True, hide_index=True, column_config={
                key: st.column_config.NumberColumn(key, format="localized")
                for key in ("Suma de cuentas", "Control extraído", "Diferencia")
            })
            final_errors = [c for c in columnas_con_diferencia if c in RAW_MONETARY_COLUMNS[4:]]
            if not final_errors:
                st.info(
                    "Las sumas de Activo, Pasivo, Pérdidas y Ganancias coinciden con el "
                    "control leído. Las diferencias mostradas están en movimientos o saldos. "
                    "No cambie importes finales que coincidan con el PDF: revise los controles "
                    "y las filas señaladas. La validación individual sigue siendo necesaria."
                )

    st.markdown("#### Qué debe hacer el analista")
    st.info(
        "Los subtotales, resultados de cierre y totales reconocidos ya están "
        "fuera de la suma de cuentas. No debe excluirlos: se conservan como "
        "controles para contrastar las cifras extraídas con el documento."
    )
    st.markdown(
        "1. Revise primero las filas señaladas en **Revisar** contra el PDF.  \n"
        "2. Si una fila es un pie de página, firma o texto legal, marque **Excluir**.  \n"
        "3. Sólo si un control aún no está reconocido, active **Subtotal/total**. "
        "Los controles ya reconocidos aparecen separados debajo de las cuentas.  \n"
        "4. Corrija una cifra únicamente cuando sea distinta de la impresa. "
        "Después pulse **Verificar y continuar**."
    )
    st.caption(
        "Los valores sugeridos se calculan desde las identidades Debe/Haber y "
        "saldo/clasificación. No se aplican automáticamente: compárelos con el PDF."
    )
    if inconsistent:
        nombres_inconsistentes = source[source["inconsistente"]][
            ["linea", "cuenta"]
        ].copy()
        st.warning(
            "Estas son las filas concretas que rompen una identidad contable:"
        )
        st.dataframe(
            nombres_inconsistentes.rename(columns={"linea": "Fila", "cuenta": "Cuenta"}),
            use_container_width=True, hide_index=True,
        )
        with st.expander("Detalle de las diferencias y posibles valores", expanded=True):
            st.dataframe(source[source["inconsistente"]][
                ["linea", "cuenta", "diagnostico", "accion_sugerida", "valores_sugeridos"]
            ].rename(columns={"linea": "Fila", "cuenta": "Cuenta",
                              "diagnostico": "Motivo", "accion_sugerida": "Qué revisar",
                              "valores_sugeridos": "Posible corrección, verificar contra PDF"}),
                         hide_index=True, use_container_width=True)
    else:
        st.info(
            "No hay filas con errores individuales detectados. El problema está en la "
            "comparación de las sumas con los controles. Revise primero el subtotal leído "
            "y si falta alguna cuenta o se repite un grupo; no hay una cifra individual "
            "identificada para corregir automáticamente."
        )
    with st.expander("Ingresar una cuenta omitida por la extracción"):
        st.caption(
            "Use esta opción sólo cuando la cuenta exista en el documento y no "
            "aparezca en la tabla. Copie sus ocho columnas tal como están impresas."
        )
        with st.form(f"manual_extraction_row_{filename}", clear_on_submit=True):
            codigo_manual = st.text_input("Código original, si existe")
            nombre_manual = st.text_input("Nombre de la cuenta")
            m1, m2, m3, m4 = st.columns(4)
            debitos_manual = m1.number_input("Debe", value=0.0, format="%.0f")
            creditos_manual = m2.number_input("Haber", value=0.0, format="%.0f")
            saldo_deudor_manual = m3.number_input("Saldo deudor", value=0.0, format="%.0f")
            saldo_acreedor_manual = m4.number_input("Saldo acreedor", value=0.0, format="%.0f")
            c1m, c2m, c3m, c4m = st.columns(4)
            activo_manual = c1m.number_input("Activo", value=0.0, format="%.0f")
            pasivo_manual = c2m.number_input("Pasivo", value=0.0, format="%.0f")
            perdida_manual = c3m.number_input("Pérdidas", value=0.0, format="%.0f")
            ganancia_manual = c4m.number_input("Ganancias", value=0.0, format="%.0f")
            total_manual = st.checkbox("Es subtotal, resultado o total impreso")
            agregar_manual = st.form_submit_button("Agregar cuenta a la extracción")
        if agregar_manual:
            try:
                nueva = _crear_cuenta_manual_extraccion(
                    resultado.cuentas,
                    codigo=codigo_manual,
                    nombre=nombre_manual,
                    montos={
                        "debitos": debitos_manual, "creditos": creditos_manual,
                        "saldo_deudor": saldo_deudor_manual,
                        "saldo_acreedor": saldo_acreedor_manual,
                        "activo": activo_manual, "pasivo": pasivo_manual,
                        "perdida": perdida_manual, "ganancia": ganancia_manual,
                    },
                    es_total=total_manual,
                )
            except ValueError as exc:
                st.error(str(exc))
            else:
                resultado.cuentas.append(nueva)
                resultado.certificacion_extraccion = certificar_extraccion_columnas(
                    resultado.cuentas, metodo="revision_humana_8_columnas",
                )
                revisions[filename] = revision + 1
                st.success("Cuenta incorporada. Revise la certificación actualizada.")
                st.rerun()
    st.caption("Puede corregir varias celdas antes de guardar. Los números conservan su precisión; los separadores se adaptan al idioma del navegador.")
    with st.form(f"extraction_batch_{filename}_{revision}"):
        controles = source[source["total"]].copy()
        detalle = source[~source["total"]].copy()
        st.markdown("#### Cuentas y filas por revisar")
        edited = st.data_editor(
            detalle,
            hide_index=True,
            use_container_width=True,
            column_order=["linea", "cuenta", "inconsistente", "activo", "pasivo", "perdida", "ganancia",
                          "debitos", "creditos", "saldo_deudor", "saldo_acreedor", "total", "excluir"],
            disabled=[
                "linea", "codigo", "cuenta", "inconsistente", "prioridad",
                "diagnostico", "accion_sugerida",
                "valores_sugeridos",
            ],
            key=f"extraction_editor_{filename}_{revision}",
            column_config={
                "linea": st.column_config.NumberColumn("Fila", format="%d"),
                "codigo": "Código",
                "cuenta": "Cuenta",
                "inconsistente": st.column_config.CheckboxColumn("Revisar"),
                "prioridad": st.column_config.NumberColumn(
                    "Prioridad", help="1: revisar primero; 2: validar tipo de fila; 3: sin señal directa.",
                ),
                "diagnostico": st.column_config.TextColumn(
                    "Qué detectó el sistema", width="large",
                ),
                "accion_sugerida": st.column_config.TextColumn(
                    "Qué debe hacer", width="large",
                ),
                "valores_sugeridos": st.column_config.TextColumn(
                    "Leído → valor contablemente posible", width="large",
                ),
                "excluir": st.column_config.CheckboxColumn(
                    "Excluir", help="Úselo sólo para pies, firmas, notas o texto que no sea una cuenta."
                ),
                "debitos": st.column_config.NumberColumn("Debe", format="localized"),
                "creditos": st.column_config.NumberColumn("Haber", format="localized"),
                "saldo_deudor": st.column_config.NumberColumn("Saldo deudor", format="localized"),
                "saldo_acreedor": st.column_config.NumberColumn("Saldo acreedor", format="localized"),
                "activo": st.column_config.NumberColumn("Activo", format="localized"),
                "pasivo": st.column_config.NumberColumn("Pasivo", format="localized"),
                "perdida": st.column_config.NumberColumn("Pérdidas", format="localized"),
                "ganancia": st.column_config.NumberColumn("Ganancias", format="localized"),
                "total": st.column_config.CheckboxColumn(
                    "Subtotal/total",
                    help="Marque filas de control; no se sumarán como cuentas.",
                ),
            },
        )
        if not controles.empty:
            with st.expander(f"Controles reconocidos: {len(controles)} (no se suman como cuentas)"):
                st.caption(
                    "No requieren exclusión. Edite un importe sólo si está mal leído. "
                    "Desmarque Subtotal/total únicamente si la fila es realmente una cuenta."
                )
                controles_editados = st.data_editor(
                    controles.drop(columns=["excluir"]),
                    hide_index=True,
                    use_container_width=True,
                    column_order=["linea", "cuenta", *RAW_MONETARY_COLUMNS, "total"],
                    key=f"extraction_controls_{filename}_{revision}",
                    disabled=[column for column in controles.columns
                              if column not in (*RAW_MONETARY_COLUMNS, "total", "excluir")],
                    column_config={
                        "linea": "Fila", "cuenta": "Control impreso",
                        **_config_importes_extraccion(),
                        "total": st.column_config.CheckboxColumn("Subtotal/total"),
                    },
                )
                controles_editados["excluir"] = False
                edited = pd.concat([edited, controles_editados], ignore_index=True)
        submitted = st.form_submit_button("🔎 Verificar correcciones y continuar", type="primary")
    if submitted:
        corrected, certification = _aplicar_correcciones_extraccion(
            resultado.cuentas, edited,
        )
        resultado.cuentas = corrected
        resultado.certificacion_extraccion = certification
        st.session_state.setdefault("extraction_certifications", {})[filename] = certification
        revisions[filename] = revision + 1
        if not _permite_clasificar_extraccion(certification):
            st.rerun()
        else:
            st.session_state.extraction_resolved[filename] = corrected
            st.session_state.extraction_pending.pop(filename, None)
            st.session_state.resultados.pop(filename, None)
            st.success("Importes para homologación validados. Iniciando clasificación.")
            st.rerun()


def _extraer_cuentas(archivo) -> tuple[list[CuentaRaw], object]:
    """Extrae cuentas y devuelve (cuentas, document_context).

    `document_context` es el DocumentProcessingContext (Sprint 31) cuando el
    archivo es PDF y el análisis documental corrió; None en caso contrario.
    """
    suffix = Path(archivo.name).suffix.lower()
    st.session_state.setdefault("processed_at", {})[archivo.name] = datetime.now().astimezone().isoformat()
    if suffix == '.pdf':
        import tempfile
        with tempfile.TemporaryDirectory(prefix="balance-seleccion-") as directory:
            tmp_path = Path(directory) / "seleccion.pdf"
            tmp_path.write_bytes(_contenido_para_extraer(archivo))
            parser = ParserPDF()
            resultado = parser.parsear(tmp_path)
        for adv in resultado.advertencias: st.warning(adv)
        document_context = getattr(resultado, 'document_context', None)
        signature = getattr(document_context, 'signature', None)
        if _documento_no_es_balance(signature):
            st.error(
                "El archivo fue identificado como anexo u otro documento, no "
                "como balance. Se detuvo la homologación para evitar interpretar "
                "sus filas con una estructura contable incorrecta."
            )
            return [], document_context
        certificacion = getattr(resultado, 'certificacion_extraccion', None)
        st.session_state.setdefault("extraction_certifications", {})[archivo.name] = certificacion
        if certificacion is not None and certificacion.estado == 'fallida' and not _permite_clasificar_extraccion(certificacion):
            st.error(
                "La extracción no reproduce los totales impresos del balance. "
                "El documento no será homologado hasta corregir filas o columnas."
            )
            if certificacion.razones:
                st.caption(" ".join(certificacion.razones))
            if certificacion.filas_inconsistentes:
                muestra = ", ".join(
                    str(linea) for linea in certificacion.filas_inconsistentes[:12]
                )
                st.caption(
                    f"Filas con inconsistencias internas: {muestra}"
                    + ("…" if len(certificacion.filas_inconsistentes) > 12 else "")
                )
            st.session_state.extraction_pending[archivo.name] = resultado
            return [], document_context
        if certificacion is not None and certificacion.estado == 'certificada':
            resultado_ejercicio = getattr(
                certificacion, 'resultado_ejercicio', None,
            )
            tipo_resultado = getattr(certificacion, 'tipo_resultado', None)
            if resultado_ejercicio is not None and tipo_resultado:
                st.success(
                    "Extracción certificada: las cuentas reproducen el subtotal "
                    f"y el cierre impreso. {tipo_resultado.capitalize()} del "
                    f"ejercicio: ${abs(resultado_ejercicio):,.0f}."
                )
            else:
                st.success(
                    "Extracción certificada: las ocho columnas reproducen los "
                    "subtotales impresos."
                )
            reconstruidas = getattr(
                certificacion, 'columnas_total_reconstruidas', [],
            )
            if reconstruidas:
                st.info(
                    "El PDF truncaba el último dígito del control de "
                    + " y ".join(reconstruidas)
                    + ". El sistema lo reconstruyó con la suma exacta de las "
                    "cuentas y verificó las otras seis columnas."
                )
        elif certificacion is not None and certificacion.estado == 'parcial':
            st.warning(
                "La ecuación final del balance está cuadrada, pero la extracción "
                "de las cuentas intermedias no está certificada. Todas las cuentas "
                "deben pasar por revisión humana antes de usar el resultado."
            )
            if certificacion.razones:
                st.caption(" ".join(certificacion.razones))
        elif certificacion is not None and certificacion.estado == 'no_evaluable':
            st.info(
                "Este formato todavía no dispone de totales independientes para "
                "certificar automáticamente la extracción. Revise sus cuentas "
                "antes de confirmar la homologación."
            )
        # GATE 4E: SAFE-R02+R03+R08 (ruido de encabezado/pie, URLs/emails,
        # duplicados) aplicado ANTES de la clasificación, ÚNICAMENTE con
        # activación explícita (env SAFE_MODE ON). Con SAFE OFF el
        # comportamiento es exactamente el previo: se devuelven todas las
        # cuentas extraídas, sin filtrado.
        if _safe_mode_enabled():
            return _safe_qualify_cuentas(resultado.cuentas), \
                document_context
        return resultado.cuentas, document_context
    else:
        cuentas = parsear_excel(archivo)
        certificacion = certificar_extraccion_columnas(
            cuentas, metodo="excel_8_columns",
        )
        st.session_state.setdefault("extraction_certifications", {})[archivo.name] = certificacion
        if certificacion.estado == "fallida" and not _permite_clasificar_extraccion(certificacion):
            st.error(
                "Las ocho columnas del Excel no reproducen sus controles "
                "impresos. La homologación queda pausada hasta revisar las filas."
            )
            if certificacion.razones:
                st.caption(" ".join(certificacion.razones))
            st.session_state.extraction_pending[archivo.name] = ResultadoParseo(
                archivo=archivo.name,
                formato_codigo=FormatoCodigo.SIN_CODIGO,
                separador_miles=".",
                requirio_ocr=False,
                rotacion_aplicada=0,
                cuentas=cuentas,
                certificacion_extraccion=certificacion,
            )
            return [], None
        if certificacion.estado == "certificada":
            st.success(
                "Excel certificado: las cuentas reproducen el subtotal y el "
                "control final de sus ocho columnas."
            )
        elif certificacion.estado == "parcial":
            st.warning(
                "El Excel cuadra en su ecuación final, pero requiere revisión "
                "humana de sus filas antes de exportar."
            )
        elif certificacion.estado == "no_evaluable":
            st.info(
                "El Excel no contiene un subtotal independiente de ocho columnas; "
                "su extracción no puede certificarse automáticamente."
            )
        return cuentas, None


def _mostrar_informacion_documento(ctx) -> None:
    """Sección solo-lectura INFORMACIÓN DEL DOCUMENTO (Sprint 31, FASE 5).

    Muestra la metadata del análisis documental ANTES de los resultados.
    No modifica ningún resultado: si `ctx` no está disponible, no dibuja nada.
    """
    if ctx is None:
        return
    try:
        info = ctx.ui_summary()
    except Exception:
        return

    with st.container(border=True):
        st.markdown("#### 🧾 INFORMACIÓN DEL DOCUMENTO")
        cols = st.columns(4)
        items = [
            ("Documento", info.get("Documento", "—")),
            ("Formato", info.get("Formato", "—")),
            ("Columnas", info.get("Columnas", "—")),
            ("Layout", info.get("Layout", "—")),
            ("OCR", info.get("OCR", "—")),
            ("Extractor", info.get("Extractor", "—")),
            ("Confianza", info.get("Confianza", "—")),
            ("Familia", info.get("Familia", "—")),
        ]
        for i, (label, value) in enumerate(items):
            with cols[i % 4]:
                st.markdown(f"**{label}**  \n`{value}`")


# ─────────────────────────────────────────────────────────────────────────────
# CONOCIMIENTO DOCUMENTAL (Sprint 32 — Document Knowledge Base)
# ─────────────────────────────────────────────────────────────────────────────

@st.cache_resource
def _cargar_document_kb():
    """Carga la DKB desde knowledge_base/document_kb.json (None si no existe)."""
    try:
        from document_intelligence.knowledge import DocumentKnowledgeBase
        path = BASE_DIR / "knowledge_base" / "document_kb.json"
        if not path.exists():
            return None
        kb = DocumentKnowledgeBase()
        kb.load(path)
        return kb
    except Exception:
        return None


def _build_fingerprint_archivo(archivo, doc_ctx):
    """Construye el fingerprint del archivo activo (con caché por nombre)."""
    if "document_fingerprints" not in st.session_state:
        st.session_state.document_fingerprints = {}
    name = archivo.name
    if name in st.session_state.document_fingerprints:
        return st.session_state.document_fingerprints[name]

    import tempfile
    from document_intelligence.knowledge.fingerprint import fingerprint_from_file
    suffix = Path(archivo.name).suffix.lower()
    archivo.seek(0)
    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
        tmp.write(archivo.read())
        tmp_path = Path(tmp.name)
    try:
        fp = fingerprint_from_file(tmp_path, signature=doc_ctx.signature)
    finally:
        archivo.seek(0)
        tmp_path.unlink(missing_ok=True)
    st.session_state.document_fingerprints[name] = fp
    return fp


def _tab_conocimiento(archivo, doc_ctx, meta_activo) -> None:
    """Sección CONOCIMIENTO DOCUMENTAL: DKB + matching del documento activo.

    Solo lectura. Si la DKB no está disponible o el matcher falla, muestra
    un aviso y NO afecta ningún resultado.
    """
    if archivo is None or doc_ctx is None:
        st.info("El análisis documental no está disponible para este archivo.")
        return

    kb = _cargar_document_kb()
    if kb is None or not getattr(kb, "profiles", None):
        st.info(
            "📖 La Document Knowledge Base aún no existe. "
            "Ejecuta `python tools/build_document_kb.py` para construirla."
        )
        return

    try:
        fp = _build_fingerprint_archivo(archivo, doc_ctx)
        company = ""
        if meta_activo is not None and getattr(meta_activo, "razon_social", ""):
            company = meta_activo.razon_social
        from document_intelligence.knowledge import Matcher
        result = Matcher().match(fp, kb.profiles, company=company)
    except Exception as exc:  # noqa: BLE001 — el matcher nunca rompe el pipeline
        st.warning(f"El matcher de la DKB no pudo ejecutarse ({exc}).")
        return

    profile = result.matched_profile
    if profile is None:
        st.info("No se encontraron perfiles similares en la DKB.")
        return

    st.markdown("#### 📖 Conocimiento Documental")
    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Perfil detectado", profile.name, help=profile.description)
    c2.metric("Empresa", profile.company)
    c3.metric("Familia", profile.family)
    c4.metric("Extractor recomendado", profile.recommended_extractor)
    c5.metric("Similitud", f"{result.similarity:.0f}%")

    st.markdown("**Top 5 perfiles similares**")
    ranking_rows = []
    for p, sim in result.ranking[:5]:
        ranking_rows.append({
            "Perfil": p.name,
            "Empresa": p.company,
            "Familia": p.family,
            "Similitud": f"{sim:.0f}%",
            "Frecuencia": p.times_seen,
        })
    st.dataframe(ranking_rows, use_container_width=True, hide_index=True)

    st.markdown("**Variantes conocidas**")
    if profile.known_variants:
        st.write(", ".join(profile.known_variants))
    else:
        st.caption("Sin variantes registradas.")

    st.markdown("**Historial**")
    h1, h2, h3 = st.columns(3)
    h1.metric("Primera aparición", profile.first_seen or "—")
    h2.metric("Última aparición", profile.last_seen or "—")
    h3.metric("Frecuencia (documentos)", profile.times_seen)


@st.cache_resource
def _cargar_mining_result():
    """Carga el resultado de minería desde document_mining.json (None si no existe)."""
    try:
        path = BASE_DIR / "knowledge_base" / "document_mining.json"
        if not path.exists():
            return None
        from document_intelligence.mining import load_analysis_result
        return load_analysis_result(path)
    except Exception:
        return None


def _tab_inteligencia() -> None:
    """Inteligencia del Dataset: minería del DKB (solo lectura).

    Muestra familias descubiertas, cobertura esperada, representantes,
    variantes y confianza. NO edita datos.
    """
    result = _cargar_mining_result()
    if result is None:
        st.info(
            "📈 La minería del dataset aún no existe. "
            "Ejecuta `python tools/run_document_mining.py` para generarla."
        )
        return

    st.markdown("#### 📈 Inteligencia del Dataset")
    st.caption("Familias descubiertas por fingerprint (sin empresa ni nombre de archivo).")

    matrix = result.get("matrix", {})
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Familias detectadas", result.get("n_families", 0))
    c2.metric("Documentos analizados", result.get("n_documents", 0))
    c3.metric("Similitud media global", f"{matrix.get('mean_similarity', 0):.0f}%")
    c4.metric("Pares comparados", f"{matrix.get('pairs_computed', 0):,}")

    coverage = result.get("coverage", {}).get("tiers", [])
    if coverage:
        st.markdown("**Cobertura esperada**")
        st.dataframe([{
            "Top N familias": t["top_n"],
            "Familias": t["families"],
            "Documentos": t["documents"],
            "% acumulado": f"{t['cumulative_pct']}%",
            "% restante": f"{t['remaining_pct']}%",
        } for t in coverage], use_container_width=True, hide_index=True)

    familias = result.get("families", [])
    st.markdown("**Top familias**")
    if familias:
        st.dataframe([{
            "Familia": f["id"],
            "Empresa principal": f.get("top_company", "") or "—",
            "Documentos": f["count"],
            "Similitud interna": f"{f['avg_similarity']:.0f}%",
            "Layout": f["dominant_layout"],
            "Código": f["dominant_code_pattern"],
            "Tipo doc": f["dominant_document_type"],
        } for f in familias[:10]], use_container_width=True, hide_index=True)

    representantes = result.get("representatives", [])
    st.markdown("**Representantes**")
    if representantes:
        st.dataframe([{
            "Familia": r["family_id"],
            "Documento representante": r["file"],
            "Similitud promedio": f"{r['avg_similarity']:.0f}%",
            "Documentos": r["n_documents"],
            "Empresa": r["company"],
        } for r in representantes[:10]], use_container_width=True, hide_index=True)

    recomendaciones = result.get("recommendations", [])
    st.markdown("**Recomendación de extractores**")
    if recomendaciones:
        top = recomendaciones[0]
        st.write(
            f"Desarrollar primero un **`{top['extractor_type']}`** para la familia "
            f"`{top['family_name']}` ({top['count']} documentos, "
            f"{top['pct_dataset']}% del dataset)."
        )
        if len(recomendaciones) > 1:
            st.caption("Siguientes candidatas: " + ", ".join(
                r["family_name"] for r in recomendaciones[1:5]
            ))
    else:
        st.caption("Aún no hay familias con volumen suficiente.")

    variantes = result.get("statistics", {}).get("top_variants", [])
    if variantes:
        st.markdown("**Top variantes (empresa · layout)**")
        st.dataframe([{
            "Empresa": v["company"],
            "Layout": v["layout"],
            "Familias": v["count"],
        } for v in variantes[:10]], use_container_width=True, hide_index=True)

    problemas = result.get("quality_issues", [])
    if problemas:
        st.markdown(f"**Problemas detectados ({len(problemas)})**")
        st.dataframe([{
            "Severidad": p["severity"],
            "Tipo": p["kind"],
            "Detalle": p["message"],
        } for p in problemas[:10]], use_container_width=True, hide_index=True)


def _tab_resumen(df: pd.DataFrame):
    col1, col2, col3, col4 = st.columns(4)
    total = len(df)
    pendientes = _pendientes_revision(df)
    sin_clasificar = (pendientes['codigo_clasificado'] == '').sum()
    requiere_rev = len(pendientes)
    confianza_prom = df.loc[df['confianza'] > 0, 'confianza'].mean()

    col1.metric("Cuentas extraídas", total)
    col2.metric("Sin clasificar", sin_clasificar)
    col3.metric("En cola de revisión", int(requiere_rev), delta=f"{100*requiere_rev/total:.0f}%" if total else None, delta_color="inverse")
    col4.metric("Confianza promedio", f"{confianza_prom:.0%}" if pd.notna(confianza_prom) else "—")

    st.subheader("Cobertura por método de clasificación")
    dist = df['metodo'].apply(lambda m: m.split('+')[0]).value_counts()
    dist_df = dist.reset_index()
    dist_df.columns = ['Método', 'Cuentas']
    
    METODO_LABELS = {
        'codigo': '0 · Código de cuenta', 'diccionario_exacto': '1 · Diccionario (exacto)',
        'diccionario_fuzzy': '1b · Diccionario (fuzzy)', 'regla_regex': '2 · Reglas regex',
        'sin_clasificar': '3-4 · Pendiente (embeddings/LLM)',
    }
    dist_df['Método'] = dist_df['Método'].map(lambda m: METODO_LABELS.get(m, m))
    st.dataframe(dist_df, use_container_width=True, hide_index=True)


def _registrar_decision(archivo_nombre, idx, row, codigo, motivo):
    """Historial de esta sesión; Neon conserva además la validación y sugerencia previa."""
    st.session_state.setdefault('historial_decisiones', []).append({
        'Archivo': archivo_nombre, 'Fila': str(idx),
        'Cuenta': row.get('nombre_original', ''),
        'Clasificación anterior': row.get('codigo_clasificado', ''),
        'Clasificación nueva': codigo, 'Método anterior': row.get('metodo', ''),
        'Motivo': motivo, 'Fecha': datetime.now().isoformat(timespec='seconds'),
    })


def _clave_revision_documento(archivo_nombre):
    doc_key = hashlib.sha1(archivo_nombre.encode('utf-8')).hexdigest()[:10]
    revision = st.session_state.get("extraction_revisions", {}).get(archivo_nombre, 0)
    return f"{doc_key}_{revision}" if revision else doc_key


def _solicitar_revision_completa(archivo_nombre):
    doc_key = _clave_revision_documento(archivo_nombre)
    st.session_state[f'revision_vista_{doc_key}'] = 'Todas (incluye confirmadas y excluidas)'
    st.session_state[f'revision_busqueda_{doc_key}'] = ''
    st.session_state['vista_trabajo_solicitada'] = '🔍 Cola de Revisión'


@st.fragment
def _tab_revision(df: pd.DataFrame, catalogo: dict, motor: MotorHibridoLocal, archivo_nombre: str):
    doc_key = _clave_revision_documento(archivo_nombre)
    vista = st.radio(
        'Cuentas a revisar', ['Pendientes', 'Todas (incluye confirmadas y excluidas)'],
        horizontal=True, key=f'revision_vista_{doc_key}',
    )
    busqueda = st.text_input('Buscar cuenta o código', key=f'revision_busqueda_{doc_key}')
    pendientes = (_pendientes_revision(df) if vista == 'Pendientes'
                  else _con_saldo_relevante(df[~df['es_total']]))
    if busqueda.strip() and not pendientes.empty:
        texto = pendientes[['nombre_original', 'codigo_original', 'codigo_clasificado']].fillna('').astype(str).agg(' '.join, axis=1)
        pendientes = pendientes[texto.str.contains(busqueda.strip(), case=False, regex=False)]
    st.caption('Puede cambiar una decisión confirmada o recuperar una cuenta excluida. Elija la nueva clasificación y pulse Confirmar; no necesita volver a cargar el documento.')

    if pendientes.empty:
        st.info("No hay cuentas en esta vista. Seleccione Todas para revisar decisiones anteriores o cambie la búsqueda.")
        return

    # Catálogo ordenado por grupos de presentación y sin cuentas no
    # seleccionables (cálculo / TOTAL). Ver catalog_selection.py.
    opciones_codigo = [''] + opciones_clasificacion(catalogo) + ['➕ NUEVA CATEGORÍA', '🚫 NO INCLUIR']

    checkbox_prefix = f"chk_{doc_key}"
    if st.session_state.get('lote_archivo') != archivo_nombre:
        st.session_state.lote_archivo = archivo_nombre
        st.session_state.lote_seleccion = set()
    elif 'lote_seleccion' not in st.session_state:
        st.session_state.lote_seleccion = set()
    st.session_state.lote_seleccion.intersection_update(pendientes.index)

    n_sel = len(st.session_state.lote_seleccion)
    with st.container(border=True):
        st.markdown(f"#### 📦 Asignación en lote — {n_sel} cuenta(s) seleccionada(s)")
        bc1, bc2, bc3 = st.columns([3, 2, 1])
        with bc1:
            cat_lote = st.selectbox(
                "Clasificar todas las seleccionadas como:", opciones_codigo,
                format_func=lambda c: f"{c} — {catalogo[c]['nombre_estandar']}" if c in catalogo else c if c else "(elegir categoría)",
                key=f"lote_categoria_{doc_key}"
            )
        with bc2:
            alcance_lote = st.radio("Alcance", ["Solo este caso", "Agregar al diccionario"], index=1, horizontal=True, key=f"lote_alcance_{doc_key}")
        with bc3:
            st.write(""); st.write("")
            confirmar_lote = st.button(f"✅ Confirmar lote ({n_sel})", disabled=(n_sel == 0 or not cat_lote), use_container_width=True)

        if confirmar_lote and n_sel > 0 and cat_lote:
            codigo_lote = '__EXCLUIR__' if cat_lote == '🚫 NO INCLUIR' else (cat_lote if cat_lote != '➕ NUEVA CATEGORÍA' else None)
            if codigo_lote:
                incompatibles = [
                    idx_lote for idx_lote in st.session_state.lote_seleccion
                    if codigo_lote != '__EXCLUIR__'
                    and not _codigo_compatible_con_origen(
                        codigo_lote,
                        df.at[idx_lote, 'origen_columna'],
                        df.at[idx_lote, 'monto'],
                        df.at[idx_lote, 'nombre_original'],
                        catalogo,
                    )
                ]
                if incompatibles:
                    st.error(
                        "La categoría elegida contradice la columna contable de "
                        f"{len(incompatibles)} cuenta(s). No se aplicó el lote."
                    )
                    st.stop()
                procesados = 0
                fallback_json_lote = False
                validaciones_lote = []
                for idx_lote in list(st.session_state.lote_seleccion):
                    _registrar_decision(archivo_nombre, idx_lote, df.loc[idx_lote].copy(), codigo_lote, 'Confirmación en lote')
                    nombre_orig = df.at[idx_lote, 'nombre_original']
                    codigo_sugerido = df.at[idx_lote, 'codigo_clasificado'] or None
                    metodo_sugerido = df.at[idx_lote, 'metodo']
                    confianza_sugerida = float(df.at[idx_lote, 'confianza'])
                    st.session_state.resultados[archivo_nombre].at[idx_lote, 'codigo_clasificado'] = codigo_lote
                    st.session_state.resultados[archivo_nombre].at[idx_lote, 'metodo'] = 'validacion_humana_lote'
                    st.session_state.resultados[archivo_nombre].at[idx_lote, 'confianza'] = 1.0
                    st.session_state.resultados[archivo_nombre].at[idx_lote, 'requiere_revision'] = False
                    st.session_state.resultados[archivo_nombre].at[idx_lote, 'origen'] = 'Manual'
                    st.session_state.resultados[archivo_nombre].at[idx_lote, 'regla'] = 'validacion_humana_lote'
                    st.session_state.resultados[archivo_nombre].at[idx_lote, 'evidencia'] = 'Asignación en lote por analista'
                    validaciones_lote.append({
                        'account_name': nombre_orig,
                        'validated_code': codigo_lote,
                        'source': 'validacion_humana_lote',
                        'add_to_dictionary': "diccionario" in alcance_lote,
                        'suggested_code': codigo_sugerido,
                        'suggested_method': metodo_sugerido,
                        'suggested_confidence': confianza_sugerida,
                        'source_file': archivo_nombre,
                    })
                    if "diccionario" in alcance_lote:
                        entrada = {'cuenta_original': nombre_orig, 'codigo_estandar': codigo_lote, 'fuente': 'validacion_humana_lote'}
                        st.session_state.diccionario.append(entrada)
                        st.session_state.correcciones.append(entrada)
                    if "diccionario" in alcance_lote:
                        propagar_clasificacion_resultados(nombre_orig, codigo_lote, 'validacion_humana_lote_propagada')
                    procesados += 1
                persistido = _persistir_validaciones_lote(validaciones_lote)
                fallback_json_lote = not persistido
                if "diccionario" in alcance_lote and fallback_json_lote:
                    with open(BASE_DIR / 'diccionario.json', 'w', encoding='utf-8') as f:
                        json.dump(st.session_state.diccionario, f, ensure_ascii=False, indent=2)
                st.session_state.lote_seleccion = set()
                st.rerun()

        qa, qb, qc = st.columns(3)
        qa.button(
            "☑️ Seleccionar todas", use_container_width=True,
            on_click=_reemplazar_seleccion_lote,
            args=(pendientes.index.tolist(), checkbox_prefix),
        )
        qb.button(
            "🟦 Seleccionar sin clasificar", use_container_width=True,
            on_click=_reemplazar_seleccion_lote,
            args=(pendientes[pendientes['codigo_clasificado'] == ''].index.tolist(), checkbox_prefix),
        )
        qc.button(
            "⬜ Limpiar selección", use_container_width=True,
            on_click=_reemplazar_seleccion_lote,
            args=([], checkbox_prefix),
        )

    pc1, pc2 = st.columns([1, 2])
    with pc1:
        page_size = st.selectbox(
            "Cuentas por página", [10, 25, 50], index=0,
            key=f"revision_page_size_{doc_key}",
        )
    total_pages = max(1, (len(pendientes) + page_size - 1) // page_size)
    current_page = min(
        max(int(st.session_state.get(f"revision_page_{doc_key}", 1)), 1),
        total_pages,
    )
    if st.session_state.get(f"revision_page_{doc_key}") != current_page:
        st.session_state[f"revision_page_{doc_key}"] = current_page
    with pc2:
        page = int(st.number_input(
            "Página", min_value=1, max_value=total_pages,
            step=1, key=f"revision_page_{doc_key}",
        ))
    start = (page - 1) * page_size
    visible = pendientes.iloc[start:start + page_size]
    st.caption(
        f"Mostrando {start + 1}–{min(start + page_size, len(pendientes))} "
        f"de {len(pendientes)} cuentas · página {page} de {total_pages}."
    )

    for idx, row in visible.iterrows():
        seleccionada = idx in st.session_state.lote_seleccion
        with st.container(border=seleccionada):
            c0, c1, c2 = st.columns([0.3, 4, 4])
            with c0:
                checkbox_key = f"{checkbox_prefix}_{idx}"
                st.checkbox(
                    "", value=seleccionada, key=checkbox_key,
                    label_visibility="collapsed",
                    on_change=_alternar_seleccion_lote,
                    args=(idx, checkbox_key),
                )

            with c1:
                col_extraida = row.get('origen_columna', 'desconocido')
                col_actual = row.get(
                    'origen_columna_efectiva',
                    _origen_efectivo(
                        col_extraida, row.get('monto'), row.get('nombre_original'),
                    ),
                ).upper()
                etiqueta_columna = _etiqueta_origen(
                    col_extraida, row.get('monto'), row.get('nombre_original'),
                )
                badge_bg = {
                    'ACTIVO': '#1E90FF', 'PASIVO': '#FF8C00',
                    'PERDIDA': '#DC143C', 'GANANCIA': '#2E8B57',
                }.get(col_actual, '#6B7280')
                st.markdown(
                    f"<span style='background:{badge_bg}; color:white; "
                    f"padding:2px 10px; border-radius:4px; font-size:0.75em; "
                    f"font-weight:600; letter-spacing:0.5px;'>{etiqueta_columna}</span>",
                    unsafe_allow_html=True,
                )
                st.caption("Columna de origen del balance")
                st.markdown(f"**{_nombre_mostrar(row)}**")
                st.caption(f"Decisión actual: {row.get('codigo_clasificado') or 'Sin clasificar'} · {row.get('metodo', '')}")
                monto_val = row['monto']
                if pd.notna(monto_val):
                    color = 'blue' if monto_val > 0 else ('red' if monto_val < 0 else 'gray')
                    st.markdown(
                        f"<span style='color:{color}; font-weight:bold;'>{monto_val:,.0f}</span>",
                        unsafe_allow_html=True,
                    )
                monto_anterior = row.get('monto_periodo_anterior')
                if pd.notna(monto_anterior):
                    st.caption(
                        f"Período seleccionado: actual · "
                        f"Anterior: {float(monto_anterior):,.0f}"
                    )
                columnas_derivadas = str(row.get('columnas_derivadas') or '').strip()
                if columnas_derivadas:
                    st.warning(
                        "Dato reconstruido contablemente: "
                        f"{columnas_derivadas}. Requiere confirmación humana."
                    )

                edit_mode = st.checkbox("Editar cuenta", key=f"edit_{doc_key}_{idx}")

                if edit_mode:
                    st.divider()
                    nuevo_nombre = st.text_input("Nombre", value=_nombre_mostrar(row), key=f"ed_nombre_{doc_key}_{idx}")
                    opciones_nat = ['ACTIVO', 'PASIVO', 'PERDIDA', 'GANANCIA']
                    columna_fisica = str(col_extraida).upper()
                    idx_nat = opciones_nat.index(columna_fisica) if columna_fisica in opciones_nat else 0
                    nueva_nat = st.selectbox("Columna contable", opciones_nat, index=idx_nat, key=f"ed_nat_{doc_key}_{idx}")
                    monto_inicial = row['monto'] if pd.notna(row['monto']) else 0.0
                    nuevo_monto = st.number_input("Monto", value=float(monto_inicial), format="%.0f", key=f"ed_monto_{doc_key}_{idx}")

                    if st.button("💾 Guardar corrección", key=f"ed_guardar_{doc_key}_{idx}", use_container_width=True):
                        df_mod = st.session_state.resultados[archivo_nombre]
                        original_col = row.get('origen_columna', '')
                        original_monto = row['monto']
                        col_changed = nueva_nat.lower() != original_col
                        monto_changed = (nuevo_monto != original_monto) if pd.notna(original_monto) else (nuevo_monto != 0)

                        if col_changed or monto_changed:
                            sel_clave = st.session_state.get(f"sel_{doc_key}_{idx}", '')
                            codigo_final = sel_clave if sel_clave not in ('', '➕ NUEVA CATEGORÍA', '🚫 NO INCLUIR') else ''
                            codigo_final = codigo_final or str(row.get('codigo_clasificado') or '')
                            if codigo_final and not _codigo_compatible_con_origen(
                                codigo_final, nueva_nat, nuevo_monto, nuevo_nombre, catalogo,
                            ):
                                st.error('La corrección contradice la clasificación actual. Seleccione una categoría compatible antes de guardarla.')
                                st.stop()
                            _registrar_decision(archivo_nombre, idx, row.copy(), codigo_final, 'Corrección de datos de la cuenta')
                            df_mod.at[idx, 'nombre_original'] = nuevo_nombre
                            df_mod.at[idx, 'nombre_revision_usuario'] = ''
                            df_mod.at[idx, 'origen_columna'] = nueva_nat.lower()
                            df_mod.at[idx, 'monto'] = nuevo_monto
                            df_mod.at[idx, 'origen_columna_efectiva'] = _origen_efectivo(
                                nueva_nat, nuevo_monto, nuevo_nombre)
                            df_mod.at[idx, 'origen_columna_display'] = _etiqueta_origen(
                                nueva_nat, nuevo_monto, nuevo_nombre)
                            if codigo_final:
                                df_mod.at[idx, 'codigo_clasificado'] = codigo_final
                            df_mod.at[idx, 'metodo'] = 'manual_revision'
                            df_mod.at[idx, 'confianza'] = 1.0
                            df_mod.at[idx, 'requiere_revision'] = True
                            df_mod.at[idx, 'tipo_revision'] = 'correccion_extraccion'
                            df_mod.at[idx, 'origen'] = 'Manual'
                            df_mod.at[idx, 'regla'] = 'manual_revision'
                            df_mod.at[idx, 'evidencia'] = 'Corrección manual de extracción'
                            # Una corrección de extracción es local y debe confirmarse.
                            st.toast(f"'{nuevo_nombre[:35]}' corregida ✅", icon="✅")
                        else:
                            df_mod.at[idx, 'nombre_revision_usuario'] = nuevo_nombre
                            df_mod.at[idx, 'tipo_revision'] = 'visual'
                            st.toast(f"'{nuevo_nombre[:35]}' nombre visual actualizado ✏️", icon="✏️")
                        st.rerun()

            with c2:
                mostrar_todas = st.checkbox(
                    "🔎 Buscar más clasificaciones",
                    key=f"mostrar_todas_{doc_key}_{idx}",
                    help=(
                        "Muestra el catálogo completo para casos contables "
                        "excepcionales que no coinciden con la columna física."
                    ),
                )
                sugerido = row['codigo_clasificado']
                if (not mostrar_todas and sugerido
                        and not _codigo_compatible_con_origen(
                            sugerido, row.get('origen_columna'), row.get('monto'),
                            _nombre_mostrar(row), catalogo)):
                    sugerido = ''
                st.write(f"Sugerido: **{sugerido or '(ninguno)'}**")

                alternativas = _alternativas_revision(
                    nombre=_nombre_mostrar(row),
                    sugerido=sugerido,
                    confianza=float(row.get('confianza') or 0.0),
                    origen_columna=row.get('origen_columna'),
                    monto=row.get('monto'),
                    catalogo=catalogo,
                    motor=motor,
                )
                if alternativas:
                    st.caption("Alternativas compatibles · selección asistida, no automática")
                    if (
                        len(alternativas) > 1
                        and alternativas[0]["codigo"] != alternativas[1]["codigo"]
                        and alternativas[0]["score"] - alternativas[1]["score"] <= 0.05
                    ):
                        st.warning(
                            "Señales contradictorias: los dos primeros candidatos "
                            "tienen relevancia similar. Requiere criterio del analista."
                        )
                    alt_cols = st.columns(len(alternativas))
                    for alt_col, alternativa in zip(alt_cols, alternativas):
                        with alt_col:
                            st.markdown(
                                f"**`{alternativa['codigo']}`**  \n"
                                f"{alternativa['nombre']}"
                            )
                            st.caption(
                                f"{alternativa['score']:.0%} · {alternativa['fuente']}  \n"
                                f"{alternativa['evidencia']}"
                            )
                            selection_key = f"sel_{doc_key}_{idx}"
                            st.button(
                                "Usar",
                                key=f"usar_alt_{doc_key}_{idx}_{alternativa['codigo']}",
                                use_container_width=True,
                                on_click=_asignar_estado_widget,
                                args=(selection_key, alternativa['codigo']),
                            )

                if mostrar_todas:
                    opciones_fila = opciones_codigo
                    st.caption("Catálogo completo habilitado para esta cuenta.")
                else:
                    opciones_fila = [opciones_codigo[0]] + [
                        codigo for codigo in opciones_codigo[1:]
                        if codigo in ('➕ NUEVA CATEGORÍA', '🚫 NO INCLUIR')
                        or _codigo_compatible_con_origen(
                            codigo, row.get('origen_columna'), row.get('monto'),
                            _nombre_mostrar(row), catalogo)
                    ]
                default_idx = (opciones_fila.index(sugerido)
                               if sugerido in opciones_fila else 0)
                seleccion = st.selectbox(
                    "Clasificación correcta",
                    opciones_fila,
                    index=default_idx,
                    format_func=lambda c: (
                        f"{c} — {catalogo[c]['nombre_estandar']}" if c in catalogo
                        else c if c else "(sin clasificar)"
                    ),
                    key=f"sel_{doc_key}_{idx}"
                )

                es_nueva_cat = seleccion == '➕ NUEVA CATEGORÍA'
                if es_nueva_cat:
                    st.info("Define la nueva categoría:")
                    nuevo_codigo = st.text_input("Código (ej: AC.10, ER.17)",
                                                  key=f"new_cod_{doc_key}_{idx}", max_chars=10)
                    nuevo_nombre_cat = st.text_input("Nombre de la categoría",
                                                  key=f"new_nom_{doc_key}_{idx}")
                    nuevo_tipo = st.selectbox("Tipo de estado",
                                              ['balance', 'resultados'],
                                              key=f"new_tipo_{doc_key}_{idx}")
                    nuevo_cat = st.selectbox(
                        "Categoría",
                        ['activo_corriente', 'activo_no_corriente',
                         'pasivo_corriente', 'pasivo_no_corriente',
                         'patrimonio', 'resultado'],
                        key=f"new_cat_{doc_key}_{idx}"
                    )
                    naturaleza_resultado = st.selectbox(
                        'Naturaleza de la nueva categoría', ['ganancia', 'perdida'],
                        key=f'new_naturaleza_{doc_key}_{idx}',
                    ) if nuevo_cat == 'resultado' else None

                if not es_nueva_cat and seleccion not in ('', '🚫 NO INCLUIR'):
                    alcance = st.radio(
                        "¿Aplicar esta clasificación?",
                        ["Solo para este caso",
                         "Agregar al diccionario (aplica a casos futuros iguales)"],
                        index=1 if row.get('requiere_revision', False) else 0,
                        key=f"alc_{doc_key}_{idx}", horizontal=True
                    )
                else:
                    alcance = "Solo para este caso"

                if st.button("✅ Confirmar", key=f"btn_{doc_key}_{idx}"):
                    codigo_final = None

                    if es_nueva_cat:
                        if nuevo_codigo and nuevo_nombre_cat:
                            nueva_entrada = {
                                'codigo_estandar': nuevo_codigo.strip().upper(),
                                'nombre_estandar': nuevo_nombre_cat.strip(),
                                'categoria': nuevo_cat,
                                'tipo_estado': nuevo_tipo,
                                'naturaleza': 'deudora' if nuevo_cat.startswith('activo') else 'acreedora',
                                'signo_normal': 1,
                                'es_deuda_financiera': False,
                                'es_activo_liquido': False,
                                'afecta_ebitda': False,
                            }
                            if naturaleza_resultado:
                                nueva_entrada['naturaleza'] = 'deudora' if naturaleza_resultado == 'perdida' else 'acreedora'
                                nueva_entrada['signo_normal'] = -1 if naturaleza_resultado == 'perdida' else 1
                            candidato = nuevo_codigo.strip().upper()
                            prefijos = {'activo_corriente': 'AC.', 'activo_no_corriente': 'ANC.',
                                        'pasivo_corriente': 'PC.', 'pasivo_no_corriente': 'PNC.',
                                        'patrimonio': 'PAT.', 'resultado': 'ER.'}
                            if candidato in catalogo or not candidato.startswith(prefijos[nuevo_cat]):
                                st.error('Use un código nuevo con el prefijo de la categoría seleccionada. No se puede redefinir una categoría existente desde esta cuenta.')
                                st.stop()
                            if not _codigo_compatible_con_origen(candidato, row.get('origen_columna'), row.get('monto'), _nombre_mostrar(row), {**catalogo, candidato: nueva_entrada}):
                                st.error('La nueva categoría contradice la naturaleza de esta cuenta. No fue creada.')
                                st.stop()
                            catalogo[nuevo_codigo.strip().upper()] = nueva_entrada
                            if not _persistir_catalogo(nueva_entrada):
                                with open(BASE_DIR / 'catalogo_maestro.json', 'w', encoding='utf-8') as f:
                                    json.dump(catalogo, f, ensure_ascii=False, indent=2)
                            codigo_final = nuevo_codigo.strip().upper()
                            st.toast(f"Nueva categoría '{nuevo_nombre_cat}' ({codigo_final}) creada ✨", icon="🆕")
                        else:
                            st.error("Debes ingresar código y nombre.")

                    elif seleccion == '🚫 NO INCLUIR':
                        _registrar_decision(archivo_nombre, idx, row.copy(), '__EXCLUIR__', 'Exclusión por analista')
                        st.session_state.resultados[archivo_nombre].at[idx, 'codigo_clasificado'] = '__EXCLUIR__'
                        st.session_state.resultados[archivo_nombre].at[idx, 'metodo'] = 'excluido_analista'
                        st.session_state.resultados[archivo_nombre].at[idx, 'confianza'] = 1.0
                        st.session_state.resultados[archivo_nombre].at[idx, 'requiere_revision'] = False
                        if "diccionario" in alcance:
                            entrada_exclusion = {
                                'cuenta_original': _nombre_mostrar(row),
                                'codigo_estandar': '__EXCLUIR__',
                                'fuente': 'excluido_analista'
                            }
                            st.session_state.diccionario.append(entrada_exclusion)
                        persistido = _persistir_validacion(
                            nombre=_nombre_mostrar(row), codigo='__EXCLUIR__',
                            fuente='excluido_analista',
                            agregar_diccionario="diccionario" in alcance,
                            sugerido=row['codigo_clasificado'] or None,
                            metodo=row['metodo'], confianza=float(row['confianza']),
                            archivo=archivo_nombre,
                        )
                        if "diccionario" in alcance and not persistido:
                            with open(BASE_DIR / 'diccionario.json', 'w', encoding='utf-8') as f:
                                json.dump(st.session_state.diccionario, f, ensure_ascii=False, indent=2)
                        if "diccionario" in alcance:
                            propagar_clasificacion_resultados(row['nombre_original'], '__EXCLUIR__', 'excluido_analista_propagado')
                        st.session_state.lote_seleccion.discard(idx)
                        st.toast(f"'{_nombre_mostrar(row)[:35]}' excluida", icon="🚫")
                        st.rerun()

                    elif seleccion:
                        codigo_final = seleccion

                    if codigo_final:
                        if not _codigo_compatible_con_origen(codigo_final, row.get('origen_columna'), row.get('monto'), _nombre_mostrar(row), catalogo):
                            st.error('La categoría contradice la naturaleza contable. No se cambió la cuenta ni se guardó en el diccionario.')
                            st.stop()
                        _registrar_decision(archivo_nombre, idx, row.copy(), codigo_final, 'Confirmación individual')
                        st.session_state.resultados[archivo_nombre].at[idx, 'codigo_clasificado'] = codigo_final
                        st.session_state.resultados[archivo_nombre].at[idx, 'metodo'] = 'validacion_humana'
                        st.session_state.resultados[archivo_nombre].at[idx, 'confianza'] = 1.0
                        st.session_state.resultados[archivo_nombre].at[idx, 'requiere_revision'] = False
                        st.session_state.lote_seleccion.discard(idx)
                        persistido = _persistir_validacion(
                            nombre=row['nombre_original'], codigo=codigo_final,
                            fuente='validacion_humana',
                            agregar_diccionario="diccionario" in alcance,
                            sugerido=row['codigo_clasificado'] or None,
                            metodo=row['metodo'], confianza=float(row['confianza']),
                            archivo=archivo_nombre,
                        )
                        if "diccionario" in alcance:
                            nuevo_dic = {
                                'cuenta_original': row['nombre_original'],
                                'codigo_estandar': codigo_final,
                                'fuente': 'validacion_humana'
                            }
                            st.session_state.diccionario.append(nuevo_dic)
                            st.session_state.correcciones.append(nuevo_dic)
                            if not persistido:
                                with open(BASE_DIR / 'diccionario.json', 'w', encoding='utf-8') as f:
                                    json.dump(st.session_state.diccionario, f, ensure_ascii=False, indent=2)
                            st.toast(f"'{_nombre_mostrar(row)[:35]}' → {codigo_final} guardado 📚", icon="✅")
                        else:
                            st.toast(f"'{_nombre_mostrar(row)[:35]}' → {codigo_final} (solo este caso)", icon="✅")
                        if "diccionario" in alcance:
                            propagar_clasificacion_resultados(row['nombre_original'], codigo_final, 'validacion_humana_propagada')
                        st.rerun()


def _diagnosticar_cuadratura(
        df: pd.DataFrame, agrupado: pd.DataFrame,
        clasificadas: pd.DataFrame, tolerancia: float = 1_000) -> dict:
    """Concilia el balance homologado y localiza causas probables del descuadre."""
    codigos = agrupado['codigo_clasificado'].fillna('').astype(str)
    activo = agrupado[codigos.str.startswith(('AC.', 'ANC.'))]['monto_total'].sum()
    pasivo_patrimonio = agrupado[
        codigos.str.startswith(('PC.', 'PNC.', 'PAT.'))
    ]['monto_total'].sum()
    diferencia = float(activo - pasivo_patrimonio)

    incompatibles = clasificadas[
        ~clasificadas.apply(
            lambda row: _codigo_compatible_con_origen(
                row.get('codigo_clasificado'), row.get('origen_columna'),
                row.get('monto'), row.get('nombre_original'),
            ),
            axis=1,
        )
    ].copy()
    if not incompatibles.empty:
        incompatibles['impacto_potencial'] = incompatibles['monto'].abs() * 2
        incompatibles['explica_diferencia'] = (
            incompatibles['impacto_potencial'] - abs(diferencia)
        ).abs() <= tolerancia

    relevantes = _con_saldo_relevante(df[~df['es_total']].copy())
    sin_clasificar = relevantes[relevantes['codigo_clasificado'] == ''].copy()
    excluidas = relevantes[relevantes['codigo_clasificado'] == '__EXCLUIR__'].copy()
    total_cuentas = len(relevantes)
    cuentas_clasificadas = total_cuentas - len(sin_clasificar) - len(excluidas)
    cobertura = cuentas_clasificadas / total_cuentas if total_cuentas else 1.0

    return {
        'activo': float(activo),
        'pasivo_patrimonio': float(pasivo_patrimonio),
        'diferencia': diferencia,
        'cuadra': abs(diferencia) <= tolerancia,
        'tolerancia': tolerancia,
        'incompatibles': incompatibles,
        'sin_clasificar': sin_clasificar,
        'excluidas': excluidas,
        'total_cuentas': total_cuentas,
        'cuentas_clasificadas': cuentas_clasificadas,
        'cobertura': cobertura,
    }


def _reabrir_incompatibles(df: pd.DataFrame, indices) -> pd.DataFrame:
    """Devuelve las clasificaciones incompatibles a la cola de revisión humana."""
    resultado = df.copy()
    indices_validos = resultado.index.intersection(indices)
    if len(indices_validos):
        resultado.loc[indices_validos, 'codigo_clasificado'] = ''
        resultado.loc[indices_validos, 'metodo'] = 'reapertura_cuadratura'
        resultado.loc[indices_validos, 'confianza'] = 0.0
        resultado.loc[indices_validos, 'requiere_revision'] = True
        if 'regla' in resultado.columns:
            resultado.loc[indices_validos, 'regla'] = 'reapertura_cuadratura'
        if 'evidencia' in resultado.columns:
            resultado.loc[indices_validos, 'evidencia'] = (
                'Clasificación incompatible detectada por conciliación contable'
            )
    return resultado


def _mostrar_resumen_cuadratura(
        diagnostico: dict, df: pd.DataFrame, archivo_nombre: str) -> None:
    """Resumen ejecutivo, causas probables y navegación a revisión humana."""
    st.subheader("✅ Control de cuadratura del balance homologado")
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Activo normalizado", f"${diagnostico['activo']:,.0f}")
    c2.metric("Pasivo + Patrimonio", f"${diagnostico['pasivo_patrimonio']:,.0f}")
    c3.metric(
        "Diferencia", f"${abs(diagnostico['diferencia']):,.0f}",
        delta="Cuadra" if diagnostico['cuadra'] else "Requiere corrección",
        delta_color="normal" if diagnostico['cuadra'] else "inverse",
    )
    c4.metric(
        "Cobertura", f"{diagnostico['cobertura']:.1%}",
        help=(f"{diagnostico['cuentas_clasificadas']} de "
              f"{diagnostico['total_cuentas']} cuentas con saldo"),
    )

    if diagnostico['cuadra']:
        st.success(
            "El activo homologado coincide con pasivo más patrimonio dentro de "
            f"la tolerancia de ${diagnostico['tolerancia']:,.0f}."
        )
        return

    st.error(
        "El balance homologado no cuadra. La descarga permanece disponible para "
        "auditoría, pero el resultado requiere revisión humana."
    )
    incompatibles = diagnostico['incompatibles']
    sin_clasificar = diagnostico['sin_clasificar']
    excluidas = diagnostico['excluidas']

    with st.container(border=True):
        st.markdown("#### Causas probables")
        if not incompatibles.empty:
            exactas = incompatibles[incompatibles['explica_diferencia']]
            if not exactas.empty:
                st.warning(
                    f"Se encontraron {len(exactas)} clasificación(es) incompatible(s) "
                    "cuyo cambio de lado explica exactamente el descuadre."
                )
            else:
                st.warning(
                    f"Se encontraron {len(incompatibles)} clasificación(es) que "
                    "contradicen la columna contable extraída."
                )
            tabla = incompatibles.copy()
            tabla['Cuenta'] = tabla['nombre_original']
            tabla['Columna extraída'] = tabla['origen_columna'].str.upper()
            tabla['Clasificación actual'] = tabla['codigo_clasificado']
            tabla['Monto'] = tabla['monto'].map(lambda x: f"${x:,.0f}")
            tabla['Impacto si cambia de lado'] = tabla['impacto_potencial'].map(
                lambda x: f"${x:,.0f}"
            )
            st.dataframe(
                tabla[['Cuenta', 'Columna extraída', 'Clasificación actual',
                       'Monto', 'Impacto si cambia de lado']],
                use_container_width=True, hide_index=True,
            )
        if not sin_clasificar.empty:
            st.warning(
                f"Hay {len(sin_clasificar)} cuenta(s) con saldo sin clasificar por "
                f"${sin_clasificar['monto'].abs().sum():,.0f}."
            )
        if not excluidas.empty:
            st.info(
                f"Hay {len(excluidas)} cuenta(s) excluida(s) por "
                f"${excluidas['monto'].abs().sum():,.0f}."
            )
        if incompatibles.empty and sin_clasificar.empty and excluidas.empty:
            st.info(
                "No se detectó una causa única. Revise signos, cuentas contra-activo "
                "y asignaciones de patrimonio cercanas al monto de la diferencia."
            )

        if not incompatibles.empty:
            if st.button(
                f"↩️ Reabrir {len(incompatibles)} cuenta(s) incompatible(s) en revisión",
                type="primary", use_container_width=True,
            ):
                st.session_state.resultados[archivo_nombre] = _reabrir_incompatibles(
                    df, incompatibles.index
                )
                st.session_state['vista_trabajo_solicitada'] = "🔍 Cola de Revisión"
                st.rerun()
        elif st.button("↩️ Volver a la clasificación humana", use_container_width=True):
            st.session_state['vista_trabajo_solicitada'] = "🔍 Cola de Revisión"
            st.rerun()


def _mostrar_control_calidad_operativo(control) -> None:
    """Presenta Structure, Coverage y Self-QA sin mezclarlo con clasificación."""
    st.subheader("🛡️ Control posterior de calidad")
    mode_label = "Aplicado a exportación" if control.mode == "enforced" else "Shadow mode"
    coverage = control.coverage
    monetary = coverage.get("monetary", {})
    semantic = coverage.get("semantic", {})
    qa_state = control.self_qa.get("approval_state", "SIN_DATOS")
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Estructura", control.structure.get("column_layout", "—"))
    c2.metric("Cobertura monetaria", f"{float(monetary.get('coverage_pct', 0)):.1%}")
    c3.metric("Cobertura de cuentas", f"{float(semantic.get('overall', 0)):.1%}")
    c4.metric("Self-QA", str(qa_state).replace("_", " "))
    st.caption(
        f"{mode_label}. Los motores auxiliares sólo leen una copia del resultado; "
        "no pueden modificar clasificaciones ni montos."
    )
    if control.requires_review:
        st.warning("Revisión requerida: " + "; ".join(control.reasons) + ".")
    else:
        st.success("Los controles posteriores no detectaron motivos para reabrir la revisión.")
    if not control.export_allowed:
        st.error(
            "Exportación bloqueada por el control posterior. Corrige los motivos "
            "indicados en la cola de revisión y vuelve a este balance."
        )


def _control_emision(df, catalogo, diagnostico, quality_control=None):
    """Control independiente, sin mutar decisiones ni importes del analista."""
    incidencias = []
    for idx, row in df[~df['es_total']].iterrows():
        codigo = str(row.get('codigo_clasificado') or '')
        monto = pd.to_numeric(row.get('monto'), errors='coerce')
        motivos = []
        if pd.isna(monto) or not math.isfinite(float(monto)):
            motivos.append('Importe ausente o no válido')
        elif monto == 0:
            continue
        if codigo == '__EXCLUIR__':
            motivos.append('Cuenta con saldo excluida; revise si debe reincorporarse')
        elif codigo not in catalogo:
            motivos.append('Sin clasificación válida')
        elif not _codigo_compatible_con_origen(
            codigo, row.get('origen_columna'), monto, row.get('nombre_original'), catalogo,
        ):
            motivos.append('La categoría contradice la naturaleza efectiva de la cuenta')
        if row.get('requiere_revision', False):
            motivos.append('Decisión pendiente de confirmación')
        if motivos:
            incidencias.append({
                'Fila': str(idx), 'Cuenta': row.get('nombre_original', ''),
                'Clasificación': codigo, 'Columna original': row.get('origen_columna', ''),
                'Monto': monto, 'Qué corregir': '; '.join(motivos),
            })
    resultado = conciliar_resultados(df.to_dict('records'), catalogo)
    motivos = list(resultado['problemas'])
    if incidencias:
        motivos.append(f'{len(incidencias)} cuenta(s) requieren corregir o confirmar su clasificación')
    if not diagnostico['cuadra']:
        motivos.append('Activo no coincide con Pasivo más Patrimonio')
    if quality_control is not None and not quality_control.export_allowed:
        motivos.extend(quality_control.reasons or ['El control posterior no autoriza la emisión definitiva'])
    return {'definitivo': not motivos, 'motivos': motivos,
            'incidencias': pd.DataFrame(incidencias), 'resultado': resultado}


def _valor_fila_periodo(row: pd.Series, periodo: str, posicion: int):
    """Recupera el importe elegido para un período sin alterar la fila fuente."""
    columna = f"monto_periodo_{periodo}"
    if columna in row.index and pd.notna(row.get(columna)):
        return float(row[columna])
    alias = "monto_periodo_actual" if posicion == 0 else "monto_periodo_anterior"
    if alias in row.index and pd.notna(row.get(alias)):
        return float(row[alias])
    if posicion == 0 and pd.notna(row.get("monto")):
        return float(row["monto"])
    return None


def _preparar_periodo_reporte(
        df: pd.DataFrame, catalogo: dict, periodo: str, posicion: int) -> dict:
    """Genera clasificación, conciliación y cuadratura independientes por período."""
    periodo_df = df.copy()
    periodo_df["monto"] = periodo_df.apply(
        lambda row: _valor_fila_periodo(row, periodo, posicion), axis=1,
    )
    clasificadas = periodo_df[
        (periodo_df["codigo_clasificado"] != "")
        & (periodo_df["codigo_clasificado"] != "__EXCLUIR__")
        & (~periodo_df["es_total"])
    ].copy()
    clasificadas = _con_saldo_relevante(clasificadas)
    clasificadas["monto"] = pd.to_numeric(
        clasificadas["monto"], errors="coerce",
    ).fillna(0)
    clasificadas["monto_presentacion"] = clasificadas.apply(
        lambda row: _monto_presentacion(
            row["codigo_clasificado"], row["monto"], row["nombre_original"],
            row.get("origen_columna"), catalogo,
        ), axis=1,
    )
    agrupado = clasificadas.groupby("codigo_clasificado").agg(
        monto_total=("monto_presentacion", "sum"),
        num_cuentas=("nombre_original", "count"),
    ).reset_index()
    agrupado["nombre_estandar"] = agrupado["codigo_clasificado"].map(
        lambda code: catalogo.get(code, {}).get("nombre_estandar", code)
    )
    agrupado["categoria"] = agrupado["codigo_clasificado"].map(
        lambda code: catalogo.get(code, {}).get("categoria", "")
    )
    conciliacion = conciliar_resultados(periodo_df.to_dict("records"), catalogo)
    resultado_periodo = conciliacion["resultado_origen"]
    if resultado_periodo is not None:
        existentes = set(agrupado["codigo_clasificado"])
        derivados = []
        for codigo_resultado in ("ER.11", "PAT.04"):
            if codigo_resultado in existentes:
                continue
            info = catalogo.get(codigo_resultado, {})
            derivados.append({
                "codigo_clasificado": codigo_resultado,
                "monto_total": (
                    conciliacion["resultado_homologado"] or 0.0
                    if codigo_resultado == "ER.11" else resultado_periodo
                ),
                "num_cuentas": 0,
                "nombre_estandar": info.get("nombre_estandar", codigo_resultado),
                "categoria": info.get("categoria", ""),
            })
        if derivados:
            agrupado = pd.concat([agrupado, pd.DataFrame(derivados)], ignore_index=True)
    diagnostico = _diagnosticar_cuadratura(periodo_df, agrupado, clasificadas)
    return {
        "periodo": periodo, "df": periodo_df, "clasificadas": clasificadas,
        "agrupado": agrupado, "conciliacion": conciliacion,
        "diagnostico": diagnostico,
    }


def _tab_balance(df: pd.DataFrame, catalogo: dict, archivo_nombre: str):
    st.button('Revisar o cambiar clasificaciones', on_click=_solicitar_revision_completa,
              args=(archivo_nombre,), use_container_width=True)
    periodos = _periodos_seleccionados()
    reportes_periodo = [
        _preparar_periodo_reporte(df, catalogo, periodo, posicion)
        for posicion, periodo in enumerate(periodos)
    ]
    principal = reportes_periodo[0]
    clasificadas = principal["clasificadas"]
    if clasificadas.empty:
        st.info("No hay cuentas clasificadas todavía.")
        return
    agrupado = principal["agrupado"].copy()
    conciliacion = principal["conciliacion"]
    resultado_periodo = conciliacion['resultado_origen']
    period_columns = [(
        periodos[0] if len(periodos) > 1 else "Importe", "monto_total",
    )]
    for reporte in reportes_periodo[1:]:
        amount_column = f"monto_total_{reporte['periodo']}"
        period_columns.append((reporte["periodo"], amount_column))
        valores = reporte["agrupado"][["codigo_clasificado", "monto_total"]].rename(
            columns={"monto_total": amount_column}
        )
        agrupado = agrupado.merge(valores, on="codigo_clasificado", how="outer")
        agrupado["nombre_estandar"] = agrupado.apply(
            lambda row: row.get("nombre_estandar")
            if pd.notna(row.get("nombre_estandar")) else catalogo.get(
                row["codigo_clasificado"], {}
            ).get("nombre_estandar", row["codigo_clasificado"]), axis=1,
        )
        agrupado["categoria"] = agrupado.apply(
            lambda row: row.get("categoria")
            if pd.notna(row.get("categoria")) else catalogo.get(
                row["codigo_clasificado"], {}
            ).get("categoria", ""), axis=1,
        )
        agrupado["num_cuentas"] = agrupado["num_cuentas"].fillna(0)
        agrupado["monto_total"] = agrupado["monto_total"].fillna(0)
        agrupado[amount_column] = agrupado[amount_column].fillna(0)

    orden_cat = ['activo_corriente', 'activo_no_corriente', 'pasivo_corriente', 'pasivo_no_corriente', 'patrimonio', 'resultado']
    agrupado['orden'] = agrupado['categoria'].map(lambda c: orden_cat.index(c) if c in orden_cat else 99)
    agrupado = agrupado.sort_values(['orden', 'codigo_clasificado'])

    diagnostico = principal["diagnostico"]
    _mostrar_resumen_cuadratura(diagnostico, df, archivo_nombre)
    if len(reportes_periodo) > 1:
        st.subheader("Control comparativo por período")
        st.dataframe(pd.DataFrame([{
            "Período": reporte["periodo"],
            "Activo": reporte["diagnostico"]["activo"],
            "Pasivo + Patrimonio": reporte["diagnostico"]["pasivo_patrimonio"],
            "Diferencia": reporte["diagnostico"]["diferencia"],
            "Cuadratura": "Cuadra" if reporte["diagnostico"]["cuadra"] else "Requiere revisión",
            "Resultado original": reporte["conciliacion"]["resultado_origen"],
            "Resultado homologado": reporte["conciliacion"]["resultado_homologado"],
            "Diferencia resultado": reporte["conciliacion"]["diferencia"],
        } for reporte in reportes_periodo]), hide_index=True, use_container_width=True)
    from pipeline.operational_quality import analyze_operational_quality
    enforce_export = os.environ.get(
        "QUALITY_CONTROL_ENFORCE_EXPORT", "false",
    ).strip().lower() in {"1", "true", "yes", "on"}
    quality_control = analyze_operational_quality(
        df, balance_squared=diagnostico['cuadra'],
        enforce_export=enforce_export,
    )
    st.session_state.setdefault("quality_controls", {})[archivo_nombre] = quality_control
    _mostrar_control_calidad_operativo(quality_control)

    emision = _control_emision(df, catalogo, diagnostico, quality_control)
    for reporte in reportes_periodo[1:]:
        if not reporte["diagnostico"]["cuadra"]:
            emision["definitivo"] = False
            emision["motivos"].append(
                f"El período {reporte['periodo']} no cuadra: diferencia "
                f"{reporte['diagnostico']['diferencia']:,.2f}"
            )
        problemas_resultado = reporte["conciliacion"].get("problemas", [])
        if problemas_resultado:
            emision["definitivo"] = False
            emision["motivos"].extend(
                f"Período {reporte['periodo']}: {problema}"
                for problema in problemas_resultado
            )
    amount_columns = [column for _, column in period_columns]
    income_available = {
        column: reporte["conciliacion"]["resultado_homologado"] is not None
        for reporte, (_, column) in zip(reportes_periodo, period_columns)
    }
    presentacion, formulas_reporte = complete_catalog(
        agrupado, catalogo, income_available, amount_columns=amount_columns,
    )
    _validar_cuadre_utilidad(emision['resultado'])
    total_cols = st.columns(3)
    for target, label, categories in (
        (total_cols[0], "Total activos", ("activo_corriente", "activo_no_corriente")),
        (total_cols[1], "Total pasivos", ("pasivo_corriente", "pasivo_no_corriente")),
        (total_cols[2], "Total patrimonio", ("patrimonio",)),
    ):
        total = presentacion[presentacion["categoria"].isin(categories)]["monto_total"].sum()
        target.metric(label, f"{total:,.2f}")
    st.caption("Se muestran todas las clasificaciones del catálogo. Cero indica que no hay importes asignados; no acredita por sí solo la integridad de la extracción.")
    if emision['definitivo']:
        st.success('Controles de emisión aprobados para este reporte.')
    else:
        st.error('Sólo se permite descargar un BORRADOR. Corrija los motivos antes de emitir el reporte definitivo.')
        for motivo in emision['motivos']:
            st.write(motivo)
        if not emision['incidencias'].empty:
            st.dataframe(emision['incidencias'], hide_index=True, use_container_width=True)
        st.caption('Pulse Revisar o cambiar clasificaciones, seleccione Todas y busque la cuenta. Puede modificar decisiones manuales y recuperar cuentas excluidas.')

    resultado_periodo_display = resultado_periodo or 0.0
    r1, r2, r3 = st.columns(3)
    ganancias = clasificadas[
        clasificadas['origen_columna_efectiva'] == 'ganancia'
    ]['monto'].abs().sum()
    perdidas = clasificadas[
        clasificadas['origen_columna_efectiva'] == 'perdida'
    ]['monto'].abs().sum()
    r1.metric("Ganancias", f"${ganancias:,.0f}")
    r2.metric("Pérdidas", f"${perdidas:,.0f}")
    r3.metric(
        "Resultado del período", f"${resultado_periodo_display:,.0f}",
        delta="Utilidad" if resultado_periodo_display >= 0 else "Pérdida",
    )
    st.divider()

    LABELS_CAT = {
        'activo_corriente': '🟦 Activo Corriente', 'activo_no_corriente': '🟦 Activo No Corriente',
        'pasivo_corriente': '🟥 Pasivo Corriente', 'pasivo_no_corriente': '🟥 Pasivo No Corriente',
        'patrimonio': '🟩 Patrimonio', 'resultado': '🟨 Estado de Resultados'
    }

    for cat in orden_cat:
        sub = presentacion[presentacion['categoria'] == cat]
        if sub.empty: continue
        st.subheader(LABELS_CAT.get(cat, cat))
        columnas_tabla = ['codigo_clasificado', 'nombre_estandar', *amount_columns, 'num_cuentas']
        tabla = sub[columnas_tabla].copy()
        nombres_montos = {column: f"Monto {label}" for label, column in period_columns}
        tabla = tabla.rename(columns={
            'codigo_clasificado': 'Código', 'nombre_estandar': 'Cuenta Estándar',
            'num_cuentas': '# Cuentas Agrupadas', **nombres_montos,
        })
        for nombre_monto in nombres_montos.values():
            tabla[nombre_monto] = tabla[nombre_monto].map(
                lambda value: f"{value:,.2f}" if pd.notna(value) else "No disponible"
            )
        st.dataframe(tabla, use_container_width=True, hide_index=True)
        if cat == 'resultado':
            st.caption('Ingresos positivos y gastos negativos. El resultado neto es un cálculo y no se suma nuevamente al detalle.')
        else:
            st.metric(f"Subtotal", f"{sub['monto_total'].sum():,.0f}")
        st.divider()

    # Ajuste Patrimonio Efectivo
    pat = agrupado[agrupado['categoria'] == 'patrimonio']
    ac06s = agrupado[agrupado['codigo_clasificado'] == 'AC.06S']
    if not pat.empty:
        monto_ac06s = ac06s['monto_total'].sum() if not ac06s.empty else 0.0
        ajuste = calcular_patrimonio_efectivo(dict(zip(pat['codigo_clasificado'], pat['monto_total'])), monto_ac06s)
        st.subheader("🎯 Patrimonio Efectivo")
        colp1, colp2, colp3 = st.columns(3)
        colp1.metric("Patrimonio contable", f"{ajuste['patrimonio_contable']:,.0f}")
        colp2.metric("Ajuste cta. socios", f"-{ajuste['ajuste_cta_socios']:,.0f}")
        colp3.metric("Patrimonio efectivo", f"{ajuste['patrimonio_efectivo']:,.0f}")

    # Detalle Excel Exporter
    import io
    from openpyxl.styles import Font, PatternFill
    
    export_df = presentacion[[
        "codigo_clasificado", "nombre_estandar", *amount_columns, "num_cuentas",
    ]].copy()
    meta = st.session_state.get('metadata_files', {}).get(archivo_nombre)
    unidad = (getattr(meta, 'moneda', None) or 'unidad original')
    columnas_exportacion = ["Código", "Cuenta Estándar"] + [
        f"Monto Total ({unidad})" if len(period_columns) == 1
        else f"Monto Total {label} ({unidad})"
        for label, _ in period_columns
    ] + ["# Cuentas Agrupadas"]
    export_df.columns = columnas_exportacion
    export_df['Tipo de fila'] = presentacion['codigo_clasificado'].map(
        lambda c: 'Calculado, no sumar al detalle' if catalogo.get(c, {}).get('clasificable') is False else 'Cuenta'
    )
    
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        FILA_INICIO_BALANCE = 7
        export_df.to_excel(writer, index=False, sheet_name="Balance Normalizado", startrow=FILA_INICIO_BALANCE - 1)
        _exportar_advertencias_auxiliares(
            writer, st.session_state.get("extraction_certifications", {}).get(archivo_nombre),
        )
        ws = writer.sheets["Balance Normalizado"]
        estado_emision = 'DEFINITIVO' if emision['definitivo'] else 'BORRADOR: REQUIERE REVISIÓN'
        ws['E1'] = 'Estado de emisión'
        ws['F1'] = estado_emision
        ws['F1'].font = Font(bold=True, color='008000' if emision['definitivo'] else 'C00000')
        control_filas = [
            {'Control': 'Estado de emisión', 'Valor': estado_emision},
            {'Control': 'Resultado según columnas originales', 'Valor': conciliacion['resultado_origen']},
            {'Control': 'Resultado según categorías homologadas', 'Valor': conciliacion['resultado_homologado']},
            {'Control': 'Diferencia de resultado', 'Valor': conciliacion['diferencia']},
            {'Control': 'Diferencia Activo menos Pasivo y Patrimonio', 'Valor': diagnostico['diferencia']},
        ] + [{'Control': 'Pendiente de resolver', 'Valor': m} for m in emision['motivos']]
        pd.DataFrame(control_filas).to_excel(writer, sheet_name='Control de emisión', index=False)
        if not emision['incidencias'].empty:
            emision['incidencias'].to_excel(writer, sheet_name='Cuentas a corregir', index=False)
        historial = [h for h in st.session_state.get('historial_decisiones', []) if h['Archivo'] == archivo_nombre]
        if historial:
            pd.DataFrame(historial).to_excel(writer, sheet_name='Decisiones de esta sesión', index=False)

        meta = st.session_state.get("metadata_files", {}).get(archivo_nombre)
        if meta:
            ws["A1"] = "Empresa:";    ws["B1"] = meta.razon_social or ""
            ws["A2"] = "RUT:";       ws["B2"] = meta.rut or ""
            ws["A3"] = "Período:";   ws["B3"] = f'{meta.periodo_desde or ""} al {meta.periodo_hasta or ""}'
            ws["A4"] = "Giro:";      ws["B4"] = meta.giro or ""
            ws["A5"] = "Moneda/unidad:"; ws["B5"] = meta.moneda or ""
            ws["A6"] = "Duración:";  ws["B6"] = (
                f"{meta.numero_meses} meses" if meta.numero_meses else ""
            )

        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 36
        ws.column_dimensions["C"].width = 22
        ws.column_dimensions["D"].width = 20

        AZUL = "1F4E79"; BLANCO = "FFFFFF"; GRIS = "F2F2F2"
        header_fill = PatternFill("solid", fgColor=AZUL)
        header_font = Font(bold=True, color=BLANCO, size=11)
        for cell in ws[FILA_INICIO_BALANCE]:
            cell.fill = header_fill
            cell.font = header_font

        for i, row in enumerate(ws.iter_rows(
                min_row=FILA_INICIO_BALANCE + 1,
                max_row=FILA_INICIO_BALANCE + len(export_df),
                max_col=len(export_df.columns)), start=0):
            if i % 2 == 0:
                for cell in row: cell.fill = PatternFill("solid", fgColor=GRIS)

        fila_sep = FILA_INICIO_BALANCE + len(export_df) + 3

        catalogo_local = catalogo
        detail_indices = pd.Index([])
        for reporte in reportes_periodo:
            detail_indices = detail_indices.union(reporte["clasificadas"].index)
        columnas_base = [
            'codigo_clasificado', 'codigo_original', 'nombre_original',
            'nombre_revision_usuario', 'origen_columna',
            'origen_columna_efectiva', 'metodo', 'confianza',
        ]
        det = df.loc[detail_indices, columnas_base].copy()

        det['nombre_visual'] = det['nombre_revision_usuario'].where(det['nombre_revision_usuario'] != '', det['nombre_original'])
        det['nombre_estandar'] = det['codigo_clasificado'].map(lambda c: catalogo_local.get(c, {}).get('nombre_estandar', c))
        columnas_detalle_periodo = []
        for posicion, periodo in enumerate(periodos):
            extraido = f"monto_extraido_{periodo}"
            normalizado = f"monto_normalizado_{periodo}"
            det[extraido] = df.loc[det.index].apply(
                lambda row, p=periodo, pos=posicion: _valor_fila_periodo(row, p, pos),
                axis=1,
            )
            det[normalizado] = det.apply(
                lambda row, col=extraido: _monto_presentacion(
                    row['codigo_clasificado'], row[col], row['nombre_visual'],
                    row.get('origen_columna'), catalogo,
                ) if pd.notna(row[col]) else None,
                axis=1,
            )
            columnas_detalle_periodo.extend([extraido, normalizado])
        detalle_completo = det[[
            'codigo_clasificado', 'nombre_estandar', 'codigo_original',
            'nombre_visual', 'origen_columna', 'origen_columna_efectiva',
            *columnas_detalle_periodo, 'metodo', 'confianza',
        ]].copy()

        # Reconstrucción de la lógica de ordenamiento nativo
        detalle_completo = detalle_completo.sort_values(
            ['codigo_clasificado', columnas_detalle_periodo[0]],
            key=lambda x: x.abs() if x.dtype.kind in 'fi' else x,
            ascending=[True, False]
        )
        nombres_detalle_periodo = []
        for periodo in periodos:
            if len(periodos) == 1:
                nombres_detalle_periodo.extend(['Monto Extraído', 'Monto Normalizado'])
            else:
                nombres_detalle_periodo.extend([
                    f'Monto Extraído {periodo}', f'Monto Normalizado {periodo}',
                ])
        detalle_completo.columns = [
            'Código Estándar', 'Nombre Estándar',
            'Cód. Original', 'Nombre', 'Columna Extraída', 'Naturaleza Efectiva',
            *nombres_detalle_periodo,
            'Método Clasificación', 'Confianza'
        ]
        detalle_completo['Confianza'] = detalle_completo['Confianza'].apply(
            lambda x: f"{x:.0%}" if pd.notna(x) else ""
        )

        ws.cell(row=fila_sep, column=1, value="APERTURA DE CUENTAS — DETALLE COMPLETO")
        title_cell = ws.cell(row=fila_sep, column=1)
        title_cell.font = Font(bold=True, size=12, color=AZUL)

        detalle_completo.to_excel(
            writer, index=False,
            sheet_name="Balance Normalizado",
            startrow=fila_sep
        )

        header_row_ap = fila_sep + 1
        naranja = "E26B0A"
        for cell in ws[header_row_ap]:
            cell.fill = PatternFill("solid", fgColor=naranja)
            cell.font = Font(bold=True, color=BLANCO, size=10)

        ws.column_dimensions["E"].width = 18
        ws.column_dimensions["F"].width = 20
        ws.column_dimensions["G"].width = 18
        ws.column_dimensions["H"].width = 20
        ws.column_dimensions["I"].width = 22
        ws.column_dimensions["J"].width = 12
        processed_at = st.session_state.setdefault("processed_at", {}).setdefault(
            archivo_nombre, datetime.now().astimezone().isoformat(),
        )
        add_report_sheets(
            writer.book, presentacion, formulas_reporte,
            account_detail=detalle_completo,
            start_row=FILA_INICIO_BALANCE, unit=unidad, meta=meta,
            processed_at=processed_at, source_name=archivo_nombre,
            pages=st.session_state.get("document_pages", {}).get(archivo_nombre),
            definitive=emision["definitivo"], reasons=emision["motivos"],
            tolerance=diagnostico["tolerancia"],
            period_columns=period_columns,
        )

    buf.seek(0)
    
    meta_state = meta
    razon_fn = (meta_state.razon_social or "empresa").replace(" ", "_")[:30] if meta_state else "empresa"
    rut_fn   = (meta_state.rut or "").replace(".", "").replace("-", "") if meta_state else ""
    nombre_archivo = f"Balance_Unificado-{razon_fn}-{rut_fn}"
    if not emision['definitivo']:
        nombre_archivo = 'BORRADOR-' + nombre_archivo

    st.download_button(
        'Descargar reporte definitivo (Excel)' if emision['definitivo'] else 'Descargar BORRADOR con diagnóstico (Excel)', data=buf.getvalue(),
        file_name=f"{nombre_archivo}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

def _validar_cuadre_utilidad(resultado):
    st.subheader('Conciliación independiente del estado de resultados')
    if resultado['resultado_origen'] is None and resultado['resultado_homologado'] is None:
        st.info('El documento no contiene detalle de ingresos y gastos para conciliar el resultado.')
        return
    c1, c2, c3 = st.columns(3)
    def mostrar(valor):
        return 'Sin detalle' if valor is None else f'{valor:,.2f}'
    c1.metric('Resultado según columnas originales', mostrar(resultado['resultado_origen']))
    c2.metric('Resultado según categorías homologadas', mostrar(resultado['resultado_homologado']))
    c3.metric('Diferencia del resultado', mostrar(resultado['diferencia']))
    if resultado['cuadra']:
        st.success('Las categorías de ingresos y gastos reproducen el resultado de las columnas originales.')
    else:
        st.error('Las categorías asignadas no reproducen el resultado de origen o falta detalle para comprobarlo. Revise las cuentas indicadas antes de emitir.')


def _tab_diccionario():
    busqueda = st.text_input("Buscar en el diccionario", "")
    catalogo_local = cargar_catalogo()
    dic = st.session_state.diccionario
    df_dic = pd.DataFrame(dic)
    
    df_dic['nombre_estandar'] = df_dic['codigo_estandar'].map(
        lambda c: catalogo_local.get(c, {}).get('nombre_estandar', '') if c else ''
    )
    df_dic['codigo_y_nombre'] = df_dic.apply(
        lambda r: f"{r['codigo_estandar']} — {r['nombre_estandar']}" if r['nombre_estandar']
        else r['codigo_estandar'], axis=1
    )
    if busqueda:
        mask = (
            df_dic['cuenta_original'].str.contains(busqueda, case=False, na=False) |
            df_dic['nombre_estandar'].str.contains(busqueda, case=False, na=False) |
            df_dic['codigo_estandar'].str.contains(busqueda, case=False, na=False)
        )
        df_dic = df_dic[mask]
    st.caption(f"{len(df_dic)} entradas encontradas")
    st.dataframe(
        df_dic[['cuenta_original', 'codigo_y_nombre', 'fuente']].rename(columns={
            'cuenta_original': 'Cuenta Original',
            'codigo_y_nombre': 'Código — Nombre Estándar',
            'fuente': 'Fuente'
        }),
        use_container_width=True, hide_index=True, height=500
    )


def _tab_aprendizaje():
    st.subheader("🧠 Autoaprendizaje — Gold Standard")
    store = NeonKnowledgeStore()
    if _neon_disponible():
        try:
            stats = store.learning_statistics()
            c1, c2, c3, c4 = st.columns(4)
            c1.metric("Diccionario activo", stats["dictionary_entries"])
            c2.metric("Aprendidas por humanos", stats["human_learned"])
            c3.metric("Validaciones registradas", stats["validations"])
            c4.metric("Correcciones", stats["corrections"])
            st.caption(
                "Fuente: Neon. Cada validación queda auditada y las entradas agregadas "
                "al diccionario son consumidas por el pipeline en futuras sesiones."
            )
            recientes = store.recent_validations(20)
            if recientes:
                st.markdown("#### Actividad reciente")
                st.dataframe(pd.DataFrame(recientes), use_container_width=True, hide_index=True)
            else:
                st.info("Neon está conectado; aún no hay validaciones registradas.")
            return
        except Exception:
            st.warning("No fue posible leer las métricas de Neon; se muestran datos locales.")
    try:
        builder = GoldBuilder()
        stats = builder.statistics()
        top = builder.top_learned()
        conflicts = builder.find_conflicts()
        builder.close()
    except Exception as e:
        st.error(f"No se pudieron cargar estadísticas: {e}")
        return

    c1, c2, c3 = st.columns(3)
    c1.metric("Registros aprendidos", stats["total_records"])
    c2.metric("Coincidencias exactas", stats["exact_hits"])
    c3.metric("Cuentas con conflicto", stats["conflicts"])

    if stats["total_records"] > 0:
        st.subheader("🏆 Top 20 cuentas más aprendidas")
        top_df = pd.DataFrame(top)
        top_df.columns = ["Cuenta", "Código Final", "Veces Usada", "Último Uso"]
        top_df["Último Uso"] = top_df["Último Uso"].str[:19]
        st.dataframe(top_df, use_container_width=True, hide_index=True)

    if conflicts:
        st.subheader("⚔️ Cuentas con conflicto (múltiples códigos asignados)")
        cf_df = pd.DataFrame(conflicts)
        cf_df.columns = ["Cuenta", "Códigos Distintos", "Códigos", "Versiones"]
        st.dataframe(cf_df, use_container_width=True, hide_index=True)
    else:
        st.info("No hay cuentas con conflictos de clasificación.")

    st.divider()
    st.subheader("🔄 Promoción al Gold Standard (Learning Loop)")
    st.caption(
        "Promueve las revisiones humanas (gold_records) a una base RUNTIME separada "
        "(`gold_standard_runtime.db`). La base del benchmark (2660/2662) NO se modifica. "
        "Ver `reports/product/P1_learning_loop_design.md`."
    )
    pc1, pc2, pc3 = st.columns(3)
    with pc1:
        if st.button("🔍 Previsualizar promoción", use_container_width=True):
            try:
                res = promover_revisiones(dry_run=True)
                st.success(
                    f"Candidatos: {res.candidates} · Promovibles: {res.promotable} · "
                    f"Conflictos: {res.conflicts} · Duplicados: {res.duplicates} · "
                    f"Reservados (total): {res.reserved}"
                )
                if res.conflict_details:
                    st.warning("Conflictos (se omiten de la promoción):")
                    for c in res.conflict_details:
                        st.write(f"- {c['name']}: candidato={c['candidate_code']} gold={c['existing_codes']}")
            except Exception as e:
                st.error(f"No se pudo previsualizar: {e}")
    with pc2:
        if st.button("✅ Aplicar a runtime", use_container_width=True):
            try:
                res = promover_revisiones(dry_run=False)
                st.success(
                    f"Promovidos: {res.promoted} · Conflictos omitidos: {res.conflicts} · "
                    f"Duplicados: {res.duplicates} · Reservados: {res.reserved}"
                )
            except Exception as e:
                st.error(f"No se pudo aplicar: {e}")
    with pc3:
        if st.button("🧹 Borrar runtime", use_container_width=True):
            try:
                rt = RuntimeGoldStorage()
                rt.close()
                rt.path.unlink(missing_ok=True)
                st.success("Runtime eliminado. El benchmark sigue intacto.")
            except Exception as e:
                st.error(f"No se pudo borrar: {e}")


def _km_usuario() -> str:
    return str(st.session_state.get("km_usuario", "analista"))


def _tab_knowledge_manager() -> None:
    """🧠 Knowledge Manager (P5).

    Administra EXCLUSIVAMENTE ``gold_standard_runtime.db`` a través de
    ``RuntimeManager`` (única capa de persistencia). La UI no ejecuta SQL ni
    abre conexiones SQLite. El benchmark (``gold_standard.db``) nunca se escribe.
    Toda acción (promover, rechazar, rollback) requiere aprobación explícita.
    """
    st.subheader("🧠 Knowledge Manager")
    if _neon_disponible():
        _tab_neon_knowledge_manager()
        return
    st.caption(
        "Administra exclusivamente `gold_standard_runtime.db` vía `RuntimeManager`. "
        "El benchmark (`gold_standard.db`) permanece intacto. Eventos auditables: "
        "PROMOTE / ROLLBACK / REJECT. Estado del candidato: PENDING / APPROVED / "
        "REJECTED / ROLLED_BACK."
    )

    runtime_path = BASE_DIR / "gold_standard_runtime.db"
    gold_path = BASE_DIR / "gold_standard.db"
    rm = RuntimeManager(runtime_path)

    tab_pend, tab_conf, tab_run, tab_hist, tab_stat, tab_an = st.tabs(
        ["📥 Promociones pendientes", "⚔️ Conflictos", "🗃️ Runtime",
         "📜 Historial", "📊 Estadísticas", "📊 Runtime Analytics"]
    )

    with tab_pend:
        _km_pendientes(rm, gold_path)
    with tab_conf:
        _km_conflictos(rm, gold_path)
    with tab_run:
        _km_runtime(rm)
    with tab_hist:
        _km_historial(rm)
    with tab_stat:
        _km_estadisticas(rm, gold_path)
    with tab_an:
        _km_runtime_analytics(rm, gold_path)


def _tab_neon_knowledge_manager() -> None:
    """Gobernanza durable del diccionario cuando Neon es la fuente activa."""
    store = NeonKnowledgeStore()
    st.caption(
        "Fuente durable: Neon. El rollback solo se permite si el cambio elegido "
        "continúa siendo el estado vigente de la cuenta."
    )
    tab_hist, tab_conf, tab_stats = st.tabs(
        ["📜 Historial", "⚔️ Conflictos", "📊 Estadísticas"]
    )
    with tab_hist:
        history = store.dictionary_history(100)
        if not history:
            st.info("Aún no hay cambios humanos en el diccionario Neon.")
        else:
            st.dataframe(pd.DataFrame(history), use_container_width=True, hide_index=True)
            options = {
                f"#{row['id']} · {row['cuenta_original']} · "
                f"{row['codigo_anterior'] or 'sin entrada'} → {row['codigo_nuevo'] or 'inactivo'}": row
                for row in history if row["accion"] != "ROLLBACK"
            }
            if options:
                selected = st.selectbox("Cambio a revertir", list(options))
                confirm = st.checkbox(
                    "Confirmo que deseo revertir este cambio vigente",
                    key="neon_rollback_confirm",
                )
                if st.button("↩️ Ejecutar rollback", disabled=not confirm):
                    row = options[selected]
                    if store.rollback_dictionary_change(row["id"], reviewer=_km_usuario()):
                        cargar_diccionario_base.clear()
                        st.success("Rollback aplicado y registrado en el historial.")
                        st.rerun()
                    else:
                        st.error(
                            "No se aplicó: el cambio ya no es el estado vigente o no existe."
                        )
    with tab_conf:
        conflicts = store.conflicts()
        if conflicts:
            st.warning(f"Se detectaron {len(conflicts)} cuentas con códigos históricos distintos.")
            st.dataframe(pd.DataFrame(conflicts), use_container_width=True, hide_index=True)
        else:
            st.success("No hay conflictos históricos registrados en Neon.")
    with tab_stats:
        stats = store.learning_statistics()
        cols = st.columns(5)
        labels = [
            ("Catálogo", "catalog_entries"), ("Diccionario", "dictionary_entries"),
            ("Aprendidas", "human_learned"), ("Validaciones", "validations"),
            ("Correcciones", "corrections"),
        ]
        for col, (label, key) in zip(cols, labels):
            col.metric(label, stats[key])


def _km_pendientes(rm: RuntimeManager, gold_path: Path) -> None:
    """TAB 1 — Promociones pendientes: aprobar/rechazar selección múltiple."""
    st.markdown("#### 📥 Promociones pendientes")
    try:
        pend = rm.get_pending_promotions(gold_path)
    except Exception as e:  # noqa: BLE001
        st.error(f"No se pudieron cargar los candidatos: {e}")
        return

    if not pend:
        st.info("No hay candidatos pendientes de promoción en `gold_records`.")
        return

    df = pd.DataFrame([{
        "Cuenta": p["account_name"],
        "Código sugerido": p["candidate_code"],
        "Origen": p["origen"] or "—",
        "Confianza": p["confidence"],
        "Fecha": (p["fecha"] or "")[:19],
        "Estado": p["state"],
        "Clasificación": p["status"],
        "source_record_id": p["source_record_id"],
    } for p in pend])
    df.insert(0, "Seleccionar", False)

    st.caption(
        "Estado actual derivado de `promotion_history`. Solo los candidatos "
        "PENDING se promueven; duplicados y conflictos se omiten automáticamente."
    )

    edited = st.data_editor(
        df,
        hide_index=True,
        disabled=[c for c in df.columns if c not in ("Seleccionar",)],
        use_container_width=True,
        key="km_pend_editor",
        column_config={
            "Seleccionar": st.column_config.CheckboxColumn("Seleccionar", default=False),
            "Confianza": st.column_config.NumberColumn("Confianza", format="%.2f"),
            "source_record_id": None,
        },
    )

    if not isinstance(edited, pd.DataFrame):
        edited = df
    seleccionadas = edited[edited["Seleccionar"] == True] if "Seleccionar" in edited.columns else edited.iloc[0:0]  # noqa: E712
    ids = [int(r["source_record_id"]) for _, r in seleccionadas.iterrows()]

    c1, c2, c3 = st.columns([1, 1, 1])
    with c1:
        aprobar = st.button("✅ Aprobar selección", use_container_width=True)
    with c2:
        rechazar = st.button("❌ Rechazar selección", use_container_width=True)
    with c3:
        st.caption(f"{len(ids)} seleccionadas · usuario: `{_km_usuario()}`")

    if aprobar:
        if not ids:
            st.warning("Selecciona al menos una fila para aprobar.")
        else:
            try:
                res = rm.promote(gold_path, dry_run=False, source_ids=ids, usuario=_km_usuario())
                st.success(
                    f"Promovidos: {res.promoted} · Conflictos omitidos: {res.conflicts} · "
                    f"Duplicados: {res.duplicates} · Reservados: {res.reserved}"
                )
            except Exception as e:  # noqa: BLE001
                st.error(f"No se pudo aprobar la selección: {e}")
    if rechazar:
        if not ids:
            st.warning("Selecciona al menos una fila para rechazar.")
        else:
            try:
                n = 0
                for _, r in seleccionadas.iterrows():
                    rm.reject_promotion(
                        source_record_id=int(r["source_record_id"]),
                        account_name=str(r["Cuenta"]),
                        candidate_code=str(r["Código sugerido"]),
                        usuario=_km_usuario(),
                    )
                    n += 1
                st.success(f"Rechazos registrados (REJECT): {n}")
            except Exception as e:  # noqa: BLE001
                st.error(f"No se pudieron registrar los rechazos: {e}")


def _km_conflictos(rm: RuntimeManager, gold_path: Path) -> None:
    """TAB 2 — Conflictos runtime vs gold."""
    st.markdown("#### ⚔️ Conflictos (runtime vs gold)")
    try:
        conf = rm.get_conflicts(gold_path)
    except Exception as e:  # noqa: BLE001
        st.error(f"No se pudieron cargar los conflictos: {e}")
        return

    if not conf:
        st.info("No hay conflictos entre runtime y gold oficial.")
        return

    st.caption(
        "Misma cuenta normalizada con código distinto en runtime y gold. "
        "Resolver: mantener runtime (no hacer nada) o rollback (prevalece gold). "
        "`gold_standard.db` nunca se modifica."
    )
    df = pd.DataFrame([{
        "Cuenta": c["account_name"],
        "Código runtime": c["codigo_runtime"],
        "Código gold": c["codigo_gold"],
    } for c in conf])
    st.dataframe(df, use_container_width=True, hide_index=True)

    for c in conf:
        with st.expander(
            f"⚔️ {c['account_name']} — runtime {c['codigo_runtime']} vs gold {c['codigo_gold']}"
        ):
            hist = rm.get_history(account_name=c["account_name"], limit=20)
            if hist:
                st.dataframe(pd.DataFrame([{
                    "Fecha": (h["fecha"] or "")[:19],
                    "Usuario": h["usuario"],
                    "Acción": h["accion"],
                    "Estado": h["state"],
                    "Código": h["codigo_nuevo"] or h["codigo_anterior"] or "",
                    "Comentario": h["comentario"],
                } for h in hist]), use_container_width=True, hide_index=True)
            else:
                st.caption("Sin historial para esta cuenta.")
            c1, c2 = st.columns(2)
            with c1:
                if st.button("🔁 Rollback (prevalece gold)", key=f"km_rb_{c['runtime_id']}"):
                    try:
                        ok = rm.rollback(
                            int(c["runtime_id"]), usuario=_km_usuario(),
                            comentario="Conflicto resuelto: prevalece gold",
                        )
                        st.success("Rollback aplicado y registrado (ROLLBACK)." if ok else "No se pudo aplicar el rollback.")
                    except Exception as e:  # noqa: BLE001
                        st.error(f"Error en rollback: {e}")
            with c2:
                if st.button("✓ Mantener runtime", key=f"km_keep_{c['runtime_id']}"):
                    st.info("Runtime mantenido (no se modifica nada).")


def _km_runtime(rm: RuntimeManager) -> None:
    """TAB 3 — Cuentas activas del runtime: filtro, buscador y rollback."""
    st.markdown("#### 🗃️ Runtime (`gold_standard_runtime.db`)")
    if not rm.path.exists():
        st.info("El runtime aún no existe. Promueve candidatos desde la pestaña 'Promociones pendientes'.")
        return
    try:
        rows = rm.load_runtime()
    except Exception as e:  # noqa: BLE001
        st.error(f"No se pudieron cargar las cuentas del runtime: {e}")
        return
    if not rows:
        st.info("El runtime existe pero no tiene cuentas activas.")
        return

    df = pd.DataFrame([{
        "id": r["id"],
        "Cuenta": r["nombre_cuenta"],
        "Código": r["codigo_estandar"],
        "Normalizado": r["normalized"],
        "Reviewer": r["reviewer"] or "—",
        "Fecha": (r["promoted_at"] or "")[:19],
    } for r in rows])

    q = st.text_input("🔍 Buscar cuenta o código", key="km_runtime_q")
    if q:
        ql = q.lower().strip()
        df = df[df["Cuenta"].str.lower().str.contains(ql)
                | df["Código"].str.lower().str.contains(ql)
                | df["Normalizado"].str.lower().str.contains(ql)]

    prefijos = sorted({str(c).split(".")[0] for c in df["Código"]})
    filtro = st.selectbox("Filtrar por prefijo de código", ["todos"] + prefijos, key="km_runtime_filtro")
    if filtro != "todos":
        df = df[df["Código"].astype(str).str.startswith(filtro)]

    st.caption(f"{len(df)} cuentas activas en runtime.")
    st.dataframe(df, use_container_width=True, hide_index=True)

    st.markdown("**Rollback / Eliminar**")
    st.caption("Rollback elimina la entrada de runtime y registra el evento ROLLBACK (prevalece gold).")
    options = {f"{r['Cuenta']} ({r['Código']})": int(r["id"]) for _, r in df.iterrows()}
    if options:
        sel_label = st.selectbox("Cuenta", list(options), key="km_runtime_sel")
        confirm = st.checkbox("Confirmo el rollback de esta cuenta", key="km_runtime_confirm")
        if st.button("🗑️ Rollback", use_container_width=True, disabled=not confirm):
            try:
                ok = rm.rollback(options[sel_label], usuario=_km_usuario(), comentario="Rollback desde Knowledge Manager")
                st.success("Rollback aplicado y registrado." if ok else "No se pudo aplicar.")
            except Exception as e:  # noqa: BLE001
                st.error(f"Error en rollback: {e}")


def _km_historial(rm: RuntimeManager) -> None:
    """TAB 4 — Historial de promotion_history (solo lectura)."""
    st.markdown("#### 📜 Historial (`promotion_history`)")
    try:
        hist = rm.get_history(limit=500)
    except Exception as e:  # noqa: BLE001
        st.error(f"No se pudo leer el historial: {e}")
        return
    if not hist:
        st.info("Sin eventos registrados en `promotion_history`.")
        return

    df = pd.DataFrame([{
        "Promotion ID": h["promotion_id"][:8],
        "Fecha": (h["fecha"] or "")[:19],
        "Usuario": h["usuario"],
        "Acción": h["accion"],
        "Estado": h["state"],
        "Cuenta": h["account_name"],
        "Código": h["codigo_nuevo"] or h["codigo_anterior"] or "",
        "Comentario": h["comentario"],
    } for h in hist])

    q = st.text_input("🔍 Buscar en historial", key="km_hist_q")
    if q:
        ql = q.lower()
        df = df[df.apply(lambda r: ql in " ".join(str(v) for v in r), axis=1)]

    st.caption("Solo lectura. El historial nunca se borra.")
    st.dataframe(df, use_container_width=True, hide_index=True)


def _km_estadisticas(rm: RuntimeManager, gold_path: Path) -> None:
    """TAB 5 — Estadísticas del runtime y del historial."""
    st.markdown("#### 📊 Estadísticas")
    try:
        s = rm.get_runtime_statistics(gold_path)
    except Exception as e:  # noqa: BLE001
        st.error(f"No se pudieron calcular las estadísticas: {e}")
        return

    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Runtime size", s["runtime_size"])
    c2.metric("Promociones (PROMOTE)", s["promotions"])
    c3.metric("Rechazos (REJECT)", s["rejects"])
    c4.metric("Rollback (ROLLBACK)", s["rollbacks"])
    c5.metric("Cobertura", f"{s['coverage']}%")

    st.caption(
        f"Gold oficial: {s['gold_size']} cuentas · Eventos en historial: "
        f"{s['history_events']} · Runtime: {s['runtime_size']} cuentas activas."
    )


def _km_runtime_analytics(rm: RuntimeManager, gold_path: Path) -> None:
    """TAB 6 — 📊 Runtime Analytics: observabilidad completa (P5.5).

    Toda la información sale de ``RuntimeStatistics`` (fuente única); la UI no
    ejecuta SQL ni abre SQLite. Las métricas de uso provienen del último
    procesamiento (``session_state``) y los eventos/cobertura de RuntimeManager.
    No promueve, no revierte, no puebla: solo observa.
    """
    st.markdown("#### 📊 Runtime Analytics")
    st.caption(
        "Observabilidad del runtime (solo lectura). Métricas de uso del último "
        "procesamiento + eventos/cobertura de `gold_standard_runtime.db`. "
        "El benchmark (`gold_standard.db`) no se modifica."
    )

    snapshot = st.session_state.get("runtime_metrics_last")
    try:
        stats = RuntimeStatistics.capture(
            metrics=snapshot,
            runtime=rm,
            gold_db=str(gold_path),
        )
    except Exception as e:  # noqa: BLE001
        st.error(f"No se pudieron calcular las métricas runtime: {e}")
        return

    c1, c2, c3, c4, c5, c6 = st.columns(6)
    c1.metric("💠 Entradas runtime", stats.runtime_size)
    c2.metric("🚀 Promociones", stats.promotion_count)
    c3.metric("🔁 Rollbacks", stats.rollback_count)
    c4.metric("❌ Rejects", stats.reject_count)
    c5.metric("🎯 Runtime exact", stats.runtime_exact_hits)
    c6.metric("🌀 Runtime fuzzy", stats.runtime_fuzzy_hits)

    c7, c8, c9, c10, c11, c12 = st.columns(6)
    c7.metric("🟢 Gold exact", stats.gold_exact_hits)
    c8.metric("🌫️ Gold fuzzy", stats.gold_fuzzy_hits)
    c9.metric("⬇️ Fallbacks", stats.fallback_to_gold)
    c10.metric("Runtime miss", stats.runtime_miss)
    c11.metric("Requests", stats.total_requests)
    c12.metric("Historial", stats.history_events)

    st.markdown("#### Cobertura")
    c13, c14, c15, c16, c17 = st.columns(5)
    c13.metric("📚 Cobertura runtime (uso)", f"{stats.runtime_usage_pct}%")
    c14.metric("🏛️ Cobertura gold (uso)", f"{stats.gold_usage_pct}%")
    c15.metric("🧠 Aprendizaje usado", f"{stats.learning_used_pct}%")
    c16.metric("📐 Catálogo runtime", f"{stats.runtime_catalog_coverage_pct}%")
    c17.metric("⬇️ Fallback", f"{stats.fallback_pct}%")

    st.markdown("#### Impacto por promoción")
    st.caption(
        "¿Qué promociones realmente generan impacto? Promoción cuyo código "
        "estándar fue usado por el runtime (hit exacto/fuzzy) durante el "
        "último procesamiento."
    )
    if not stats.promotion_impact:
        st.info("Sin promociones registradas o sin actividad de uso todavía.")
    else:
        df = pd.DataFrame(stats.promotion_impact)
        df["impacto"] = df["impactful"].map({True: "✅ Impacto", False: "—"})
        st.dataframe(
            df[["account_name", "code", "hits", "impacto", "promotion_id"]],
            use_container_width=True, hide_index=True,
        )

    if stats.promotion_count:
        st.caption(
            f"Promociones con impacto real: **{stats.impactful_promotions}** "
            f"de {len(stats.promotion_impact)} promociones registradas. "
            "Promociones sin hits pendientes de adopción."
        )


def _mostrar_resumen_catalogo(catalogo: dict):
    st.subheader("Catálogo Maestro de Homologación")
    df = pd.DataFrame.from_dict(catalogo, orient='index')
    df = df[['codigo_estandar', 'nombre_estandar', 'categoria', 'tipo_estado']]
    df.columns = ['Código', 'Nombre Estándar', 'Categoría', 'Tipo Estado']
    st.dataframe(df, use_container_width=True, hide_index=True)


# ─────────────────────────────────────────────────────────────────────────────
# INVOCACIÓN RESTRINGIDA (ESCUDO DE EJECUCIÓN GLOBAL)
# ─────────────────────────────────────────────────────────────────────────────

class TaxFolder:
    """Clase de marcador de posición para evitar que el orquestador falle al importar"""
    pass
    
if __name__ == '__main__':
    main()
