import json
import time
from collections import Counter
from pathlib import Path
from typing import Any

from parser_quality.gatekeeper import GatekeeperResult
from parser_quality.rejection_reasons import RejectionReason


class ParserQualityReport:
    """Genera todos los reportes de calidad del parser."""

    def __init__(self, output_dir: str | Path = "reports/parser_quality") -> None:
        self._output_dir = Path(output_dir)

    def generate(
        self,
        results: list[GatekeeperResult],
        ocr_count: int = 0,
        native_count: int = 0,
        total_docs: int = 0,
    ) -> dict[str, Any]:
        self._output_dir.mkdir(parents=True, exist_ok=True)
        metrics = self._compute_metrics(results, ocr_count, native_count, total_docs)

        self._save_statistics(metrics)
        self._save_quality_report(metrics, results)
        self._save_quality_xlsx(results)
        self._save_rejections_xlsx(results)
        self._save_reviews_xlsx(results)
        self._save_confidence_distribution_xlsx(results)
        self._save_examples_xlsx(results)

        return metrics

    def _compute_metrics(
        self,
        results: list[GatekeeperResult],
        ocr_count: int,
        native_count: int,
        total_docs: int,
    ) -> dict[str, Any]:
        total = len(results)
        accepted = [r for r in results if r.status == "ACCEPT"]
        reviews = [r for r in results if r.status == "REVIEW"]
        rejected = [r for r in results if r.status == "REJECT"]

        # Top rejection reasons
        all_reasons: Counter[str] = Counter()
        for r in rejected:
            for reason in r.reasons:
                all_reasons[str(reason)] += 1

        # Top docs by rejections
        doc_rejections: Counter[str] = Counter()
        for r in rejected:
            doc_rejections[r.file_name] += 1

        # Top docs by lowest avg confidence
        doc_confidences: dict[str, list[float]] = {}
        for r in results:
            doc_confidences.setdefault(r.file_name, []).append(r.confidence)
        doc_avg_conf = {
            doc: sum(confs) / len(confs)
            for doc, confs in doc_confidences.items()
        }
        lowest_conf_docs = sorted(doc_avg_conf.items(), key=lambda x: x[1])[:20]

        # Confidence distribution
        bins = [(0.0, 0.25), (0.25, 0.50), (0.50, 0.70), (0.70, 0.85), (0.85, 1.0)]
        dist = {
            f"{l:.2f}-{u:.2f}": sum(1 for r in results if l <= r.confidence < u)
            for l, u in bins
        }
        dist["1.00"] = sum(1 for r in results if r.confidence == 1.0)

        # OCR vs native
        ocr_results = [r for r in results if r.requirio_ocr]
        native_results = [r for r in results if not r.requirio_ocr]

        metrics = {
            "total_candidates": total,
            "total_documents": total_docs,
            "accepted": len(accepted),
            "accepted_pct": round(len(accepted) / total * 100, 2) if total else 0,
            "review": len(reviews),
            "review_pct": round(len(reviews) / total * 100, 2) if total else 0,
            "rejected": len(rejected),
            "rejected_pct": round(len(rejected) / total * 100, 2) if total else 0,
            "avg_confidence": round(
                sum(r.confidence for r in results) / total, 4
            ) if total else 0,
            "top_rejection_reasons": all_reasons.most_common(20),
            "top_docs_by_rejections": doc_rejections.most_common(20),
            "lowest_confidence_docs": [
                (doc, round(conf, 4)) for doc, conf in lowest_conf_docs
            ],
            "confidence_distribution": dist,
            "ocr": {
                "count": ocr_count,
                "results": len(ocr_results),
                "avg_confidence": round(
                    sum(r.confidence for r in ocr_results) / len(ocr_results), 4
                ) if ocr_results else 0,
            },
            "native": {
                "count": native_count,
                "results": len(native_results),
                "avg_confidence": round(
                    sum(r.confidence for r in native_results) / len(native_results), 4
                ) if native_results else 0,
            },
        }
        return metrics

    def _save_statistics(self, metrics: dict[str, Any]) -> None:
        path = self._output_dir / "parser_statistics.json"
        path.write_text(json.dumps(metrics, indent=2, ensure_ascii=False, default=str))
        print(f"  Guardado: {path}")

    def _save_quality_report(
        self, metrics: dict[str, Any], results: list[GatekeeperResult]
    ) -> None:
        lines: list[str] = []
        lines.append("# Reporte de Calidad del Parser — Parser Gatekeeper")
        lines.append("")
        lines.append(f"**Fecha:** {time.strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append("")
        lines.append("---")
        lines.append("")
        lines.append("## Resumen General")
        lines.append("")
        lines.append(f"- **Documentos procesados:** {metrics['total_documents']}")
        lines.append(f"- **Candidatos evaluados:** {metrics['total_candidates']}")
        lines.append(f"- **Confianza promedio:** {metrics['avg_confidence']:.3f}")
        lines.append("")
        lines.append("## Decisiones del Gatekeeper")
        lines.append("")
        lines.append("| Decisión | Cuentas | % |")
        lines.append("|----------|---------|---|")
        a, rv, rj = metrics['accepted'], metrics['review'], metrics['rejected']
        at, rt, rjt = metrics['accepted_pct'], metrics['review_pct'], metrics['rejected_pct']
        lines.append(f"| ✅ ACCEPT | {a} | {at:.1f}% |")
        lines.append(f"| ⚠️ REVIEW | {rv} | {rt:.1f}% |")
        lines.append(f"| ❌ REJECT | {rj} | {rjt:.1f}% |")
        lines.append("")
        lines.append("## Confianza por Tipo de Documento")
        lines.append("")
        lines.append("| Tipo | Documentos | Candidatos | Confianza Promedio |")
        lines.append("|------|------------|------------|--------------------|")
        lines.append(f"| 📄 Texto nativo | {metrics['native']['count']} | {metrics['native']['results']} | {metrics['native']['avg_confidence']:.3f} |")
        lines.append(f"| 🖼️ OCR | {metrics['ocr']['count']} | {metrics['ocr']['results']} | {metrics['ocr']['avg_confidence']:.3f} |")
        lines.append("")
        lines.append("## Top 10 Razones de Rechazo")
        lines.append("")
        lines.append("| # | Razón | Ocurrencias |")
        lines.append("|---|-------|-------------|")
        for i, (reason, count) in enumerate(metrics['top_rejection_reasons'][:10], 1):
            lines.append(f"| {i} | `{reason}` | {count} |")
        lines.append("")
        lines.append("## Top 10 Documentos con Más Rechazos")
        lines.append("")
        lines.append("| # | Documento | Rechazos |")
        lines.append("|---|----------|----------|")
        for i, (doc, cnt) in enumerate(metrics['top_docs_by_rejections'][:10], 1):
            lines.append(f"| {i} | {doc} | {cnt} |")
        lines.append("")
        lines.append("## Top 10 Documentos con Menor Confianza Promedio")
        lines.append("")
        lines.append("| # | Documento | Confianza Prom. |")
        lines.append("|---|-----------|-----------------|")
        for i, (doc, conf) in enumerate(metrics['lowest_confidence_docs'][:10], 1):
            lines.append(f"| {i} | {doc} | {conf:.3f} |")
        lines.append("")
        lines.append("## Distribución de Confianza")
        lines.append("")
        lines.append("| Rango | Cuentas |")
        lines.append("|-------|---------|")
        for rng, cnt in sorted(metrics['confidence_distribution'].items()):
            bar = "█" * max(1, cnt // max(1, max(metrics['confidence_distribution'].values()) // 40))
            lines.append(f"| {rng} | {cnt} {bar}")
        lines.append("")

        path = self._output_dir / "parser_quality.md"
        path.write_text("\n".join(lines))
        print(f"  Guardado: {path}")

    def _save_quality_xlsx(self, results: list[GatekeeperResult]) -> None:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ParserQuality"

        headers = ["Archivo", "Línea", "Nombre", "Código", "Monto",
                    "Confianza", "Estado", "Razones", "OCR"]
        hfont = Font(bold=True, color="FFFFFF")
        hfill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = hfont
            cell.fill = hfill

        accept_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        review_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        reject_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        for r, res in enumerate(results, 2):
            ws.cell(row=r, column=1, value=res.file_name)
            ws.cell(row=r, column=2, value=res.line_number)
            ws.cell(row=r, column=3, value=res.account_name)
            ws.cell(row=r, column=4, value=res.account_code or "")
            ws.cell(row=r, column=5, value=res.monto)
            ws.cell(row=r, column=6, value=res.confidence)
            ws.cell(row=r, column=7, value=res.status)
            ws.cell(row=r, column=8, value="; ".join(str(r) for r in res.reasons))
            ws.cell(row=r, column=9, value="SÍ" if res.requirio_ocr else "no")

            fill = accept_fill if res.status == "ACCEPT" else (
                review_fill if res.status == "REVIEW" else reject_fill
            )
            for c in range(1, 10):
                ws.cell(row=r, column=c).fill = fill

        for c in range(1, len(headers) + 1):
            max_len = max(len(str(ws.cell(row=r, column=c).value or "")) for r in range(1, len(results) + 2))
            ws.column_dimensions[chr(64 + c) if c <= 26 else "A"].width = min(max_len + 3, 60)

        path = self._output_dir / "parser_quality.xlsx"
        wb.save(path)
        print(f"  Guardado: {path}")

    def _save_rejections_xlsx(self, results: list[GatekeeperResult]) -> None:
        rejected = [r for r in results if r.status == "REJECT"]
        if not rejected:
            return
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Rejections"
        headers = ["Archivo", "Línea", "Nombre", "Confianza", "Razones"]
        for c, h in enumerate(headers, 1):
            ws.cell(row=1, column=c, value=h).font = openpyxl.styles.Font(bold=True)
        for r, res in enumerate(rejected, 2):
            ws.cell(row=r, column=1, value=res.file_name)
            ws.cell(row=r, column=2, value=res.line_number)
            ws.cell(row=r, column=3, value=res.account_name)
            ws.cell(row=r, column=4, value=res.confidence)
            ws.cell(row=r, column=5, value="; ".join(str(rr) for rr in res.reasons))
        path = self._output_dir / "parser_rejections.xlsx"
        wb.save(path)
        print(f"  Guardado: {path}")

    def _save_reviews_xlsx(self, results: list[GatekeeperResult]) -> None:
        reviews = [r for r in results if r.status == "REVIEW"]
        if not reviews:
            return
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reviews"
        headers = ["Archivo", "Línea", "Nombre", "Código", "Confianza", "Razones"]
        for c, h in enumerate(headers, 1):
            ws.cell(row=1, column=c, value=h).font = openpyxl.styles.Font(bold=True)
        for r, res in enumerate(reviews, 2):
            ws.cell(row=r, column=1, value=res.file_name)
            ws.cell(row=r, column=2, value=res.line_number)
            ws.cell(row=r, column=3, value=res.account_name)
            ws.cell(row=r, column=4, value=res.account_code or "")
            ws.cell(row=r, column=5, value=res.confidence)
            ws.cell(row=r, column=6, value="; ".join(str(rr) for rr in res.reasons))
        path = self._output_dir / "parser_reviews.xlsx"
        wb.save(path)
        print(f"  Guardado: {path}")

    def _save_confidence_distribution_xlsx(self, results: list[GatekeeperResult]) -> None:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ConfidenceDist"
        headers = ["Rango", "Cuentas"]
        for c, h in enumerate(headers, 1):
            ws.cell(row=1, column=c, value=h).font = openpyxl.styles.Font(bold=True)
        bins = [(0.0, 0.25), (0.25, 0.50), (0.50, 0.70), (0.70, 0.85), (0.85, 1.0), (1.0, 1.01)]
        for r, (l, u) in enumerate(bins, 2):
            label = f"{l:.2f}-{u:.2f}" if u <= 1.0 else "1.00"
            cnt = sum(1 for res in results if l <= res.confidence < u)
            ws.cell(row=r, column=1, value=label)
            ws.cell(row=r, column=2, value=cnt)
        path = self._output_dir / "parser_confidence_distribution.xlsx"
        wb.save(path)
        print(f"  Guardado: {path}")

    def _save_examples_xlsx(self, results: list[GatekeeperResult]) -> None:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Examples"

        accepted = [r for r in results if r.status == "ACCEPT"]
        rejected = [r for r in results if r.status == "REJECT"]
        reviews = [r for r in results if r.status == "REVIEW"]

        headers = ["Tipo", "Archivo", "Nombre", "Confianza", "Razones"]
        for c, h in enumerate(headers, 1):
            ws.cell(row=1, column=c, value=h).font = openpyxl.styles.Font(bold=True)

        row = 2
        for label, group in [("ACCEPT", accepted[:20]), ("REVIEW", reviews[:20]),
                              ("REJECT", rejected[:20])]:
            for res in group:
                ws.cell(row=row, column=1, value=label)
                ws.cell(row=row, column=2, value=res.file_name)
                ws.cell(row=row, column=3, value=res.account_name)
                ws.cell(row=row, column=4, value=res.confidence)
                ws.cell(row=row, column=5, value="; ".join(str(r) for r in res.reasons))
                row += 1

        path = self._output_dir / "parser_examples.xlsx"
        wb.save(path)
        print(f"  Guardado: {path}")
