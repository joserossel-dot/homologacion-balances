"""Presentación completa del catálogo. No modifica las cuentas ni su clasificación."""
import pandas as pd

CATEGORY_ORDER = (
    "activo_corriente", "activo_no_corriente", "pasivo_corriente",
    "pasivo_no_corriente", "patrimonio", "resultado",
)
CATEGORY_LABELS = dict(zip(CATEGORY_ORDER, (
    "Activo corriente", "Activo no corriente", "Pasivo corriente",
    "Pasivo no corriente", "Patrimonio", "Estado de resultados",
)))
ER_ORDER = (
    "ER.01", "ER.02", "ER.03", "ER.04", "ER.05", "ER.06", "ER.07",
    "ER.08", "ER.12", "ER.09", "ER.13", "ER.14", "ER.15", "ER.16",
    "ER.17", "ER.18", "ER.19", "ER.10", "ER.11", "ER.20", "ER.21",
)
CALCULATED = {"ER.03", "ER.06", "ER.08", "ER.19", "ER.11"}


def complete_catalog(grouped, catalog, has_income_detail, amount_columns=None):
    amount_columns = list(amount_columns or ["monto_total"])
    existing = {r["codigo_clasificado"]: r for r in grouped.to_dict("records")}
    rows = []
    for code in dict.fromkeys([*catalog, *existing]):
        entry = catalog.get(code, {})
        row = dict(existing.get(code, {}))
        row.update(codigo_clasificado=code,
                   nombre_estandar=entry.get("nombre_estandar", code),
                   categoria=entry.get("categoria", "sin_catalogo"))
        for amount_column in amount_columns:
            row.setdefault(amount_column, 0.0)
        row.setdefault("num_cuentas", 0)
        row["estado_presentacion"] = "Con cuentas" if row["num_cuentas"] else "Sin cuentas asignadas"
        if entry.get("clasificable") is False:
            row["estado_presentacion"] = "Calculado" if code in existing else "No disponible"
        rows.append(row)
    by_code = {r["codigo_clasificado"]: r for r in rows}
    detail_codes = [k for k, r in by_code.items() if r["categoria"] == "resultado"
                    and catalog.get(k, {}).get("clasificable") is not False
                    and catalog.get(k, {}).get("aditivo_resultado") is not False
                    and k not in CALCULATED]
    formulas = {
        "ER.03": ["ER.01", "ER.02"],
        "ER.06": ["ER.01", "ER.02", "ER.04", "ER.05"],
        "ER.08": ["ER.01", "ER.02", "ER.04", "ER.05", "ER.07"],
        "ER.19": [k for k in detail_codes if k != "ER.10"],
        "ER.11": detail_codes,
    }
    for amount_column in amount_columns:
        amounts = {k: float(r[amount_column]) for k, r in by_code.items()}
        income_available = (
            has_income_detail.get(amount_column, False)
            if isinstance(has_income_detail, dict) else bool(has_income_detail)
        )
        for code, dependencies in formulas.items():
            if code in by_code:
                by_code[code][amount_column] = (
                    sum(amounts.get(k, 0) for k in dependencies)
                    if income_available else None
                )
                by_code[code][f"estado_{amount_column}"] = (
                    "Calculado" if income_available else "No disponible"
                )
    for code in formulas:
        if code not in by_code:
            continue
        states = [by_code[code].get(f"estado_{column}") for column in amount_columns]
        by_code[code]["estado_presentacion"] = (
            "Calculado" if "Calculado" in states else "No disponible"
        )
    def order(row):
        category = row["categoria"]
        code = row["codigo_clasificado"]
        # Categorías nuevas de resultado se presentan antes del cierre.
        er_index = ER_ORDER.index(code) if code in ER_ORDER else ER_ORDER.index("ER.19") - 0.5
        return (CATEGORY_ORDER.index(category) if category in CATEGORY_ORDER else 99,
                er_index if category == "resultado" else 0, code)
    return pd.DataFrame(sorted(rows, key=order)), formulas


def add_report_sheets(workbook, complete, formulas, *, account_detail, start_row, unit,
                      meta, processed_at, source_name, pages, definitive, reasons,
                      tolerance=1000, period_columns=None):
    """Extiende el exportador operativo, con fórmulas auditables y datos separados."""
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.worksheet.table import Table, TableStyleInfo

    blue, grey = "1F4E79", "EAF0F6"
    numeric_format = '#,##0.00;[Red](#,##0.00);0.00'
    balance = workbook["Balance Normalizado"]
    period_columns = list(period_columns or [("Importe", "monto_total")])
    rows = complete.to_dict("records")
    index = {r["codigo_clasificado"]: start_row + i + 1 for i, r in enumerate(rows)}
    amount_positions = {
        amount_column: 3 + position
        for position, (_, amount_column) in enumerate(period_columns)
    }
    def ref(code, amount_column="monto_total"):
        column = get_column_letter(amount_positions[amount_column])
        return f"'Balance Normalizado'!{column}{index[code]}"
    for code, dependencies in formulas.items():
        if code in index:
            record = complete.loc[complete.codigo_clasificado == code].iloc[0]
            for _, amount_column in period_columns:
                if record.get(f"estado_{amount_column}") != "Calculado":
                    continue
                refs = [ref(k, amount_column) for k in dependencies if k in index]
                balance.cell(
                    index[code], amount_positions[amount_column],
                    "=SUM(" + ",".join(refs) + ")" if refs else "=0",
                )
    for i, record in enumerate(rows, start=start_row+1):
        group_column = 5 + len(period_columns)
        balance.cell(i, group_column, CATEGORY_LABELS.get(record["categoria"], record["categoria"]))
        balance.cell(i, group_column + 1, record["estado_presentacion"])
    balance.cell(start_row, 5 + len(period_columns), "Grupo")
    balance.cell(start_row, 6 + len(period_columns), "Estado")
    balance["E2"], balance["F2"] = "Fecha de proceso", processed_at
    balance["E3"], balance["F3"] = "Analista", None

    summary = workbook.create_sheet("Resumen")
    workbook.move_sheet(summary, offset=-len(workbook.sheetnames) + 1)
    metadata = [
        ("Informe de homologación", "DEFINITIVO" if definitive else "BORRADOR"),
        ("Empresa", getattr(meta, "razon_social", "") or ""),
        ("RUT", getattr(meta, "rut", "") or ""),
        ("Período", f'{getattr(meta, "periodo_desde", "") or ""} al {getattr(meta, "periodo_hasta", "") or ""}'),
        ("Moneda/unidad", unit), ("Fecha de proceso", processed_at),
        ("Analista", None), ("Documento fuente", source_name),
        ("Páginas analizadas", ", ".join(map(str, pages)) if pages else "Documento completo"),
        ("Criterio", "Cero = sin cuentas asignadas. No acredita que falte o no exista una cuenta en el original."),
    ]
    for entry in metadata:
        summary.append(entry)
    summary.append(["Control del balance", *[
        f"{label} ({unit})" for label, _ in period_columns
    ]])
    category_rows = {}
    for category in CATEGORY_ORDER[:-1]:
        row = summary.max_row + 1
        category_rows[category] = row
        formulas_periodo = []
        for _, amount_column in period_columns:
            refs = [ref(r["codigo_clasificado"], amount_column) for r in rows if r["categoria"] == category]
            formulas_periodo.append("=SUM(" + ",".join(refs) + ")" if refs else "=0")
        summary.append([CATEGORY_LABELS[category], *formulas_periodo])
    period_count = len(period_columns)
    def row_formulas(builder):
        return [builder(get_column_letter(2 + position)) for position in range(period_count)]
    summary.append(["TOTAL ACTIVOS", *row_formulas(
        lambda col: f"=SUM({col}{category_rows['activo_corriente']}:{col}{category_rows['activo_no_corriente']})"
    )])
    asset_row = summary.max_row
    summary.append(["TOTAL PASIVOS", *row_formulas(
        lambda col: f"=SUM({col}{category_rows['pasivo_corriente']}:{col}{category_rows['pasivo_no_corriente']})"
    )])
    liability_row = summary.max_row
    summary.append(["TOTAL PATRIMONIO", *row_formulas(
        lambda col: f"={col}{category_rows['patrimonio']}"
    )])
    equity_row = summary.max_row
    summary.append(["PASIVO + PATRIMONIO", *row_formulas(
        lambda col: f"=SUM({col}{liability_row}:{col}{equity_row})"
    )])
    right_row = summary.max_row
    summary.append(["Diferencia de cuadratura", *row_formulas(
        lambda col: f"={col}{asset_row}-{col}{right_row}"
    )])
    delta_row = summary.max_row
    summary.append(["Tolerancia de cuadratura", *([tolerance] * period_count)])
    tolerance_row = summary.max_row
    summary.append(["Cuadratura aritmética", *row_formulas(
        lambda col: f'=IF(ABS({col}{delta_row})<={col}{tolerance_row},"CUADRA DENTRO DE TOLERANCIA","NO CUADRA")'
    )])
    summary.append(["Estado de emisión", "DEFINITIVO" if definitive else "BORRADOR: REQUIERE REVISIÓN"])
    summary.append(["Advertencia", "La cuadratura no sustituye la revisión de clasificación y extracción. Consulte Control de emisión para los pendientes."])

    # Contrato rígido para integraciones: encabezado siempre en la fila 26 y
    # una fila por cuenta fuente, sin subtotales ni celdas combinadas.
    detail_headers = [
        "codigo_homologado", "nombre_homologado", "codigo_original",
        "nombre_original",
    ]
    for label, _ in period_columns:
        suffix = "" if len(period_columns) == 1 else f"_{label}"
        detail_headers.extend([f"valor_extraido{suffix}", f"valor_homologado{suffix}"])
    while summary.max_row < 25:
        summary.append([])
    summary.append(detail_headers)
    detail_columns = ["Código Estándar", "Nombre Estándar", "Cód. Original", "Nombre"]
    for label, _ in period_columns:
        if len(period_columns) == 1:
            detail_columns.extend(["Monto Extraído", "Monto Normalizado"])
        else:
            detail_columns.extend([f"Monto Extraído {label}", f"Monto Normalizado {label}"])
    detail_rows = account_detail.reindex(columns=detail_columns).copy()
    detail_rows = detail_rows.where(pd.notna(detail_rows), None)
    for values in detail_rows.itertuples(index=False, name=None):
        summary.append(list(values))
    last_detail_row = max(26, 26 + len(detail_rows))
    account_table = Table(
        displayName="CuentasClasificadas",
        ref=f"A26:{get_column_letter(len(detail_headers))}{last_detail_row}",
    )
    account_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    summary.add_table(account_table)

    income = workbook.create_sheet("Estado de Resultados")
    income.append(["Estado de resultados", unit, "Ingresos positivos, gastos negativos"])
    income.append(["Código", "Clasificación", *[
        f"{label} ({unit})" for label, _ in period_columns
    ], "Tipo"])
    for record in rows:
        if record["categoria"] != "resultado":
            continue
        code = record["codigo_clasificado"]
        values = [
            "=" + ref(code, amount_column)
            if record.get(f"estado_{amount_column}", record["estado_presentacion"]) != "No disponible"
            else None
            for _, amount_column in period_columns
        ]
        income.append([code, record["nombre_estandar"], *values, record["estado_presentacion"]])
    income.append(["", "No sumar los resultados calculados nuevamente al detalle."])
    income.append(["", "EBITDA: ventas + costo de ventas + gastos de administración + gastos de venta (importes con signo)."])
    # La validez del resultado se contrasta en Control de emisión, no se deduce
    # de la mera igualdad entre dos celdas calculadas desde la misma fuente.

    for sheet in (summary, income, balance):
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = "C3" if sheet is income else ("C11" if sheet is summary else f"C{start_row+1}")
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.print_options.horizontalCentered = True
        for row in sheet:
            for cell in row:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if isinstance(cell.value, (int, float)) or (cell.data_type == "f" and "IF(" not in cell.value):
                    cell.number_format = numeric_format
                    cell.alignment = Alignment(horizontal="right", vertical="center")
        for column in range(1, sheet.max_column+1):
            sheet.column_dimensions[get_column_letter(column)].width = 24
    summary.column_dimensions["A"].width = 33
    summary.column_dimensions["B"].width = 92
    income.column_dimensions["B"].width = 66
    income.column_dimensions[get_column_letter(3 + len(period_columns))].width = 25
    balance.column_dimensions["B"].width = 48
    balance.column_dimensions["F"].width = 40
    accounts_column = 3 + len(period_columns)
    for row in range(start_row + 1, start_row + len(rows) + 1):
        balance.cell(row, accounts_column).number_format = "0"
    for sheet, headers in ((summary, (1, 11)), (income, (1, 2)), (balance, (start_row,))):
        for header in headers:
            for cell in sheet[header]:
                cell.fill = PatternFill("solid", fgColor=blue)
                cell.font = Font(bold=True, color="FFFFFF")
        sheet.print_title_rows = f"1:{max(headers)}"
    for row in range(asset_row, delta_row+1):
        for cell in summary[row]:
            cell.fill = PatternFill("solid", fgColor=grey)
            cell.font = Font(bold=True)
            cell.border = Border(top=Side(style="thin", color=blue))
    for row in range(27, last_detail_row + 1):
        for column in range(5, 5 + 2 * period_count):
            summary.cell(row, column).number_format = numeric_format
            summary.cell(row, column).alignment = Alignment(horizontal="right")
    summary.column_dimensions["C"].width = 20
    summary.column_dimensions["D"].width = 44
    for column in range(5, 5 + 2 * period_count):
        summary.column_dimensions[get_column_letter(column)].width = 22
    for position in range(period_count):
        column = get_column_letter(2 + position)
        summary.conditional_formatting.add(
            f"{column}{delta_row}", FormulaRule(
                formula=[f"ABS({column}{delta_row})>{column}{tolerance_row}"],
                fill=PatternFill("solid", fgColor="FFC7CE")))
    for i, record in enumerate([r for r in rows if r["categoria"] == "resultado"], start=3):
        if record["codigo_clasificado"] in CALCULATED:
            for cell in income[i]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor=grey)
